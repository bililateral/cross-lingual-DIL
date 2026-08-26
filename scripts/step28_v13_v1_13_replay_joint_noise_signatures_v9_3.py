#!/usr/bin/env python3
"""Independently replay a V9.3 joint-noise signature from pinned raw sources.

This verifier intentionally does not import the producer.  It reconstructs the
aggregate from the workbook, manifest, and train-side allow-list and compares
the complete derived tables byte-for-byte at the object level.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
import csv
import hashlib
import json
from pathlib import Path
import re
import unicodedata
from typing import Any, Mapping

from openpyxl import load_workbook


VERSION = "2026-08-25-step28-v13-v1-13-independent-signature-replay-v9-3"
PAYLOAD_VERSION = "2026-08-25-step28-v13-v1-13-joint-noise-signatures-v9-3"
ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "market_item.xlsx"
WORKBOOK_SHA256 = "3625a226974827a0441ec87c54688f85ed0ff93a1c4c687532ef550ec1640187"
MANIFEST_PATH = ROOT / "reports" / "step2_content_item_manifest.csv"
MANIFEST_SHA256 = "89307835c82683abe53561a56a39fb28419cbf46671d34463701d73179e3869f"
ALLOWLIST_PATH = (
    ROOT
    / "reports"
    / "step28_synthetic_chinese_dataset"
    / "v13_dev_smoke_v1_20260727"
    / "reference"
    / "style_source_train_sellers.csv"
)
ALLOWLIST_SHA256 = "b57cbfc3908ae3f36983c7c9e6d4dd974b59222f6617d214cec090524c3eb6de"
EXPECTED_SELLERS = 676
EXPECTED_ROWS = 3439
SLOT_COUNT = 28
WHITESPACE_RE = re.compile(r"\s+")


class IndependentReplayError(ValueError):
    """Raised when raw-source replay does not exactly reproduce the payload."""


def canonical_json_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: object) -> str:
    return hashlib.sha256(canonical_json_bytes(value)).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clean(value: object) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(
        " ", unicodedata.normalize("NFKC", str(value))
    ).strip()


def _read_allowlist() -> list[str]:
    with ALLOWLIST_PATH.open("r", encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream)
        if list(reader.fieldnames or []) != ["seller_uid"]:
            raise IndependentReplayError("Allow-list schema drift")
        sellers = [row["seller_uid"].strip() for row in reader]
    if (
        len(sellers) != EXPECTED_SELLERS
        or sellers != sorted(set(sellers), key=lambda value: value.encode("utf-8"))
    ):
        raise IndependentReplayError("Allow-list order/cardinality drift")
    return sellers


def _read_manifest(allowed: set[str]) -> dict[int, str]:
    selected: dict[int, str] = {}
    required = {
        "seller_uid",
        "source_dataset",
        "source_row_number",
        "data_bucket",
        "eligibility_status",
    }
    with MANIFEST_PATH.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        if not required <= set(reader.fieldnames or []):
            raise IndependentReplayError("Manifest schema drift")
        for row in reader:
            if not (
                row["source_dataset"] == "market_item.xlsx"
                and row["data_bucket"] == "zh_target_strict"
                and row["eligibility_status"] == "target_eval_candidate"
                and row["seller_uid"] in allowed
            ):
                continue
            source_row = int(row["source_row_number"])
            if source_row < 2 or source_row in selected:
                raise IndependentReplayError("Manifest selected-row drift")
            selected[source_row] = row["seller_uid"]
    if len(selected) != EXPECTED_ROWS or set(selected.values()) != allowed:
        raise IndependentReplayError("Manifest coverage/cardinality drift")
    return selected


def _read_presence(selected: Mapping[int, str]) -> dict[str, list[tuple[int, bool, bool]]]:
    observed: dict[str, list[tuple[int, bool, bool]]] = defaultdict(list)
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            seller = selected.get(row_number)
            if seller is None:
                continue
            if len(row) != 7:
                raise IndependentReplayError("Workbook schema drift")
            observed[seller].append(
                (row_number, bool(_clean(row[2])), bool(_clean(row[3])))
            )
    finally:
        workbook.close()
    if sum(map(len, observed.values())) != EXPECTED_ROWS:
        raise IndependentReplayError("Workbook selected-row count drift")
    return dict(observed)


def _signature(rows: list[tuple[int, bool, bool]]) -> dict[str, Any]:
    if rows != sorted(rows) or not rows:
        raise IndependentReplayError("Seller row order drift")
    states = [(title, description) for _row, title, description in rows]
    if len(states) == 1:
        states = [states[0], states[0]]
    states = states[:8]
    return {
        "item_count": len(states),
        "title_present_mask": "".join("1" if title else "0" for title, _ in states),
        "description_present_mask": "".join(
            "1" if description else "0" for _, description in states
        ),
        "joint_empty_mask": "".join(
            "1" if not title and not description else "0"
            for title, description in states
        ),
    }


def _integerize(
    counts: Counter[bytes], signatures: Mapping[bytes, dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    floor_total = 0
    for key in sorted(counts):
        count = int(counts[key])
        floor, remainder = divmod(count * SLOT_COUNT, EXPECTED_SELLERS)
        floor_total += floor
        rows.append(
            {
                "signature": signatures[key],
                "seller_count": count,
                "probability_numerator": count,
                "probability_denominator": EXPECTED_SELLERS,
                "floor_slot_count": floor,
                "remainder_numerator": remainder,
                "allocated_slot_count": floor,
            }
        )
    ranking = sorted(
        range(len(rows)),
        key=lambda index: (
            -int(rows[index]["remainder_numerator"]),
            canonical_json_bytes(rows[index]["signature"]),
        ),
    )
    for index in ranking[: SLOT_COUNT - floor_total]:
        rows[index]["allocated_slot_count"] += 1
    slots = [
        dict(row["signature"])
        for row in rows
        for _copy in range(int(row["allocated_slot_count"]))
    ]
    if len(slots) != SLOT_COUNT:
        raise IndependentReplayError("Integerized slot count drift")
    return rows, slots


def replay(payload: Mapping[str, Any]) -> dict[str, Any]:
    if payload.get("version") != PAYLOAD_VERSION:
        raise IndependentReplayError("Payload version drift")
    projection = dict(payload)
    supplied_self = projection.pop("canonical_self_sha256", None)
    if supplied_self != canonical_sha256(projection):
        raise IndependentReplayError("Payload canonical self-hash drift")
    expected_pins = {
        WORKBOOK_PATH: WORKBOOK_SHA256,
        MANIFEST_PATH: MANIFEST_SHA256,
        ALLOWLIST_PATH: ALLOWLIST_SHA256,
    }
    for path, expected in expected_pins.items():
        if not path.is_file() or sha256_file(path) != expected:
            raise IndependentReplayError(f"Pinned source bytes drift: {path}")

    sellers = _read_allowlist()
    selected = _read_manifest(set(sellers))
    presence = _read_presence(selected)
    if set(presence) != set(sellers):
        raise IndependentReplayError("Workbook seller coverage drift")
    counts: Counter[bytes] = Counter()
    signatures: dict[bytes, dict[str, Any]] = {}
    raw_histogram: Counter[int] = Counter()
    for seller in sellers:
        raw_histogram[len(presence[seller])] += 1
        signature = _signature(presence[seller])
        key = canonical_json_bytes(signature)
        signatures.setdefault(key, signature)
        counts[key] += 1
    frequency, slots = _integerize(counts, signatures)
    eligibility = []
    for slot, signature in enumerate(slots):
        title = signature["title_present_mask"]
        description = signature["description_present_mask"]
        eligibility.append(
            {
                "noise_slot": slot,
                "title_present_logical_item_ordinals": [
                    index for index, bit in enumerate(title) if bit == "1"
                ],
                "title_and_description_present_logical_item_ordinals": [
                    index
                    for index, bits in enumerate(zip(title, description))
                    if bits == ("1", "1")
                ],
            }
        )
    expected = {
        "raw_item_count_histogram": {
            str(key): value for key, value in sorted(raw_histogram.items())
        },
        "observed_signature_count": len(frequency),
        "signature_frequency_and_integerization": frequency,
        "noise_slot_multiset": [
            {"noise_slot": index, "signature": signature}
            for index, signature in enumerate(slots)
        ],
        "slot_eligibility": eligibility,
    }
    for field, value in expected.items():
        if payload.get(field) != value:
            raise IndependentReplayError(f"Raw-source replay drift: {field}")
    return {
        "version": VERSION,
        "payload_canonical_self_sha256": supplied_self,
        "replayed_derived_tables_sha256": canonical_sha256(expected),
        "seller_count": len(sellers),
        "selected_item_row_count": len(selected),
        "status": "PASS_INDEPENDENT_RAW_SOURCE_REPLAY_ONLY_NOT_METHOD_OR_TRAINING_QUALIFIED",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("payload", type=Path)
    args = parser.parse_args()
    value = json.loads(args.payload.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise IndependentReplayError("Payload root must be a JSON object")
    print(canonical_json_bytes(replay(value)).decode("utf-8"))


if __name__ == "__main__":
    main()
