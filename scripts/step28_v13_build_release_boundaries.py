#!/usr/bin/env python3
"""Build label-free anti-copy and identity-value deny boundaries for Step 28-v13."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import step28_v13_common as common


WHITESPACE_RE = re.compile(r"\s+")


def normalized_identity_value(value: object) -> str:
    output = unicodedata.normalize("NFC", str(value).strip().casefold())
    if not output:
        raise common.ContractError("Identity deny source contains an empty projected value")
    return output


def identity_value_hash(value: object) -> str:
    return common.sha256_bytes(normalized_identity_value(value).encode("utf-8"))


def normalized_real_text(value: object) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(
        " ",
        unicodedata.normalize("NFKC", str(value)),
    ).strip().casefold()


def _source_path_and_pin(spec: dict[str, Any]) -> Path:
    return common.verify_file_pin(spec, label=f"identity deny source {spec['path']}")


def _real_value_hashes(path: Path) -> tuple[set[str], int, int]:
    output: set[str] = set()
    row_count = 0
    value_count = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if "normalized_value" not in (reader.fieldnames or []):
            raise common.ContractError(f"Real deny source lacks normalized_value: {path}")
        for row in reader:
            row_count += 1
            value = str(row["normalized_value"]).strip()
            if not value:
                continue
            output.add(identity_value_hash(value))
            value_count += 1
    return output, row_count, value_count


def _prior_value_hashes(path: Path) -> tuple[set[str], int, int]:
    output: set[str] = set()
    row_count = 0
    value_count = 0
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                raise common.ContractError(
                    f"Blank JSONL line in prior identity source: {path}:{line_number}"
                )
            row = json.loads(line)
            values = row.get("identifier_values")
            if not isinstance(values, list):
                raise common.ContractError(
                    f"Prior identity source lacks identifier_values list: {path}:{line_number}"
                )
            row_count += 1
            for value in values:
                if not isinstance(value, str) or not value.strip():
                    raise common.ContractError(
                        f"Invalid prior identity value: {path}:{line_number}"
                    )
                output.add(identity_value_hash(value))
                value_count += 1
    return output, row_count, value_count


def build_identity_deny_artifact(policy: dict[str, Any]) -> dict[str, Any]:
    salt = policy["identity_design"]["identity_value_generation"]["salt_selection"]
    records: list[dict[str, Any]] = []
    complete: set[str] = set()
    sources = sorted(
        salt["deny_hash_artifact"]["exact_sources"],
        key=lambda value: (
            str(value["source_class"]).encode("utf-8"),
            str(value["path"]).encode("utf-8"),
        ),
    )
    for spec in sources:
        path = _source_path_and_pin(spec)
        if spec["source_class"] == "real":
            hashes, row_count, value_count = _real_value_hashes(path)
        elif spec["source_class"] == "prior_synthetic":
            hashes, row_count, value_count = _prior_value_hashes(path)
        else:
            raise common.ContractError(
                f"Unknown identity deny source class: {spec['source_class']}"
            )
        complete.update(hashes)
        records.append(
            {
                "path": str(spec["path"]),
                "sha256": str(spec["sha256"]),
                "source_class": str(spec["source_class"]),
                "row_count": row_count,
                "projected_nonempty_value_count": value_count,
                "source_unique_value_hash_count": len(hashes),
            }
        )
    artifact: dict[str, Any] = {
        "version": "2026-07-27-step28-v13-identity-value-deny-v1",
        "status": "PASS_BOUNDARY_ONLY",
        "scientific_metrics_produced": False,
        "normalization": "SHA256(UTF-8(NFC(casefold(strip(value)))))",
        "source_records": records,
        "value_hashes": common.utf8_sort(complete),
        "unique_value_hash_count": len(complete),
    }
    artifact["canonical_self_hash"] = common.canonical_sha256(artifact)
    return artifact


def _digest_lines(lines: list[bytes]) -> str:
    digest = hashlib.sha256()
    for line in lines:
        digest.update(line)
        digest.update(b"\n")
    return digest.hexdigest()


def build_real_text_boundary(policy: dict[str, Any]) -> dict[str, Any]:
    spec = policy["frozen_inputs"]["raw_chinese_items"]
    path = common.verify_file_pin(spec, label="raw Chinese anti-copy source")
    counts: Counter[str] = Counter()
    field_count = 0
    row_count = 0
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if len(row) != 7:
                raise common.ContractError(
                    f"Anti-copy workbook row is not seven cells: {row_number}"
                )
            row_count += 1
            for value in (row[2], row[3]):
                normalized = normalized_real_text(value)
                if not normalized:
                    continue
                counts[normalized] += 1
                field_count += 1
    finally:
        workbook.close()
    if not counts or not field_count:
        raise common.ContractError("Real anti-copy corpus is empty")

    ordered = sorted(counts, key=lambda value: value.encode("utf-8"))
    field_hash_lines = [
        common.canonical_json_bytes(
            {"field_sha256": common.sha256_bytes(value.encode("utf-8"))}
        )
        for value in ordered
    ]
    multiplicity_lines = [
        common.canonical_json_bytes({"multiplicity": counts[value]})
        for value in ordered
    ]
    boundary = policy["text_generation"]["real_anticopy_boundary"]
    artifact: dict[str, Any] = {
        "version": "2026-07-27-step28-v13-real-text-anticopy-boundary-v1",
        "status": "PASS_BOUNDARY_ONLY",
        "scientific_metrics_produced": False,
        "source_path": str(spec["path"]),
        "source_sha256": str(spec["sha256"]),
        "workbook_data_row_count": row_count,
        "normalization": str(boundary["normalization"]),
        "nonempty_field_count": field_count,
        "unique_field_count": len(ordered),
        "ordered_normalized_field_sha256_digest": _digest_lines(field_hash_lines),
        "ordered_multiplicity_sha256_digest": _digest_lines(multiplicity_lines),
    }
    artifact["canonical_self_hash"] = common.canonical_sha256(artifact)
    return artifact


def run(policy_path: Path) -> tuple[Path, Path]:
    policy = common.load_policy(policy_path, mode="development_smoke")
    deny = build_identity_deny_artifact(policy)
    anti_copy = build_real_text_boundary(policy)

    deny_path = common.repo_path(
        policy["identity_design"]["identity_value_generation"]["salt_selection"][
            "deny_hash_artifact"
        ]["path"]
    )
    anti_copy_path = common.repo_path(
        policy["text_generation"]["real_anticopy_boundary"]["manifest"]["path"]
    )
    common.write_json(deny_path, deny)
    common.write_json(anti_copy_path, anti_copy)
    return deny_path, anti_copy_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    common.add_policy_argument(parser)
    parser.add_argument("--validate-config-only", action="store_true")
    args = parser.parse_args()
    policy = common.load_policy(args.policy, mode="development_smoke")
    if args.validate_config_only:
        for spec in policy["identity_design"]["identity_value_generation"][
            "salt_selection"
        ]["deny_hash_artifact"]["exact_sources"]:
            _source_path_and_pin(spec)
        common.verify_file_pin(
            policy["frozen_inputs"]["raw_chinese_items"],
            label="raw Chinese anti-copy source",
        )
        print("Step28-v13 release-boundary configuration is valid")
        return
    deny_path, anti_copy_path = run(args.policy)
    print(
        "Step28-v13 release boundaries materialized:",
        deny_path.relative_to(common.ROOT).as_posix(),
        anti_copy_path.relative_to(common.ROOT).as_posix(),
    )


if __name__ == "__main__":
    main()
