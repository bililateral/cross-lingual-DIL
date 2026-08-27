#!/usr/bin/env python3
"""Rebuild the 28 V9.4 training-ready nuisance signatures from source data."""

from __future__ import annotations

from collections import Counter, defaultdict
from collections.abc import Mapping
import csv
from dataclasses import dataclass
import hashlib
import json
from pathlib import Path
import re
from types import MappingProxyType
from typing import Any
import unicodedata

from openpyxl import load_workbook


VERSION = "2026-08-27-step28-v13-v1-13-joint-noise-signatures-v9-4"
ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "market_item.xlsx"
WORKBOOK_SHA256 = "3625a226974827a0441ec87c54688f85ed0ff93a1c4c687532ef550ec1640187"
MANIFEST_PATH = ROOT / "reports" / "step2_content_item_manifest.csv"
MANIFEST_SHA256 = "89307835c82683abe53561a56a39fb28419cbf46671d34463701d73179e3869f"
ALLOWLIST_PATH = (
    ROOT
    / "reports"
    / "step28_synthetic_chinese_dataset"
    / "v13_dev_smoke_v1_20260727"
    / "reference"
    / "style_source_train_sellers.csv"
)
ALLOWLIST_SHA256 = "b57cbfc3908ae3f36983c7c9e6d4dd974b59222f6617d214cec090524c3eb6de"
SOURCE_SELLER_COUNT = 676
SOURCE_SELECTED_ITEM_ROWS = 3439
ELIGIBLE_SELLER_COUNT = 648
EXCLUDED_SELLER_COUNT = 28
ELIGIBLE_ITEM_ROWS = 3354
EXCLUDED_ITEM_ROWS = 85
SLOT_COUNT = 28
SOURCE_DATASET = "market_item.xlsx"
DATA_BUCKET = "zh_target_strict"
ELIGIBILITY_STATUS = "target_eval_candidate"
SIGNATURE_FIELDS = (
    "noise_slot",
    "item_count",
    "title_present_mask",
    "description_present_mask",
    "joint_empty_mask",
)
EXPECTED_SIGNATURE_ROWS_SHA256 = (
    "9c8a5d27adf0ac4e5f65d241f1e7c83ee80f1c4caa03eae9e7a0b7f0fd74e763"
)
EXPECTED_SIGNATURE_SET_COMMITMENT_SHA256 = (
    "15da977c89f620383f14805566ca54a1901a8323416310344609db8003f0eb39"
)
WHITESPACE_RE = re.compile(r"\s+")


class JointNoiseSignaturesV94Error(ValueError):
    """Raised when the label-free nuisance-signature derivation drifts."""


@dataclass(frozen=True)
class NoiseSignatureSet:
    rows: tuple[Mapping[str, Any], ...]
    commitment: Mapping[str, Any]


def _canonical_sha256(value: Any) -> str:
    return hashlib.sha256(
        json.dumps(
            value,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _verify_pin(path: Path, expected: str, name: str) -> None:
    if not path.is_file() or _sha256_file(path) != expected:
        raise JointNoiseSignaturesV94Error(f"Pinned {name} drift")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(
        " ", unicodedata.normalize("NFKC", str(value))
    ).strip()


def _signature(rows: list[tuple[int, bool, bool]]) -> dict[str, Any]:
    if not rows:
        raise JointNoiseSignaturesV94Error("Empty seller signature source")
    states = [(title, description) for _, title, description in rows]
    if len(states) == 1:
        states.append(states[0])
    states = states[:8]
    title = "".join("1" if value[0] else "0" for value in states)
    description = "".join("1" if value[1] else "0" for value in states)
    return {
        "item_count": len(states),
        "title_present_mask": title,
        "description_present_mask": description,
        "joint_empty_mask": "".join(
            "1" if not title_value and not description_value else "0"
            for title_value, description_value in states
        ),
    }


def _signature_key(signature: Mapping[str, Any]) -> bytes:
    return json.dumps(
        dict(signature),
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")


def _largest_remainder(
    counts: Counter[bytes],
    signatures: Mapping[bytes, dict[str, Any]],
) -> list[dict[str, Any]]:
    if sum(counts.values()) != ELIGIBLE_SELLER_COUNT:
        raise JointNoiseSignaturesV94Error("Eligible seller total drift")
    allocations: list[list[Any]] = []
    allocated = 0
    for key in sorted(counts):
        floor_slots, remainder = divmod(
            int(counts[key]) * SLOT_COUNT,
            ELIGIBLE_SELLER_COUNT,
        )
        allocated += floor_slots
        allocations.append([key, floor_slots, remainder])
    remaining = SLOT_COUNT - allocated
    ranking = sorted(
        range(len(allocations)),
        key=lambda index: (-int(allocations[index][2]), allocations[index][0]),
    )
    for index in ranking[:remaining]:
        allocations[index][1] += 1
    slots: list[dict[str, Any]] = []
    for key, slot_count, _ in allocations:
        slots.extend(dict(signatures[key]) for _ in range(int(slot_count)))
    if len(slots) != SLOT_COUNT:
        raise JointNoiseSignaturesV94Error("Noise slot allocation drift")
    return slots


def _read_sources() -> tuple[list[str], dict[str, list[tuple[int, bool, bool]]]]:
    with ALLOWLIST_PATH.open("r", encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream)
        if list(reader.fieldnames or ()) != ["seller_uid"]:
            raise JointNoiseSignaturesV94Error("Seller allow-list schema drift")
        sellers = [row["seller_uid"].strip() for row in reader]
    if (
        len(sellers) != SOURCE_SELLER_COUNT
        or sellers != sorted(set(sellers), key=lambda value: value.encode("utf-8"))
    ):
        raise JointNoiseSignaturesV94Error("Seller allow-list closure drift")
    selected: dict[int, str] = {}
    selected_counts: Counter[str] = Counter()
    with MANIFEST_PATH.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        required = {
            "seller_uid",
            "source_dataset",
            "source_row_number",
            "data_bucket",
            "eligibility_status",
        }
        if not required <= set(reader.fieldnames or ()):
            raise JointNoiseSignaturesV94Error("Item manifest schema drift")
        for row in reader:
            if (
                row["source_dataset"] != SOURCE_DATASET
                or row["data_bucket"] != DATA_BUCKET
                or row["eligibility_status"] != ELIGIBILITY_STATUS
                or row["seller_uid"] not in sellers
            ):
                continue
            source_row = int(row["source_row_number"])
            if source_row < 2 or source_row in selected:
                raise JointNoiseSignaturesV94Error("Selected item row drift")
            selected[source_row] = row["seller_uid"]
            selected_counts[row["seller_uid"]] += 1
    if (
        len(selected) != SOURCE_SELECTED_ITEM_ROWS
        or set(selected_counts) != set(sellers)
    ):
        raise JointNoiseSignaturesV94Error("Selected item boundary drift")
    presence: dict[str, list[tuple[int, bool, bool]]] = defaultdict(list)
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            seller_uid = selected.get(row_number)
            if seller_uid is None:
                continue
            if len(row) != 7:
                raise JointNoiseSignaturesV94Error("Workbook column count drift")
            presence[seller_uid].append((
                row_number,
                bool(_clean_text(row[2])),
                bool(_clean_text(row[3])),
            ))
    finally:
        workbook.close()
    if (
        set(presence) != set(sellers)
        or sum(len(value) for value in presence.values())
        != SOURCE_SELECTED_ITEM_ROWS
    ):
        raise JointNoiseSignaturesV94Error("Workbook selection replay drift")
    for seller in sellers:
        if presence[seller] != sorted(presence[seller]):
            raise JointNoiseSignaturesV94Error("Seller item order drift")
    return sellers, dict(presence)


def build_noise_signatures() -> NoiseSignatureSet:
    for path, expected, name in (
        (WORKBOOK_PATH, WORKBOOK_SHA256, "Chinese item workbook"),
        (MANIFEST_PATH, MANIFEST_SHA256, "item manifest"),
        (ALLOWLIST_PATH, ALLOWLIST_SHA256, "training seller allow-list"),
    ):
        _verify_pin(path, expected, name)
    sellers, presence = _read_sources()
    counts: Counter[bytes] = Counter()
    signatures: dict[bytes, dict[str, Any]] = {}
    eligible_sellers = 0
    eligible_items = 0
    for seller in sellers:
        signature = _signature(presence[seller])
        if (
            signature["title_present_mask"].count("1") < 1
            or signature["description_present_mask"].count("1") < 2
        ):
            continue
        eligible_sellers += 1
        eligible_items += len(presence[seller])
        key = _signature_key(signature)
        signatures.setdefault(key, signature)
        counts[key] += 1
    if (
        eligible_sellers != ELIGIBLE_SELLER_COUNT
        or len(sellers) - eligible_sellers != EXCLUDED_SELLER_COUNT
        or eligible_items != ELIGIBLE_ITEM_ROWS
        or SOURCE_SELECTED_ITEM_ROWS - eligible_items != EXCLUDED_ITEM_ROWS
    ):
        raise JointNoiseSignaturesV94Error("Training-ready source filter drift")
    slots = _largest_remainder(counts, signatures)
    rows = tuple(
        MappingProxyType({"noise_slot": noise_slot, **signature})
        for noise_slot, signature in enumerate(slots)
    )
    payload = [dict(row) for row in rows]
    source_pins = (
        ("market_item.xlsx", WORKBOOK_SHA256),
        ("reports/step2_content_item_manifest.csv", MANIFEST_SHA256),
        (
            "reports/step28_synthetic_chinese_dataset/"
            "v13_dev_smoke_v1_20260727/reference/"
            "style_source_train_sellers.csv",
            ALLOWLIST_SHA256,
        ),
    )
    commitment_payload = {
        "version": VERSION,
        "source_seller_count": SOURCE_SELLER_COUNT,
        "eligible_seller_count": ELIGIBLE_SELLER_COUNT,
        "excluded_seller_count": EXCLUDED_SELLER_COUNT,
        "source_item_row_count": SOURCE_SELECTED_ITEM_ROWS,
        "eligible_item_row_count": ELIGIBLE_ITEM_ROWS,
        "excluded_item_row_count": EXCLUDED_ITEM_ROWS,
        "source_pins": source_pins,
        "signature_rows_sha256": _canonical_sha256(payload),
    }
    commitment_payload["signature_set_commitment_sha256"] = _canonical_sha256(
        commitment_payload
    )
    result = NoiseSignatureSet(
        rows=rows,
        commitment=MappingProxyType(commitment_payload),
    )
    verify_noise_signatures(result)
    return result


def verify_noise_signatures(signatures: NoiseSignatureSet) -> None:
    expected_fields = (
        "version",
        "source_seller_count",
        "eligible_seller_count",
        "excluded_seller_count",
        "source_item_row_count",
        "eligible_item_row_count",
        "excluded_item_row_count",
        "source_pins",
        "signature_rows_sha256",
        "signature_set_commitment_sha256",
    )
    if (
        type(signatures) is not NoiseSignatureSet
        or type(signatures.commitment) is not MappingProxyType
        or tuple(signatures.commitment) != expected_fields
        or len(signatures.rows) != SLOT_COUNT
        or signatures.commitment["version"] != VERSION
        or signatures.commitment["source_seller_count"] != SOURCE_SELLER_COUNT
        or signatures.commitment["eligible_seller_count"] != ELIGIBLE_SELLER_COUNT
        or signatures.commitment["excluded_seller_count"] != EXCLUDED_SELLER_COUNT
        or signatures.commitment["source_item_row_count"]
        != SOURCE_SELECTED_ITEM_ROWS
        or signatures.commitment["eligible_item_row_count"] != ELIGIBLE_ITEM_ROWS
        or signatures.commitment["excluded_item_row_count"] != EXCLUDED_ITEM_ROWS
        or signatures.commitment["signature_rows_sha256"]
        != EXPECTED_SIGNATURE_ROWS_SHA256
        or signatures.commitment["signature_set_commitment_sha256"]
        != EXPECTED_SIGNATURE_SET_COMMITMENT_SHA256
    ):
        raise JointNoiseSignaturesV94Error("Noise signature capability drift")
    payload: list[dict[str, Any]] = []
    for noise_slot, row in enumerate(signatures.rows):
        if (
            type(row) is not MappingProxyType
            or tuple(row) != SIGNATURE_FIELDS
            or row["noise_slot"] != noise_slot
            or type(row["item_count"]) is not int
            or not 2 <= row["item_count"] <= 8
            or any(
                type(row[field]) is not str
                or len(row[field]) != row["item_count"]
                or not set(row[field]) <= {"0", "1"}
                for field in SIGNATURE_FIELDS[2:]
            )
            or row["title_present_mask"].count("1") < 1
            or row["description_present_mask"].count("1") < 2
            or any(
                row["joint_empty_mask"][index]
                != (
                    "1"
                    if row["title_present_mask"][index] == "0"
                    and row["description_present_mask"][index] == "0"
                    else "0"
                )
                for index in range(row["item_count"])
            )
        ):
            raise JointNoiseSignaturesV94Error("Noise signature row drift")
        payload.append(dict(row))
    source_pins = (
        ("market_item.xlsx", WORKBOOK_SHA256),
        ("reports/step2_content_item_manifest.csv", MANIFEST_SHA256),
        (
            "reports/step28_synthetic_chinese_dataset/"
            "v13_dev_smoke_v1_20260727/reference/"
            "style_source_train_sellers.csv",
            ALLOWLIST_SHA256,
        ),
    )
    expected_without_self = {
        key: signatures.commitment[key] for key in expected_fields[:-1]
    }
    if (
        signatures.commitment["source_pins"] != source_pins
        or signatures.commitment["signature_rows_sha256"]
        != _canonical_sha256(payload)
        or signatures.commitment["signature_set_commitment_sha256"]
        != _canonical_sha256(expected_without_self)
    ):
        raise JointNoiseSignaturesV94Error("Noise signature commitment drift")


def signature_dicts(signatures: NoiseSignatureSet) -> list[dict[str, Any]]:
    verify_noise_signatures(signatures)
    return [dict(row) for row in signatures.rows]


def contract_payload() -> dict[str, Any]:
    return {
        "version": VERSION,
        "source_role": "label_free_real_chinese_training_side",
        "statistical_unit": "seller_equal_weight",
        "slot_count": SLOT_COUNT,
        "minimum_nonempty_title_count": 1,
        "minimum_nonempty_description_count": 2,
        "reads_pair_truth": False,
        "reads_model_output": False,
    }
