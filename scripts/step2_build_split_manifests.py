from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from html import unescape
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
POLICY_PATH = ROOT / "schema" / "step2_split_policy.json"
HTML_RIPS_DIR = ROOT / "html-rips"
TARGET_MARKETS = {"中文暗网交易市场", "茶马古道"}
CJK_RE = re.compile(r"[\u4e00-\u9fff]")
WS_RE = re.compile(r"\s+")
HEX_RE = re.compile(r"[^0-9A-Fa-f]+")
HTML_TAG_RE = re.compile(r"<[^>]+>")
HTML_VENDOR_NAME_RE = re.compile(r'<div class="h1"\s*>\s*([^<]+?)\s*<small', re.I)
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,24}")
TELEGRAM_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:telegram|tg|电报)(?![A-Za-z0-9])(?:\s*(?:id|account|acct|handle|账号|号))?[\s:：=@._-]{0,6}([A-Za-z0-9_]{5,32})\b",
    re.I,
)
WICKR_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:wickr(?:\s*me|\s*id)?|威克)(?![A-Za-z0-9])[\s:：=@._-]{0,6}([A-Za-z0-9_.-]{3,32})\b",
    re.I,
)
WECHAT_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:微信|vx|wechat)(?![A-Za-z0-9])(?:\s*(?:id|号|账号))?[\s:：=@._-]{0,6}([A-Za-z][A-Za-z0-9_-]{4,32})\b",
    re.I,
)
QQ_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:qq)(?![A-Za-z0-9])(?:\s*(?:id|号|账号))?[\s:：=@._-]{0,6}([1-9]\d{4,11})",
    re.I,
)
PHONE_CONTEXT_RE = re.compile(
    r"(?:whatsapp|phone|tel|电话|手机|联系|联系方式)[^0-9A-Za-z+]{0,8}([+]?\d[\d\-\s()]{7,}\d)",
    re.I,
)
PGP_BLOCK_RE = re.compile(r"-----BEGIN PGP PUBLIC KEY BLOCK-----(.*?)-----END PGP PUBLIC KEY BLOCK-----", re.I | re.S)
CONTACT_NOISE_STOPWORDS = {
    "account",
    "address",
    "app",
    "contact",
    "email",
    "group",
    "mail",
    "message",
    "messages",
    "messenger",
    "private",
    "support",
    "telegram",
    "wechat",
    "whatsapp",
    "wickr",
}
HTML_HEADER_SPLIT_TOKENS = (
    'panel-title text-primary strong">Listings',
    "<!--end of grams reviews-->",
    "<h2>Recent Market Reviews</h2>",
)
VENDOR_SQL_RE = re.compile(
    r"^\((\d+), '((?:[^'\\]|\\.)*)', (\d+), (\d+), (\d+), '(.*?)', "
    r"'((?:[^'\\]|\\.)*)', (\d+), (\d+), (\d+), (\d+)\),?$"
)
PGP_SQL_RE = re.compile(
    r"^\((\d+), '((?:[^']|'')*)', '((?:[^']|'')*)', '((?:[^']|'')*)', 0x[0-9A-Fa-f]+, "
    r"'((?:[^']|'')*)', '((?:[^']|'')*)', (\d+), ([0-9.]+)\),?;?$"
)


def load_policy() -> dict:
    with POLICY_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def normalize_alias(raw: object) -> str:
    if raw is None:
        return ""
    return str(raw).strip().casefold()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return WS_RE.sub(" ", str(value)).strip()


def sql_unescape(raw: str) -> str:
    return raw.replace("''", "'")


def normalize_fingerprint(raw: object) -> str:
    value = HEX_RE.sub("", str(raw or "")).upper()
    if not value:
        return ""
    return value


def normalize_fingerprint_short(raw: object) -> str:
    value = normalize_fingerprint(raw)
    if not value:
        return ""
    if len(value) > 16:
        value = value[-16:]
    return f"0x{value}"


def parse_vendor_ids(raw: object) -> set[str]:
    return {token for token in re.findall(r"\d+", str(raw or ""))}


def normalize_contact_value(contact_type: str, value: str) -> str:
    raw = clean_text(value).strip("[](){}<>,;:'\"")
    if not raw:
        return ""

    if contact_type == "email":
        matches = re.findall(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,24}", raw.lower())
        valid_matches = []
        for match in matches:
            local_part = match.split("@", 1)[0]
            if local_part.startswith(".") or local_part.endswith(".") or ".." in local_part:
                continue
            valid_matches.append(match)
        unique_matches = sorted(set(valid_matches))
        if len(unique_matches) == 1:
            return unique_matches[0]
        return ""

    if contact_type in {"telegram", "wickr", "wechat"}:
        raw = raw.lower().lstrip("@")
        raw = re.sub(r"[^a-z0-9_.\-]", "", raw)
        raw = raw.strip("._-")
        if len(raw) < 5 or len(raw) > 32:
            return ""
        if raw in CONTACT_NOISE_STOPWORDS or raw.isdigit():
            return ""
        if not re.search(r"[a-z]", raw):
            return ""
        return raw

    if contact_type == "qq":
        digits = re.sub(r"\D", "", raw)
        if re.fullmatch(r"[1-9]\d{4,11}", digits or ""):
            return digits
        return ""

    if contact_type == "phone":
        digits = re.sub(r"\D", "", raw)
        if 7 <= len(digits) <= 15 and len(set(digits)) >= 3:
            return digits
        return ""

    return ""


def extract_contacts(*fields: object) -> dict[str, set[str]]:
    text = "\n".join(str(field or "") for field in fields if field is not None)
    text_without_pgp = PGP_BLOCK_RE.sub(" ", text)
    contact_map = {
        "email": {normalize_contact_value("email", match) for match in EMAIL_RE.findall(text)},
        "telegram": {normalize_contact_value("telegram", match) for match in TELEGRAM_RE.findall(text_without_pgp)},
        "wickr": {normalize_contact_value("wickr", match) for match in WICKR_RE.findall(text_without_pgp)},
        "wechat": {normalize_contact_value("wechat", match) for match in WECHAT_RE.findall(text_without_pgp)},
        "qq": {normalize_contact_value("qq", match) for match in QQ_RE.findall(text_without_pgp)},
        "phone": {normalize_contact_value("phone", match) for match in PHONE_CONTEXT_RE.findall(text_without_pgp)},
    }
    return {
        contact_type: {value for value in values if value}
        for contact_type, values in contact_map.items()
    }


def html_header_visible_text(raw_html: str) -> str:
    header_html = raw_html
    for token in HTML_HEADER_SPLIT_TOKENS:
        if token in header_html:
            header_html = header_html.split(token, 1)[0]
            break
    header_html = re.sub(r"(?i)<br\s*/?>", "\n", header_html)
    header_html = re.sub(r"(?i)</(p|div|li|td|tr|pre|h1|h2|h3|h4|ul|ol)>", "\n", header_html)
    header_html = HTML_TAG_RE.sub(" ", header_html)
    return clean_text(unescape(header_html))


def extract_html_vendor_alias(raw_html: str) -> str:
    match = HTML_VENDOR_NAME_RE.search(raw_html)
    if not match:
        return ""
    return normalize_alias(unescape(match.group(1)))


def stable_hash(*parts: object) -> str:
    payload = "\x1f".join("" if part is None else str(part) for part in parts)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def item_uid(prefix: str, *parts: object) -> str:
    return f"{prefix}|{stable_hash(*parts)}"


def seller_uid(source_dataset: str, source_market_raw: str, source_seller_raw: str, source_seller_id_raw: str = "") -> str:
    if source_seller_raw:
        return f"{source_dataset}|{source_market_raw}|seller_raw:{source_seller_raw}"
    return f"{source_dataset}|{source_market_raw}|seller_id:{source_seller_id_raw}"


def contains_cjk(*fields: object) -> bool:
    return any(CJK_RE.search(str(field or "")) for field in fields)


def write_csv_header(path: Path, fieldnames: list[str]) -> tuple[csv.DictWriter, object]:
    handle = path.open("w", encoding="utf-8-sig", newline="")
    writer = csv.DictWriter(handle, fieldnames=fieldnames)
    writer.writeheader()
    return writer, handle


def append_seller_row(stats: dict, item_row: dict) -> None:
    stats["item_count_total"] += 1
    bucket = item_row["data_bucket"]
    if bucket == "en_content_train_pool":
        stats["item_count_en_content_eligible"] += 1
    elif bucket == "zh_target_strict":
        stats["item_count_zh_target_strict"] += 1
    elif bucket == "zh_target_aux":
        stats["item_count_zh_target_aux"] += 1
    elif bucket == "excluded":
        reason = item_row["exclusion_reason"]
        if reason == "alias_overlap_with_en_gold":
            stats["item_count_excluded_alias_overlap"] += 1
        elif reason == "pgp_fingerprint_overlap_via_aux_alias":
            stats["item_count_excluded_pgp_overlap"] += 1
        elif reason == "contact_overlap_with_en_gold":
            stats["item_count_excluded_contact_overlap"] += 1
        elif reason == "non_target_contains_cjk":
            stats["item_count_excluded_non_target_cjk"] += 1
        else:
            stats["item_count_excluded_other"] += 1


def seller_primary_bucket(stats: dict) -> tuple[str, str]:
    if stats["item_count_zh_target_strict"] > 0:
        return "zh_target_strict", ""
    if stats["item_count_zh_target_aux"] > 0:
        return "zh_target_aux", ""
    if stats["item_count_en_content_eligible"] > 0:
        if (
            stats["item_count_excluded_alias_overlap"] > 0
            or stats["item_count_excluded_pgp_overlap"] > 0
            or stats["item_count_excluded_contact_overlap"] > 0
            or stats["item_count_excluded_non_target_cjk"] > 0
        ):
            return "en_content_train_pool", "partial_item_exclusion"
        return "en_content_train_pool", ""
    if stats["item_count_excluded_alias_overlap"] > 0:
        return "excluded", "alias_overlap_with_en_gold"
    if stats["item_count_excluded_pgp_overlap"] > 0:
        return "excluded", "pgp_fingerprint_overlap_via_aux_alias"
    if stats["item_count_excluded_contact_overlap"] > 0:
        return "excluded", "contact_overlap_with_en_gold"
    if stats["item_count_excluded_non_target_cjk"] > 0:
        return "excluded", "no_eligible_items_after_cjk_filter"
    return "excluded", "no_eligible_items"


def build_en_gold_alias_ledger(
    benchmark_manifest_writer: csv.DictWriter,
) -> tuple[dict[str, dict[str, int]], Counter, dict]:
    ledger: dict[str, dict[str, int]] = defaultdict(lambda: {
        "tijkc3xx": 0,
        "suspected_strong": 0,
        "suspected_weak": 0,
        "suspected_imposter": 0,
    })
    benchmark_counts = Counter()
    benchmark_state = {
        "aliases": set(),
        "vendor_ids": set(),
        "fingerprints": set(),
        "fingerprint_shorts": set(),
    }

    # tijkc3xx.sql
    with (ROOT / "tijkc3xx.sql").open("r", encoding="utf-8", errors="ignore") as handle:
        for line_number, line in enumerate(handle, start=1):
            line = line.strip()
            if not line.startswith("("):
                continue
            match = VENDOR_SQL_RE.match(line)
            if not match:
                continue
            vendor_id_raw, user_name_raw, market_id_raw, user_id_raw, link_raw, profile_raw, vendor_link_raw, *_rest, imposter_flag = match.groups()
            alias_normalized = normalize_alias(user_name_raw)
            if alias_normalized:
                ledger[alias_normalized]["tijkc3xx"] += 1
                benchmark_state["aliases"].add(alias_normalized)
            benchmark_manifest_writer.writerow({
                "benchmark_uid": f"tijkc3xx|market:{market_id_raw}|vendor:{vendor_id_raw}",
                "source_dataset": "tijkc3xx.sql",
                "record_type": "vendor_registry",
                "benchmark_bucket": "en_gold_benchmark",
                "benchmark_role": "vendor_registry",
                "market_id_raw": market_id_raw,
                "vendor_id_raw": vendor_id_raw,
                "user_name_raw": user_name_raw,
                "alias_normalized": alias_normalized,
                "evidence_key_raw": "",
                "fingerprint_raw": "",
                "fingerprint_short_raw": "",
                "group_size_raw": "",
                "source_row_number": line_number,
                "aux_info": f"vendor_link_host={urlparse(vendor_link_raw).netloc.lower()}|imposter_flag={imposter_flag}",
            })
            benchmark_counts["tijkc3xx_rows"] += 1

    def process_group_csv(source_name: str, source_key: str, benchmark_role: str, evidence_key_field: str, group_size_field: str) -> None:
        nonlocal ledger, benchmark_counts
        with (ROOT / source_name).open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for source_row_number, row in enumerate(reader, start=2):
                user_name_raw = row.get("user_name", "")
                alias_normalized = normalize_alias(user_name_raw)
                if alias_normalized:
                    ledger[alias_normalized][source_key] += 1
                    benchmark_state["aliases"].add(alias_normalized)
                key_alias_normalized = normalize_alias(row.get(evidence_key_field, ""))
                if key_alias_normalized:
                    ledger[key_alias_normalized][source_key] += 1
                    benchmark_state["aliases"].add(key_alias_normalized)
                vendor_id_raw = clean_text(row.get("vendor_id", ""))
                if vendor_id_raw:
                    benchmark_state["vendor_ids"].add(vendor_id_raw)
                fingerprint_raw = normalize_fingerprint(row.get("fingerprint", ""))
                fingerprint_short_raw = normalize_fingerprint_short(row.get("fingerprint_f", "") or fingerprint_raw)
                if fingerprint_raw:
                    benchmark_state["fingerprints"].add(fingerprint_raw)
                if fingerprint_short_raw:
                    benchmark_state["fingerprint_shorts"].add(fingerprint_short_raw)
                benchmark_manifest_writer.writerow({
                    "benchmark_uid": item_uid(source_name, source_row_number, row.get("market_id"), row.get("vendor_id"), row.get(evidence_key_field)),
                    "source_dataset": source_name,
                    "record_type": "benchmark_evidence",
                    "benchmark_bucket": "en_gold_benchmark",
                    "benchmark_role": benchmark_role,
                    "market_id_raw": row.get("market_id", ""),
                    "vendor_id_raw": vendor_id_raw,
                    "user_name_raw": user_name_raw,
                    "alias_normalized": alias_normalized,
                    "evidence_key_raw": row.get(evidence_key_field, ""),
                    "fingerprint_raw": fingerprint_raw,
                    "fingerprint_short_raw": fingerprint_short_raw,
                    "group_size_raw": row.get(group_size_field, ""),
                    "source_row_number": source_row_number,
                    "aux_info": "",
                })
                benchmark_counts[f"{source_key}_rows"] += 1

    process_group_csv(
        "suspected_sockpuppet_strong.csv",
        "suspected_strong",
        "strong_sockpuppet_group",
        "key_alias",
        "strong_group_distinct_user_count",
    )
    process_group_csv(
        "suspected_sockpuppet_weak.csv",
        "suspected_weak",
        "weak_sockpuppet_group",
        "key_alias",
        "weak_group_distinct_user_count",
    )

    with (ROOT / "suspected_imposter_rows.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for source_row_number, row in enumerate(reader, start=2):
            user_name_raw = row.get("user_name", "")
            alias_normalized = normalize_alias(user_name_raw)
            if alias_normalized:
                ledger[alias_normalized]["suspected_imposter"] += 1
                benchmark_state["aliases"].add(alias_normalized)
            vendor_id_raw = clean_text(row.get("vendor_id", ""))
            if vendor_id_raw:
                benchmark_state["vendor_ids"].add(vendor_id_raw)
            benchmark_manifest_writer.writerow({
                "benchmark_uid": item_uid("suspected_imposter_rows.csv", source_row_number, row.get("market_id"), row.get("vendor_id"), row.get("imposter")),
                "source_dataset": "suspected_imposter_rows.csv",
                "record_type": "benchmark_evidence",
                "benchmark_bucket": "en_gold_benchmark",
                "benchmark_role": "imposter_flag",
                "market_id_raw": row.get("market_id", ""),
                "vendor_id_raw": vendor_id_raw,
                "user_name_raw": user_name_raw,
                "alias_normalized": alias_normalized,
                "evidence_key_raw": row.get("imposter", ""),
                "fingerprint_raw": "",
                "fingerprint_short_raw": "",
                "group_size_raw": "",
                "source_row_number": source_row_number,
                "aux_info": "",
            })
            benchmark_counts["suspected_imposter_rows"] += 1

    return ledger, benchmark_counts, benchmark_state


def write_aux_pgp_manifest() -> tuple[list[dict], Counter]:
    output_path = ROOT / "reports" / "step2_aux_pgp_evidence_manifest.csv"
    writer, handle = write_csv_header(output_path, [
        "evidence_uid",
        "source_dataset",
        "pgp_id_raw",
        "alias_raw",
        "alias_normalized",
        "fingerprint_short_raw",
        "fingerprint_raw",
        "fingerprint_short_normalized",
        "fingerprint_raw_normalized",
        "vendor_ids_raw",
        "vendor_id_count",
        "user_hash_raw",
        "review_count_raw",
        "star_rate_raw",
    ])
    counts = Counter()
    rows: list[dict] = []
    with (ROOT / "3z669jwe.sql").open("r", encoding="utf-8", errors="ignore") as source:
        for line_number, line in enumerate(source, start=1):
            line = line.strip()
            if not line.startswith("("):
                continue
            match = PGP_SQL_RE.match(line)
            if not match:
                continue
            pgp_id_raw, alias_raw, fingerprint_short_raw, fingerprint_raw, vendor_ids_raw, user_hash_raw, review_count_raw, star_rate_raw = match.groups()
            alias_raw = sql_unescape(alias_raw)
            fingerprint_short_raw = sql_unescape(fingerprint_short_raw)
            fingerprint_raw = sql_unescape(fingerprint_raw)
            vendor_ids_raw = sql_unescape(vendor_ids_raw)
            user_hash_raw = sql_unescape(user_hash_raw)
            fingerprint_raw_normalized = normalize_fingerprint(fingerprint_raw)
            fingerprint_short_normalized = normalize_fingerprint_short(fingerprint_short_raw or fingerprint_raw_normalized)
            vendor_id_set = parse_vendor_ids(vendor_ids_raw)
            writer.writerow({
                "evidence_uid": f"3z669jwe|pgp:{pgp_id_raw}",
                "source_dataset": "3z669jwe.sql",
                "pgp_id_raw": pgp_id_raw,
                "alias_raw": alias_raw,
                "alias_normalized": normalize_alias(alias_raw),
                "fingerprint_short_raw": fingerprint_short_raw,
                "fingerprint_raw": fingerprint_raw,
                "fingerprint_short_normalized": fingerprint_short_normalized,
                "fingerprint_raw_normalized": fingerprint_raw_normalized,
                "vendor_ids_raw": vendor_ids_raw,
                "vendor_id_count": len(vendor_id_set),
                "user_hash_raw": user_hash_raw,
                "review_count_raw": review_count_raw,
                "star_rate_raw": star_rate_raw,
            })
            rows.append({
                "pgp_id_raw": pgp_id_raw,
                "alias_raw": alias_raw,
                "alias_normalized": normalize_alias(alias_raw),
                "fingerprint_short_raw": fingerprint_short_raw,
                "fingerprint_raw": fingerprint_raw,
                "fingerprint_short_normalized": fingerprint_short_normalized,
                "fingerprint_raw_normalized": fingerprint_raw_normalized,
                "vendor_ids_raw": vendor_ids_raw,
                "vendor_ids": vendor_id_set,
                "user_hash_raw": user_hash_raw,
            })
            counts["aux_pgp_rows"] += 1
    handle.close()
    return rows, counts


def build_gold_identity_closure(
    alias_ledger: dict[str, dict[str, int]],
    benchmark_state: dict,
    aux_rows: list[dict],
) -> dict:
    strong_aliases = set(benchmark_state["aliases"])
    strong_vendor_ids = set(benchmark_state["vendor_ids"])
    strong_fingerprints = set(benchmark_state["fingerprints"])
    strong_fingerprint_shorts = set(benchmark_state["fingerprint_shorts"])
    alias_to_fingerprints: dict[str, set[str]] = defaultdict(set)

    for row in aux_rows:
        alias = row["alias_normalized"]
        if not alias:
            continue
        if row["fingerprint_raw_normalized"]:
            alias_to_fingerprints[alias].add(row["fingerprint_raw_normalized"])
        if row["fingerprint_short_normalized"]:
            alias_to_fingerprints[alias].add(row["fingerprint_short_normalized"])

    iterations = 0
    changed = True
    while changed:
        changed = False
        iterations += 1
        for row in aux_rows:
            alias = row["alias_normalized"]
            row_vendor_ids = row["vendor_ids"]
            row_active = (
                (alias and alias in strong_aliases)
                or (row["fingerprint_raw_normalized"] and row["fingerprint_raw_normalized"] in strong_fingerprints)
                or (row["fingerprint_short_normalized"] and row["fingerprint_short_normalized"] in strong_fingerprint_shorts)
                or bool(row_vendor_ids & strong_vendor_ids)
            )
            if not row_active:
                continue

            if alias and alias not in strong_aliases:
                strong_aliases.add(alias)
                changed = True
            if row["fingerprint_raw_normalized"] and row["fingerprint_raw_normalized"] not in strong_fingerprints:
                strong_fingerprints.add(row["fingerprint_raw_normalized"])
                changed = True
            if row["fingerprint_short_normalized"] and row["fingerprint_short_normalized"] not in strong_fingerprint_shorts:
                strong_fingerprint_shorts.add(row["fingerprint_short_normalized"])
                changed = True
            new_vendor_ids = row_vendor_ids - strong_vendor_ids
            if new_vendor_ids:
                strong_vendor_ids.update(new_vendor_ids)
                changed = True

    expanded_aliases = strong_aliases - set(alias_ledger.keys())
    return {
        "strong_aliases": strong_aliases,
        "strong_vendor_ids": strong_vendor_ids,
        "strong_fingerprints": strong_fingerprints,
        "strong_fingerprint_shorts": strong_fingerprint_shorts,
        "alias_to_fingerprints": alias_to_fingerprints,
        "closure_iterations": iterations,
        "expanded_alias_count": len(expanded_aliases),
        "expanded_aliases": expanded_aliases,
    }


def extract_benchmark_linked_contacts(identity_closure: dict) -> dict:
    contact_index: dict[tuple[str, str], dict] = {}
    page_count = 0

    for page_path in sorted(HTML_RIPS_DIR.glob("0x*")):
        fingerprint_short = normalize_fingerprint_short(page_path.name)
        if not fingerprint_short or fingerprint_short not in identity_closure["strong_fingerprint_shorts"]:
            continue
        raw_html = page_path.read_text(encoding="utf-8", errors="ignore")
        header_text = html_header_visible_text(raw_html)
        header_contacts = extract_contacts(header_text)
        page_alias = extract_html_vendor_alias(raw_html)
        page_count += 1

        for contact_type, values in header_contacts.items():
            for value in values:
                key = (contact_type, value)
                record = contact_index.setdefault(key, {
                    "contact_type": contact_type,
                    "contact_value": value,
                    "page_aliases": set(),
                    "fingerprint_shorts": set(),
                    "page_count": 0,
                })
                record["page_aliases"].add(page_alias)
                record["fingerprint_shorts"].add(fingerprint_short)
                record["page_count"] += 1

    return {
        "page_count": page_count,
        "contact_index": contact_index,
        "strong_contacts": set(contact_index.keys()),
    }


def write_contact_exclusion_manifest(contact_bundle: dict) -> Counter:
    output_path = ROOT / "reports" / "step2_en_gold_contact_exclusion_list.csv"
    writer, handle = write_csv_header(output_path, [
        "contact_type",
        "contact_value",
        "linked_page_count",
        "alias_sample_count",
        "alias_samples",
        "fingerprint_short_count",
        "fingerprint_short_samples",
    ])
    counts = Counter()
    for (_contact_type, _value), record in sorted(contact_bundle["contact_index"].items()):
        alias_samples = sorted(filter(None, record["page_aliases"]))
        fingerprint_samples = sorted(record["fingerprint_shorts"])
        writer.writerow({
            "contact_type": record["contact_type"],
            "contact_value": record["contact_value"],
            "linked_page_count": record["page_count"],
            "alias_sample_count": len(alias_samples),
            "alias_samples": " || ".join(alias_samples[:5]),
            "fingerprint_short_count": len(fingerprint_samples),
            "fingerprint_short_samples": " || ".join(fingerprint_samples[:5]),
        })
        counts["benchmark_contact_values"] += 1
    handle.close()
    counts["benchmark_contact_pages"] = contact_bundle["page_count"]
    return counts


def write_en_gold_exclusion_list(alias_ledger: dict[str, dict[str, int]], identity_closure: dict) -> Counter:
    output_path = ROOT / "reports" / "step2_en_gold_alias_exclusion_list.csv"
    writer, handle = write_csv_header(output_path, [
        "alias_normalized",
        "from_tijkc3xx_count",
        "from_suspected_strong_count",
        "from_suspected_weak_count",
        "from_suspected_imposter_count",
        "source_presence_count",
        "is_official_benchmark_alias",
        "is_aux_graph_expanded_alias",
        "linked_aux_fingerprint_count",
        "linked_aux_fingerprints",
    ])
    counts = Counter()
    strong_aliases = identity_closure["strong_aliases"]
    expanded_aliases = identity_closure["expanded_aliases"]
    alias_to_fingerprints = identity_closure["alias_to_fingerprints"]
    for alias in sorted(strong_aliases):
        row = alias_ledger.get(alias, {
            "tijkc3xx": 0,
            "suspected_strong": 0,
            "suspected_weak": 0,
            "suspected_imposter": 0,
        })
        fingerprint_values = sorted(alias_to_fingerprints.get(alias, set()))
        writer.writerow({
            "alias_normalized": alias,
            "from_tijkc3xx_count": row["tijkc3xx"],
            "from_suspected_strong_count": row["suspected_strong"],
            "from_suspected_weak_count": row["suspected_weak"],
            "from_suspected_imposter_count": row["suspected_imposter"],
            "source_presence_count": sum(1 for value in row.values() if value > 0),
            "is_official_benchmark_alias": str(alias in alias_ledger).lower(),
            "is_aux_graph_expanded_alias": str(alias in expanded_aliases).lower(),
            "linked_aux_fingerprint_count": len(fingerprint_values),
            "linked_aux_fingerprints": " || ".join(fingerprint_values[:5]),
        })
        counts["strong_alias_rows"] += 1
        if alias in expanded_aliases:
            counts["expanded_alias_rows"] += 1
    handle.close()
    return counts


def process_content_sources(identity_closure: dict, benchmark_contacts: dict) -> dict:
    item_manifest_path = ROOT / "reports" / "step2_content_item_manifest.csv"
    seller_manifest_path = ROOT / "reports" / "step2_content_seller_manifest.csv"
    strong_aliases = identity_closure["strong_aliases"]
    strong_fingerprints = identity_closure["strong_fingerprints"] | identity_closure["strong_fingerprint_shorts"]
    aux_alias_to_fingerprints = identity_closure["alias_to_fingerprints"]
    strong_contacts = benchmark_contacts["strong_contacts"]

    item_writer, item_handle = write_csv_header(item_manifest_path, [
        "item_uid",
        "seller_uid",
        "source_dataset",
        "source_row_number",
        "source_market_raw",
        "source_seller_raw",
        "source_seller_id_raw",
        "source_item_id_raw",
        "alias_normalized",
        "text_script_bucket",
        "data_bucket",
        "eligibility_status",
        "exclusion_reason",
        "exclusion_evidence",
    ])

    summary = {
        "en_content_raw_item_count": 0,
        "en_content_eligible_item_count": 0,
        "en_content_excluded_alias_overlap_item_count": 0,
        "en_content_excluded_pgp_overlap_item_count": 0,
        "en_content_excluded_contact_overlap_item_count": 0,
        "en_content_excluded_non_target_cjk_item_count": 0,
        "zh_target_strict_item_count": 0,
        "zh_target_aux_item_count": 0,
        "zh_target_strict_by_market": Counter(),
        "en_content_raw_seller_uids": set(),
        "en_content_eligible_seller_uids": set(),
    }
    seller_stats: dict[str, dict] = {}

    def ensure_seller(item_row: dict) -> dict:
        seller_key = item_row["seller_uid"]
        if seller_key not in seller_stats:
            seller_stats[seller_key] = {
                "seller_uid": seller_key,
                "source_dataset": item_row["source_dataset"],
                "source_market_raw": item_row["source_market_raw"],
                "source_seller_raw": item_row["source_seller_raw"],
                "source_seller_id_raw": item_row["source_seller_id_raw"],
                "alias_normalized": item_row["alias_normalized"],
                "item_count_total": 0,
                "item_count_en_content_eligible": 0,
                "item_count_zh_target_strict": 0,
                "item_count_zh_target_aux": 0,
                "item_count_excluded_alias_overlap": 0,
                "item_count_excluded_pgp_overlap": 0,
                "item_count_excluded_contact_overlap": 0,
                "item_count_excluded_non_target_cjk": 0,
                "item_count_excluded_other": 0,
            }
        return seller_stats[seller_key]

    # products_data.csv -> zh_target_aux
    with (ROOT / "products_data.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for source_row_number, row in enumerate(reader, start=2):
            source_seller_id_raw = row.get("卖家ID", "")
            seller_key = seller_uid("products_data.csv", "__unknown__", "", source_seller_id_raw)
            item_row = {
                "item_uid": f"products_data.csv|__unknown__|seller_id:{source_seller_id_raw}|item_id:{row.get('交易编号', '')}|row:{source_row_number}",
                "seller_uid": seller_key,
                "source_dataset": "products_data.csv",
                "source_row_number": source_row_number,
                "source_market_raw": "__unknown__",
                "source_seller_raw": "",
                "source_seller_id_raw": source_seller_id_raw,
                "source_item_id_raw": row.get("交易编号", ""),
                "alias_normalized": "",
                "text_script_bucket": "contains_cjk" if contains_cjk(row.get("标题"), row.get("商品描述"), row.get("类别")) else "no_cjk_detected",
                "data_bucket": "zh_target_aux",
                "eligibility_status": "target_aux_only",
                "exclusion_reason": "",
                "exclusion_evidence": "",
            }
            item_writer.writerow(item_row)
            stats = ensure_seller(item_row)
            append_seller_row(stats, item_row)
            summary["zh_target_aux_item_count"] += 1

    # market_item.xlsx
    workbook = load_workbook(ROOT / "market_item.xlsx", read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for source_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        source_seller_raw, ship_from_raw, title_raw, description_raw, price_raw, category_raw, source_market_raw = row
        source_seller_raw = "" if source_seller_raw is None else str(source_seller_raw)
        source_market_raw = "" if source_market_raw is None else str(source_market_raw)
        alias_normalized = normalize_alias(source_seller_raw)
        seller_key = seller_uid("market_item.xlsx", source_market_raw, source_seller_raw)
        row_contains_cjk = contains_cjk(title_raw, description_raw, category_raw)
        if source_market_raw in TARGET_MARKETS:
            data_bucket = "zh_target_strict"
            eligibility_status = "target_eval_candidate"
            exclusion_reason = ""
            exclusion_evidence = ""
            summary["zh_target_strict_item_count"] += 1
            summary["zh_target_strict_by_market"][source_market_raw] += 1
        else:
            summary["en_content_raw_item_count"] += 1
            summary["en_content_raw_seller_uids"].add(seller_key)
            row_contacts = extract_contacts(title_raw, description_raw, category_raw, source_seller_raw)
            row_contact_hits = sorted(
                f"{contact_type}:{value}"
                for contact_type, values in row_contacts.items()
                for value in values
                if (contact_type, value) in strong_contacts
            )
            row_fingerprint_hits = sorted(aux_alias_to_fingerprints.get(alias_normalized, set()) & strong_fingerprints)
            if alias_normalized and alias_normalized in strong_aliases:
                data_bucket = "excluded"
                eligibility_status = "excluded"
                exclusion_reason = "alias_overlap_with_en_gold"
                exclusion_evidence = alias_normalized
                summary["en_content_excluded_alias_overlap_item_count"] += 1
            elif row_fingerprint_hits:
                data_bucket = "excluded"
                eligibility_status = "excluded"
                exclusion_reason = "pgp_fingerprint_overlap_via_aux_alias"
                exclusion_evidence = " || ".join(row_fingerprint_hits[:5])
                summary["en_content_excluded_pgp_overlap_item_count"] += 1
            elif row_contact_hits:
                data_bucket = "excluded"
                eligibility_status = "excluded"
                exclusion_reason = "contact_overlap_with_en_gold"
                exclusion_evidence = " || ".join(row_contact_hits[:5])
                summary["en_content_excluded_contact_overlap_item_count"] += 1
            elif row_contains_cjk:
                data_bucket = "excluded"
                eligibility_status = "excluded"
                exclusion_reason = "non_target_contains_cjk"
                exclusion_evidence = ""
                summary["en_content_excluded_non_target_cjk_item_count"] += 1
            else:
                data_bucket = "en_content_train_pool"
                eligibility_status = "content_train_eligible"
                exclusion_reason = ""
                exclusion_evidence = ""
                summary["en_content_eligible_item_count"] += 1
                summary["en_content_eligible_seller_uids"].add(seller_key)
        item_row = {
            "item_uid": item_uid("market_item.xlsx", source_row_number, source_market_raw, source_seller_raw, title_raw, description_raw, price_raw, category_raw),
            "seller_uid": seller_key,
            "source_dataset": "market_item.xlsx",
            "source_row_number": source_row_number,
            "source_market_raw": source_market_raw,
            "source_seller_raw": source_seller_raw,
            "source_seller_id_raw": "",
            "source_item_id_raw": "",
            "alias_normalized": alias_normalized,
            "text_script_bucket": "contains_cjk" if row_contains_cjk else "no_cjk_detected",
            "data_bucket": data_bucket,
            "eligibility_status": eligibility_status,
            "exclusion_reason": exclusion_reason,
            "exclusion_evidence": exclusion_evidence,
        }
        item_writer.writerow(item_row)
        stats = ensure_seller(item_row)
        append_seller_row(stats, item_row)
    workbook.close()

    # Agora -> en_content candidate
    workbook = load_workbook(ROOT / "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx", read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for source_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        source_seller_raw = "" if row[0] is None else str(row[0]).strip()
        category_raw = row[1]
        title_raw = row[2]
        description_raw = row[3]
        price_raw = row[4]
        source_market_raw = "agora"
        alias_normalized = normalize_alias(source_seller_raw)
        seller_key = seller_uid("2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx", source_market_raw, source_seller_raw)
        row_contains_cjk = contains_cjk(title_raw, description_raw, category_raw)
        summary["en_content_raw_item_count"] += 1
        summary["en_content_raw_seller_uids"].add(seller_key)
        row_contacts = extract_contacts(title_raw, description_raw, category_raw, source_seller_raw)
        row_contact_hits = sorted(
            f"{contact_type}:{value}"
            for contact_type, values in row_contacts.items()
            for value in values
            if (contact_type, value) in strong_contacts
        )
        row_fingerprint_hits = sorted(aux_alias_to_fingerprints.get(alias_normalized, set()) & strong_fingerprints)
        if alias_normalized and alias_normalized in strong_aliases:
            data_bucket = "excluded"
            eligibility_status = "excluded"
            exclusion_reason = "alias_overlap_with_en_gold"
            exclusion_evidence = alias_normalized
            summary["en_content_excluded_alias_overlap_item_count"] += 1
        elif row_fingerprint_hits:
            data_bucket = "excluded"
            eligibility_status = "excluded"
            exclusion_reason = "pgp_fingerprint_overlap_via_aux_alias"
            exclusion_evidence = " || ".join(row_fingerprint_hits[:5])
            summary["en_content_excluded_pgp_overlap_item_count"] += 1
        elif row_contact_hits:
            data_bucket = "excluded"
            eligibility_status = "excluded"
            exclusion_reason = "contact_overlap_with_en_gold"
            exclusion_evidence = " || ".join(row_contact_hits[:5])
            summary["en_content_excluded_contact_overlap_item_count"] += 1
        elif row_contains_cjk:
            data_bucket = "excluded"
            eligibility_status = "excluded"
            exclusion_reason = "non_target_contains_cjk"
            exclusion_evidence = ""
            summary["en_content_excluded_non_target_cjk_item_count"] += 1
        else:
            data_bucket = "en_content_train_pool"
            eligibility_status = "content_train_eligible"
            exclusion_reason = ""
            exclusion_evidence = ""
            summary["en_content_eligible_item_count"] += 1
            summary["en_content_eligible_seller_uids"].add(seller_key)
        item_row = {
            "item_uid": item_uid("2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx", source_row_number, source_seller_raw, title_raw, description_raw, price_raw, category_raw),
            "seller_uid": seller_key,
            "source_dataset": "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx",
            "source_row_number": source_row_number,
            "source_market_raw": source_market_raw,
            "source_seller_raw": source_seller_raw,
            "source_seller_id_raw": "",
            "source_item_id_raw": "",
            "alias_normalized": alias_normalized,
            "text_script_bucket": "contains_cjk" if row_contains_cjk else "no_cjk_detected",
            "data_bucket": data_bucket,
            "eligibility_status": eligibility_status,
            "exclusion_reason": exclusion_reason,
            "exclusion_evidence": exclusion_evidence,
        }
        item_writer.writerow(item_row)
        stats = ensure_seller(item_row)
        append_seller_row(stats, item_row)
    workbook.close()
    item_handle.close()

    seller_writer, seller_handle = write_csv_header(seller_manifest_path, [
        "seller_uid",
        "source_dataset",
        "source_market_raw",
        "source_seller_raw",
        "source_seller_id_raw",
        "alias_normalized",
        "primary_data_bucket",
        "seller_level_note",
        "item_count_total",
        "item_count_en_content_eligible",
        "item_count_zh_target_strict",
        "item_count_zh_target_aux",
        "item_count_excluded_alias_overlap",
        "item_count_excluded_pgp_overlap",
        "item_count_excluded_contact_overlap",
        "item_count_excluded_non_target_cjk",
        "item_count_excluded_other",
    ])
    seller_summary_counts = Counter()
    post_en_content_aliases = set()
    post_en_content_fingerprints = set()
    for seller_key in sorted(seller_stats):
        stats = seller_stats[seller_key]
        primary_bucket, seller_level_note = seller_primary_bucket(stats)
        if primary_bucket == "en_content_train_pool":
            seller_summary_counts["en_content_eligible_sellers"] += 1
            if stats["alias_normalized"]:
                post_en_content_aliases.add(stats["alias_normalized"])
                post_en_content_fingerprints.update(aux_alias_to_fingerprints.get(stats["alias_normalized"], set()))
        elif primary_bucket == "zh_target_strict":
            seller_summary_counts["zh_target_strict_sellers"] += 1
        elif primary_bucket == "zh_target_aux":
            seller_summary_counts["zh_target_aux_sellers"] += 1
        else:
            if seller_level_note == "alias_overlap_with_en_gold":
                seller_summary_counts["excluded_alias_overlap_sellers"] += 1
            elif seller_level_note == "pgp_fingerprint_overlap_via_aux_alias":
                seller_summary_counts["excluded_pgp_overlap_sellers"] += 1
            elif seller_level_note == "contact_overlap_with_en_gold":
                seller_summary_counts["excluded_contact_overlap_sellers"] += 1
            elif seller_level_note == "no_eligible_items_after_cjk_filter":
                seller_summary_counts["excluded_cjk_only_sellers"] += 1
            else:
                seller_summary_counts["excluded_other_sellers"] += 1
        seller_writer.writerow({
            "seller_uid": stats["seller_uid"],
            "source_dataset": stats["source_dataset"],
            "source_market_raw": stats["source_market_raw"],
            "source_seller_raw": stats["source_seller_raw"],
            "source_seller_id_raw": stats["source_seller_id_raw"],
            "alias_normalized": stats["alias_normalized"],
            "primary_data_bucket": primary_bucket,
            "seller_level_note": seller_level_note,
            "item_count_total": stats["item_count_total"],
            "item_count_en_content_eligible": stats["item_count_en_content_eligible"],
            "item_count_zh_target_strict": stats["item_count_zh_target_strict"],
            "item_count_zh_target_aux": stats["item_count_zh_target_aux"],
            "item_count_excluded_alias_overlap": stats["item_count_excluded_alias_overlap"],
            "item_count_excluded_pgp_overlap": stats["item_count_excluded_pgp_overlap"],
            "item_count_excluded_contact_overlap": stats["item_count_excluded_contact_overlap"],
            "item_count_excluded_non_target_cjk": stats["item_count_excluded_non_target_cjk"],
            "item_count_excluded_other": stats["item_count_excluded_other"],
        })
    seller_handle.close()

    summary["en_content_raw_seller_count"] = len(summary["en_content_raw_seller_uids"])
    summary["en_content_eligible_seller_count"] = len(summary["en_content_eligible_seller_uids"])
    summary["en_content_postfilter_alias_overlap_count"] = len(post_en_content_aliases & strong_aliases)
    summary["en_content_postfilter_aux_fingerprint_overlap_count"] = len(post_en_content_fingerprints & strong_fingerprints)
    summary["en_content_prefilter_alias_overlap_count"] = len({
        normalize_alias(uid.split("seller_raw:", 1)[1]) for uid in summary["en_content_raw_seller_uids"] if "seller_raw:" in uid
    } & strong_aliases)
    summary["seller_summary_counts"] = dict(seller_summary_counts)
    summary["zh_target_strict_by_market"] = dict(summary["zh_target_strict_by_market"])
    summary["en_content_raw_seller_uids"] = None
    summary["en_content_eligible_seller_uids"] = None
    return summary


def main() -> None:
    policy = load_policy()
    reports_dir = ROOT / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    benchmark_manifest_writer, benchmark_manifest_handle = write_csv_header(
        reports_dir / "step2_en_gold_benchmark_manifest.csv",
        [
            "benchmark_uid",
            "source_dataset",
            "record_type",
            "benchmark_bucket",
            "benchmark_role",
            "market_id_raw",
            "vendor_id_raw",
            "user_name_raw",
            "alias_normalized",
            "evidence_key_raw",
            "fingerprint_raw",
            "fingerprint_short_raw",
            "group_size_raw",
            "source_row_number",
            "aux_info",
        ],
    )

    alias_ledger, benchmark_counts, benchmark_state = build_en_gold_alias_ledger(benchmark_manifest_writer)
    benchmark_manifest_handle.close()

    aux_rows, aux_counts = write_aux_pgp_manifest()
    identity_closure = build_gold_identity_closure(alias_ledger, benchmark_state, aux_rows)
    exclusion_counts = write_en_gold_exclusion_list(alias_ledger, identity_closure)
    benchmark_contacts = extract_benchmark_linked_contacts(identity_closure)
    contact_counts = write_contact_exclusion_manifest(benchmark_contacts)
    content_summary = process_content_sources(identity_closure, benchmark_contacts)

    summary = {
        "policy_path": str(POLICY_PATH.relative_to(ROOT)),
        "official_gold_alias_exclusion_count": len(alias_ledger),
        "strong_gold_alias_exclusion_count": len(identity_closure["strong_aliases"]),
        "strong_gold_fingerprint_exclusion_count": len(identity_closure["strong_fingerprint_shorts"]),
        "benchmark_manifest_counts": dict(benchmark_counts),
        "auxiliary_identity_evidence_counts": dict(aux_counts),
        "exclusion_ledger_counts": dict(exclusion_counts),
        "benchmark_contact_counts": dict(contact_counts),
        "identity_closure": {
            "closure_iterations": identity_closure["closure_iterations"],
            "expanded_alias_count": identity_closure["expanded_alias_count"],
            "strong_vendor_id_count": len(identity_closure["strong_vendor_ids"]),
            "strong_fingerprint_short_count": len(identity_closure["strong_fingerprint_shorts"]),
            "strong_fingerprint_raw_count": len(identity_closure["strong_fingerprints"]),
        },
        "content_split_summary": content_summary,
        "acceptance_checks": {
            "en_content_postfilter_alias_overlap_count": content_summary["en_content_postfilter_alias_overlap_count"],
            "en_content_postfilter_alias_overlap_pass": content_summary["en_content_postfilter_alias_overlap_count"] == 0,
            "en_content_postfilter_aux_fingerprint_overlap_count": content_summary["en_content_postfilter_aux_fingerprint_overlap_count"],
            "en_content_postfilter_aux_fingerprint_overlap_pass": content_summary["en_content_postfilter_aux_fingerprint_overlap_count"] == 0,
            "zh_target_strict_markets": sorted(TARGET_MARKETS),
            "products_data_bucket": "zh_target_aux",
        },
        "residual_risks": [
            "Identity-closure exclusion is intentionally strict and may over-remove legitimate English content sellers linked via reused fingerprints or contacts.",
            "CJK filtering removes the strongest observed target-language contamination, but non-CJK non-English rows are not separately filtered at this step.",
            "Benchmark-linked contacts are extracted from local Grams HTML header regions and remain incomplete when a seller has no matching snapshot page.",
            "products_data.csv remains auxiliary because market provenance is unresolved in the raw file.",
        ],
    }

    with (reports_dir / "step2_split_summary.json").open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)

    print(f"Wrote {reports_dir / 'step2_split_summary.json'}")


if __name__ == "__main__":
    main()
