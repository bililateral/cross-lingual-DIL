#!/usr/bin/env python3
"""Build train-only pseudo-alias views from disjoint real Chinese seller items."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

import step3_build_seller_profiles as step3
import step15_build_v7_clean_embedding_cache as redaction


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_POLICY = ROOT / "schema" / "step22_same_seller_split_policy.json"
WS_RE = re.compile(r"\s+")


def resolve(value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else ROOT / path


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def canonical_hash(value: object) -> str:
    payload = json.dumps(value, ensure_ascii=True, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def bool_value(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes"}


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    return WS_RE.sub(" ", text).strip()


def content_signature(item: dict) -> str:
    title = normalize_text(item.get("title", "")).casefold()
    description = normalize_text(item.get("description", "")).casefold()
    return canonical_hash([title, description]) if title or description else ""


def deterministic_key(seed: int, *parts: object) -> str:
    payload = "\x1f".join([str(seed), *(str(part) for part in parts)])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def split_values(value: object) -> set[str]:
    return {token.strip() for token in re.split(r"[|,;]+", str(value or "")) if token.strip()}


def heldout_entities(assignments: list[dict], exclusions: list[dict], excluded_splits: set[str]) -> tuple[set[str], set[str], dict[str, str]]:
    heldout_sellers: set[str] = set()
    heldout_aliases: set[str] = set()
    train_components_by_seller: dict[str, set[str]] = defaultdict(set)
    for row in assignments:
        if row.get("dataset") != "zh_target_strict":
            continue
        split = row.get("split_name", "")
        sellers = {row.get("seller_uid_left", ""), row.get("seller_uid_right", "")}
        sellers.discard("")
        if split in excluded_splits:
            heldout_sellers.update(sellers)
        elif split == "train":
            for seller_uid in sellers:
                train_components_by_seller[seller_uid].add(row["recomputed_component_id"])
    for row in exclusions:
        source_splits = split_values(row.get("source_splits", ""))
        if not source_splits.intersection(excluded_splits):
            continue
        if row.get("dataset") != "zh_target_strict":
            continue
        if row.get("entity_type") == "seller_uid":
            heldout_sellers.add(row.get("entity_id", ""))
        elif row.get("entity_type") == "seller_alias":
            heldout_aliases.add(normalize_text(row.get("entity_id", "")).casefold())
    component_index = {}
    for seller_uid, components in train_components_by_seller.items():
        if len(components) != 1:
            raise ValueError(f"Train seller spans multiple recomputed components: {seller_uid}")
        component_index[seller_uid] = next(iter(components))
    return heldout_sellers, heldout_aliases, component_index


def load_strict_item_metadata(path: Path, policy: dict, heldout_sellers: set[str], heldout_aliases: set[str]) -> tuple[dict[int, dict], Counter]:
    eligibility = policy["eligibility"]
    selected: dict[int, dict] = {}
    counters: Counter = Counter()
    candidates = []
    discovered_heldout_aliases: set[str] = set()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if row.get("data_bucket") != eligibility["required_data_bucket"]:
                continue
            if row.get("source_dataset") != eligibility["required_source_dataset"]:
                continue
            counters["strict_manifest_rows"] += 1
            if row.get("seller_uid") in heldout_sellers:
                counters["excluded_heldout_seller_rows"] += 1
                alias = normalize_text(row.get("alias_normalized", "")).casefold()
                if alias:
                    discovered_heldout_aliases.add(alias)
                continue
            candidates.append(row)
    heldout_aliases.update(discovered_heldout_aliases)
    for row in candidates:
        alias = normalize_text(row.get("alias_normalized", "")).casefold()
        if alias and alias in heldout_aliases:
            counters["excluded_heldout_alias_rows"] += 1
            continue
        selected[int(row["source_row_number"])] = row
    return selected, counters


def redact_item(item: dict, literals: list[str]) -> tuple[dict, Counter]:
    output = dict(item)
    diagnostics: Counter = Counter()
    for field in ("title", "description", "category", "ship_from"):
        clean, diag = redaction.redact_identifiers(str(item.get(field, "") or ""), literals)
        redaction.assert_no_known_identifier_residue(clean, literals, item["seller_uid"])
        output[field] = clean
        diagnostics.update(diag)
    output["content_signature"] = content_signature(output)
    return output, diagnostics


def load_redacted_items(workbook_path: Path, metadata: dict[int, dict], signal_literals: dict[str, list[str]]) -> tuple[dict[str, list[dict]], Counter]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    items_by_seller: dict[str, list[dict]] = defaultdict(list)
    diagnostics: Counter = Counter()
    for source_row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        meta = metadata.get(source_row_number)
        if meta is None:
            continue
        vendor, ship_from, title, description, price, category, market = values
        item = {
            "item_uid": meta["item_uid"],
            "seller_uid": meta["seller_uid"],
            "source_row_number": source_row_number,
            "source_market_raw": normalize_text(market),
            "source_seller_raw": normalize_text(vendor),
            "title": normalize_text(title),
            "description": normalize_text(description),
            "price": normalize_text(price),
            "category": normalize_text(category),
            "ship_from": normalize_text(ship_from),
        }
        literals = list(signal_literals.get(item["seller_uid"], []))
        alias_literal = redaction.safe_signal_literal("seller_alias", meta.get("alias_normalized", ""))
        if alias_literal:
            literals.append(alias_literal)
        clean, diag = redact_item(item, sorted(set(literals), key=lambda value: (-len(value), value.casefold())))
        diagnostics.update(diag)
        if clean["content_signature"]:
            items_by_seller[clean["seller_uid"]].append(clean)
    workbook.close()
    return items_by_seller, diagnostics


def content_groups(items: list[dict]) -> list[list[dict]]:
    eligible = [item for item in items if item.get("content_signature") or content_signature(item)]
    parent = list(range(len(eligible)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root != right_root:
            parent[max(left_root, right_root)] = min(left_root, right_root)

    key_owner: dict[tuple[str, str], int] = {}
    for index, item in enumerate(eligible):
        title = normalize_text(item.get("title", "")).casefold()
        description = normalize_text(item.get("description", "")).casefold()
        keys = []
        if title:
            keys.append(("title", title))
        if description:
            keys.append(("description", description))
        for key in keys:
            if key in key_owner:
                union(index, key_owner[key])
            else:
                key_owner[key] = index
    grouped: dict[int, list[dict]] = defaultdict(list)
    for index, item in enumerate(eligible):
        grouped[find(index)].append(item)
    groups = [sorted(group, key=lambda row: row["item_uid"]) for group in grouped.values()]
    return sorted(groups, key=lambda group: group[0]["item_uid"])


def partition_positive_items(items: list[dict], seller_uid: str, seed: int, cfg: dict) -> tuple[list[dict], list[dict]] | None:
    groups = content_groups(items)
    if len(items) < int(cfg["minimum_source_items"]) or len(groups) < int(cfg["minimum_unique_content_groups"]):
        return None
    max_groups = int(cfg["maximum_content_groups_per_positive_parent"])
    groups = sorted(groups, key=lambda group: deterministic_key(seed, seller_uid, group[0]["content_signature"]))[:max_groups]
    by_category: dict[str, list[list[dict]]] = defaultdict(list)
    for group in groups:
        category = normalize_text(group[0].get("category", "")).casefold()
        by_category[category].append(group)
    left_groups: list[list[dict]] = []
    right_groups: list[list[dict]] = []
    for category, category_groups in sorted(by_category.items()):
        ordered = sorted(category_groups, key=lambda group: deterministic_key(seed, seller_uid, category, group[0]["content_signature"]))
        offset = int(deterministic_key(seed, seller_uid, category)[:2], 16) % 2
        for index, group in enumerate(ordered):
            (left_groups if (index + offset) % 2 == 0 else right_groups).append(group)
    all_groups = left_groups + right_groups
    min_groups = int(cfg["minimum_content_groups_per_positive_view"])
    if len(left_groups) < min_groups or len(right_groups) < min_groups:
        ordered = sorted(all_groups, key=lambda group: deterministic_key(seed, seller_uid, "rebalance", group[0]["content_signature"]))
        left_groups = ordered[::2]
        right_groups = ordered[1::2]
    left = [item for group in left_groups for item in group]
    right = [item for group in right_groups for item in group]
    if len(left_groups) < min_groups or len(right_groups) < min_groups:
        return None
    if len(left) < int(cfg["minimum_items_per_positive_view"]) or len(right) < int(cfg["minimum_items_per_positive_view"]):
        return None
    left_titles = {normalize_text(item["title"]).casefold() for item in left if normalize_text(item["title"])}
    right_titles = {normalize_text(item["title"]).casefold() for item in right if normalize_text(item["title"])}
    left_descriptions = {normalize_text(item["description"]).casefold() for item in left if normalize_text(item["description"])}
    right_descriptions = {normalize_text(item["description"]).casefold() for item in right if normalize_text(item["description"])}
    if left_titles.intersection(right_titles) or left_descriptions.intersection(right_descriptions):
        raise ValueError(f"Positive view exact-content overlap survived grouping: {seller_uid}")
    return left, right


def select_negative_view(items: list[dict], key: str, seed: int, cfg: dict) -> list[dict] | None:
    groups = content_groups(items)
    minimum = int(cfg["minimum_content_groups_per_negative_view"])
    if len(items) < int(cfg["minimum_items_per_negative_view"]) or len(groups) < minimum:
        return None
    maximum = int(cfg["maximum_content_groups_per_negative_view"])
    selected = sorted(groups, key=lambda group: deterministic_key(seed, key, group[0]["content_signature"]))[:maximum]
    view = [item for group in selected for item in group]
    return view if len(view) >= int(cfg["minimum_items_per_negative_view"]) else None


def numeric_stats(values: list[float]) -> dict:
    if not values:
        return {"count": 0, "min": None, "median": None, "max": None}
    ordered = sorted(values)
    middle = len(ordered) // 2
    median = ordered[middle] if len(ordered) % 2 else (ordered[middle - 1] + ordered[middle]) / 2
    return {"count": len(values), "min": min(values), "median": median, "max": max(values)}


def top_values(values: list[str], limit: int) -> list[dict]:
    counter = Counter(value for value in values if value)
    first = {value: values.index(value) for value in counter}
    return [
        {"value": value, "count": count}
        for value, count in sorted(counter.items(), key=lambda item: (-item[1], first[item[0]], item[0]))[:limit]
    ]


def build_profile(items: list[dict], pseudo_uid: str, marker: str, limit: int) -> dict:
    categories = [normalize_text(item.get("category", "")) for item in items]
    titles = [normalize_text(item.get("title", "")) for item in items]
    descriptions = [step3.description_snippet(item.get("description", "")) for item in items]
    prices = [step3.parse_first_number(item.get("price", "")) for item in items]
    prices = [float(value) for value in prices if value is not None and math.isfinite(float(value))]
    style = [step3.style_snapshot(item.get("title", ""), item.get("description", "")) for item in items]
    top_categories = top_values(categories, limit)
    top_titles = top_values(titles, limit)
    top_descriptions = top_values(descriptions, limit)
    category_concat = " || ".join(row["value"] for row in top_categories)
    title_concat = " || ".join(row["value"] for row in top_titles)
    description_concat = " || ".join(row["value"] for row in top_descriptions)
    # Match the v7 clean encoder: raw configured fields joined by newlines,
    # without synthetic-only section labels or provenance markers.
    sections = [value for value in (category_concat, title_concat, description_concat) if value]
    profile_text = "\n".join(sections)
    if not profile_text.strip():
        raise ValueError(f"Step22 pseudo profile is empty: {pseudo_uid}")
    count = len(items)
    return {
        "seller_uid": pseudo_uid,
        "data_bucket": "zh_synthetic_train_only",
        "source_dataset": "synthetic_same_seller_item_split",
        "source_market_raw": marker,
        "source_seller_raw": "",
        "source_seller_id_raw": "",
        "alias_normalized": "",
        "item_count": count,
        "first_source_row_number": "",
        "last_source_row_number": "",
        "unique_title_count": len({value.casefold() for value in titles if value}),
        "unique_description_snippet_count": len({value.casefold() for value in descriptions if value}),
        "unique_category_count": len({value.casefold() for value in categories if value}),
        "cjk_item_count": sum(step3.contains_cjk_item(item.get("title", ""), item.get("description", ""), item.get("category", "")) for item in items),
        "parsed_price_count": len(prices),
        "parsed_rating_count": 0,
        "contact_type_count": 0,
        "contact_token_count_total": 0,
        "title_length_stats": numeric_stats([float(row["title_len"]) for row in style]),
        "description_length_stats": numeric_stats([float(row["description_len"]) for row in style]),
        "style_stats": {
            "digit_ratio_mean": sum(row["digit_ratio"] for row in style) / count,
            "punct_ratio_mean": sum(row["punct_ratio"] for row in style) / count,
            "uppercase_ratio_mean": sum(row["uppercase_ratio"] for row in style) / count,
            "newline_count_mean": sum(row["newline_count"] for row in style) / count,
            "repeated_title_share": 1.0 - len({value.casefold() for value in titles if value}) / count,
            "repeated_description_share": 1.0 - len({value.casefold() for value in descriptions if value}) / count,
            "max_category_share": max(Counter(value for value in categories if value).values(), default=0) / count,
        },
        "price_numeric_approx_stats": numeric_stats(prices),
        "rating_numeric_stats": numeric_stats([]),
        "source_specific_numeric_stats": {},
        "top_categories": top_categories,
        "signature_titles": [],
        "top_titles": top_titles,
        "signature_description_segments": [],
        "top_description_snippets": top_descriptions,
        "top_price_strings": [],
        "top_ship_from_values": [],
        "contact_signals": {contact_type: [] for contact_type in step3.CONTACT_TYPES},
        "structured_snapshot_examples": [],
        "category_concat_top": category_concat,
        "signature_title_concat": "",
        "title_concat_top": title_concat,
        "signature_description_concat": "",
        "description_concat_top": description_concat,
        "contact_concat_top": "",
        "structured_snapshot_concat_top": "",
        "profile_text": profile_text,
        "synthetic_train_only": True,
        "benchmark_eligible": False,
    }


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        if fields:
            writer.writeheader()
            writer.writerows(rows)


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--policy", default=str(DEFAULT_POLICY))
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    policy_path = resolve(args.policy)
    policy = load_json(policy_path)
    inputs = {name: resolve(value) for name, value in policy["inputs"].items()}
    output_root = resolve(policy["outputs_root"])
    summary_path = output_root / policy["outputs"]["summary"]
    manifest_path = output_root / policy["outputs"]["manifest"]
    producer_path = Path(__file__).resolve()
    input_manifest = {
        name: {"path": str(path.relative_to(ROOT)).replace("\\", "/"), "size_bytes": path.stat().st_size, "sha256": sha256(path)}
        for name, path in inputs.items()
    }
    expected_identity = {
        "policy_sha256": sha256(policy_path),
        "producer_sha256": sha256(producer_path),
        "dependency_sha256": {
            "step3_build_seller_profiles": sha256(Path(step3.__file__).resolve()),
            "step15_build_v7_clean_embedding_cache": sha256(Path(redaction.__file__).resolve()),
        },
        "inputs": input_manifest,
    }
    if output_root.exists() and not args.force:
        if not summary_path.is_file() or not manifest_path.is_file():
            raise FileExistsError(f"Incomplete Step22 root exists: {output_root}")
        existing = load_json(manifest_path)
        if {key: existing.get(key) for key in expected_identity} != expected_identity:
            raise FileExistsError("Existing Step22 root belongs to different code, policy, or data")
        for record in existing.get("outputs", []):
            path = resolve(record["path"])
            if not path.is_file() or path.stat().st_size != int(record["size_bytes"]) or sha256(path) != record["sha256"]:
                raise ValueError(f"Existing Step22 generation artifact drift: {path}")
        print(summary_path.read_text(encoding="utf-8"))
        return

    assignments = load_csv(inputs["component_assignments"])
    exclusions = load_csv(inputs["permanent_exclusions"])
    excluded_splits = set(policy["eligibility"]["excluded_canonical_splits"])
    heldout_sellers, heldout_aliases, train_component_index = heldout_entities(assignments, exclusions, excluded_splits)
    metadata, manifest_counters = load_strict_item_metadata(inputs["item_manifest"], policy, heldout_sellers, heldout_aliases)
    signal_literals, signal_summary = redaction.signal_literals_by_seller(inputs["item_identity_signals"])
    items_by_seller, redaction_diagnostics = load_redacted_items(inputs["market_item_workbook"], metadata, signal_literals)

    eligibility = policy["eligibility"]
    seed = int(policy["global_seed"])
    marker = policy["generation"]["synthetic_market_marker"]
    prefix = policy["generation"]["synthetic_uid_prefix"].rstrip("/")
    limit = int(policy["generation"]["profile_text_item_limit"])
    positive_candidates = []
    rejected_positive = Counter()
    for seller_uid in sorted(items_by_seller, key=lambda value: deterministic_key(seed, "positive", value)):
        partition = partition_positive_items(items_by_seller[seller_uid], seller_uid, seed, eligibility)
        if partition is None:
            rejected_positive["insufficient_disjoint_item_content"] += 1
            continue
        positive_candidates.append((seller_uid, partition))
    max_positive = int(eligibility.get("maximum_positive_parent_sellers", 0) or 0)
    if max_positive:
        positive_candidates = positive_candidates[:max_positive]

    labels = load_csv(inputs["frozen_labels"])
    evidence_index = {row["pair_uid"]: row for row in load_csv(inputs["evidence_labels"])}
    assignment_index = {row["pair_uid"]: row for row in assignments if row.get("dataset") == "zh_target_strict"}
    negative_candidates = []
    rejected_negative = Counter()
    for row in sorted(labels, key=lambda item: item["pair_uid"]):
        if row.get("split_name") != "train" or row.get("review_label") != "negative":
            continue
        if eligibility["require_negative_parent_usable_for_supervision"] and not bool_value(row.get("usable_for_supervision")):
            continue
        if eligibility["require_negative_parent_usable_for_core_transfer"] and not bool_value(row.get("usable_for_core_transfer")):
            continue
        assignment = assignment_index.get(row["pair_uid"])
        if assignment is None or assignment.get("split_name") != "train" or bool_value(assignment.get("cross_split_component_leakage")):
            raise ValueError(f"Invalid Step16I negative parent assignment: {row['pair_uid']}")
        left_items = select_negative_view(items_by_seller.get(row["seller_uid_left"], []), row["pair_uid"] + "|left", seed, eligibility)
        right_items = select_negative_view(items_by_seller.get(row["seller_uid_right"], []), row["pair_uid"] + "|right", seed, eligibility)
        if left_items is None or right_items is None:
            rejected_negative["insufficient_item_content"] += 1
            continue
        negative_candidates.append((row, assignment, left_items, right_items, evidence_index.get(row["pair_uid"], {})))
    max_negative = int(eligibility.get("maximum_negative_parent_pairs", 0) or 0)
    if max_negative:
        negative_candidates = negative_candidates[:max_negative]
    if len(positive_candidates) < 100:
        raise ValueError(f"Step22 has too few independent same-seller parents: {len(positive_candidates)}")
    if len(negative_candidates) < 20:
        raise ValueError(f"Step22 has too few reviewed hard-negative parents: {len(negative_candidates)}")

    profiles = []
    pair_rows = []
    pair_lineage = []
    item_lineage = []
    for index, (seller_uid, (left_items, right_items)) in enumerate(positive_candidates):
        key = f"positive/p{index:05d}"
        left_uid = f"{prefix}/{key}/left"
        right_uid = f"{prefix}/{key}/right"
        pair_uid = f"{left_uid}||{right_uid}"
        component = train_component_index.get(seller_uid, f"step22_unlabeled_seller_{canonical_hash(seller_uid)[:20]}")
        profiles.extend([build_profile(left_items, left_uid, marker, limit), build_profile(right_items, right_uid, marker, limit)])
        pair_rows.append({
            "pair_uid": pair_uid, "seller_uid_left": left_uid, "seller_uid_right": right_uid,
            "review_label": "positive", "split_name": "train", "split_component_id": component,
            "synthetic_class": "same_seller_split_positive", "synthetic_train_only": "1",
            "benchmark_eligible": "0", "usable_for_supervision": "1", "usable_for_core_transfer": "0",
            "training_sample_weight": "1.000000000000",
        })
        pair_lineage.append({
            "synthetic_pair_uid": pair_uid, "synthetic_class": "same_seller_split_positive",
            "source_parent_seller_uid_left": seller_uid, "source_parent_seller_uid_right": seller_uid,
            "source_parent_pair_uid": "", "parent_component_id": component,
            "parent_evidence_type": "same_observed_seller_account_item_split",
            "independent_source_unit": seller_uid, "real_cross_account_sockpuppet_relation": "0",
            "synthetic_train_only": "1", "benchmark_eligible": "0",
            "left_item_count": len(left_items), "right_item_count": len(right_items),
        })
        for side, uid, selected_items in (("left", left_uid, left_items), ("right", right_uid, right_items)):
            for item in selected_items:
                item_lineage.append({
                    "synthetic_pair_uid": pair_uid, "synthetic_seller_uid": uid, "side": side,
                    "source_seller_uid": seller_uid, "source_item_uid": item["item_uid"],
                    "source_row_number": item["source_row_number"], "content_signature": item["content_signature"],
                })

    for index, (parent, assignment, left_items, right_items, evidence) in enumerate(negative_candidates):
        key = f"negative/n{index:05d}"
        left_uid = f"{prefix}/{key}/left"
        right_uid = f"{prefix}/{key}/right"
        pair_uid = f"{left_uid}||{right_uid}"
        profiles.extend([build_profile(left_items, left_uid, marker, limit), build_profile(right_items, right_uid, marker, limit)])
        pair_rows.append({
            "pair_uid": pair_uid, "seller_uid_left": left_uid, "seller_uid_right": right_uid,
            "review_label": "negative", "split_name": "train", "split_component_id": assignment["recomputed_component_id"],
            "synthetic_class": "reviewed_negative_item_view", "synthetic_train_only": "1",
            "benchmark_eligible": "0", "usable_for_supervision": "1", "usable_for_core_transfer": "0",
            "training_sample_weight": "1.000000000000",
        })
        pair_lineage.append({
            "synthetic_pair_uid": pair_uid, "synthetic_class": "reviewed_negative_item_view",
            "source_parent_seller_uid_left": parent["seller_uid_left"], "source_parent_seller_uid_right": parent["seller_uid_right"],
            "source_parent_pair_uid": parent["pair_uid"], "parent_component_id": assignment["recomputed_component_id"],
            "parent_evidence_type": evidence.get("evidence_type", "ordinary_negative"),
            "independent_source_unit": parent["pair_uid"], "real_cross_account_sockpuppet_relation": "0",
            "synthetic_train_only": "1", "benchmark_eligible": "0",
            "left_item_count": len(left_items), "right_item_count": len(right_items),
        })
        for side, uid, source_uid, selected_items in (
            ("left", left_uid, parent["seller_uid_left"], left_items),
            ("right", right_uid, parent["seller_uid_right"], right_items),
        ):
            for item in selected_items:
                item_lineage.append({
                    "synthetic_pair_uid": pair_uid, "synthetic_seller_uid": uid, "side": side,
                    "source_seller_uid": source_uid, "source_item_uid": item["item_uid"],
                    "source_row_number": item["source_row_number"], "content_signature": item["content_signature"],
                })

    if len({row["seller_uid"] for row in profiles}) != len(profiles):
        raise ValueError("Step22 pseudo seller UIDs are not unique")
    if len({row["pair_uid"] for row in pair_rows}) != len(pair_rows):
        raise ValueError("Step22 pseudo pair UIDs are not unique")
    if any(row["split_name"] != "train" or row["benchmark_eligible"] != "0" for row in pair_rows):
        raise ValueError("Step22 emitted a benchmark-eligible or non-train row")
    source_items_by_pair_side: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in item_lineage:
        source_items_by_pair_side[(row["synthetic_pair_uid"], row["side"])].add(row["source_item_uid"])
    for row in pair_rows:
        if row["review_label"] == "positive" and source_items_by_pair_side[(row["pair_uid"], "left")].intersection(source_items_by_pair_side[(row["pair_uid"], "right")]):
            raise ValueError(f"Step22 positive views share source items: {row['pair_uid']}")

    paths = {name: output_root / policy["outputs"][name] for name in ("profiles", "pair_labels", "pair_lineage", "item_lineage")}
    write_jsonl(paths["profiles"], profiles)
    write_csv(paths["pair_labels"], pair_rows)
    write_csv(paths["pair_lineage"], pair_lineage)
    write_csv(paths["item_lineage"], item_lineage)
    summary = {
        "step": "step22_same_seller_split_generation",
        "policy_version": policy["version"],
        "status": "generated_train_only_not_benchmark",
        "same_seller_positive_parent_count": len(positive_candidates),
        "reviewed_negative_parent_count": len(negative_candidates),
        "synthetic_positive_pair_count": sum(row["review_label"] == "positive" for row in pair_rows),
        "synthetic_negative_pair_count": sum(row["review_label"] == "negative" for row in pair_rows),
        "synthetic_profile_count": len(profiles),
        "source_item_lineage_count": len(item_lineage),
        "heldout_seller_exclusion_count": len(heldout_sellers),
        "heldout_alias_exclusion_count": len(heldout_aliases),
        "manifest_filter_counts": dict(sorted(manifest_counters.items())),
        "positive_rejection_counts": dict(sorted(rejected_positive.items())),
        "negative_rejection_counts": dict(sorted(rejected_negative.items())),
        "redaction_diagnostics": dict(sorted(redaction_diagnostics.items())),
        "signal_source_summary": signal_summary,
        "input_manifest_sha256": canonical_hash(input_manifest),
        "scientific_interpretation": {
            "new_real_cross_account_positive_count": 0,
            "independent_real_source_seller_count": len(positive_candidates),
            "may_be_used_for_training_augmentation": True,
            "may_be_used_for_validation_or_test": False,
        },
        "output_paths": {name: str(path.relative_to(ROOT)).replace("\\", "/") for name, path in paths.items()},
    }
    write_json(summary_path, summary)
    output_files = [*paths.values(), summary_path]
    manifest = {
        **expected_identity,
        "policy": str(policy_path.relative_to(ROOT)).replace("\\", "/"),
        "producer": str(producer_path.relative_to(ROOT)).replace("\\", "/"),
        "outputs": [
            {"path": str(path.relative_to(ROOT)).replace("\\", "/"), "size_bytes": path.stat().st_size, "sha256": sha256(path)}
            for path in sorted(output_files)
        ],
    }
    write_json(manifest_path, manifest)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
