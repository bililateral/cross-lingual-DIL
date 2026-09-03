from __future__ import annotations

import csv
import hashlib
import html
import itertools
import json
import heapq
import math
import platform
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from importlib.metadata import version as package_version
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
POLICY_PATH = ROOT / "schema" / "step7_v5_english_source_dataset_policy.json"

VENDOR_SQL_RE = re.compile(
    r"^\((\d+), '((?:[^'\\]|\\.)*)', (\d+), (\d+), (\d+), '(.*?)', "
    r"'((?:[^'\\]|\\.)*)', (\d+), (\d+), (\d+), (\d+)\),?;?$"
)
PGP_SQL_RE = re.compile(
    r"^\((\d+), '((?:[^']|'')*)', '((?:[^']|'')*)', "
    r"'((?:[^']|'')*)', 0x[0-9A-Fa-f]+, '((?:[^']|'')*)', "
    r"'((?:[^']|'')*)', (\d+), ([0-9.]+)\),?;?$"
)
WS_RE = re.compile(r"\s+")
TAG_RE = re.compile(r"<[^>]+>")
WORD_RE = re.compile(r"[A-Za-z0-9]+")
PGP_ARMOR_RE = re.compile(
    r"-----BEGIN PGP (?:PUBLIC KEY BLOCK|SIGNED MESSAGE|SIGNATURE)-----.*?"
    r"-----END PGP (?:PUBLIC KEY BLOCK|SIGNATURE)-----",
    re.IGNORECASE | re.DOTALL,
)
URL_RE = re.compile(
    r"(?:(?:https?|ftp)://|www\.)\S+|\b[a-z2-7]{16,56}\.onion\b\S*",
    re.IGNORECASE,
)
EMAIL_RE = re.compile(
    r"\b[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,24}\b",
    re.IGNORECASE,
)
HANDLE_RE = re.compile(
    r"(?:\b(?:telegram|wickr|wickrme|jabber|tox|skype|whatsapp)\b\s*"
    r"(?:id|me|at|is|:|=|@)?\s*[A-Za-z0-9_.@+\-]{3,})|"
    r"(?<![\w.])@[A-Za-z0-9_][A-Za-z0-9_.\-]{2,}",
    re.IGNORECASE,
)
PHONE_RE = re.compile(
    r"\b(?:phone|tel|telephone|whatsapp|call|text)\b\s*[:=]?\s*"
    r"\+?\d[\d\s().\-]{7,}\d",
    re.IGNORECASE,
)
HEX_ID_RE = re.compile(
    r"(?<![0-9A-Fa-f])(?:0x)?[0-9A-Fa-f]{16,40}(?![0-9A-Fa-f])"
)
CRYPTO_RE = re.compile(
    r"\b(?:bc1[a-z0-9]{20,90}|[13][a-km-zA-HJ-NP-Z1-9]{25,34})\b"
)
HEXLIKE_SPAN_RE = re.compile(
    r"(?<![0-9A-Fa-f])[0-9A-Fa-f](?:[0-9A-Fa-f\s:\-]{6,100}[0-9A-Fa-f])?(?![0-9A-Fa-f])"
)
STYLE_TOKEN_RE = re.compile(r"[A-Za-z]+|[\u3400-\u9fff]+|\d+(?:[.,]\d+)*")

AGORA_HEADER = (
    "Vendor",
    " Category",
    " Item",
    " Item Description",
    " Price",
    " Origin",
    " Destination",
    " Rating",
    " Remarks",
)

MATCH_FEATURES = (
    "log_min_items",
    "log_max_items",
    "absolute_log_item_ratio",
    "log_min_tokens",
    "log_max_tokens",
    "absolute_log_token_ratio",
    "category_jaccard",
    "token_jaccard",
    "absolute_title_fraction_difference",
    "mean_description_fraction",
)
STRUCTURAL_FEATURES = tuple(
    name for name in MATCH_FEATURES if name not in {"category_jaccard", "token_jaccard"}
)
TOPIC_DIAGNOSTIC_FEATURES = ("category_jaccard", "token_jaccard")


class DatasetConstructionError(RuntimeError):
    pass


def normalize_text(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def normalize_alias(value: object) -> str:
    return normalize_text(value).casefold()


def normalize_fingerprint(value: object) -> str:
    return re.sub(r"[^0-9A-Fa-f]", "", str(value or "")).upper()


def compact_alias(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_alias(value))


def stable_hash(*parts: object) -> str:
    payload = "\x1f".join(str(part) for part in parts).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_policy() -> dict:
    with POLICY_PATH.open("r", encoding="utf-8") as handle:
        policy = json.load(handle)
    for relative, expected in policy["inputs"].items():
        path = ROOT / relative
        if not path.is_file():
            raise FileNotFoundError(path)
        observed = file_sha256(path)
        if observed != expected:
            raise DatasetConstructionError(
                f"Frozen input drift: {relative}; expected={expected}; observed={observed}"
            )
    return policy


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def parse_sql_tuple_line(line: str, expected_fields: int) -> list[str]:
    stripped = line.strip()
    if stripped.endswith(";"):
        stripped = stripped[:-1]
    if stripped.endswith(","):
        stripped = stripped[:-1]
    if not (stripped.startswith("(") and stripped.endswith(")")):
        raise ValueError("not a SQL tuple")
    fields = next(
        csv.reader(
            [stripped[1:-1]],
            delimiter=",",
            quotechar="'",
            doublequote=True,
            escapechar="\\",
            skipinitialspace=True,
            strict=True,
        )
    )
    if len(fields) != expected_fields:
        raise ValueError(f"expected {expected_fields} fields, observed {len(fields)}")
    return fields


def parse_vendor_registry() -> tuple[dict[tuple[str, str], dict], dict]:
    records: dict[tuple[str, str], dict] = {}
    tuple_lines = 0
    with (ROOT / "tijkc3xx.sql").open("r", encoding="utf-8", errors="strict") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.lstrip().startswith("("):
                continue
            try:
                fields = parse_sql_tuple_line(line, 11)
            except (csv.Error, ValueError) as error:
                if line.lstrip().startswith("("):
                    raise DatasetConstructionError(
                        f"Unparsed vendor SQL tuple at line {line_number}: {error}"
                    ) from error
            tuple_lines += 1
            (
                vendor_id,
                user_name,
                market_id,
                user_id,
                link_id,
                profile,
                vendor_link,
                _,
                _,
                _,
                imposter,
            ) = fields
            records[(market_id, vendor_id)] = {
                "user_name": normalize_alias(user_name),
                "user_id": user_id,
                "link_id": link_id,
                "profile": profile,
                "vendor_link": vendor_link,
                "imposter": int(imposter),
            }
    return records, {"parsed_tuple_lines": tuple_lines, "unparsed_tuple_lines": 0}


def parse_auxiliary_pgp() -> tuple[
    dict[str, set[str]], dict[str, set[str]], dict[str, set[str]], dict
]:
    fingerprint_vendor_ids: dict[str, set[str]] = defaultdict(set)
    vendor_fingerprints: dict[str, set[str]] = defaultdict(set)
    fingerprint_aliases: dict[str, set[str]] = defaultdict(set)
    tuple_lines = 0
    with (ROOT / "3z669jwe.sql").open("r", encoding="utf-8", errors="strict") as handle:
        for line in handle:
            stripped = line.strip()
            if not stripped.startswith("("):
                continue
            try:
                fields = parse_sql_tuple_line(stripped, 9)
            except (csv.Error, ValueError) as error:
                raise DatasetConstructionError(
                    f"Unparsed auxiliary PGP SQL tuple after {tuple_lines} parsed tuples: {error}"
                ) from error
            tuple_lines += 1
            _, alias, _, fingerprint, _, vendor_ids, _, _, _ = fields
            fingerprint = normalize_fingerprint(fingerprint)
            if len(fingerprint) != 40:
                continue
            alias = normalize_alias(alias)
            if alias:
                fingerprint_aliases[fingerprint].add(alias)
            for vendor_id in re.findall(r"\d+", vendor_ids):
                fingerprint_vendor_ids[fingerprint].add(vendor_id)
                vendor_fingerprints[vendor_id].add(fingerprint)
    return (
        fingerprint_vendor_ids,
        vendor_fingerprints,
        fingerprint_aliases,
        {"parsed_tuple_lines": tuple_lines, "unparsed_tuple_lines": 0},
    )


def compile_alias_pattern(aliases: Iterable[str], minimum_length: int) -> re.Pattern[str]:
    retained = sorted(
        {
            normalize_alias(alias)
            for alias in aliases
            if len(normalize_alias(alias)) >= minimum_length
        },
        key=lambda value: (-len(value), value),
    )
    if not retained:
        return re.compile(r"(?!x)x")
    return re.compile(
        r"(?<![\w])(?:" + "|".join(re.escape(value) for value in retained) + r")(?![\w])",
        re.IGNORECASE,
    )


def compile_account_alias_pattern(
    aliases: Iterable[str], minimum_length: int
) -> re.Pattern[str]:
    """Match one account's source aliases even after spacing/punctuation changes.

    This pattern is applied only to the account that owns the aliases.  It is
    therefore safe to remove embedded or de-punctuated variants without
    globally deleting ordinary words that happen to equal another seller's
    alias.  The alias set is derived before strong labels are read.
    """

    patterns = set()
    for alias in aliases:
        compact = compact_alias(alias)
        if len(compact) < minimum_length:
            continue
        core = r"[^A-Za-z0-9]*".join(re.escape(character) for character in compact)
        if len(compact) < 5:
            core = r"(?<![A-Za-z0-9])" + core + r"(?![A-Za-z0-9])"
        patterns.add(core)
    if not patterns:
        return re.compile(r"(?!x)x")
    return re.compile(
        "(?:" + "|".join(sorted(patterns, key=lambda value: (-len(value), value))) + ")",
        re.IGNORECASE,
    )


def source_alias_residuals(
    value: str, aliases: Iterable[str], minimum_length: int
) -> set[str]:
    """Independently detect source aliases after punctuation is removed."""

    normalized = normalize_alias(value)
    compact_text = re.sub(r"[^a-z0-9]", "", normalized)
    hits = set()
    for alias in aliases:
        normalized_alias = normalize_alias(alias)
        compact = compact_alias(alias)
        if len(compact) < minimum_length:
            continue
        if normalized_alias in normalized or compact in compact_text:
            hits.add(normalized_alias)
    return hits


def clean_visible_text(value: object, alias_pattern: re.Pattern[str]) -> str:
    text = normalize_text(html.unescape(str(value or "")))
    for pattern in (
        PGP_ARMOR_RE,
        URL_RE,
        EMAIL_RE,
        HANDLE_RE,
        PHONE_RE,
        HEX_ID_RE,
        CRYPTO_RE,
        alias_pattern,
        TAG_RE,
    ):
        text = pattern.sub(" ", text)
    return WS_RE.sub(" ", text).strip()


def exact_text_key(value: str) -> str:
    return WS_RE.sub(" ", normalize_text(value).casefold()).strip()


def text_tokens(value: str) -> list[str]:
    return WORD_RE.findall(value.casefold())


def style_projection(value: str) -> str:
    """Remove lexical identity without reading labels, pairs, or evidence classes."""

    def replace(match: re.Match[str]) -> str:
        token = match.group(0)
        if token[0].isdigit():
            return "N" + re.sub(r"\d", "0", token)
        if "\u3400" <= token[0] <= "\u9fff":
            return f"H{min(len(token), 99)}"
        if token.isupper():
            case_shape = "U"
        elif token.islower():
            case_shape = "L"
        elif token[:1].isupper() and token[1:].islower():
            case_shape = "T"
        else:
            case_shape = "M"
        return f"W{min(len(token), 99)}{case_shape}"

    return STYLE_TOKEN_RE.sub(replace, value)


def source_fingerprint_residuals(value: str, known_key_ids: set[str]) -> set[str]:
    hits = set()
    for match in HEXLIKE_SPAN_RE.finditer(value):
        compact = normalize_fingerprint(match.group(0))
        if len(compact) < 8:
            continue
        for width in (8, 16, 40):
            if len(compact) >= width:
                candidate = compact[-width:]
                if candidate in known_key_ids:
                    hits.add(candidate)
    return hits


def read_and_clean_agora(
    alias_pattern: re.Pattern[str],
    account_alias_patterns: dict[str, re.Pattern[str]],
) -> tuple[dict[str, list[dict]], dict]:
    path = ROOT / "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = tuple(next(rows))
    if header != AGORA_HEADER:
        workbook.close()
        raise DatasetConstructionError(f"Unexpected Agora header: {header!r}")

    items: dict[str, list[dict]] = defaultdict(list)
    title_owners: dict[str, set[str]] = defaultdict(set)
    description_owners: dict[str, set[str]] = defaultdict(set)
    for source_row_number, row in enumerate(rows, start=2):
        alias = normalize_alias(row[0])
        if not alias:
            continue
        title = clean_visible_text(row[2], alias_pattern)
        description = clean_visible_text(row[3], alias_pattern)
        own_alias_pattern = account_alias_patterns.get(alias)
        if own_alias_pattern is not None:
            title = WS_RE.sub(" ", own_alias_pattern.sub(" ", title)).strip()
            description = WS_RE.sub(
                " ", own_alias_pattern.sub(" ", description)
            ).strip()
        category = clean_visible_text(row[1], alias_pattern)
        title_key = exact_text_key(title)
        description_key = exact_text_key(description)
        if title_key:
            title_owners[title_key].add(alias)
        if description_key:
            description_owners[description_key].add(alias)
        items[alias].append(
            {
                "source_row_number": source_row_number,
                "title_clean": title,
                "description_clean": description,
                "category_clean": category,
                "title_key": title_key,
                "description_key": description_key,
            }
        )
    workbook.close()

    shared_titles = {value for value, owners in title_owners.items() if len(owners) > 1}
    shared_descriptions = {
        value for value, owners in description_owners.items() if len(owners) > 1
    }
    removed_title_rows = 0
    removed_description_rows = 0
    for seller_items in items.values():
        for item in seller_items:
            if item["title_key"] in shared_titles:
                item["title_clean"] = ""
                item["title_key"] = ""
                removed_title_rows += 1
            if item["description_key"] in shared_descriptions:
                item["description_clean"] = ""
                item["description_key"] = ""
                removed_description_rows += 1

    return items, {
        "raw_seller_count": len(items),
        "raw_item_count": sum(len(value) for value in items.values()),
        "cross_seller_exact_title_values_removed": len(shared_titles),
        "cross_seller_exact_description_values_removed": len(shared_descriptions),
        "title_rows_cleared": removed_title_rows,
        "description_rows_cleared": removed_description_rows,
    }


def read_agora_aliases() -> set[str]:
    path = ROOT / "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = tuple(next(rows))
    if header != AGORA_HEADER:
        workbook.close()
        raise DatasetConstructionError(f"Unexpected Agora header: {header!r}")
    aliases = {normalize_alias(row[0]) for row in rows if normalize_alias(row[0])}
    workbook.close()
    return aliases


def compact_aliases_too_similar(left_compact: str, right_compact: str) -> bool:
    if not left_compact or not right_compact:
        return True
    if left_compact == right_compact:
        return True
    shorter = min(len(left_compact), len(right_compact))
    if shorter >= 5 and (
        left_compact in right_compact or right_compact in left_compact
    ):
        return True
    return shorter >= 5 and SequenceMatcher(None, left_compact, right_compact).ratio() >= 0.9


def jaccard(left: set[str], right: set[str]) -> float:
    union = left | right
    if not union:
        return 0.0
    return len(left & right) / len(union)


def account_summary(items: Sequence[dict]) -> dict:
    title_tokens = [
        token for item in items for token in text_tokens(item["title_clean"])
    ]
    description_tokens = [
        token for item in items for token in text_tokens(item["description_clean"])
    ]
    all_tokens = title_tokens + description_tokens
    return {
        "item_count": len(items),
        "token_count": len(all_tokens),
        "title_token_count": len(title_tokens),
        "description_token_count": len(description_tokens),
        "token_set": set(all_tokens),
        "category_set": {
            exact_text_key(item["category_clean"])
            for item in items
            if exact_text_key(item["category_clean"])
        },
    }


def build_identity_conflict_components(accounts: dict[str, dict]) -> dict[str, str]:
    parent = {uid: uid for uid in accounts}

    def find(uid: str) -> str:
        while parent[uid] != uid:
            parent[uid] = parent[parent[uid]]
            uid = parent[uid]
        return uid

    def union(left: str, right: str) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root == right_root:
            return
        if left_root < right_root:
            parent[right_root] = left_root
        else:
            parent[left_root] = right_root

    evidence_owner: dict[tuple[str, str], str] = {}
    for uid in sorted(accounts):
        account = accounts[uid]
        tokens = {
            *(('auxiliary_pgp', value) for value in account["aux_fingerprints"]),
            *(('weak_pgp', value) for value in account["weak_fingerprints"]),
            *(('strong_pgp', value) for value in account["strong_fingerprints"]),
            *(('strong_key_alias', value) for value in account["strong_key_aliases"]),
        }
        for token in sorted(tokens):
            owner = evidence_owner.setdefault(token, uid)
            union(uid, owner)

    members: dict[str, list[str]] = defaultdict(list)
    for uid in sorted(accounts):
        members[find(uid)].append(uid)
    component_by_account = {}
    for values in members.values():
        component_uid = stable_hash("step7-v5-identity-conflict-component-v2", *values)
        for uid in values:
            component_by_account[uid] = component_uid
    return component_by_account


def pair_covariates(left: dict, right: dict) -> dict[str, float]:
    item_counts = sorted((left["item_count"], right["item_count"]))
    token_counts = sorted((left["token_count"], right["token_count"]))
    left_total = max(1, left["token_count"])
    right_total = max(1, right["token_count"])
    left_title_fraction = left["title_token_count"] / left_total
    right_title_fraction = right["title_token_count"] / right_total
    left_description_fraction = left["description_token_count"] / left_total
    right_description_fraction = right["description_token_count"] / right_total
    return {
        "log_min_items": math.log1p(item_counts[0]),
        "log_max_items": math.log1p(item_counts[1]),
        "absolute_log_item_ratio": abs(math.log((item_counts[0] + 1) / (item_counts[1] + 1))),
        "log_min_tokens": math.log1p(token_counts[0]),
        "log_max_tokens": math.log1p(token_counts[1]),
        "absolute_log_token_ratio": abs(
            math.log((token_counts[0] + 1) / (token_counts[1] + 1))
        ),
        "category_jaccard": jaccard(left["category_set"], right["category_set"]),
        "token_jaccard": jaccard(left["token_set"], right["token_set"]),
        "absolute_title_fraction_difference": abs(
            left_title_fraction - right_title_fraction
        ),
        "mean_description_fraction": (
            left_description_fraction + right_description_fraction
        ) / 2,
    }


def assign_controller_folds(
    accounts_by_controller: dict[str, list[str]],
    summaries: dict[str, dict],
    fold_count: int,
    seed: str,
) -> dict[str, int]:
    controller_stats = []
    for controller, accounts in accounts_by_controller.items():
        positive_pairs = len(accounts) * (len(accounts) - 1) // 2
        controller_stats.append(
            (
                controller,
                positive_pairs,
                len(accounts),
                sum(summaries[account]["item_count"] for account in accounts),
                stable_hash(seed, controller),
            )
        )
    controller_stats.sort(key=lambda value: (-value[1], -value[2], -value[3], value[4]))
    fold_totals = [
        {"positive_pairs": 0, "accounts": 0, "items": 0} for _ in range(fold_count)
    ]
    assignments: dict[str, int] = {}
    for controller, positive_pairs, account_count, item_count, _ in controller_stats:
        fold = min(
            range(fold_count),
            key=lambda index: (
                fold_totals[index]["positive_pairs"],
                fold_totals[index]["accounts"],
                fold_totals[index]["items"],
                index,
            ),
        )
        assignments[controller] = fold
        fold_totals[fold]["positive_pairs"] += positive_pairs
        fold_totals[fold]["accounts"] += account_count
        fold_totals[fold]["items"] += item_count
    return assignments


def matching_cost(
    target: dict[str, float],
    candidate: dict[str, float],
    scales: dict[str, float],
    tie_breaker: float,
) -> float:
    return sum(
        ((target[name] - candidate[name]) / scales[name]) ** 2
        for name in MATCH_FEATURES
    ) + tie_breaker


ROUGH_MATCH_SCALES = {
    "log_min_items": 1.5,
    "log_max_items": 1.5,
    "absolute_log_item_ratio": 1.0,
    "log_min_tokens": 2.0,
    "log_max_tokens": 2.0,
    "absolute_log_token_ratio": 1.0,
    "category_jaccard": 0.15,
    "token_jaccard": 0.08,
    "absolute_title_fraction_difference": 0.12,
    "mean_description_fraction": 0.12,
}


def shortlist_negative_candidates(
    accounts: dict[str, dict],
    positive_pairs: list[dict],
    per_positive_limit: int,
    seed: str,
) -> tuple[list[dict], Counter[str], int]:
    """Keep the nearest metadata-screened contrasts for every positive pair.

    The complete same-fold pair pool can contain hundreds of thousands of pairs.
    This streaming top-k step is deterministic and uses only the same visible
    structural covariates later audited as potential shortcuts.
    """

    positive_by_fold: dict[int, list[dict]] = defaultdict(list)
    for row in positive_pairs:
        positive_by_fold[row["fold_id"]].append(row)
    account_uids_by_fold: dict[int, list[str]] = defaultdict(list)
    for uid, account in accounts.items():
        account_uids_by_fold[account["fold_id"]].append(uid)
    for values in account_uids_by_fold.values():
        values.sort()

    heaps: dict[str, list[tuple[float, str, dict]]] = {
        row["raw_pair_key"]: [] for row in positive_pairs
    }
    screen_counts: Counter[str] = Counter()
    eligible_candidate_count = 0
    for fold, positive_rows in sorted(positive_by_fold.items()):
        for left_uid, right_uid in itertools.combinations(account_uids_by_fold[fold], 2):
            left = accounts[left_uid]
            right = accounts[right_uid]
            if left["conflict_component_uid"] == right["conflict_component_uid"]:
                screen_counts["same_identity_conflict_component"] += 1
                continue
            if left["weak_fingerprints"] & right["weak_fingerprints"]:
                screen_counts["shared_weak_fingerprint"] += 1
                continue
            if left["aux_fingerprints"] & right["aux_fingerprints"]:
                screen_counts["shared_auxiliary_fingerprint"] += 1
                continue
            if left["strong_key_aliases"] & right["strong_key_aliases"]:
                screen_counts["shared_key_alias"] += 1
                continue
            if compact_aliases_too_similar(
                left["alias_compact"], right["alias_compact"]
            ):
                screen_counts["similar_alias"] += 1
                continue
            eligible_candidate_count += 1
            covariates = pair_covariates(left["summary"], right["summary"])
            raw_pair_key = stable_hash("raw-pair", left_uid, right_uid)
            candidate = {
                "left_uid": left_uid,
                "right_uid": right_uid,
                "fold_id": fold,
                "label": 0,
                "evidence_class": "screened_distinct_full_pgp_contrast",
                "raw_pair_key": raw_pair_key,
                "covariates": covariates,
            }
            for positive in positive_rows:
                rough_cost = matching_cost(
                    positive["covariates"],
                    covariates,
                    ROUGH_MATCH_SCALES,
                    int(stable_hash(seed, positive["raw_pair_key"], raw_pair_key)[:12], 16)
                    / (16**12)
                    * 1e-9,
                )
                heap = heaps[positive["raw_pair_key"]]
                entry = (-rough_cost, raw_pair_key, candidate)
                if len(heap) < per_positive_limit:
                    heapq.heappush(heap, entry)
                elif entry > heap[0]:
                    heapq.heapreplace(heap, entry)

    shortlisted: dict[str, dict] = {}
    for heap in heaps.values():
        for _, raw_pair_key, candidate in heap:
            shortlisted[raw_pair_key] = candidate
    return list(shortlisted.values()), screen_counts, eligible_candidate_count


def select_matched_negatives(
    positive_pairs: list[dict],
    negative_candidates: list[dict],
    ratio: int,
    seed: str,
) -> list[dict]:
    try:
        import numpy as np
        from scipy.optimize import linear_sum_assignment
    except ImportError as error:
        raise DatasetConstructionError(
            "NumPy and SciPy are required for deterministic minimum-cost negative matching"
        ) from error

    selected: list[dict] = []
    for fold in sorted({row["fold_id"] for row in positive_pairs}):
        positives = [row for row in positive_pairs if row["fold_id"] == fold]
        candidates = [row for row in negative_candidates if row["fold_id"] == fold]
        slots = [row for row in positives for _ in range(ratio)]
        if len(candidates) < len(slots):
            raise DatasetConstructionError(
                f"Fold {fold} has {len(candidates)} negative candidates for {len(slots)} slots"
            )
        all_rows = positives + candidates
        scales: dict[str, float] = {}
        for name in MATCH_FEATURES:
            values = np.asarray([row["covariates"][name] for row in all_rows], dtype=float)
            scale = float(values.std())
            scales[name] = scale if scale > 1e-12 else 1.0
        costs = np.empty((len(slots), len(candidates)), dtype=float)
        for row_index, target in enumerate(slots):
            for column_index, candidate in enumerate(candidates):
                tie_hash = stable_hash(seed, fold, row_index, candidate["raw_pair_key"])
                tie_breaker = int(tie_hash[:12], 16) / (16**12) * 1e-9
                costs[row_index, column_index] = matching_cost(
                    target["covariates"], candidate["covariates"], scales, tie_breaker
                )
        row_indices, column_indices = linear_sum_assignment(costs)
        if len(row_indices) != len(slots):
            raise DatasetConstructionError(f"Fold {fold} did not fill every negative slot")
        for slot_index, candidate_index in zip(row_indices.tolist(), column_indices.tolist()):
            row = dict(candidates[candidate_index])
            row["matched_positive_raw_pair_key"] = slots[slot_index]["raw_pair_key"]
            row["matching_cost"] = float(costs[slot_index, candidate_index])
            selected.append(row)
    return selected


def standardized_mean_differences(
    rows: Sequence[dict], feature_names: Sequence[str] = MATCH_FEATURES
) -> dict[str, float]:
    import numpy as np

    result = {}
    for name in feature_names:
        positives = np.asarray(
            [row["covariates"][name] for row in rows if row["label"] == 1], dtype=float
        )
        negatives = np.asarray(
            [row["covariates"][name] for row in rows if row["label"] == 0], dtype=float
        )
        pooled = math.sqrt((float(positives.var()) + float(negatives.var())) / 2)
        difference = float(positives.mean() - negatives.mean())
        if pooled <= 1e-12:
            result[name] = 0.0 if abs(difference) <= 1e-12 else math.inf
        else:
            result[name] = difference / pooled
    return result


def proxy_audit(
    rows: Sequence[dict], fold_count: int, feature_names: Sequence[str], protocol_name: str
) -> dict:
    try:
        import numpy as np
        from sklearn.linear_model import LogisticRegression
        from sklearn.metrics import average_precision_score, roc_auc_score
        from sklearn.pipeline import make_pipeline
        from sklearn.preprocessing import StandardScaler
    except ImportError as error:
        raise DatasetConstructionError(
            "NumPy and scikit-learn are required for the structural shortcut audit"
        ) from error

    matrix = np.asarray(
        [[row["covariates"][name] for name in feature_names] for row in rows],
        dtype=float,
    )
    labels = np.asarray([row["label"] for row in rows], dtype=int)
    folds = np.asarray([row["fold_id"] for row in rows], dtype=int)
    probabilities = np.zeros(len(rows), dtype=float)
    fold_metrics = []
    for fold in range(fold_count):
        train = folds != fold
        test = folds == fold
        if len(set(labels[train].tolist())) != 2 or len(set(labels[test].tolist())) != 2:
            raise DatasetConstructionError(f"Structural audit fold {fold} lacks both labels")
        model = make_pipeline(
            StandardScaler(),
            LogisticRegression(C=1.0, solver="lbfgs", max_iter=5000, random_state=0),
        )
        model.fit(matrix[train], labels[train])
        probabilities[test] = model.predict_proba(matrix[test])[:, 1]
        fold_metrics.append(
            {
                "fold_id": fold,
                "row_count": int(test.sum()),
                "positive_count": int(labels[test].sum()),
                "roc_auc": float(roc_auc_score(labels[test], probabilities[test])),
                "average_precision": float(
                    average_precision_score(labels[test], probabilities[test])
                ),
            }
        )
    prevalence = float(labels.mean())
    auc = float(roc_auc_score(labels, probabilities))
    average_precision = float(average_precision_score(labels, probabilities))
    return {
        "feature_names": list(feature_names),
        "protocol": protocol_name,
        "prevalence": prevalence,
        "roc_auc": auc,
        "bidirectional_roc_auc": max(auc, 1.0 - auc),
        "average_precision": average_precision,
        "average_precision_lift_over_prevalence": average_precision - prevalence,
        "folds": fold_metrics,
    }


def write_csv(path: Path, fieldnames: Sequence[str], rows: Iterable[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def write_json(path: Path, payload: object) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def build() -> dict:
    policy = load_policy()
    construction = policy["construction"]
    market_id = policy["source_market"]["raw_market_id"]
    strong_all = read_csv_rows(ROOT / "suspected_sockpuppet_strong.csv")
    strong_market = [row for row in strong_all if row["market_id"] == market_id]
    weak_all = read_csv_rows(ROOT / "suspected_sockpuppet_weak.csv")
    imposter_rows = read_csv_rows(ROOT / "suspected_imposter_rows.csv")
    registry, vendor_sql_parse_audit = parse_vendor_registry()
    (
        fingerprint_vendor_ids,
        vendor_fingerprints,
        fingerprint_aliases,
        pgp_sql_parse_audit,
    ) = parse_auxiliary_pgp()

    imposter_keys = {
        (row["market_id"], row["vendor_id"])
        for row in imposter_rows
        if row.get("imposter") == "1"
    }
    imposter_aliases = {
        normalize_alias(row["user_name"])
        for row in imposter_rows
        if row["market_id"] == market_id and row.get("imposter") == "1"
    }
    agora_aliases = read_agora_aliases()
    registry_by_alias: dict[str, list[tuple[str, dict]]] = defaultdict(list)
    for (row_market_id, vendor_id), registry_row in registry.items():
        if row_market_id == market_id:
            registry_by_alias[registry_row["user_name"]].append((vendor_id, registry_row))

    candidate_metadata: dict[str, dict] = {}
    pre_exclusion_counts: Counter[str] = Counter()
    for alias in sorted(agora_aliases):
        if (
            len(compact_alias(alias))
            < construction["alias_minimum_redaction_length"]
        ):
            pre_exclusion_counts["alias_too_short_for_safe_redaction"] += 1
            continue
        registry_rows = registry_by_alias.get(alias, [])
        if len(registry_rows) != 1:
            pre_exclusion_counts[
                "no_unique_market_registry_match"
                if not registry_rows
                else "ambiguous_market_registry_match"
            ] += 1
            continue
        vendor_id, registry_row = registry_rows[0]
        if (
            registry_row["imposter"]
            or (market_id, vendor_id) in imposter_keys
            or alias in imposter_aliases
        ):
            pre_exclusion_counts["imposter"] += 1
            continue
        fingerprints = vendor_fingerprints.get(vendor_id, set())
        if len(fingerprints) != 1:
            pre_exclusion_counts[
                "no_auxiliary_pgp" if not fingerprints else "multiple_auxiliary_pgp"
            ] += 1
            continue
        fingerprint = next(iter(fingerprints))
        if len(fingerprint) != 40 or vendor_id not in fingerprint_vendor_ids[fingerprint]:
            pre_exclusion_counts["invalid_auxiliary_pgp"] += 1
            continue
        source_aliases = {alias, *fingerprint_aliases.get(fingerprint, set())}
        if any(
            len(compact_alias(source_alias))
            < construction["alias_minimum_redaction_length"]
            for source_alias in source_aliases
        ):
            pre_exclusion_counts["source_alias_too_short_for_safe_redaction"] += 1
            continue
        candidate_metadata[alias] = {
            "vendor_id": vendor_id,
            "fingerprint": fingerprint,
            "registry": registry_row,
        }

    candidate_fingerprints = {row["fingerprint"] for row in candidate_metadata.values()}
    known_key_ids = {
        fingerprint[-width:]
        for fingerprint in fingerprint_vendor_ids
        for width in (8, 16, 40)
    }
    aliases_to_redact = set(candidate_metadata)
    source_aliases_by_account: dict[str, set[str]] = {}
    for fingerprint in candidate_fingerprints:
        aliases_to_redact.update(fingerprint_aliases.get(fingerprint, set()))
    for alias, metadata in candidate_metadata.items():
        source_aliases_by_account[alias] = {
            alias,
            *fingerprint_aliases.get(metadata["fingerprint"], set()),
        }
    alias_pattern = compile_alias_pattern(
        aliases_to_redact, construction["alias_minimum_redaction_length"]
    )
    account_alias_patterns = {
        alias: compile_account_alias_pattern(
            aliases, construction["alias_minimum_redaction_length"]
        )
        for alias, aliases in source_aliases_by_account.items()
    }
    all_items, text_cleaning_audit = read_and_clean_agora(
        alias_pattern, account_alias_patterns
    )

    weak_fingerprints_by_alias: dict[str, set[str]] = defaultdict(set)
    for row in weak_all:
        if row["market_id"] != market_id:
            continue
        alias = normalize_alias(row["user_name"])
        metadata = candidate_metadata.get(alias)
        if metadata is None or metadata["vendor_id"] != row["vendor_id"]:
            continue
        fingerprint = normalize_fingerprint(row["fingerprint"])
        if alias and fingerprint:
            weak_fingerprints_by_alias[alias].add(fingerprint)

    strong_by_alias: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in strong_market:
        strong_by_alias[normalize_alias(row["user_name"])].append(row)

    strong_fingerprints_by_alias: dict[str, set[str]] = defaultdict(set)
    strong_key_aliases_by_alias: dict[str, set[str]] = defaultdict(set)
    for alias, rows in strong_by_alias.items():
        strong_fingerprints_by_alias[alias].update(
            fingerprint
            for fingerprint in (
                normalize_fingerprint(row["fingerprint"]) for row in rows
            )
            if fingerprint
        )
        strong_key_aliases_by_alias[alias].update(
            key_alias
            for key_alias in (
                normalize_alias(row.get("key_alias", "")) for row in rows
            )
            if key_alias
        )

    verified_strong: dict[str, dict] = {}
    strong_verification_exclusions: Counter[str] = Counter()
    for alias, rows in sorted(strong_by_alias.items()):
        identities = {
            (row["vendor_id"], normalize_fingerprint(row["fingerprint"])) for row in rows
        }
        if len(identities) != 1:
            strong_verification_exclusions["ambiguous_strong_alias"] += 1
            continue
        row = rows[0]
        metadata = candidate_metadata.get(alias)
        if metadata is None:
            strong_verification_exclusions["not_unique_pgp_account_candidate"] += 1
            continue
        if (
            metadata["vendor_id"] != row["vendor_id"]
            or metadata["fingerprint"] != normalize_fingerprint(row["fingerprint"])
        ):
            strong_verification_exclusions["strong_cross_source_mismatch"] += 1
            continue
        verified_strong[alias] = {
            "vendor_id": row["vendor_id"],
            "fingerprint": normalize_fingerprint(row["fingerprint"]),
            "key_aliases": set(strong_key_aliases_by_alias[alias]),
        }

    identity_graph_accounts: dict[str, dict] = {}
    for alias, metadata in sorted(candidate_metadata.items()):
        account_uid = stable_hash(
            construction["account_uid_namespace"], "agora", alias
        )
        identity_graph_accounts[account_uid] = {
            "aux_fingerprints": set(
                vendor_fingerprints.get(metadata["vendor_id"], set())
            ),
            "weak_fingerprints": set(weak_fingerprints_by_alias.get(alias, set())),
            "strong_fingerprints": set(
                strong_fingerprints_by_alias.get(alias, set())
            ),
            "strong_key_aliases": set(
                strong_key_aliases_by_alias.get(alias, set())
            ),
        }
    full_component_by_account = build_identity_conflict_components(
        identity_graph_accounts
    )

    accounts: dict[str, dict] = {}
    exclusion_counts: Counter[str] = pre_exclusion_counts.copy()
    invalid_strong_aliases = set(strong_by_alias) - set(verified_strong)
    for alias in sorted(candidate_metadata):
        if alias in invalid_strong_aliases:
            exclusion_counts["strong_evidence_conflict_or_ambiguity"] += 1
            continue
        metadata = candidate_metadata[alias]
        vendor_id = metadata["vendor_id"]
        fingerprint = metadata["fingerprint"]
        raw_items = all_items.get(alias, [])
        if not raw_items:
            exclusion_counts["no_agora_items"] += 1
            continue
        clean_items = [
            item
            for item in raw_items
            if item["title_clean"] or item["description_clean"]
        ]
        summary = account_summary(clean_items)
        if summary["item_count"] < construction["minimum_clean_items_per_account"]:
            exclusion_counts["too_few_clean_items"] += 1
            continue
        if summary["token_count"] < construction["minimum_clean_tokens_per_account"]:
            exclusion_counts["too_few_clean_tokens"] += 1
            continue
        account_uid = stable_hash(construction["account_uid_namespace"], "agora", alias)
        controller_uid = stable_hash(
            construction["controller_uid_namespace"], fingerprint
        )
        accounts[account_uid] = {
            "account_uid": account_uid,
            "controller_uid": controller_uid,
            "alias": alias,
            "alias_compact": compact_alias(alias),
            "vendor_id": vendor_id,
            "fingerprint": fingerprint,
            "strong_key_aliases": set(
                strong_key_aliases_by_alias.get(alias, set())
            ),
            "strong_fingerprints": set(
                strong_fingerprints_by_alias.get(alias, set())
            ),
            "positive_evidence_eligible": alias in verified_strong,
            "weak_fingerprints": weak_fingerprints_by_alias.get(alias, set()),
            "aux_fingerprints": vendor_fingerprints.get(vendor_id, set()),
            "items": clean_items,
            "summary": summary,
        }

    if not accounts:
        raise DatasetConstructionError("No eligible English accounts")
    accounts_by_controller: dict[str, list[str]] = defaultdict(list)
    accounts_by_conflict_component: dict[str, list[str]] = defaultdict(list)
    positive_accounts_by_controller: dict[str, list[str]] = defaultdict(list)
    for account_uid, account in accounts.items():
        account["conflict_component_uid"] = full_component_by_account[account_uid]
        accounts_by_controller[account["controller_uid"]].append(account_uid)
        accounts_by_conflict_component[account["conflict_component_uid"]].append(
            account_uid
        )
        if account["positive_evidence_eligible"]:
            positive_accounts_by_controller[account["controller_uid"]].append(account_uid)
    for values in accounts_by_controller.values():
        values.sort()
    for values in accounts_by_conflict_component.values():
        values.sort()
    for values in positive_accounts_by_controller.values():
        values.sort()

    fold_by_component = assign_controller_folds(
        accounts_by_conflict_component,
        {uid: account["summary"] for uid, account in accounts.items()},
        construction["fold_count"],
        construction["fold_seed"],
    )
    for account in accounts.values():
        account["fold_id"] = fold_by_component[account["conflict_component_uid"]]

    positive_pairs: list[dict] = []
    for controller_uid, account_uids in positive_accounts_by_controller.items():
        for left_uid, right_uid in itertools.combinations(account_uids, 2):
            left = accounts[left_uid]
            right = accounts[right_uid]
            positive_pairs.append(
                {
                    "left_uid": left_uid,
                    "right_uid": right_uid,
                    "fold_id": left["fold_id"],
                    "label": 1,
                    "evidence_class": "cross_source_verified_full_pgp_reuse_strong_silver",
                    "raw_pair_key": stable_hash("raw-pair", left_uid, right_uid),
                    "covariates": pair_covariates(left["summary"], right["summary"]),
                }
            )

    negative_candidates, negative_screen_counts, eligible_negative_candidate_count = (
        shortlist_negative_candidates(
            accounts,
            positive_pairs,
            construction["negative_shortlist_per_positive"],
            construction["ordering_seed"],
        )
    )

    selected_negatives = select_matched_negatives(
        positive_pairs,
        negative_candidates,
        construction["negative_per_positive"],
        construction["ordering_seed"],
    )
    pair_rows = positive_pairs + selected_negatives
    seen_pairs: dict[str, int] = {}
    for row in pair_rows:
        canonical = "|".join(sorted((row["left_uid"], row["right_uid"])))
        if canonical in seen_pairs and seen_pairs[canonical] != row["label"]:
            raise DatasetConstructionError("Conflicting pair labels")
        if canonical in seen_pairs:
            raise DatasetConstructionError("Duplicate pair")
        seen_pairs[canonical] = row["label"]
        row["pair_uid"] = stable_hash(construction["pair_uid_namespace"], canonical)
        if int(stable_hash(construction["ordering_seed"], canonical, "side")[:2], 16) % 2:
            row["left_uid"], row["right_uid"] = row["right_uid"], row["left_uid"]
        row["sort_key"] = stable_hash(
            construction["ordering_seed"], row["pair_uid"], "row"
        )
    pair_rows.sort(key=lambda row: row["sort_key"])

    all_smd = standardized_mean_differences(pair_rows)
    structural_smd = {
        name: all_smd[name] for name in STRUCTURAL_FEATURES
    }
    topic_smd = {
        name: all_smd[name] for name in TOPIC_DIAGNOSTIC_FEATURES
    }
    proxy = proxy_audit(
        pair_rows,
        construction["fold_count"],
        STRUCTURAL_FEATURES,
        "fixed logistic regression; identity-component-isolated five-fold out-of-fold; structural features only",
    )
    topic_proxy = proxy_audit(
        pair_rows,
        construction["fold_count"],
        TOPIC_DIAGNOSTIC_FEATURES,
        "diagnostic only: fixed logistic regression; identity-component-isolated five-fold out-of-fold; category and lexical overlap",
    )

    selected_account_uids = {
        uid for row in pair_rows for uid in (row["left_uid"], row["right_uid"])
    }
    public_items_full_clean: list[dict] = []
    public_items_style: list[dict] = []
    identity_residual_counts: Counter[str] = Counter()
    exact_owners: dict[tuple[str, str], set[str]] = defaultdict(set)
    for account_uid in sorted(selected_account_uids):
        account = accounts[account_uid]
        for item in account["items"]:
            for field in ("title_clean", "description_clean"):
                value = item[field]
                audit_aliases = set(source_aliases_by_account[account["alias"]])
                audit_aliases.update(account["strong_key_aliases"])
                alias_hits = source_alias_residuals(
                    value,
                    audit_aliases,
                    construction["alias_minimum_redaction_length"],
                )
                if alias_hits:
                    identity_residual_counts["source_known_alias"] += len(alias_hits)
                if PGP_ARMOR_RE.search(value) or HEX_ID_RE.search(value):
                    identity_residual_counts["pgp_or_long_hex"] += 1
                source_fingerprint_hits = source_fingerprint_residuals(
                    value, known_key_ids
                )
                if source_fingerprint_hits:
                    identity_residual_counts[
                        "source_known_fingerprint_or_key_id"
                    ] += len(source_fingerprint_hits)
                if URL_RE.search(value) or EMAIL_RE.search(value):
                    identity_residual_counts["url_or_email"] += 1
                if (
                    HANDLE_RE.search(value)
                    or PHONE_RE.search(value)
                    or CRYPTO_RE.search(value)
                ):
                    identity_residual_counts["contact_or_crypto"] += 1
            for field in ("title_clean", "description_clean"):
                key = exact_text_key(item[field])
                if key:
                    exact_owners[(field, key)].add(account_uid)
            item_uid = stable_hash(
                construction["item_uid_namespace"],
                account_uid,
                item["source_row_number"],
                item["title_clean"],
                item["description_clean"],
            )
            public_items_full_clean.append(
                {
                    "item_uid": item_uid,
                    "account_uid": account_uid,
                    "title_clean": item["title_clean"],
                    "description_clean": item["description_clean"],
                }
            )
            public_items_style.append(
                {
                    "item_uid": item_uid,
                    "account_uid": account_uid,
                    "title_style": style_projection(item["title_clean"]),
                    "description_style": style_projection(item["description_clean"]),
                }
            )
    cross_account_exact_values = sum(
        1 for owners in exact_owners.values() if len(owners) > 1
    )
    component_fold_sets: dict[str, set[int]] = defaultdict(set)
    for account in accounts.values():
        if account["account_uid"] in selected_account_uids:
            component_fold_sets[account["conflict_component_uid"]].add(
                account["fold_id"]
            )
    controller_fold_overlaps = sum(
        1 for folds in component_fold_sets.values() if len(folds) > 1
    )

    gates = policy["quality_gates"]
    positive_count = sum(row["label"] == 1 for row in pair_rows)
    negative_count = sum(row["label"] == 0 for row in pair_rows)
    positive_controllers = len(
        {
            accounts[row["left_uid"]]["controller_uid"]
            for row in pair_rows
            if row["label"] == 1
        }
    )
    gate_results = {
        "minimum_positive_pairs": positive_count >= gates["minimum_positive_pairs"],
        "minimum_positive_controller_groups": positive_controllers
        >= gates["minimum_positive_controller_groups"],
        "identity_residual_hits": sum(identity_residual_counts.values())
        <= gates["maximum_identity_residual_hits"],
        "cross_account_exact_text_values": cross_account_exact_values
        <= gates["maximum_cross_account_exact_text_values"],
        "identity_component_fold_overlaps": controller_fold_overlaps
        <= gates["maximum_identity_component_fold_overlaps"],
        "duplicate_or_conflicting_pairs": len(seen_pairs) == len(pair_rows),
        "structural_standardized_mean_difference": max(
            abs(value) for value in structural_smd.values()
        )
        <= gates["maximum_absolute_structural_standardized_mean_difference"],
        "structural_proxy_roc_auc": proxy["bidirectional_roc_auc"]
        <= gates["maximum_structural_proxy_roc_auc"],
        "structural_proxy_average_precision_lift": proxy[
            "average_precision_lift_over_prevalence"
        ]
        <= gates["maximum_structural_proxy_ap_lift_over_prevalence"],
        "negative_ratio": negative_count
        == positive_count * construction["negative_per_positive"],
    }
    quality_status = "PASSED" if all(gate_results.values()) else "FAILED"

    output_directory = ROOT / policy["output_directory"]
    building_directory = output_directory.with_name(output_directory.name + ".building")
    if output_directory.exists():
        raise DatasetConstructionError(f"Publication directory already exists: {output_directory}")
    if building_directory.exists():
        shutil.rmtree(building_directory)
    building_directory.mkdir(parents=True)
    try:
        write_csv(
            building_directory / "public_pairs.csv",
            ("pair_uid", "fold_id", "account_left_uid", "account_right_uid"),
            (
                {
                    "pair_uid": row["pair_uid"],
                    "fold_id": row["fold_id"],
                    "account_left_uid": row["left_uid"],
                    "account_right_uid": row["right_uid"],
                }
                for row in pair_rows
            ),
        )
        with (building_directory / "public_items_full_clean.jsonl").open(
            "w", encoding="utf-8", newline=""
        ) as handle:
            for row in sorted(
                public_items_full_clean, key=lambda value: value["item_uid"]
            ):
                handle.write(
                    json.dumps(row, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                    + "\n"
                )
        with (building_directory / "public_items_style_projection.jsonl").open(
            "w", encoding="utf-8", newline=""
        ) as handle:
            for row in sorted(public_items_style, key=lambda value: value["item_uid"]):
                handle.write(
                    json.dumps(row, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                    + "\n"
                )
        write_csv(
            building_directory / "labels.csv",
            ("pair_uid", "label"),
            (
                {
                    "pair_uid": row["pair_uid"],
                    "label": row["label"],
                }
                for row in pair_rows
            ),
        )
        quality_audit = {
            "status": quality_status,
            "gate_results": gate_results,
            "counts": {
                "strong_rows_in_market": len(strong_market),
                "unique_pgp_account_candidates_before_text_filter": len(candidate_metadata),
                "full_identity_graph_nodes": len(identity_graph_accounts),
                "full_identity_conflict_components": len(
                    set(full_component_by_account.values())
                ),
                "identity_bridge_nodes_not_text_eligible": len(identity_graph_accounts)
                - len(accounts),
                "eligible_accounts_before_pair_selection": len(accounts),
                "published_accounts": len(selected_account_uids),
                "published_items_per_view": len(public_items_full_clean),
                "positive_pairs": positive_count,
                "negative_pairs": negative_count,
                "positive_controller_groups": positive_controllers,
                "all_published_identity_conflict_components": len(component_fold_sets),
                "eligible_negative_candidates_before_shortlist": eligible_negative_candidate_count,
                "shortlisted_negative_candidates": len(negative_candidates),
            },
            "exclusion_counts": dict(sorted(exclusion_counts.items())),
            "strong_verification_exclusions": dict(
                sorted(strong_verification_exclusions.items())
            ),
            "negative_screen_counts": dict(sorted(negative_screen_counts.items())),
            "text_cleaning": text_cleaning_audit,
            "sql_parse_coverage": {
                "vendor_registry": vendor_sql_parse_audit,
                "auxiliary_pgp": pgp_sql_parse_audit,
            },
            "identity_residual_counts": dict(sorted(identity_residual_counts.items())),
            "cross_account_exact_text_values": cross_account_exact_values,
            "identity_component_fold_overlaps": controller_fold_overlaps,
            "structural_standardized_mean_differences": structural_smd,
            "maximum_absolute_structural_standardized_mean_difference": max(
                abs(value) for value in structural_smd.values()
            ),
            "structural_proxy": proxy,
            "topic_lexical_diagnostics": {
                "hard_gate": False,
                "reason": "Category and lexical overlap mix product topic with potentially intended author vocabulary; causal attribution is handled by preregistered model controls.",
                "standardized_mean_differences": topic_smd,
                "proxy": topic_proxy,
            },
            "fold_counts": {
                str(fold): {
                    "pairs": sum(row["fold_id"] == fold for row in pair_rows),
                    "positives": sum(
                        row["fold_id"] == fold and row["label"] == 1
                        for row in pair_rows
                    ),
                    "negatives": sum(
                        row["fold_id"] == fold and row["label"] == 0
                        for row in pair_rows
                    ),
                    "accounts": sum(
                        account["fold_id"] == fold
                        and account["account_uid"] in selected_account_uids
                        for account in accounts.values()
                    ),
                }
                for fold in range(construction["fold_count"])
            },
            "limitations": [
                "Positive labels are cross-source-verified strong-silver PGP reuse, not manual controller ground truth.",
                "Negative labels are screened contrast pairs, not proof that key rotation or undocumented control sharing is absent.",
                "This source dataset is small and cannot independently certify universal English performance.",
            ],
            "label_semantics": policy["label_semantics"],
        }
        write_json(building_directory / "quality_audit.json", quality_audit)
        file_records = []
        for path in sorted(building_directory.iterdir(), key=lambda value: value.name):
            file_records.append(
                {
                    "path": path.name,
                    "size_bytes": path.stat().st_size,
                    "sha256": file_sha256(path),
                }
            )
        manifest = {
            "version": policy["version"],
            "status": quality_status,
            "policy_path": POLICY_PATH.relative_to(ROOT).as_posix(),
            "policy_sha256": file_sha256(POLICY_PATH),
            "builder_path": Path(__file__).resolve().relative_to(ROOT).as_posix(),
            "builder_sha256": file_sha256(Path(__file__).resolve()),
            "input_sha256": policy["inputs"],
            "runtime": {
                "python": sys.version.split()[0],
                "implementation": platform.python_implementation(),
                "numpy": package_version("numpy"),
                "scipy": package_version("scipy"),
                "scikit_learn": package_version("scikit-learn"),
                "openpyxl": package_version("openpyxl"),
            },
            "training_projection": {
                "join_metadata_only": [
                    "pair_uid",
                    "account_left_uid",
                    "account_right_uid",
                    "account_uid",
                    "item_uid",
                    "fold_id"
                ],
                "label_field": "label",
                "full_clean_model_fields": ["title_clean", "description_clean"],
                "style_projection_model_fields": ["title_style", "description_style"],
                "forbidden_fields": [
                    "controller_uid",
                    "conflict_component_uid",
                    "evidence_class",
                    "clean_item_count",
                    "clean_token_count",
                    "source_row_number",
                    "category_clean"
                ]
            },
            "counts": quality_audit["counts"],
            "files": file_records,
        }
        write_json(building_directory / "manifest.json", manifest)
        if quality_status != "PASSED":
            raise DatasetConstructionError(
                "English source dataset failed quality gates: "
                + ", ".join(name for name, passed in gate_results.items() if not passed)
                + "; max_abs_smd="
                + f"{quality_audit['maximum_absolute_structural_standardized_mean_difference']:.6f}"
                + "; structural_auc="
                + f"{proxy['bidirectional_roc_auc']:.6f}"
                + "; structural_ap_lift="
                + f"{proxy['average_precision_lift_over_prevalence']:.6f}"
                + "; smd="
                + json.dumps(structural_smd, sort_keys=True)
            )
        building_directory.replace(output_directory)
        return manifest
    except Exception:
        if building_directory.exists():
            shutil.rmtree(building_directory)
        raise


def main() -> None:
    manifest = build()
    print(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
