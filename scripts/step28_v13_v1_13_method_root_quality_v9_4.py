#!/usr/bin/env python3
"""Audit the V9.4 method root before any M0/M1/M2/M3 training."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Mapping, Sequence

import numpy as np
from scipy import sparse
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import average_precision_score, roc_auc_score
from sklearn.preprocessing import StandardScaler

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import step28_v13_v1_13_balanced_world_schedule_v9_4 as schedule_v94
import step28_v13_v1_13_joint_noise_signatures_v9_4 as noise_v94
import step28_v13_v1_13_method_root_builder_v9_4 as builder
import step28_v13_v1_13_model_visible_prebuild_source_v9_4 as prebuild_v94
import step28_v13_v1_13_model_visible_public_replay_v9_4 as replay_v94
import step28_v13_v1_13_style_derangement as style_derangement


VERSION = "2026-08-28-step28-v13-v1-13-method-root-quality-v9-4-v1"
AP_BASELINE = 20.0 / 378.0
TEXT_AP_BASELINE = 20.0 / 372.0
COUNTERFACTUAL_MAX_AUC = 0.53
COUNTERFACTUAL_MAX_AP_UPLIFT = 0.01
IDENTITY_MIN_AUC = 0.65
IDENTITY_MIN_AP = 0.15
VISIBLE_ARTIFICIAL_CODE = re.compile(r"Q[A-P]{10}")


class MethodRootQualityError(ValueError):
    pass


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(json.dumps(
        value, ensure_ascii=False, separators=(",", ":"), sort_keys=True
    ).encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as stream:
        value = json.load(stream)
    if not isinstance(value, dict):
        raise MethodRootQualityError(f"JSON object required: {path}")
    return value


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as stream:
        return [json.loads(line) for line in stream if line.strip()]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=False)
    with path.open("w", encoding="utf-8", newline="\n") as stream:
        json.dump(value, stream, ensure_ascii=False, indent=2, sort_keys=True)
        stream.write("\n")


def _formal_key(name: str, policy: Mapping[str, Any]) -> bytes:
    auth = read_json(ROOT / str(policy["formal_authorization_path"]))
    spec = auth["key_files"][name]
    path = ROOT / str(spec["path"])
    data = path.read_bytes() if path.is_file() else b""
    if len(data) != 32 or hashlib.sha256(data).hexdigest() != spec["commitment_sha256"]:
        raise MethodRootQualityError(f"Quality authority drift: {name}")
    return data


def _time_key(policy: Mapping[str, Any], smoke: bool) -> bytes:
    if smoke:
        return builder.smoke_authorities().time
    path = builder.TIME_KEY_PATH
    data = path.read_bytes() if path.is_file() else b""
    if (
        len(data) != 32
        or hashlib.sha256(data).hexdigest()
        != "b99fe117617313ec2cda0228d8d40d56ccea8f63891425fe5b2332dc5b338c82"
    ):
        raise MethodRootQualityError("Retained time key drift during quality audit")
    return data


def verify_manifests(public_root: Path, private_root: Path) -> dict[str, Any]:
    manifest = read_json(public_root / "root_manifest.json")
    claimed = manifest.get("canonical_self_hash")
    payload = dict(manifest)
    payload.pop("canonical_self_hash", None)
    if claimed != canonical_sha256(payload):
        raise MethodRootQualityError("Root manifest self-hash drift")
    for record in manifest["public_files"]:
        path = public_root / record["path"]
        if (
            not path.is_file()
            or path.stat().st_size != record["size_bytes"]
            or sha256_file(path) != record["sha256"]
        ):
            raise MethodRootQualityError(f"Public file manifest drift: {record['path']}")
    for record in manifest["private_file_commitments"]:
        path = private_root / record["path"]
        if (
            not path.is_file()
            or path.stat().st_size != record["size_bytes"]
            or sha256_file(path) != record["sha256"]
        ):
            raise MethodRootQualityError(f"Private file commitment drift: {record['path']}")
    return manifest


def public_surface_audit(
    public_root: Path, manifest: Mapping[str, Any]
) -> tuple[dict[str, Any], dict[str, dict[str, list[dict[str, Any]]]]]:
    expected_counts = manifest["world_counts"]
    rows: dict[str, dict[str, list[dict[str, Any]]]] = {}
    all_item_documents: set[str] = set()
    all_profile_documents: set[str] = set()
    all_item_uids: set[str] = set()
    all_seller_uids: set[str] = set()
    forbidden_hits = 0
    artificial_codes = 0
    exact_item_collisions = 0
    exact_profile_collisions = 0
    totals = Counter()
    for split in builder.SPLITS:
        observed = public_root / split / "observed"
        worlds = read_jsonl(observed / "worlds.jsonl")
        sellers = read_jsonl(observed / "sellers.jsonl")
        items = read_jsonl(observed / "items.jsonl")
        replay_items = read_jsonl(observed / "model_visible_replay_items.jsonl")
        redacted = read_jsonl(observed / "redacted_items.jsonl")
        profiles = read_jsonl(observed / "model_seller_profiles.jsonl")
        endpoints = read_csv(observed / "complete_model_pair_endpoints.csv")
        identity33 = read_csv(observed / "identity33_all_pairs.csv")
        if (
            len(worlds) != int(expected_counts[split])
            or len(sellers) != 28 * len(worlds)
            or len(endpoints) != 378 * len(worlds)
            or len(identity33) != len(endpoints)
            or len(replay_items) != len(items)
            or len(redacted) != len(items)
            or len(profiles) != len(sellers)
        ):
            raise MethodRootQualityError(f"Public cardinality drift: {split}")
        item_index = {row["item_uid"]: row for row in items}
        redacted_index = {row["item_uid"]: row for row in redacted}
        replay_index = {row["item_uid"]: row for row in replay_items}
        if set(item_index) != set(redacted_index) or set(item_index) != set(replay_index):
            raise MethodRootQualityError(f"Item projection keyset drift: {split}")
        if all_item_uids.intersection(item_index):
            raise MethodRootQualityError("Cross-split item UID reuse")
        all_item_uids.update(item_index)
        split_sellers = {row["seller_uid"] for row in sellers}
        if all_seller_uids.intersection(split_sellers) or len(split_sellers) != len(sellers):
            raise MethodRootQualityError("Cross-split seller UID reuse")
        all_seller_uids.update(split_sellers)
        for item_uid, item in item_index.items():
            clean = redacted_index[item_uid]
            replay = replay_index[item_uid]
            visible = str(item["title"]) + "\n" + str(item["description"])
            artificial_codes += int(VISIBLE_ARTIFICIAL_CODE.search(visible) is not None)
            forbidden_hits += int(builder.FORBIDDEN_VISIBLE_RE.search(visible) is not None)
            if item["title"] != clean["title"] or replay["title"] != item["title"]:
                raise MethodRootQualityError("Title projection drift")
            if not str(item["description"]).startswith(str(clean["description"])):
                raise MethodRootQualityError("Redacted description is not the raw prefix")
            document_hash = hashlib.sha256(
                (str(clean["title"]) + "\0" + str(clean["description"])).encode("utf-8")
            ).hexdigest()
            exact_item_collisions += int(document_hash in all_item_documents)
            all_item_documents.add(document_hash)
        for profile in profiles:
            document_hash = hashlib.sha256(json.dumps(
                profile, ensure_ascii=False, separators=(",", ":"), sort_keys=True
            ).encode("utf-8")).hexdigest()
            exact_profile_collisions += int(document_hash in all_profile_documents)
            all_profile_documents.add(document_hash)
        totals.update({
            "worlds": len(worlds), "sellers": len(sellers),
            "items": len(items), "pairs": len(endpoints),
        })
        rows[split] = {
            "worlds": worlds, "sellers": sellers, "items": items,
            "replay_items": replay_items, "redacted": redacted,
            "profiles": profiles, "endpoints": endpoints, "identity33": identity33,
        }
    return {
        "totals": dict(totals),
        "artificial_item_code_count": artificial_codes,
        "forbidden_internal_marker_count": forbidden_hits,
        "exact_redacted_item_document_collision_count": exact_item_collisions,
        "exact_model_profile_collision_count": exact_profile_collisions,
        "split_isolation_passed": True,
    }, rows


def source_copy_audit(rows: Mapping[str, Mapping[str, Sequence[Mapping[str, Any]]]]) -> dict[str, Any]:
    from openpyxl import load_workbook

    def clean(value: object) -> str:
        if value is None:
            return ""
        return " ".join(unicodedata.normalize("NFKC", str(value)).split())

    source_values: set[str] = set()
    workbook = load_workbook(ROOT / "market_item.xlsx", read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for value in (row[2], row[3]):
                normalized = clean(value)
                if normalized:
                    source_values.add(normalized)
    finally:
        workbook.close()
    copied = []
    for split in builder.SPLITS:
        for row in rows[split]["redacted"]:
            for field in ("title", "description"):
                value = clean(row[field])
                if value and value in source_values:
                    copied.append((split, row["item_uid"], field))
    return {
        "real_source_field_count": len(source_values),
        "exact_real_source_text_copy_count": len(copied),
        "first_copy": list(copied[0]) if copied else None,
    }


def exact_v94_replay_audit(
    rows: Mapping[str, Mapping[str, Sequence[Mapping[str, Any]]]], time_key: bytes,
    smoke: bool,
) -> dict[str, Any]:
    signatures = [dict(row) for row in noise_v94.build_noise_signatures().rows]
    replayed = 0
    split_commitments: dict[str, str] = {}
    for split in ("train", "development"):
        if smoke:
            schedules = [builder._smoke_world(split)]
            public_worlds = [{
                "split": split, "world_ordinal": 0,
                "world_uid": schedules[0].world_uid,
                "seller_uids": list(schedules[0].seller_uids),
                "noise_slot_by_seller_slot": list(schedules[0].noise_slots),
            }]
            split_commitments[split] = "smoke"
        else:
            schedule = schedule_v94.build_split_schedule(split)
            public_worlds = [{
                "split": split, "world_ordinal": int(world["world_ordinal"]),
                "world_uid": str(world["world_uid"]),
                "seller_uids": list(world["seller_uids"]),
                "noise_slot_by_seller_slot": list(world["noise_slot_by_seller_slot"]),
            } for world in schedule.public_worlds]
            split_commitments[split] = str(
                schedule.commitment["split_schedule_commitment_sha256"]
            )
        endpoints_by_world: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        items_by_world: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for endpoint in rows[split]["endpoints"]:
            endpoints_by_world[str(endpoint["world_uid"])].append({
                "world_uid": endpoint["world_uid"],
                "canonical_pair_uid": endpoint["canonical_pair_uid"],
                "seller_uid_left": endpoint["seller_uid_left"],
                "seller_uid_right": endpoint["seller_uid_right"],
            })
        for item in rows[split]["replay_items"]:
            items_by_world[str(item["world_uid"])].append(dict(item))
        for public in public_worlds:
            registered = prebuild_v94.build_truth_free_world_projection(
                world=public, noise_signatures=signatures, time_key_hex=time_key.hex()
            )
            replay_v94.require_exact_replay(
                registered_rows=registered,
                public_endpoint_rows=endpoints_by_world[public["world_uid"]],
                public_item_rows=items_by_world[public["world_uid"]],
            )
            replayed += 1
    return {
        "world_count": replayed,
        "split_schedule_commitments": split_commitments,
        "exact_public_14d_replay_passed": True,
    }


def _labels_and_matrix(
    public_root: Path, private_root: Path, split: str,
    rows: Mapping[str, Mapping[str, Sequence[Mapping[str, Any]]]],
) -> tuple[np.ndarray, np.ndarray]:
    labels = read_csv(private_root / split / "pair_labels.csv")
    identity = rows[split]["identity33"]
    label_index = {
        (row["world_uid"], row["canonical_pair_uid"]): int(row["label"])
        for row in labels
    }
    keys = [(row["world_uid"], row["canonical_pair_uid"]) for row in identity]
    if set(keys) != set(label_index) or len(keys) != len(label_index):
        raise MethodRootQualityError(f"Identity/label join drift: {split}")
    names = [name for name in identity[0] if name not in {"canonical_pair_uid", "world_uid"}]
    matrix = np.asarray(
        [[float(row[name]) for name in names] for row in identity], dtype=np.float64
    )
    return matrix, np.asarray([label_index[key] for key in keys], dtype=np.int8)


def identity_positive_control(
    public_root: Path, private_root: Path,
    rows: Mapping[str, Mapping[str, Sequence[Mapping[str, Any]]]],
) -> dict[str, Any]:
    train_x, train_y = _labels_and_matrix(public_root, private_root, "train", rows)
    dev_x, dev_y = _labels_and_matrix(public_root, private_root, "development", rows)
    scaler = StandardScaler().fit(train_x)
    model = LogisticRegression(
        C=1.0, penalty="l2", solver="lbfgs", max_iter=10000,
        tol=1e-10, random_state=281320828,
    ).fit(scaler.transform(train_x), train_y)
    score = model.predict_proba(scaler.transform(dev_x))[:, 1]
    auc = float(roc_auc_score(dev_y, score))
    ap = float(average_precision_score(dev_y, score))
    return {
        "model": "fixed_l2_logistic_identity33_train_to_development",
        "development_roc_auc": auc,
        "development_average_precision": ap,
        "minimum_roc_auc": IDENTITY_MIN_AUC,
        "minimum_average_precision": IDENTITY_MIN_AP,
        "passed": auc >= IDENTITY_MIN_AUC and ap >= IDENTITY_MIN_AP,
        "score_sha256": hashlib.sha256(np.asarray(score, dtype="<f8").tobytes()).hexdigest(),
    }


def _seller_documents(
    split_rows: Mapping[str, Sequence[Mapping[str, Any]]]
) -> tuple[list[str], list[str]]:
    by_seller: defaultdict[str, list[tuple[str, str]]] = defaultdict(list)
    for row in split_rows["redacted"]:
        by_seller[str(row["seller_uid"])].append((str(row["title"]), str(row["description"])))
    sellers = sorted(by_seller, key=lambda value: value.encode("utf-8"))
    documents = [
        "\n".join("\n".join(pair) for pair in by_seller[seller])
        for seller in sellers
    ]
    return sellers, documents


def _pair_cosines(
    matrix: sparse.csr_matrix, sellers: Sequence[str],
    endpoints: Sequence[Mapping[str, Any]],
) -> np.ndarray:
    index = {seller: position for position, seller in enumerate(sellers)}
    left = np.asarray([index[str(row["seller_uid_left"])] for row in endpoints])
    right = np.asarray([index[str(row["seller_uid_right"])] for row in endpoints])
    return np.asarray(matrix[left].multiply(matrix[right]).sum(axis=1)).ravel()


def _text_exclusions(private_root: Path, split: str) -> set[tuple[str, str]]:
    excluded: set[tuple[str, str]] = set()
    for audit in read_jsonl(private_root / split / "generation_audit.jsonl"):
        world_uid = str(audit["world_uid"])
        for row in audit["registered_negative_controls"]:
            excluded.add((world_uid, str(row["canonical_pair_uid"])))
    return excluded


def _fit_text_probe(
    *, train_sellers: Sequence[str], train_docs: Sequence[str],
    dev_sellers: Sequence[str], dev_docs: Sequence[str],
    rows: Mapping[str, Mapping[str, Sequence[Mapping[str, Any]]]],
    private_root: Path, view: str,
) -> dict[str, Any]:
    vectorizer = TfidfVectorizer(
        analyzer="char", ngram_range=(3, 3), lowercase=True,
        min_df=3, max_features=30000, norm="l2", dtype=np.float32,
    )
    train_matrix = vectorizer.fit_transform(train_docs).tocsr()
    dev_matrix = vectorizer.transform(dev_docs).tocsr()
    train_excluded = _text_exclusions(private_root, "train")
    dev_excluded = _text_exclusions(private_root, "development")
    train_keep = [
        index for index, row in enumerate(rows["train"]["endpoints"])
        if (str(row["world_uid"]), str(row["canonical_pair_uid"])) not in train_excluded
    ]
    dev_keep = [
        index for index, row in enumerate(rows["development"]["endpoints"])
        if (str(row["world_uid"]), str(row["canonical_pair_uid"])) not in dev_excluded
    ]
    train_all = _pair_cosines(train_matrix, train_sellers, rows["train"]["endpoints"])
    dev_all = _pair_cosines(dev_matrix, dev_sellers, rows["development"]["endpoints"])
    train_score = train_all[np.asarray(train_keep)]
    dev_score = dev_all[np.asarray(dev_keep)]
    train_labels = read_csv(private_root / "train" / "pair_labels.csv")
    dev_labels = read_csv(private_root / "development" / "pair_labels.csv")
    train_y = np.asarray([int(train_labels[index]["label"]) for index in train_keep], dtype=np.int8)
    dev_y = np.asarray([int(dev_labels[index]["label"]) for index in dev_keep], dtype=np.int8)
    calibrator = LogisticRegression(
        C=1.0, solver="lbfgs", max_iter=10000, tol=1e-10,
        random_state=281320829,
    ).fit(train_score.reshape(-1, 1), train_y)
    predicted = calibrator.predict_proba(dev_score.reshape(-1, 1))[:, 1]
    auc = float(roc_auc_score(dev_y, predicted))
    ap = float(average_precision_score(dev_y, predicted))
    return {
        "view": view,
        "vocabulary_size": len(vectorizer.vocabulary_),
        "train_pair_count": len(train_keep),
        "development_pair_count": len(dev_keep),
        "development_roc_auc": auc,
        "development_symmetric_roc_auc": max(auc, 1.0 - auc),
        "development_average_precision": ap,
        "development_average_precision_uplift": ap - TEXT_AP_BASELINE,
        "score_sha256": hashlib.sha256(np.asarray(predicted, dtype="<f8").tobytes()).hexdigest(),
    }


def text_probe(
    rows: Mapping[str, Mapping[str, Sequence[Mapping[str, Any]]]],
    private_root: Path,
) -> dict[str, Any]:
    train_sellers, train_docs = _seller_documents(rows["train"])
    dev_sellers, dev_docs = _seller_documents(rows["development"])
    return _fit_text_probe(
        train_sellers=train_sellers, train_docs=train_docs,
        dev_sellers=dev_sellers, dev_docs=dev_docs,
        rows=rows, private_root=private_root,
        view="original_redacted_full_seller_text_character_trigram_cosine",
    )


def counterfactual_text_probe(
    *, rows: Mapping[str, Mapping[str, Sequence[Mapping[str, Any]]]],
    private_root: Path, smoke: bool, text_key: bytes,
) -> dict[str, Any]:
    template = read_json(builder.TEMPLATE_PATH)
    docs_by_split: dict[str, tuple[list[str], list[str]]] = {}
    replay_exact = True
    derangement_hashes: list[tuple[str, str]] = []
    for split in ("train", "development"):
        if smoke:
            worlds = [builder._smoke_world(split)]
        else:
            schedule = schedule_v94.build_split_schedule(split)
            worlds = [
                builder.PublicWorld(
                    split=split,
                    ordinal=int(public["world_ordinal"]),
                    world_uid=str(public["world_uid"]),
                    seller_uids=tuple(public["seller_uids"]),
                    noise_slots=tuple(public["noise_slot_by_seller_slot"]),
                    controller_groups=tuple(tuple(group) for group in groups),
                )
                for public, groups in zip(
                    schedule.public_worlds,
                    schedule.controller_groups_by_world,
                    strict=True,
                )
            ]
        audit_by_world = {
            str(row["world_uid"]): row
            for row in read_jsonl(private_root / split / "generation_audit.jsonl")
        }
        redacted_by_item = {
            str(row["item_uid"]): row for row in rows[split]["redacted"]
        }
        replay_by_world: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows[split]["replay_items"]:
            replay_by_world[str(row["world_uid"])].append(dict(row))
        counterfactual_by_seller: defaultdict[str, list[tuple[int, str, str]]] = defaultdict(list)
        for world in worlds:
            audit = audit_by_world.get(world.world_uid)
            if audit is None:
                raise MethodRootQualityError("Counterfactual generation audit is missing")
            styles = {
                str(row["seller_uid"]): row for row in audit["style_assignments"]
            }
            derangement = style_derangement.build_style_source_derangement(
                split=split, world_uid=world.world_uid, seller_uids=world.seller_uids
            )
            style_source = derangement.as_mapping()
            derangement_hashes.append((world.world_uid, derangement.mapping_sha256))
            noise_by_seller = {
                seller: world.noise_slots[index]
                for index, seller in enumerate(world.seller_uids)
            }
            original_render: dict[str, tuple[str, str]] = {}
            counterfactual_render: dict[str, tuple[str, str]] = {}
            item_owner: dict[str, tuple[str, int]] = {}
            for persisted in replay_by_world[world.world_uid]:
                seller = str(persisted["seller_uid"])
                ordinal = int(persisted["logical_item_ordinal"])
                item_uid = str(persisted["item_uid"])
                clean = redacted_by_item[item_uid]
                original_title, original_description, _category, _components = builder.render_base_item(
                    world=world, seller_uid=seller,
                    noise_slot=noise_by_seller[seller], ordinal=ordinal,
                    title_nonempty=bool(clean["title"]),
                    description_nonempty=bool(clean["description"]),
                    style=styles[seller], template=template, key=text_key,
                )
                cf_title, cf_description, _category, _components = builder.render_base_item(
                    world=world, seller_uid=seller,
                    noise_slot=noise_by_seller[seller], ordinal=ordinal,
                    title_nonempty=bool(clean["title"]),
                    description_nonempty=bool(clean["description"]),
                    style=styles[style_source[seller]], template=template, key=text_key,
                )
                original_render[item_uid] = (original_title, original_description)
                counterfactual_render[item_uid] = (cf_title, cf_description)
                item_owner[item_uid] = (seller, ordinal)
            for control in audit["registered_negative_controls"]:
                source = str(control["source_item_uid"])
                target = str(control["target_item_uid"])
                suffix = "" if control["control_type"] == "exact_title_clone_negative" else " 配置另选"
                original_render[target] = (
                    original_render[source][0] + suffix,
                    original_render[target][1],
                )
                counterfactual_render[target] = (
                    counterfactual_render[source][0] + suffix,
                    counterfactual_render[target][1],
                )
            for item_uid, (title, description) in original_render.items():
                clean = redacted_by_item[item_uid]
                if title != clean["title"] or description != clean["description"]:
                    replay_exact = False
                    raise MethodRootQualityError(
                        "Original redacted text cannot be independently re-rendered"
                    )
            for item_uid, (title, description) in counterfactual_render.items():
                seller, ordinal = item_owner[item_uid]
                counterfactual_by_seller[seller].append((ordinal, title, description))
        sellers = sorted(counterfactual_by_seller, key=lambda value: value.encode("utf-8"))
        documents = [
            "\n".join(
                title + "\n" + description
                for _ordinal, title, description in sorted(counterfactual_by_seller[seller])
            )
            for seller in sellers
        ]
        docs_by_split[split] = (sellers, documents)
    probe = _fit_text_probe(
        train_sellers=docs_by_split["train"][0],
        train_docs=docs_by_split["train"][1],
        dev_sellers=docs_by_split["development"][0],
        dev_docs=docs_by_split["development"][1],
        rows=rows, private_root=private_root,
        view="controller_blind_style_deranged_redacted_text_character_trigram_cosine",
    )
    probe.update({
        "original_renderer_exact_replay": replay_exact,
        "derangement_world_count": len(derangement_hashes),
        "derangement_mapping_commitment_sha256": canonical_sha256(derangement_hashes),
        "maximum_symmetric_roc_auc": COUNTERFACTUAL_MAX_AUC,
        "maximum_average_precision_uplift": COUNTERFACTUAL_MAX_AP_UPLIFT,
    })
    probe["passed"] = (
        probe["development_symmetric_roc_auc"] <= COUNTERFACTUAL_MAX_AUC
        and probe["development_average_precision_uplift"]
        <= COUNTERFACTUAL_MAX_AP_UPLIFT
        and replay_exact
    )
    return probe


def registered_private_audit(private_root: Path, world_counts: Mapping[str, int]) -> dict[str, Any]:
    mechanism_counts: Counter[str] = Counter()
    identity_types: set[str] = set()
    value_hashes: set[str] = set()
    controls = Counter()
    active_worlds = 0
    parser_rows = 0
    for split in ("train", "development"):
        audits = read_jsonl(private_root / split / "generation_audit.jsonl")
        plans = read_jsonl(private_root / split / "identity_plan.jsonl")
        parser = read_jsonl(private_root / split / "parsed_identity_occurrences.jsonl")
        if len(audits) != int(world_counts[split]):
            raise MethodRootQualityError("Generation audit world count drift")
        parser_rows += len(parser)
        for audit in audits:
            mechanism_counts.update(
                row["mechanism"] for row in audit["mechanism_assignments"]
            )
            controls.update(
                row["control_type"] for row in audit["registered_negative_controls"]
            )
            active_worlds += int(audit["identity33_active_pair_count"] > 0)
        for plan in plans:
            identity_types.add(str(plan["identity_type"]))
            value_hash = str(plan["value_sha256"])
            if value_hash in value_hashes:
                raise MethodRootQualityError("Cross-world identity value reuse")
            value_hashes.add(value_hash)
    expected_worlds = int(world_counts["train"]) + int(world_counts["development"])
    return {
        "train_development_world_count": expected_worlds,
        "identity33_active_world_count": active_worlds,
        "identity_types": sorted(identity_types),
        "unique_identity_value_hash_count": len(value_hashes),
        "mechanism_counts": dict(sorted(mechanism_counts.items())),
        "registered_negative_control_counts": dict(sorted(controls.items())),
        "production_parser_row_count": parser_rows,
        "all_eight_identity_types_present": identity_types == set(builder.IDENTITY_TYPES),
        "all_worlds_identity33_non_degenerate": active_worlds == expected_worlds,
        "two_clone_and_four_semantic_controls_per_world": (
            controls["exact_title_clone_negative"] == 2 * expected_worlds
            and controls["high_semantic_similarity_negative"] == 4 * expected_worlds
        ),
    }


def audit(*, public_root: Path, private_root: Path, output_root: Path, smoke: bool) -> dict[str, Any]:
    if output_root.exists():
        raise MethodRootQualityError("Quality output path already exists")
    manifest = verify_manifests(public_root, private_root)
    surface, rows = public_surface_audit(public_root, manifest)
    source_copy = source_copy_audit(rows)
    replay = exact_v94_replay_audit(rows, _time_key(read_json(builder.POLICY_PATH), smoke), smoke)
    identity = identity_positive_control(public_root, private_root, rows)
    text_original = text_probe(rows, private_root)
    policy = read_json(builder.POLICY_PATH)
    text_key = builder.smoke_authorities().text if smoke else _formal_key("text", policy)
    text_counterfactual = counterfactual_text_probe(
        rows=rows, private_root=private_root, smoke=smoke, text_key=text_key
    )
    private = registered_private_audit(private_root, manifest["world_counts"])
    hard_gates = {
        "manifest_and_cardinality": True,
        "split_isolation": surface["split_isolation_passed"],
        "artificial_item_code_zero": surface["artificial_item_code_count"] == 0,
        "forbidden_internal_marker_zero": surface["forbidden_internal_marker_count"] == 0,
        "redacted_item_document_collision_zero": surface[
            "exact_redacted_item_document_collision_count"
        ] == 0,
        "model_profile_collision_zero": surface["exact_model_profile_collision_count"] == 0,
        "real_source_text_copy_zero": source_copy["exact_real_source_text_copy_count"] == 0,
        "exact_v9_4_14d_replay": replay["exact_public_14d_replay_passed"],
        "identity_positive_control": identity["passed"],
        "all_eight_identity_types": private["all_eight_identity_types_present"],
        "identity33_all_worlds_non_degenerate": private["all_worlds_identity33_non_degenerate"],
        "registered_text_negative_controls": private[
            "two_clone_and_four_semantic_controls_per_world"
        ],
        "audit_truth_unopened": manifest["audit_truth_read_counts"] == {
            "audit_a": 0, "audit_b": 0
        },
        "counterfactual_text_shortcut_gate": text_counterfactual["passed"],
    }
    passed = all(hard_gates.values())
    result: dict[str, Any] = {
        "version": VERSION,
        "status": "PASSED_METHOD_ROOT_QUALITY"
        if passed else "FAILED_METHOD_ROOT_QUALITY",
        "public_root": public_root.relative_to(ROOT).as_posix(),
        "private_root_commitment_only": private_root.relative_to(ROOT).as_posix(),
        "manifest_canonical_self_hash": manifest["canonical_self_hash"],
        "surface_audit": surface,
        "source_copy_audit": source_copy,
        "exact_v9_4_replay": replay,
        "identity_positive_control": identity,
        "original_text_descriptive_probe": text_original,
        "counterfactual_text_shortcut_probe": text_counterfactual,
        "registered_private_audit_train_development_only": private,
        "hard_gates": hard_gates,
        "audit_truth_reads": {"audit_a": 0, "audit_b": 0},
        "training_qualified": passed,
    }
    result["canonical_self_hash"] = canonical_sha256(result)
    write_json(output_root / "quality_result.json", result)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--smoke", action="store_true")
    parser.add_argument("--public-root", type=Path)
    parser.add_argument("--private-root", type=Path)
    parser.add_argument("--output-root", type=Path)
    args = parser.parse_args()
    policy = read_json(builder.POLICY_PATH)
    public_root = args.public_root or ROOT / str(policy["formal_output_root"])
    private_root = args.private_root or ROOT / str(policy["formal_private_root"])
    output_root = args.output_root or ROOT / str(policy["formal_quality_output_root"])
    result = audit(
        public_root=public_root.resolve(), private_root=private_root.resolve(),
        output_root=output_root.resolve(), smoke=args.smoke,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
