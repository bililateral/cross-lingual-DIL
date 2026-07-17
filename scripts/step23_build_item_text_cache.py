#!/usr/bin/env python3
"""Build a label-blind, identifier-redacted cache of real train items for Step23."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import unicodedata
from collections import Counter, defaultdict, deque
from pathlib import Path

from openpyxl import load_workbook

import step15_build_v7_clean_embedding_cache as redaction
import step7_build_semantic_pair_features as semantic


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_POLICY = ROOT / "schema" / "step23_item_multi_instance_policy.json"


def resolve(value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else ROOT / path


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def load_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def normalize(value: object) -> str:
    return " ".join(str(value or "").replace("\x00", " ").split()).strip()


def bool_value(value: object) -> bool:
    return str(value or "").strip().casefold() in {"1", "true", "yes"}


def style_features(text: str) -> dict[str, float]:
    length = max(len(text), 1)
    digits = sum(character.isdigit() for character in text)
    punctuation = sum(unicodedata.category(character).startswith("P") for character in text)
    cjk = sum("\u3400" <= character <= "\u9fff" for character in text)
    return {
        "length_log": math.log1p(len(text)),
        "digit_ratio": digits / length,
        "punct_ratio": punctuation / length,
        "cjk_ratio": cjk / length,
    }


def train_scope(pool_name: str, pool_cfg: dict, selection_cfg: dict) -> tuple[set[str], dict]:
    labels_path = resolve(pool_cfg["frozen_labels"])
    rows = load_csv(labels_path)
    allowed_labels = set(selection_cfg["allowed_review_labels"])
    train_rows = [
        row
        for row in rows
        if row.get("split_name") == selection_cfg["allowed_split"]
        and row.get("review_label") in allowed_labels
        and (
            not selection_cfg["require_usable_for_supervision"]
            or bool_value(row.get("usable_for_supervision"))
        )
    ]
    heldout_rows = [row for row in rows if row.get("split_name") in {"valid", "test"}]
    train_sellers = {
        row[key]
        for row in train_rows
        for key in ("seller_uid_left", "seller_uid_right")
        if row.get(key)
    }
    heldout_sellers = {
        row[key]
        for row in heldout_rows
        for key in ("seller_uid_left", "seller_uid_right")
        if row.get(key)
    }
    overlap = sorted(train_sellers & heldout_sellers)
    if overlap:
        raise ValueError(f"Step23 train/heldout seller overlap in {pool_name}: {overlap[0]}")
    if not train_sellers:
        raise ValueError(f"Step23 found no train sellers in {pool_name}")
    return train_sellers, {
        "canonical_binary_rows": len(rows),
        "eligible_train_pair_count": len(train_rows),
        "train_seller_count": len(train_sellers),
        "heldout_seller_count": len(heldout_sellers),
        "train_heldout_seller_overlap": 0,
    }


def seller_literals(pool_cfg: dict, train_sellers: set[str]) -> tuple[dict[str, list[str]], dict]:
    signals_path = resolve(pool_cfg["item_identity_signals"])
    profiles_path = resolve(pool_cfg["seller_profiles"])
    literals, diagnostics = redaction.signal_literals_by_seller(signals_path)
    profiles = {row["seller_uid"]: row for row in semantic.load_jsonl(profiles_path)}
    missing = sorted(train_sellers - set(profiles))
    if missing:
        raise ValueError(f"Step23 missing seller profile: {missing[0]}")
    output = {}
    for seller_uid in train_sellers:
        values = list(literals.get(seller_uid, []))
        for field in ("source_seller_raw", "alias_normalized"):
            literal = redaction.safe_signal_literal("seller_alias", profiles[seller_uid].get(field, ""))
            if literal:
                values.append(literal)
        output[seller_uid] = sorted(set(values), key=lambda value: (-len(value), value.casefold()))
    return output, diagnostics


def redact_field(value: object, literals: list[str], seller_uid: str) -> tuple[str, dict]:
    clean, diagnostics = redaction.redact_identifiers(normalize(value), literals)
    redaction.assert_no_known_identifier_residue(clean, literals, seller_uid)
    return clean, diagnostics


def build_item(meta: dict, title: object, description: object, category: object, literals: list[str], cfg: dict) -> tuple[dict, Counter]:
    diagnostics: Counter = Counter()
    clean_category, category_diag = redact_field(category, literals, meta["seller_uid"])
    clean_title, title_diag = redact_field(title, literals, meta["seller_uid"])
    clean_description, description_diag = redact_field(description, literals, meta["seller_uid"])
    for diag in (category_diag, title_diag, description_diag):
        diagnostics.update(diag)
    field_text = "\n".join(value for value in (clean_category, clean_title, clean_description) if value)
    text, cross_field_diag = redact_field(field_text, literals, meta["seller_uid"])
    diagnostics.update(cross_field_diag)
    cross_field_changed = text != field_text
    if cross_field_changed:
        diagnostics["cross_field_redaction_item_count"] += 1
    if not text:
        text = cfg["empty_text_fallback"]
        diagnostics["empty_text_fallback_count"] += 1
    # Exact-content deduplication uses the complete redacted text. Truncation is
    # an encoder/style budget and must not merge items with a shared long prefix.
    content_signature = sha256_text(text.casefold())
    text = text[: int(cfg["maximum_clean_text_characters"])]
    normalized_category = clean_category.casefold() or "__uncategorized__"
    title_hash = sha256_text(clean_title.casefold()) if clean_title and not cross_field_changed else ""
    description_hash = (
        sha256_text(clean_description.casefold())
        if clean_description and not cross_field_changed
        else ""
    )
    # Deduplication must retain the final redacted content even when exact field
    # overlap is disabled after a cross-field identifier match.
    if cross_field_changed:
        diagnostics["exact_overlap_disabled_item_count"] += 1
    return {
        "item_uid": meta["item_uid"],
        "pool": meta["pool"],
        "domain": meta["domain"],
        "seller_uid": meta["seller_uid"],
        "source_dataset": meta["source_dataset"],
        "source_row_number": int(meta["source_row_number"]),
        "category": clean_category,
        "category_key": normalized_category,
        "title_hash": title_hash,
        "description_hash": description_hash,
        "content_signature": content_signature,
        "exact_overlap_eligible": not cross_field_changed,
        "cross_field_redaction_applied": cross_field_changed,
        "clean_text": text,
        **style_features(text),
        "identifier_redacted": True,
        "synthetic": False,
        "split_name": "train",
    }, diagnostics


def category_round_robin(items: list[dict], maximum: int) -> list[dict]:
    seller_uids = {item["seller_uid"] for item in items}
    if len(seller_uids) > 1:
        raise ValueError("Step23 seller-local selection received items from multiple sellers")
    seller_uid = next(iter(seller_uids), "")
    unique = {}
    for item in sorted(items, key=lambda row: (row["source_row_number"], row["item_uid"])):
        unique.setdefault(item["content_signature"], item)
    queues: dict[str, deque] = defaultdict(deque)
    for item in sorted(unique.values(), key=lambda row: (row["category_key"], row["content_signature"], row["item_uid"])):
        queues[item["category_key"]].append(item)
    selected = []
    category_keys = sorted(
        queues,
        key=lambda category_key: sha256_text(f"{seller_uid}|{category_key}"),
    )
    while len(selected) < maximum and any(queues.values()):
        for category_key in category_keys:
            if queues[category_key] and len(selected) < maximum:
                selected.append(queues[category_key].popleft())
    return selected


def assert_disjoint_pool_sellers(pool_sellers: dict[str, set[str]]) -> None:
    pool_names = sorted(pool_sellers)
    for left_index, left_name in enumerate(pool_names):
        for right_name in pool_names[left_index + 1 :]:
            overlap = sorted(pool_sellers[left_name] & pool_sellers[right_name])
            if overlap:
                raise ValueError(
                    f"Step23 cross-domain train seller overlap between {left_name} and "
                    f"{right_name}: {overlap[0]}"
                )


def immutable_write(path: Path, content: bytes) -> None:
    if path.exists():
        if path.read_bytes() != content:
            raise ValueError(f"Refusing to overwrite a different Step23 item artifact: {path}")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--policy", default=str(DEFAULT_POLICY))
    parser.add_argument("--validate-config-only", action="store_true")
    args = parser.parse_args()

    policy_path = resolve(args.policy)
    policy = json.loads(policy_path.read_text(encoding="utf-8"))
    selection_cfg = policy["item_selection"]
    if selection_cfg["selection_uses_labels"] or selection_cfg["selection_uses_cross_seller_statistics"]:
        raise ValueError("Step23 item selection must remain label-blind and seller-local")
    if not selection_cfg["identifier_redaction_required"]:
        raise ValueError("Step23 requires identifier redaction")
    if args.validate_config_only:
        print(json.dumps({"status": "pass", "policy_version": policy["version"]}, indent=2))
        return

    output_root = resolve(policy["outputs_root"])
    output_root.mkdir(parents=True, exist_ok=True)
    output_cfg = policy["outputs"]
    item_path = output_root / output_cfg["selected_items"]
    summary_path = output_root / output_cfg["item_selection_summary"]
    manifest_path = output_root / output_cfg["item_selection_manifest"]

    pool_sellers = {}
    pool_diagnostics = {}
    literals_by_pool = {}
    signal_diagnostics = {}
    for pool_name, pool_cfg in policy["pools"].items():
        sellers, diagnostics = train_scope(pool_name, pool_cfg, selection_cfg)
        pool_sellers[pool_name] = sellers
        pool_diagnostics[pool_name] = diagnostics
        literals_by_pool[pool_name], signal_diagnostics[pool_name] = seller_literals(pool_cfg, sellers)
    assert_disjoint_pool_sellers(pool_sellers)

    manifest_path_input = resolve(policy["inputs"]["item_manifest"])
    relevant_by_dataset: dict[str, dict[int, dict]] = defaultdict(dict)
    relevant_manifest_count = Counter()
    with manifest_path_input.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            pool_name = row["data_bucket"]
            if pool_name not in pool_sellers or row["seller_uid"] not in pool_sellers[pool_name]:
                continue
            source_row = int(row["source_row_number"])
            if source_row in relevant_by_dataset[row["source_dataset"]]:
                raise ValueError(f"Duplicate Step2 source row: {row['source_dataset']}:{source_row}")
            relevant_by_dataset[row["source_dataset"]][source_row] = {
                "item_uid": row["item_uid"],
                "pool": pool_name,
                "domain": policy["pools"][pool_name]["domain"],
                "seller_uid": row["seller_uid"],
                "source_dataset": row["source_dataset"],
                "source_row_number": source_row,
            }
            relevant_manifest_count[pool_name] += 1

    supported = {"market_item.xlsx", "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx"}
    unsupported = sorted(dataset for dataset, rows in relevant_by_dataset.items() if rows and dataset not in supported)
    if unsupported:
        raise ValueError(f"Step23 has unsupported relevant source datasets: {unsupported}")

    items_by_seller: dict[str, list[dict]] = defaultdict(list)
    redaction_counts: Counter = Counter()
    loaded_source_rows = set()
    market_rows = relevant_by_dataset.get("market_item.xlsx", {})
    if market_rows:
        workbook = load_workbook(resolve(policy["inputs"]["market_item_workbook"]), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            meta = market_rows.get(row_number)
            if meta is None:
                continue
            _vendor, _ship_from, title, description, _price, category, _market = values
            item, diagnostics = build_item(
                meta, title, description, category, literals_by_pool[meta["pool"]][meta["seller_uid"]], selection_cfg
            )
            items_by_seller[meta["seller_uid"]].append(item)
            redaction_counts.update(diagnostics)
            loaded_source_rows.add((meta["source_dataset"], row_number))
        workbook.close()

    agora_dataset = "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx"
    agora_rows = relevant_by_dataset.get(agora_dataset, {})
    if agora_rows:
        workbook = load_workbook(resolve(policy["inputs"]["agora_workbook"]), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            meta = agora_rows.get(row_number)
            if meta is None:
                continue
            _vendor, category, title, description, _price, _origin, _destination, _rating, _remarks = values
            item, diagnostics = build_item(
                meta, title, description, category, literals_by_pool[meta["pool"]][meta["seller_uid"]], selection_cfg
            )
            items_by_seller[meta["seller_uid"]].append(item)
            redaction_counts.update(diagnostics)
            loaded_source_rows.add((meta["source_dataset"], row_number))
        workbook.close()

    expected_source_rows = {
        (dataset, source_row) for dataset, rows in relevant_by_dataset.items() for source_row in rows
    }
    missing_source_rows = sorted(expected_source_rows - loaded_source_rows)
    if missing_source_rows:
        raise ValueError(f"Step23 failed to load a relevant raw item: {missing_source_rows[0]}")

    selected_items = []
    selection_stats = {}
    maximum = int(selection_cfg["maximum_items_per_seller"])
    minimum = int(selection_cfg["minimum_nonempty_items_per_seller"])
    for pool_name, sellers in pool_sellers.items():
        raw_count = 0
        selected_count = 0
        content_group_count = 0
        missing_sellers = []
        for seller_uid in sorted(sellers):
            raw_items = items_by_seller.get(seller_uid, [])
            raw_count += len(raw_items)
            content_group_count += len({item["content_signature"] for item in raw_items})
            chosen = category_round_robin(raw_items, maximum)
            if len(chosen) < minimum:
                missing_sellers.append(seller_uid)
                continue
            for rank, item in enumerate(chosen):
                item["seller_item_rank"] = rank
                selected_items.append(item)
            selected_count += len(chosen)
        if missing_sellers:
            raise ValueError(f"Step23 seller has no usable real item in {pool_name}: {missing_sellers[0]}")
        selection_stats[pool_name] = {
            "seller_count": len(sellers),
            "manifest_item_count": relevant_manifest_count[pool_name],
            "raw_item_count": raw_count,
            "unique_content_group_count": content_group_count,
            "selected_item_count": selected_count,
            "selected_item_mean_per_seller": selected_count / len(sellers),
            "selected_item_maximum_per_seller": maximum,
        }

    selected_items.sort(key=lambda row: (row["pool"], row["seller_uid"], row["seller_item_rank"], row["item_uid"]))
    if len({row["item_uid"] for row in selected_items}) != len(selected_items):
        raise ValueError("Step23 selected item UIDs are not unique")
    item_bytes = "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in selected_items).encode("utf-8")
    immutable_write(item_path, item_bytes)

    input_hashes = {
        "policy": sha256_file(policy_path),
        "producer": sha256_file(Path(__file__)),
        "redaction_dependency": sha256_file(Path(redaction.__file__).resolve()),
        "semantic_dependency": sha256_file(Path(semantic.__file__).resolve()),
        "item_manifest": sha256_file(manifest_path_input),
        "market_item_workbook": sha256_file(resolve(policy["inputs"]["market_item_workbook"])),
        "agora_workbook": sha256_file(resolve(policy["inputs"]["agora_workbook"])),
    }
    for pool_name, pool_cfg in policy["pools"].items():
        for key in ("frozen_labels", "evidence_labels", "seller_profiles", "item_identity_signals"):
            input_hashes[f"{pool_name}:{key}"] = sha256_file(resolve(pool_cfg[key]))
    summary = {
        "step": "step23_real_train_item_selection",
        "policy_version": policy["version"],
        "status": "real_train_items_only_heldout_labels_used_for_seller_exclusion_only",
        "selection_strategy": selection_cfg["strategy"],
        "selected_item_count": len(selected_items),
        "pool_scope": pool_diagnostics,
        "selection_stats": selection_stats,
        "cross_pool_train_seller_overlap": 0,
        "signal_diagnostics": signal_diagnostics,
        "redaction_diagnostics": dict(sorted(redaction_counts.items())),
        "valid_test_items_encoded": False,
        "synthetic_item_count": 0,
        "input_hashes": input_hashes,
        "selected_items_sha256": sha256_file(item_path),
    }
    summary_bytes = (json.dumps(summary, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    immutable_write(summary_path, summary_bytes)
    manifest = {
        "step": "step23_item_selection_manifest",
        "policy_version": policy["version"],
        "input_hashes": input_hashes,
        "outputs": {
            item_path.relative_to(ROOT).as_posix(): sha256_file(item_path),
            summary_path.relative_to(ROOT).as_posix(): sha256_file(summary_path),
        },
    }
    immutable_write(manifest_path, (json.dumps(manifest, ensure_ascii=False, indent=2) + "\n").encode("utf-8"))
    print(json.dumps({
        "status": summary["status"],
        "selected_item_count": len(selected_items),
        "selection_stats": selection_stats,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
