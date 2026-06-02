from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SCHEMA_PATH = ROOT / "schema" / "step3_seller_profile_schema.json"
STEP2_SUMMARY_PATH = ROOT / "reports" / "step2_split_summary.json"
ITEM_MANIFEST_PATH = ROOT / "reports" / "step2_content_item_manifest.csv"
ELIGIBLE_BUCKETS = {"en_content_train_pool", "zh_target_strict", "zh_target_aux"}
CONTACT_TYPES = ("email", "telegram", "wickr", "wechat", "qq", "phone")
ITEM_IDENTITY_TYPES = (
    "email",
    "telegram",
    "wickr",
    "wechat",
    "jabber",
    "qq",
    "phone",
    "crypto_wallet",
    "pgp_public_key",
    "pgp_fingerprint",
    "bat",
    "external_url",
)
DIRECT_ITEM_IDENTITY_TYPES = set(ITEM_IDENTITY_TYPES) - {"external_url"}

WS_RE = re.compile(r"\s+")
NUMBER_RE = re.compile(r"[-+]?\d+(?:[.,]\d+)?")
GOOD_RISK_RE = re.compile(r"(\d+)\s*[∶:]\s*(\d+)")
CJK_RE = re.compile(r"[\u4e00-\u9fff]")
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
TELEGRAM_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:(?:电报|纸飞机|飞机)(?:\s*(?:账号|号|联系))?[\s:：=@._\-－—–~～]{0,8}@?|"
    r"(?:telegram|tg|tele)(?:\s*(?:id|account|acct|handle|账号|号|联系))?[\s:：=@._\-－—–~～]{1,8}@?)"
    r"([A-Za-z][A-Za-z0-9_]{4,31})\b",
    re.I,
)
WICKR_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:wickr(?:\s*me|\s*id)?|威克)(?![A-Za-z0-9])[\s:：=@._-]{0,6}([A-Za-z0-9_.-]{3,32})\b",
    re.I,
)
WECHAT_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:微信|威信|weixin|wechat|vx|wx|v信|v[:：])(?:\s*(?:id|号|账号))?[\s:：=@._\-－—–~～]{0,8}([A-Za-z][A-Za-z0-9_-]{4,32})\b",
    re.I,
)
QQ_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:qq)(?:\s*(?:id|号|账号))?[\s:：=@._\-－—–~～]{0,8}([1-9]\d{4,11})",
    re.I,
)
PHONE_CONTEXT_RE = re.compile(
    r"(?:whatsapp|phone|tel|电话|手机|联系|联系方式)[^0-9A-Za-z+]{0,8}([+]?\d[\d\-\s()]{7,}\d)",
    re.I,
)
PGP_BLOCK_RE = re.compile(r"-----BEGIN PGP PUBLIC KEY BLOCK-----(.*?)-----END PGP PUBLIC KEY BLOCK-----", re.I | re.S)
PGP_FINGERPRINT_RE = re.compile(
    r"(?:pgp|fingerprint|指纹|公钥)[^\n\rA-Fa-f0-9]{0,24}([A-Fa-f0-9]{4}(?:[\s:.-]?[A-Fa-f0-9]{4}){9})",
    re.I,
)
SENTENCE_SPLIT_RE = re.compile(r"(?:[。！？!?；;]+|\n+)")
CLAUSE_SPLIT_RE = re.compile(r"(?:[，,]+)")
SELLER_CONTACT_CUE_RE = re.compile(
    r"(contact|message|dm\b|reach|support|telegram|(?<![A-Za-z0-9])tg(?![A-Za-z0-9])|电报|飞机|纸飞机|tele|wickr|威克|"
    r"wechat|weixin|微信|威信|\bvx\b|\bwx\b|(?<![A-Za-z0-9])v[:：]|jabber|xmpp|qq|企鹅|扣扣|蝙蝠|bat|"
    r"whatsapp|phone|tel|联系|聯繫|联糸|联络|联系方式|客服|咨询|合作|拍前|找我|加我|留下|私聊|"
    r"钱包|收款|付款|地址|usdt|btc|bitcoin|eth|tron|trc20|erc20|pgp|公钥)",
    re.I,
)
PRODUCT_DATA_RISK_RE = re.compile(
    r"(数据|脱库|库|可验证|反向验证|注册页|账号|帐号|会员|博彩|彩票|网贷|贷款|客户|"
    r"数据库|手机号|电话号码|电子邮箱|email|mailbox|full_name|password|passwd|pwd|"
    r"username|account\|full_name|reg_ip|格式|字段|样例|sample|dump|leak|credential|cvv|"
    r"信用卡|身份证|通讯录|邮箱库|购买者邮箱|收货|快递)",
    re.I,
)
WALLET_CUE_RE = re.compile(r"(钱包|收款|付款|充值|地址|usdt|trc20|erc20|btc|bitcoin|eth|tron|wallet|address)", re.I)
URL_RE = re.compile(r"\b((?:https?://|www\.)[^\s\]）)>,，。；;'\"`]{6,180})", re.I)
BARE_DOMAIN_RE = re.compile(
    r"(?<![@\w.-])([a-z0-9][a-z0-9\-]{1,40}\.(?:com|net|org|vip|cn|cc|top|xyz|onion)"
    r"(?:/[^\s\]）)>,，。；;'\"`]{1,120})?)",
    re.I,
)
JABBER_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:jabber|xmpp)(?![A-Za-z0-9])[\s:：=@._-]{0,6}"
    r"([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,24})",
    re.I,
)
BAT_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:(?:蝙蝠)(?:\s*(?:id|账号|号))?[\s:：=@._\-－—–~～]{0,8}|"
    r"(?:bat)(?:\s*(?:id|账号|号))?[\s:：=@._\-－—–~～]{1,8})"
    r"([A-Za-z0-9][A-Za-z0-9_.-]{4,32})\b",
    re.I,
)
TELEGRAM_ITEM_PATTERNS = (
    ("telegram_cue", re.compile(
        r"(?<![A-Za-z0-9])(?:(?:电报|纸飞机|飞机)(?:\s*(?:账号|号|联系))?[\s:：=@._\-－—–~～]{0,12}@?|"
        r"(?:telegram|tg|tele)(?:\s*(?:id|account|acct|handle|账号|号|联系))?[\s:：=@._\-－—–~～]{1,12}@?)"
        r"([A-Za-z][A-Za-z0-9_]{4,31})\b",
        re.I,
    )),
    ("telegram_url", re.compile(r"(?:t\.me|telegram\.me)\s*/\s*([A-Za-z][A-Za-z0-9_]{4,31})\b", re.I)),
    ("telegram_at_handle_with_context", re.compile(r"(?<![A-Za-z0-9._%+-])@([A-Za-z][A-Za-z0-9_]{4,31})(?![A-Za-z0-9._%+-])")),
)
WECHAT_ITEM_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:微信|威信|weixin|wechat|vx|wx|v信|v[:：])"
    r"(?:\s*(?:id|号|账号|联系))?[\s:：=@._\-－—–~～]{0,10}([A-Za-z][A-Za-z0-9_-]{4,32})\b",
    re.I,
)
QQ_ITEM_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:qq|企鹅|扣扣|q\s*q?|q号)"
    r"(?:\s*(?:id|号|账号|联系))?[\s:：=@._\-－—–~～]{0,10}([1-9]\d{4,11})",
    re.I,
)
CRYPTO_WALLET_RE = re.compile(
    r"\b(bc1[ac-hj-np-z02-9]{11,71}|[13][a-km-zA-HJ-NP-Z1-9]{25,34}|0x[a-fA-F0-9]{40}|T[A-HJ-NP-Za-km-z1-9]{33})\b"
)
TOKEN_EXCLUDE_SET = {
    "admin",
    "contact",
    "customer",
    "service",
    "support",
    "telegram",
    "wechat",
    "wickr",
    "wallet",
    "address",
    "group",
    "channel",
    "username",
    "password",
    "account",
    "phone",
    "mobile",
    "email",
    "mail",
    "tel",
    "tg",
    "qq",
    "vx",
    "wx",
    "id",
    "handle",
    "video",
    "photo",
}
EXTERNAL_URL_EXCLUDED_HOSTS = {
    "youtube.com",
    "www.youtube.com",
    "telegram.org",
    "t.me",
    "telegram.me",
    "mega.nz",
    "nordvpn.com",
    "www.nordvpn.com",
    "expressvpn.com",
    "www.expressvpn.com",
}


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def clean_text(value: object) -> str:
    return WS_RE.sub(" ", to_text(value)).strip()


def normalize_identity_match_text(value: object) -> str:
    return unicodedata.normalize("NFKC", clean_text(value))


def description_snippet(value: object, limit: int = 280) -> str:
    text = clean_text(value)
    if len(text) <= limit:
        return text
    return text[:limit].rstrip()


def normalize_signature_text(value: object) -> str:
    text = clean_text(value).casefold()
    if not text:
        return ""
    if CJK_RE.search(text):
        return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "", text)
    return " ".join(re.findall(r"[a-z0-9]+", text))


def is_contentful_signature_segment(text: str) -> bool:
    if not text:
        return False
    cjk_count = len(CJK_RE.findall(text))
    alnum_chars = len(re.findall(r"[A-Za-z0-9]", text))
    word_count = len(re.findall(r"[A-Za-z0-9]+", text))
    return cjk_count >= 10 or alnum_chars >= 40 or word_count >= 7


def extract_description_segments(value: object, max_segments: int = 12) -> list[str]:
    text = clean_text(value)
    if not text:
        return []

    segments: list[str] = []
    for sentence in SENTENCE_SPLIT_RE.split(text):
        sentence = sentence.strip()
        if not sentence:
            continue
        clauses = [clause.strip() for clause in CLAUSE_SPLIT_RE.split(sentence) if clause.strip()]
        if len(sentence) > 240 and len(clauses) > 1:
            segments.extend(clauses)
        else:
            segments.append(sentence)

    ranked = sorted(
        (segment for segment in segments if is_contentful_signature_segment(segment)),
        key=lambda item: (-len(item), item),
    )
    unique_segments = []
    seen = set()
    for segment in ranked:
        norm = normalize_signature_text(segment)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        unique_segments.append(segment)
        if len(unique_segments) >= max_segments:
            break

    if not unique_segments and is_contentful_signature_segment(text):
        unique_segments.append(text)
    return unique_segments


def stats_from_values(values: list[float | int]) -> dict:
    if not values:
        return {"count": 0, "min": None, "median": None, "max": None}
    return {
        "count": len(values),
        "min": min(values),
        "median": median(values),
        "max": max(values),
    }


def parse_first_number(raw: object) -> float | None:
    text = to_text(raw).strip()
    if not text:
        return None
    match = NUMBER_RE.search(text)
    if not match:
        return None
    token = match.group(0)
    if "," in token and "." not in token:
        token = token.replace(",", ".")
    else:
        token = token.replace(",", "")
    try:
        return float(token)
    except ValueError:
        return None


def parse_int_like(raw: object) -> int | None:
    value = parse_first_number(raw)
    if value is None:
        return None
    return int(value)


def parse_online_hours(raw: object) -> int | None:
    text = clean_text(raw)
    if not text:
        return None
    match = re.search(r"(\d+)\s*小时", text)
    if match:
        return int(match.group(1))
    return parse_int_like(text)


def parse_good_risk(raw: object) -> tuple[int | None, int | None]:
    text = clean_text(raw)
    if not text:
        return None, None
    match = GOOD_RISK_RE.search(text)
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def safe_ratio(numerator: float, denominator: float) -> float:
    if denominator == 0:
        return 0.0
    return round(numerator / denominator, 6)


def style_snapshot(title_raw: object, description_raw: object) -> dict:
    title_text = to_text(title_raw)
    desc_text = to_text(description_raw)
    combined = title_text + " " + desc_text
    visible_chars = [ch for ch in combined if not ch.isspace()]
    alpha_chars = [ch for ch in combined if ch.isalpha()]
    digit_count = sum(ch.isdigit() for ch in visible_chars)
    punct_count = sum((not ch.isalnum()) for ch in visible_chars)
    upper_count = sum(ch.isupper() for ch in alpha_chars)
    return {
        "title_len": len(clean_text(title_text)),
        "description_len": len(clean_text(desc_text)),
        "digit_ratio": safe_ratio(digit_count, len(visible_chars)),
        "punct_ratio": safe_ratio(punct_count, len(visible_chars)),
        "uppercase_ratio": safe_ratio(upper_count, len(alpha_chars)),
        "newline_count": title_text.count("\n") + desc_text.count("\n"),
    }


def normalize_contact(contact_type: str, token: str) -> str:
    token = token.strip()
    if not token:
        return ""
    if contact_type in {"email", "telegram", "wickr", "wechat"}:
        return token.casefold()
    if contact_type == "qq":
        return re.sub(r"\D", "", token)
    if contact_type == "phone":
        digits = re.sub(r"\D", "", token)
        if len(digits) < 8:
            return ""
        if token.strip().startswith("+"):
            return f"+{digits}"
        return digits
    return token


def normalize_external_url(raw_value: str) -> str:
    raw = raw_value.strip("[](){}<>,;:'\"，。；、")
    if not raw:
        return ""
    parse_value = raw if re.match(r"https?://", raw, re.I) else f"http://{raw}"
    parsed = urlparse(parse_value)
    host = parsed.netloc.lower().split("@")[-1].split(":")[0].strip(".")
    if not host or "." not in host:
        return ""
    if host in EXTERNAL_URL_EXCLUDED_HOSTS or any(host.endswith("." + item) for item in EXTERNAL_URL_EXCLUDED_HOSTS):
        return ""
    path = re.sub(r"/+", "/", parsed.path or "").strip("/")
    if host.endswith(".onion") and re.match(r"(viewtopic|index|forum|thread)", path, re.I):
        return ""
    segments = [segment for segment in path.split("/") if segment][:2]
    normalized = host + ("/" + "/".join(segments) if segments else "")
    normalized = normalized.lower().rstrip("/")
    if len(normalized) < 6 or len(normalized) > 160:
        return ""
    return normalized


def normalize_identity_value(contact_type: str, raw_value: str) -> str:
    raw = raw_value.strip()
    if not raw:
        return ""
    if contact_type in {"email", "jabber"}:
        lowered = raw.lower().strip("[](){}<>,;:'\"，。；、")
        match = re.fullmatch(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,24}", lowered)
        if not match:
            return ""
        local = lowered.split("@", 1)[0]
        if local.startswith(".") or local.endswith(".") or ".." in local:
            return ""
        return lowered
    if contact_type == "bat":
        lowered = raw.lower().lstrip("@")
        lowered = re.sub(r"[^a-z0-9_.\-]", "", lowered).strip("._-")
        if len(lowered) < 5 or len(lowered) > 32:
            return ""
        if lowered in TOKEN_EXCLUDE_SET:
            return ""
        if re.fullmatch(r"[1-9]\d{4,11}", lowered):
            return lowered
        if not lowered.isdigit() and re.search(r"[a-z]", lowered):
            return lowered
        return ""
    if contact_type in {"telegram", "wickr", "wechat"}:
        lowered = raw.lower().lstrip("@")
        lowered = re.sub(r"[^a-z0-9_.\-]", "", lowered).strip("._-")
        if len(lowered) < 5 or len(lowered) > 32:
            return ""
        if lowered in TOKEN_EXCLUDE_SET or lowered.isdigit() or not re.search(r"[a-z]", lowered):
            return ""
        return lowered
    if contact_type == "qq":
        digits = re.sub(r"\D", "", raw)
        if re.fullmatch(r"[1-9]\d{4,11}", digits or ""):
            return digits
        return ""
    if contact_type == "phone":
        digits = re.sub(r"\D", "", raw)
        if 7 <= len(digits) <= 15 and len(set(digits)) >= 3:
            return digits
        return ""
    if contact_type == "crypto_wallet":
        wallet = re.sub(r"[^A-Za-z0-9]", "", raw)
        if re.fullmatch(r"bc1[ac-hj-np-z02-9]{11,71}", wallet, re.I):
            return wallet.lower()
        if re.fullmatch(r"[13][a-km-zA-HJ-NP-Z1-9]{25,34}", wallet):
            return wallet
        if re.fullmatch(r"0x[a-fA-F0-9]{40}", wallet):
            return wallet.lower()
        if re.fullmatch(r"T[A-HJ-NP-Za-km-z1-9]{33}", wallet):
            return wallet
        return ""
    if contact_type in {"pgp_public_key", "pgp_fingerprint"}:
        return raw.lower()
    if contact_type == "external_url":
        return normalize_external_url(raw)
    return ""


def context_window(text: str, start: int, end: int, radius: int = 90) -> str:
    return clean_text(text[max(0, start - radius) : min(len(text), end + radius)])


def source_field_texts(
    *,
    title_raw: object,
    description_raw: object,
    structured_snapshot: str = "",
    extra_fields: dict[str, object] | None = None,
) -> list[tuple[str, str]]:
    fields = [
        ("title", to_text(title_raw)),
        ("description", to_text(description_raw)),
        ("structured_snapshot", to_text(structured_snapshot)),
    ]
    if extra_fields:
        for key, value in extra_fields.items():
            fields.append((key, to_text(value)))
    return [(name, text) for name, text in fields if text]


def signal_flags(contact_type: str, evidence_level: str, context: str) -> tuple[int, int, int]:
    seller_facing = bool(SELLER_CONTACT_CUE_RE.search(context))
    product_data_risk = bool(PRODUCT_DATA_RISK_RE.search(context))
    if evidence_level in {"telegram_url", "pgp_public_key_block", "pgp_fingerprint"}:
        seller_facing = True
    if contact_type == "crypto_wallet":
        seller_facing = seller_facing or bool(WALLET_CUE_RE.search(context))
    support_only = contact_type == "external_url"
    direct_eligible = int(
        contact_type in DIRECT_ITEM_IDENTITY_TYPES
        and seller_facing
        and not product_data_risk
    )
    return int(seller_facing), int(product_data_risk), int(support_only or not direct_eligible)


def item_signal_row(
    *,
    meta: dict,
    source_field: str,
    contact_type: str,
    normalized_value: str,
    raw_value: str,
    evidence_level: str,
    context: str,
    title_raw: object,
    description_raw: object,
) -> dict:
    seller_facing, product_data_risk, support_only = signal_flags(contact_type, evidence_level, context)
    direct_eligible = int(contact_type in DIRECT_ITEM_IDENTITY_TYPES and seller_facing and not product_data_risk)
    signal_uid_raw = "|".join(
        [
            str(meta["source_dataset"]),
            str(meta["source_row_number"]),
            str(contact_type),
            str(normalized_value),
            str(source_field),
        ]
    )
    return {
        "signal_uid": hashlib.sha1(signal_uid_raw.encode("utf-8")).hexdigest(),
        "data_bucket": meta["data_bucket"],
        "source_dataset": meta["source_dataset"],
        "source_row_number": meta["source_row_number"],
        "seller_uid": meta["seller_uid"],
        "source_market_raw": meta["source_market_raw"],
        "source_seller_raw": meta["source_seller_raw"],
        "source_seller_id_raw": meta["source_seller_id_raw"],
        "alias_normalized": meta["alias_normalized"],
        "source_field": source_field,
        "contact_type": contact_type,
        "normalized_value": normalized_value,
        "raw_value": step3_capped_text(raw_value, 180),
        "evidence_level": evidence_level,
        "seller_facing_context": seller_facing,
        "product_data_risk_context": product_data_risk,
        "direct_identity_eligible": direct_eligible,
        "support_only": support_only,
        "context": step3_capped_text(context, 260),
        "title_snippet": step3_capped_text(clean_text(title_raw), 160),
        "description_snippet": step3_capped_text(clean_text(description_raw), 220),
    }


def step3_capped_text(value: str, limit: int) -> str:
    value = clean_text(value)
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def extract_item_identity_signals(
    meta: dict,
    *,
    title_raw: object,
    description_raw: object,
    structured_snapshot: str = "",
    extra_fields: dict[str, object] | None = None,
) -> list[dict]:
    rows: list[dict] = []
    seen: set[tuple[str, str, str]] = set()

    def add_match(
        *,
        source_field: str,
        field_text: str,
        contact_type: str,
        raw_value: str,
        evidence_level: str,
        start: int,
        end: int,
        normalized_value: str | None = None,
    ) -> None:
        token = normalized_value if normalized_value is not None else normalize_identity_value(contact_type, raw_value)
        if not token:
            return
        key = (source_field, contact_type, token)
        if key in seen:
            return
        seen.add(key)
        context = context_window(field_text, start, end)
        if evidence_level == "telegram_at_handle_with_context" and not SELLER_CONTACT_CUE_RE.search(context):
            return
        rows.append(
            item_signal_row(
                meta=meta,
                source_field=source_field,
                contact_type=contact_type,
                normalized_value=token,
                raw_value=raw_value,
                evidence_level=evidence_level,
                context=context,
                title_raw=title_raw,
                description_raw=description_raw,
            )
        )

    for source_field, field_text in source_field_texts(
        title_raw=title_raw,
        description_raw=description_raw,
        structured_snapshot=structured_snapshot,
        extra_fields=extra_fields,
    ):
        match_text = normalize_identity_match_text(field_text)
        text_without_pgp = PGP_BLOCK_RE.sub(" ", match_text)

        for match in PGP_BLOCK_RE.finditer(field_text):
            canonical = re.sub(r"\s+", "", match.group(0))
            token = "pgp_sha256:" + hashlib.sha256(canonical.encode("utf-8")).hexdigest()
            add_match(
                source_field=source_field,
                field_text=field_text,
                contact_type="pgp_public_key",
                raw_value="PGP_PUBLIC_KEY_BLOCK",
                evidence_level="pgp_public_key_block",
                start=match.start(),
                end=match.end(),
                normalized_value=token,
            )

        for match in PGP_FINGERPRINT_RE.finditer(text_without_pgp):
            fingerprint = re.sub(r"[^A-Fa-f0-9]", "", match.group(1)).lower()
            if len(fingerprint) == 40:
                add_match(
                    source_field=source_field,
                    field_text=text_without_pgp,
                    contact_type="pgp_fingerprint",
                    raw_value=match.group(1),
                    evidence_level="pgp_fingerprint",
                    start=match.start(1),
                    end=match.end(1),
                    normalized_value=fingerprint,
                )

        for match in EMAIL_RE.finditer(text_without_pgp):
            add_match(
                source_field=source_field,
                field_text=text_without_pgp,
                contact_type="email",
                raw_value=match.group(0),
                evidence_level="plain_email",
                start=match.start(),
                end=match.end(),
            )

        for evidence_level, pattern in TELEGRAM_ITEM_PATTERNS:
            for match in pattern.finditer(text_without_pgp):
                add_match(
                    source_field=source_field,
                    field_text=text_without_pgp,
                    contact_type="telegram",
                    raw_value=match.group(1),
                    evidence_level=evidence_level,
                    start=match.start(1),
                    end=match.end(1),
                )

        for contact_type, evidence_level, pattern in (
            ("wickr", "wickr_cue", WICKR_RE),
            ("wechat", "wechat_cue", WECHAT_ITEM_RE),
            ("jabber", "jabber_cue", JABBER_RE),
            ("qq", "qq_cue", QQ_ITEM_RE),
            ("phone", "phone_cue", PHONE_CONTEXT_RE),
            ("bat", "bat_cue", BAT_RE),
            ("crypto_wallet", "crypto_wallet", CRYPTO_WALLET_RE),
        ):
            for match in pattern.finditer(text_without_pgp):
                add_match(
                    source_field=source_field,
                    field_text=text_without_pgp,
                    contact_type=contact_type,
                    raw_value=match.group(1),
                    evidence_level=evidence_level,
                    start=match.start(1),
                    end=match.end(1),
                )

        for evidence_level, pattern in (("external_url", URL_RE), ("bare_domain", BARE_DOMAIN_RE)):
            for match in pattern.finditer(text_without_pgp):
                add_match(
                    source_field=source_field,
                    field_text=text_without_pgp,
                    contact_type="external_url",
                    raw_value=match.group(1),
                    evidence_level=evidence_level,
                    start=match.start(1),
                    end=match.end(1),
                )

    return rows


def extract_contacts(*fields: object) -> dict[str, list[str]]:
    text = "\n".join(to_text(field) for field in fields if field is not None)
    text_without_pgp = PGP_BLOCK_RE.sub(" ", text)
    results = {
        "email": [normalize_contact("email", match) for match in EMAIL_RE.findall(text)],
        "telegram": [normalize_contact("telegram", match) for match in TELEGRAM_RE.findall(text_without_pgp)],
        "wickr": [normalize_contact("wickr", match) for match in WICKR_RE.findall(text_without_pgp)],
        "wechat": [normalize_contact("wechat", match) for match in WECHAT_RE.findall(text_without_pgp)],
        "qq": [normalize_contact("qq", match) for match in QQ_RE.findall(text_without_pgp)],
        "phone": [normalize_contact("phone", match) for match in PHONE_CONTEXT_RE.findall(text_without_pgp)],
    }
    for contact_type in results:
        results[contact_type] = [token for token in results[contact_type] if token]
    return results


def sorted_top(counter: Counter, first_seen: dict[str, int], limit: int) -> list[dict]:
    ranked = sorted(counter.items(), key=lambda item: (-item[1], first_seen[item[0]], item[0]))
    return [{"value": value, "count": count} for value, count in ranked[:limit]]


def concat_top(items: list[dict]) -> str:
    return " || ".join(item["value"] for item in items if item["value"])


def ensure_profile(profiles: dict[str, dict], meta: dict) -> dict:
    seller_uid = meta["seller_uid"]
    if seller_uid not in profiles:
        profiles[seller_uid] = {
            "seller_uid": seller_uid,
            "data_bucket": meta["data_bucket"],
            "source_dataset": meta["source_dataset"],
            "source_market_raw": meta["source_market_raw"],
            "source_seller_raw": meta["source_seller_raw"],
            "source_seller_id_raw": meta["source_seller_id_raw"],
            "alias_normalized": meta["alias_normalized"],
            "item_count": 0,
            "first_source_row_number": meta["source_row_number"],
            "last_source_row_number": meta["source_row_number"],
            "title_counter": Counter(),
            "title_first_seen": {},
            "description_counter": Counter(),
            "description_first_seen": {},
            "description_segment_counter": Counter(),
            "description_segment_first_seen": {},
            "category_counter": Counter(),
            "category_first_seen": {},
            "price_counter": Counter(),
            "price_first_seen": {},
            "ship_from_counter": Counter(),
            "ship_from_first_seen": {},
            "structured_counter": Counter(),
            "structured_first_seen": {},
            "contact_counters": {contact_type: Counter() for contact_type in CONTACT_TYPES},
            "contact_first_seen": {contact_type: {} for contact_type in CONTACT_TYPES},
            "item_sequence": 0,
            "title_lengths": [],
            "description_lengths": [],
            "digit_ratio_sum": 0.0,
            "punct_ratio_sum": 0.0,
            "uppercase_ratio_sum": 0.0,
            "newline_count_sum": 0,
            "cjk_item_count": 0,
            "parsed_prices": [],
            "parsed_ratings": [],
            "source_numeric_values": defaultdict(list),
        }
    return profiles[seller_uid]


def add_ranked(counter: Counter, first_seen: dict[str, int], value: str, order_index: int) -> None:
    if not value:
        return
    counter[value] += 1
    first_seen.setdefault(value, order_index)


def update_profile(
    profile: dict,
    meta: dict,
    *,
    title_raw: object,
    description_raw: object,
    category_raw: object,
    price_raw: object,
    ship_from_raw: object = "",
    structured_snapshot: str = "",
    parsed_rating: float | None = None,
    source_specific_numeric: dict[str, float | int | None] | None = None,
) -> None:
    profile["item_count"] += 1
    profile["item_sequence"] += 1
    profile["last_source_row_number"] = meta["source_row_number"]

    title_text = clean_text(title_raw)
    desc_text = description_snippet(description_raw)
    category_text = clean_text(category_raw)
    price_text = clean_text(price_raw)
    ship_from_text = clean_text(ship_from_raw)
    structured_text = clean_text(structured_snapshot)

    style = style_snapshot(title_raw, description_raw)
    profile["title_lengths"].append(style["title_len"])
    profile["description_lengths"].append(style["description_len"])
    profile["digit_ratio_sum"] += style["digit_ratio"]
    profile["punct_ratio_sum"] += style["punct_ratio"]
    profile["uppercase_ratio_sum"] += style["uppercase_ratio"]
    profile["newline_count_sum"] += style["newline_count"]
    if contains_cjk_item(title_raw, description_raw, category_raw):
        profile["cjk_item_count"] += 1

    add_ranked(profile["title_counter"], profile["title_first_seen"], title_text, profile["item_sequence"])
    add_ranked(profile["description_counter"], profile["description_first_seen"], desc_text, profile["item_sequence"])
    for segment_index, segment_text in enumerate(extract_description_segments(description_raw), start=1):
        add_ranked(
            profile["description_segment_counter"],
            profile["description_segment_first_seen"],
            segment_text,
            profile["item_sequence"] * 100 + segment_index,
        )
    add_ranked(profile["category_counter"], profile["category_first_seen"], category_text, profile["item_sequence"])
    add_ranked(profile["price_counter"], profile["price_first_seen"], price_text, profile["item_sequence"])
    add_ranked(profile["ship_from_counter"], profile["ship_from_first_seen"], ship_from_text, profile["item_sequence"])
    add_ranked(profile["structured_counter"], profile["structured_first_seen"], structured_text, profile["item_sequence"])

    parsed_price = parse_first_number(price_raw)
    if parsed_price is not None:
        profile["parsed_prices"].append(parsed_price)
    if parsed_rating is not None:
        profile["parsed_ratings"].append(parsed_rating)

    if source_specific_numeric:
        for key, value in source_specific_numeric.items():
            if value is not None:
                profile["source_numeric_values"][key].append(value)

    contacts = extract_contacts(title_raw, description_raw, structured_snapshot)
    for contact_type, tokens in contacts.items():
        for token in tokens:
            add_ranked(
                profile["contact_counters"][contact_type],
                profile["contact_first_seen"][contact_type],
                token,
                profile["item_sequence"],
            )


def contains_cjk_item(*fields: object) -> bool:
    return any(CJK_RE.search(to_text(field)) for field in fields if field is not None)


def load_eligible_row_lookup() -> tuple[dict[str, dict[int, dict]], dict]:
    source_lookup: dict[str, dict[int, dict]] = defaultdict(dict)
    counts = {
        "item_count_by_bucket": Counter(),
        "seller_uids_by_bucket": defaultdict(set),
    }
    with ITEM_MANIFEST_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            bucket = row["data_bucket"]
            if bucket not in ELIGIBLE_BUCKETS:
                continue
            source_row_number = int(row["source_row_number"])
            source_lookup[row["source_dataset"]][source_row_number] = {
                "seller_uid": row["seller_uid"],
                "data_bucket": bucket,
                "source_dataset": row["source_dataset"],
                "source_market_raw": row["source_market_raw"],
                "source_seller_raw": row["source_seller_raw"],
                "source_seller_id_raw": row["source_seller_id_raw"],
                "alias_normalized": row["alias_normalized"],
                "source_row_number": source_row_number,
            }
            counts["item_count_by_bucket"][bucket] += 1
            counts["seller_uids_by_bucket"][bucket].add(row["seller_uid"])
    counts["seller_count_by_bucket"] = {
        bucket: len(counts["seller_uids_by_bucket"][bucket]) for bucket in counts["seller_uids_by_bucket"]
    }
    counts["seller_uids_by_bucket"] = None
    return source_lookup, counts


def process_products_data(source_rows: dict[int, dict], profiles: dict[str, dict], item_signals: list[dict]) -> None:
    if not source_rows:
        return
    with (ROOT / "products_data.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for source_row_number, row in enumerate(reader, start=2):
            meta = source_rows.get(source_row_number)
            if not meta:
                continue
            profile = ensure_profile(profiles, meta)
            positive, risk = parse_good_risk(row.get("好评_风险", ""))
            structured_snapshot = " | ".join(
                part
                for part in [
                    f"sales_amount={clean_text(row.get('成交额', ''))}" if clean_text(row.get("成交额", "")) else "",
                    f"sales_volume={clean_text(row.get('成交量', ''))}" if clean_text(row.get("成交量", "")) else "",
                    f"inquiry_count={clean_text(row.get('咨询数', ''))}" if clean_text(row.get("咨询数", "")) else "",
                    f"heat={clean_text(row.get('热度', ''))}" if clean_text(row.get("热度", "")) else "",
                    f"online_hours={clean_text(row.get('卖家在线时长', ''))}" if clean_text(row.get("卖家在线时长", "")) else "",
                    f"reputation={clean_text(row.get('好评_风险', ''))}" if clean_text(row.get("好评_风险", "")) else "",
                ]
                if part
            )
            update_profile(
                profile,
                meta,
                title_raw=row.get("标题", ""),
                description_raw=row.get("商品描述", ""),
                category_raw=row.get("类别", ""),
                price_raw=row.get("单价(USD)", ""),
                structured_snapshot=structured_snapshot,
                source_specific_numeric={
                    "sales_amount": parse_first_number(row.get("成交额", "")),
                    "sales_volume": parse_first_number(row.get("成交量", "")),
                    "inquiry_count": parse_first_number(row.get("咨询数", "")),
                    "heat_score": parse_first_number(row.get("热度", "")),
                    "online_hours": parse_online_hours(row.get("卖家在线时长", "")),
                    "reputation_positive": positive,
                    "reputation_risk": risk,
                },
            )
            item_signals.extend(
                extract_item_identity_signals(
                    meta,
                    title_raw=row.get("标题", ""),
                    description_raw=row.get("商品描述", ""),
                    structured_snapshot=structured_snapshot,
                )
            )


def process_market_item(source_rows: dict[int, dict], profiles: dict[str, dict], item_signals: list[dict]) -> None:
    if not source_rows:
        return
    workbook = load_workbook(ROOT / "market_item.xlsx", read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    for source_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        meta = source_rows.get(source_row_number)
        if not meta:
            continue
        vendor, ship_from, title, description, price, category, market = row
        structured_snapshot = ""
        if clean_text(ship_from):
            structured_snapshot = f"ship_from={clean_text(ship_from)}"
        profile = ensure_profile(profiles, meta)
        update_profile(
            profile,
            meta,
            title_raw=title,
            description_raw=description,
            category_raw=category,
            price_raw=price,
            ship_from_raw=ship_from,
            structured_snapshot=structured_snapshot,
        )
        item_signals.extend(
            extract_item_identity_signals(
                meta,
                title_raw=title,
                description_raw=description,
                structured_snapshot=structured_snapshot,
            )
        )
    workbook.close()


def process_agora(source_rows: dict[int, dict], profiles: dict[str, dict], item_signals: list[dict]) -> None:
    if not source_rows:
        return
    workbook = load_workbook(
        ROOT / "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx",
        read_only=True,
        data_only=True,
    )
    worksheet = workbook[workbook.sheetnames[0]]
    for source_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        meta = source_rows.get(source_row_number)
        if not meta:
            continue
        vendor, category, title, description, price, origin, destination, rating, remarks = row
        structured_snapshot = " | ".join(
            part
            for part in [
                f"origin={clean_text(origin)}" if clean_text(origin) else "",
                f"destination={clean_text(destination)}" if clean_text(destination) else "",
                f"rating={clean_text(rating)}" if clean_text(rating) else "",
                f"remarks={clean_text(remarks)}" if clean_text(remarks) else "",
            ]
            if part
        )
        profile = ensure_profile(profiles, meta)
        update_profile(
            profile,
            meta,
            title_raw=title,
            description_raw=description,
            category_raw=category,
            price_raw=price,
            ship_from_raw=origin,
            structured_snapshot=structured_snapshot,
            parsed_rating=parse_first_number(rating),
            source_specific_numeric={
                "rating_numeric": parse_first_number(rating),
            },
        )
        item_signals.extend(
            extract_item_identity_signals(
                meta,
                title_raw=title,
                description_raw=description,
                structured_snapshot=structured_snapshot,
                extra_fields={"remarks": remarks},
            )
        )
    workbook.close()


def build_profile_text(profile: dict) -> str:
    sections = []
    if profile["alias_normalized"] or profile["source_seller_raw"]:
        sections.append(f"[SELLER] {profile['source_seller_raw'] or profile['alias_normalized']}")
    sections.append(f"[MARKET] {profile['source_market_raw']}")
    if profile["category_concat_top"]:
        sections.append(f"[CATEGORIES] {profile['category_concat_top']}")
    if profile["signature_title_concat"]:
        sections.append(f"[SIGNATURE_TITLES] {profile['signature_title_concat']}")
    if profile["title_concat_top"]:
        sections.append(f"[TITLES] {profile['title_concat_top']}")
    if profile["signature_description_concat"]:
        sections.append(f"[SIGNATURE_DESCRIPTIONS] {profile['signature_description_concat']}")
    if profile["description_concat_top"]:
        sections.append(f"[DESCRIPTIONS] {profile['description_concat_top']}")
    if profile["contact_concat_top"]:
        sections.append(f"[CONTACTS] {profile['contact_concat_top']}")
    if profile["structured_snapshot_concat_top"]:
        sections.append(f"[STRUCTURE] {profile['structured_snapshot_concat_top']}")
    return "\n".join(section for section in sections if section.strip())


def select_signature_items(
    counter: Counter,
    first_seen: dict[str, int],
    df_counter: Counter,
    seller_count: int,
    limit: int,
) -> list[dict]:
    ranked = []
    for value, count in counter.items():
        norm = normalize_signature_text(value)
        if not norm:
            continue
        df_value = df_counter.get(norm, 1)
        idf = math.log((seller_count + 1) / (df_value + 1)) + 1.0
        length_value = max(len(value), len(CJK_RE.findall(value)) * 4)
        score = idf * math.log1p(min(length_value, 600)) * (1.0 + 0.1 * math.log1p(count))
        ranked.append((score, count, -first_seen[value], value, df_value))

    ranked.sort(key=lambda item: (-item[0], -item[1], item[2], item[3]))
    return [
        {
            "value": value,
            "count": count,
            "specificity_score": round(score, 6),
            "seller_df": df_value,
        }
        for score, count, _neg_first_seen, value, df_value in ranked[:limit]
    ]


def finalize_profile(raw_profile: dict, compression: dict, specificity_catalog: dict) -> dict:
    item_count = raw_profile["item_count"]
    top_categories = sorted_top(raw_profile["category_counter"], raw_profile["category_first_seen"], compression["category_top_k"])
    top_titles = sorted_top(raw_profile["title_counter"], raw_profile["title_first_seen"], compression["title_top_k"])
    top_descriptions = sorted_top(
        raw_profile["description_counter"],
        raw_profile["description_first_seen"],
        compression["description_top_k"],
    )
    top_price_strings = sorted_top(raw_profile["price_counter"], raw_profile["price_first_seen"], compression["price_top_k"])
    top_ship_from_values = sorted_top(
        raw_profile["ship_from_counter"],
        raw_profile["ship_from_first_seen"],
        compression["origin_top_k"],
    )
    structured_examples = sorted_top(
        raw_profile["structured_counter"],
        raw_profile["structured_first_seen"],
        compression["structured_snapshot_top_k"],
    )
    signature_titles = select_signature_items(
        raw_profile["title_counter"],
        raw_profile["title_first_seen"],
        specificity_catalog["title_df"],
        specificity_catalog["seller_count"],
        compression.get("signature_title_top_k", 8),
    )
    signature_description_segments = select_signature_items(
        raw_profile["description_segment_counter"],
        raw_profile["description_segment_first_seen"],
        specificity_catalog["description_segment_df"],
        specificity_catalog["seller_count"],
        compression.get("signature_description_top_k", 10),
    )
    contact_signals = {}
    for contact_type in CONTACT_TYPES:
        contact_signals[contact_type] = sorted_top(
            raw_profile["contact_counters"][contact_type],
            raw_profile["contact_first_seen"][contact_type],
            compression["contact_top_k_per_type"],
        )
    category_concat_top = concat_top(top_categories)
    signature_title_concat = concat_top(signature_titles)
    title_concat_top = concat_top(top_titles)
    signature_description_concat = concat_top(signature_description_segments)
    description_concat_top = concat_top(top_descriptions)
    structured_snapshot_concat_top = concat_top(structured_examples)
    contact_parts = []
    for contact_type in CONTACT_TYPES:
        for token in contact_signals[contact_type]:
            contact_parts.append(f"{contact_type}:{token['value']}")
    contact_concat_top = " || ".join(contact_parts)

    profile = {
        "seller_uid": raw_profile["seller_uid"],
        "data_bucket": raw_profile["data_bucket"],
        "source_dataset": raw_profile["source_dataset"],
        "source_market_raw": raw_profile["source_market_raw"],
        "source_seller_raw": raw_profile["source_seller_raw"],
        "source_seller_id_raw": raw_profile["source_seller_id_raw"],
        "alias_normalized": raw_profile["alias_normalized"],
        "item_count": item_count,
        "first_source_row_number": raw_profile["first_source_row_number"],
        "last_source_row_number": raw_profile["last_source_row_number"],
        "unique_title_count": len(raw_profile["title_counter"]),
        "unique_description_snippet_count": len(raw_profile["description_counter"]),
        "unique_category_count": len(raw_profile["category_counter"]),
        "cjk_item_count": raw_profile["cjk_item_count"],
        "parsed_price_count": len(raw_profile["parsed_prices"]),
        "parsed_rating_count": len(raw_profile["parsed_ratings"]),
        "contact_type_count": sum(1 for contact_type in CONTACT_TYPES if raw_profile["contact_counters"][contact_type]),
        "contact_token_count_total": sum(len(raw_profile["contact_counters"][contact_type]) for contact_type in CONTACT_TYPES),
        "title_length_stats": stats_from_values(raw_profile["title_lengths"]),
        "description_length_stats": stats_from_values(raw_profile["description_lengths"]),
        "style_stats": {
            "digit_ratio_mean": round(raw_profile["digit_ratio_sum"] / item_count, 6),
            "punct_ratio_mean": round(raw_profile["punct_ratio_sum"] / item_count, 6),
            "uppercase_ratio_mean": round(raw_profile["uppercase_ratio_sum"] / item_count, 6),
            "newline_count_mean": round(raw_profile["newline_count_sum"] / item_count, 6),
            "repeated_title_share": round(1 - (len(raw_profile["title_counter"]) / item_count), 6),
            "repeated_description_share": round(1 - (len(raw_profile["description_counter"]) / item_count), 6),
            "max_category_share": round(
                max(raw_profile["category_counter"].values()) / item_count if raw_profile["category_counter"] else 0.0,
                6,
            ),
        },
        "price_numeric_approx_stats": stats_from_values(raw_profile["parsed_prices"]),
        "rating_numeric_stats": stats_from_values(raw_profile["parsed_ratings"]),
        "source_specific_numeric_stats": {
            key: stats_from_values(values)
            for key, values in sorted(raw_profile["source_numeric_values"].items())
        },
        "top_categories": top_categories,
        "signature_titles": signature_titles,
        "top_titles": top_titles,
        "signature_description_segments": signature_description_segments,
        "top_description_snippets": top_descriptions,
        "top_price_strings": top_price_strings,
        "top_ship_from_values": top_ship_from_values,
        "contact_signals": contact_signals,
        "structured_snapshot_examples": structured_examples,
        "category_concat_top": category_concat_top,
        "signature_title_concat": signature_title_concat,
        "title_concat_top": title_concat_top,
        "signature_description_concat": signature_description_concat,
        "description_concat_top": description_concat_top,
        "contact_concat_top": contact_concat_top,
        "structured_snapshot_concat_top": structured_snapshot_concat_top,
    }
    profile["profile_text"] = build_profile_text(profile)
    return profile


def expected_step2_counts() -> dict:
    summary = load_json(STEP2_SUMMARY_PATH)
    content_summary = summary["content_split_summary"]
    return {
        "seller_count_by_bucket": {
            "en_content_train_pool": content_summary["seller_summary_counts"]["en_content_eligible_sellers"],
            "zh_target_strict": content_summary["seller_summary_counts"]["zh_target_strict_sellers"],
            "zh_target_aux": content_summary["seller_summary_counts"]["zh_target_aux_sellers"],
        },
        "item_count_by_bucket": {
            "en_content_train_pool": content_summary["en_content_eligible_item_count"],
            "zh_target_strict": content_summary["zh_target_strict_item_count"],
            "zh_target_aux": content_summary["zh_target_aux_item_count"],
        },
    }


def build_specificity_catalog(raw_profiles: dict[str, dict]) -> dict:
    title_df: Counter = Counter()
    description_segment_df: Counter = Counter()

    for raw_profile in raw_profiles.values():
        title_norms = {
            normalize_signature_text(value)
            for value in raw_profile["title_counter"].keys()
            if normalize_signature_text(value)
        }
        description_segment_norms = {
            normalize_signature_text(value)
            for value in raw_profile["description_segment_counter"].keys()
            if normalize_signature_text(value)
        }
        title_df.update(title_norms)
        description_segment_df.update(description_segment_norms)

    return {
        "seller_count": len(raw_profiles),
        "title_df": title_df,
        "description_segment_df": description_segment_df,
    }


def write_profiles(profiles: list[dict]) -> dict:
    reports_dir = ROOT / "reports"
    bucket_paths = {
        bucket: reports_dir / f"step3_seller_profiles.{bucket}.jsonl" for bucket in ELIGIBLE_BUCKETS
    }
    handles = {bucket: bucket_paths[bucket].open("w", encoding="utf-8") for bucket in bucket_paths}

    observed = {
        "seller_count_by_bucket": Counter(),
        "item_count_by_bucket": Counter(),
        "profiles_with_contacts_by_bucket": Counter(),
        "profiles_with_parsed_price_by_bucket": Counter(),
        "profiles_with_signature_titles_by_bucket": Counter(),
        "profiles_with_signature_descriptions_by_bucket": Counter(),
        "profile_text_nonempty_count": 0,
        "profile_text_max_length": 0,
    }

    for profile in profiles:
        line = json.dumps(profile, ensure_ascii=False)
        handles[profile["data_bucket"]].write(line + "\n")

        bucket = profile["data_bucket"]
        observed["seller_count_by_bucket"][bucket] += 1
        observed["item_count_by_bucket"][bucket] += profile["item_count"]
        if profile["contact_type_count"] > 0:
            observed["profiles_with_contacts_by_bucket"][bucket] += 1
        if profile["parsed_price_count"] > 0:
            observed["profiles_with_parsed_price_by_bucket"][bucket] += 1
        if profile["signature_titles"]:
            observed["profiles_with_signature_titles_by_bucket"][bucket] += 1
        if profile["signature_description_segments"]:
            observed["profiles_with_signature_descriptions_by_bucket"][bucket] += 1
        if profile["profile_text"]:
            observed["profile_text_nonempty_count"] += 1
            observed["profile_text_max_length"] = max(observed["profile_text_max_length"], len(profile["profile_text"]))

    for handle in handles.values():
        handle.close()

    observed["seller_count_by_bucket"] = dict(observed["seller_count_by_bucket"])
    observed["item_count_by_bucket"] = dict(observed["item_count_by_bucket"])
    observed["profiles_with_contacts_by_bucket"] = dict(observed["profiles_with_contacts_by_bucket"])
    observed["profiles_with_parsed_price_by_bucket"] = dict(observed["profiles_with_parsed_price_by_bucket"])
    observed["profiles_with_signature_titles_by_bucket"] = dict(observed["profiles_with_signature_titles_by_bucket"])
    observed["profiles_with_signature_descriptions_by_bucket"] = dict(observed["profiles_with_signature_descriptions_by_bucket"])
    return observed


ITEM_SIGNAL_FIELDS = [
    "signal_uid",
    "data_bucket",
    "source_dataset",
    "source_row_number",
    "seller_uid",
    "source_market_raw",
    "source_seller_raw",
    "source_seller_id_raw",
    "alias_normalized",
    "source_field",
    "contact_type",
    "normalized_value",
    "raw_value",
    "evidence_level",
    "seller_facing_context",
    "product_data_risk_context",
    "direct_identity_eligible",
    "support_only",
    "context",
    "title_snippet",
    "description_snippet",
]


def write_item_identity_signals(item_signals: list[dict]) -> dict:
    reports_dir = ROOT / "reports"
    bucket_paths = {
        bucket: reports_dir / f"step3_item_identity_signals.{bucket}.csv" for bucket in ELIGIBLE_BUCKETS
    }
    handles = {bucket: bucket_paths[bucket].open("w", encoding="utf-8-sig", newline="") for bucket in bucket_paths}
    writers = {}
    for bucket, handle in handles.items():
        writer = csv.DictWriter(handle, fieldnames=ITEM_SIGNAL_FIELDS)
        writer.writeheader()
        writers[bucket] = writer

    signal_counts_by_bucket = Counter()
    direct_counts_by_bucket = Counter()
    seller_facing_counts_by_bucket = Counter()
    risk_counts_by_bucket = Counter()
    type_counts = Counter()
    direct_type_counts = Counter()
    seller_sets_by_type: dict[str, set[str]] = defaultdict(set)
    direct_seller_sets_by_type: dict[str, set[str]] = defaultdict(set)
    direct_token_sellers: dict[tuple[str, str], set[str]] = defaultdict(set)

    for row in sorted(item_signals, key=lambda item: (item["data_bucket"], item["source_dataset"], int(item["source_row_number"]), item["contact_type"], item["normalized_value"])):
        bucket = row["data_bucket"]
        writers[bucket].writerow({field: row.get(field, "") for field in ITEM_SIGNAL_FIELDS})
        signal_counts_by_bucket[bucket] += 1
        type_counts[row["contact_type"]] += 1
        seller_sets_by_type[row["contact_type"]].add(row["seller_uid"])
        if int(row["seller_facing_context"]):
            seller_facing_counts_by_bucket[bucket] += 1
        if int(row["product_data_risk_context"]):
            risk_counts_by_bucket[bucket] += 1
        if int(row["direct_identity_eligible"]):
            direct_counts_by_bucket[bucket] += 1
            direct_type_counts[row["contact_type"]] += 1
            direct_seller_sets_by_type[row["contact_type"]].add(row["seller_uid"])
            direct_token_sellers[(row["contact_type"], row["normalized_value"])].add(row["seller_uid"])

    for handle in handles.values():
        handle.close()

    shared_direct_tokens = {
        contact_type: sum(
            1
            for (token_type, _token), seller_uids in direct_token_sellers.items()
            if token_type == contact_type and len(seller_uids) >= 2
        )
        for contact_type in sorted(DIRECT_ITEM_IDENTITY_TYPES)
    }
    return {
        "output_files": {bucket: str(path.relative_to(ROOT)) for bucket, path in bucket_paths.items()},
        "signal_count_total": len(item_signals),
        "signal_counts_by_bucket": dict(signal_counts_by_bucket),
        "direct_identity_eligible_counts_by_bucket": dict(direct_counts_by_bucket),
        "seller_facing_counts_by_bucket": dict(seller_facing_counts_by_bucket),
        "product_data_risk_counts_by_bucket": dict(risk_counts_by_bucket),
        "signal_counts_by_type": dict(type_counts),
        "direct_identity_eligible_counts_by_type": dict(direct_type_counts),
        "seller_counts_by_type": {key: len(value) for key, value in sorted(seller_sets_by_type.items())},
        "direct_identity_seller_counts_by_type": {key: len(value) for key, value in sorted(direct_seller_sets_by_type.items())},
        "shared_direct_token_counts_by_type": shared_direct_tokens,
        "hard_rule": "Item-level identity signals are extraction evidence only; they are not Step 5 labels or ground truth.",
    }


def build_summary(profiles: list[dict], observed: dict, expected: dict, item_identity_summary: dict) -> dict:
    acceptance_checks = {
        "seller_count_match": {},
        "item_count_match": {},
        "all_profile_text_nonempty": observed["profile_text_nonempty_count"] == len(profiles),
    }
    for bucket in ELIGIBLE_BUCKETS:
        acceptance_checks["seller_count_match"][bucket] = (
            observed["seller_count_by_bucket"].get(bucket, 0) == expected["seller_count_by_bucket"].get(bucket, 0)
        )
        acceptance_checks["item_count_match"][bucket] = (
            observed["item_count_by_bucket"].get(bucket, 0) == expected["item_count_by_bucket"].get(bucket, 0)
        )

    return {
        "schema_path": str(SCHEMA_PATH.relative_to(ROOT)),
        "input_dependencies": [
            str(ITEM_MANIFEST_PATH.relative_to(ROOT)),
            str(STEP2_SUMMARY_PATH.relative_to(ROOT)),
        ],
        "profile_count_total": len(profiles),
        "expected_counts_from_step2": expected,
        "observed_counts": observed,
        "item_identity_signal_summary": item_identity_summary,
        "acceptance_checks": acceptance_checks,
        "residual_risks": [
            "price_numeric_approx_stats are approximate and not currency-normalized across markets.",
            "contact extraction uses strict context cues for phones and may under-recall some identifiers by design.",
            "profile_text now preserves long-tail signature fields, but it is still a compressed profile rather than a full seller archive.",
            "Item-level identity signals are parser evidence for review queues; product/victim-data risk flags must be honored before any positive labeling.",
        ],
    }


def main() -> None:
    schema = load_json(SCHEMA_PATH)
    compression = schema["compression_policy"]
    source_lookup, manifest_counts = load_eligible_row_lookup()

    profiles: dict[str, dict] = {}
    item_signals: list[dict] = []
    process_products_data(source_lookup.get("products_data.csv", {}), profiles, item_signals)
    process_market_item(source_lookup.get("market_item.xlsx", {}), profiles, item_signals)
    process_agora(
        source_lookup.get("2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx", {}),
        profiles,
        item_signals,
    )

    specificity_catalog = build_specificity_catalog(profiles)
    finalized_profiles = []
    for seller_uid in sorted(profiles):
        finalized_profiles.append(finalize_profile(profiles[seller_uid], compression, specificity_catalog))

    observed = write_profiles(finalized_profiles)
    item_identity_summary = write_item_identity_signals(item_signals)
    expected = expected_step2_counts()
    summary = build_summary(finalized_profiles, observed, expected, item_identity_summary)
    summary["manifest_counts"] = manifest_counts

    with (ROOT / "reports" / "step3_seller_profile_summary.json").open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)

    print(f"Wrote {ROOT / 'reports' / 'step3_seller_profile_summary.json'}")


if __name__ == "__main__":
    main()
