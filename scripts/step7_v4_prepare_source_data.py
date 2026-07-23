#!/usr/bin/env python3
"""Prepare label-free raw-item text and separate Step 7-v4 development labels."""

from __future__ import annotations

import argparse
import csv
import hashlib
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

import step7_v3_1_prepare_source_data as parent_prepare
import step7_v3_1_source_data as source
import step7_v4_common as common


PREPARATION_SCRIPT = Path(__file__).resolve()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Prepare Step7-v4 complete raw-item title/description data. Public "
            "preparation never reads labels; private train/valid projection is separate."
        )
    )
    parser.add_argument("--policy", type=Path, default=common.DEFAULT_POLICY)
    parser.add_argument(
        "--stage",
        choices=("public", "private-labels", "validate-existing"),
        default="public",
    )
    parser.add_argument("--validate-config-only", action="store_true")
    return parser.parse_args()


def _parent_policy(policy: dict) -> dict:
    path = common.resolve(policy["inputs"]["parent_source_policy"]["path"])
    parent = source.load_json(path)
    source.validate_policy(parent)
    return parent


def _semantic_rows_match(left: list[dict], right: list[dict], role: str) -> None:
    if len(left) != len(right):
        raise ValueError(f"Step7-v4 parent {role} row-count drift")
    for left_row, right_row in zip(left, right, strict=True):
        if list(left_row) != list(right_row):
            raise ValueError(f"Step7-v4 parent {role} schema/order drift")
        for name in left_row:
            left_value = left_row[name]
            right_value = right_row[name]
            try:
                difference = abs(float(left_value) - float(right_value))
            except (TypeError, ValueError):
                if str(left_value) != str(right_value):
                    raise ValueError(
                        f"Step7-v4 parent {role} value drift: {name}"
                    )
            else:
                if not math.isfinite(difference) or difference > 1e-15:
                    raise ValueError(
                        f"Step7-v4 parent {role} numeric drift: {name}"
                    )


def project_label_free_parent_quarantine(
    policy: dict, pair_rows: list[dict], safe_rows: list[dict]
) -> tuple[list[dict], list[dict], set[str], dict]:
    """Remove only the exact raw-structure-invalid parent component.

    This function is deliberately label-free.  The pair is identified by
    immutable public pair metadata and SHA-256 values, and must be a two-seller
    isolated validation component before it can be removed.
    """

    cfg = policy["parent_fragment_quarantine"]
    if len(pair_rows) != int(cfg["expected_parent_pair_count"]):
        raise ValueError("Step7-v4 parent pair count drift before quarantine")
    matches = [
        row
        for row in pair_rows
        if common.sha256_text(row["pair_uid"]) == cfg["pair_uid_sha256"]
    ]
    if len(matches) != int(cfg["excluded_pair_count"]):
        raise ValueError("Step7-v4 quarantined parent pair identity drift")
    excluded = matches[0]
    excluded_sellers = {
        excluded["seller_uid_left"],
        excluded["seller_uid_right"],
    }
    observed_seller_hashes = sorted(
        common.sha256_text(value) for value in excluded_sellers
    )
    if (
        excluded["split_name"] != cfg["split_name"]
        or excluded["component_id"] != cfg["component_id"]
        or observed_seller_hashes != cfg["seller_uid_sha256"]
    ):
        raise ValueError("Step7-v4 quarantined parent pair metadata drift")
    component_rows = [
        row
        for row in pair_rows
        if row["component_id"] == excluded["component_id"]
    ]
    seller_degrees = Counter(
        row[endpoint]
        for row in pair_rows
        for endpoint in ("seller_uid_left", "seller_uid_right")
        if row[endpoint] in excluded_sellers
    )
    if (
        cfg["must_be_an_isolated_component"] is not True
        or component_rows != [excluded]
        or seller_degrees != Counter({seller: 1 for seller in excluded_sellers})
    ):
        raise ValueError("Step7-v4 quarantined pair is not an isolated component")

    pair_uids = [row["pair_uid"] for row in pair_rows]
    safe_uids = [row["pair_uid"] for row in safe_rows]
    if safe_uids != pair_uids:
        raise ValueError("Step7-v4 parent safe-feature order drift before quarantine")
    effective_pairs = [row for row in pair_rows if row is not excluded]
    effective_safe = [
        row for row in safe_rows if row["pair_uid"] != excluded["pair_uid"]
    ]
    if (
        len(effective_pairs) != len(pair_rows) - int(cfg["excluded_pair_count"])
        or [row["pair_uid"] for row in effective_safe]
        != [row["pair_uid"] for row in effective_pairs]
    ):
        raise ValueError("Step7-v4 effective parent projection drift")
    common.validate_pair_manifest(policy, effective_pairs)
    isolation = parent_prepare.validate_split_isolation(effective_pairs)
    boundary = policy["supervision_boundary"]
    if (
        isolation["component_count_by_split"]
        != boundary["expected_component_count_by_split"]
        or isolation["seller_count_by_split"]
        != boundary["expected_seller_count_by_split"]
    ):
        raise ValueError("Step7-v4 effective split isolation drift")
    audit = {
        "status": "pass_exact_label_free_raw_fragment_component_quarantined",
        "decision_basis": cfg["decision_basis"],
        "reason": cfg["reason"],
        "parent_pair_count": len(pair_rows),
        "effective_pair_count": len(effective_pairs),
        "excluded_pair_count": int(cfg["excluded_pair_count"]),
        "excluded_pair_uid_sha256": cfg["pair_uid_sha256"],
        "excluded_seller_uid_sha256": observed_seller_hashes,
        "excluded_split_name": excluded["split_name"],
        "excluded_component_id": excluded["component_id"],
        "isolated_component_verified": True,
        "labels_or_evidence_types_read": False,
    }
    return effective_pairs, effective_safe, excluded_sellers, audit


def protected_identity_collision_compacts(parent_policy: dict) -> set[str]:
    quality = parent_policy["clean_text_contract"]["quality_gates"]
    return {
        source.compact_identifier(term)
        for term in (
            *quality["protected_content_words"],
            *quality["protected_identity_collision_terms"],
        )
    }


def split_seller_local_content_collisions(
    policy: dict,
    parent_policy: dict,
    seller_uids: set[str],
    seller_literals: dict[str, list[str]],
    seller_phrases: dict[str, set[str]],
) -> tuple[dict[str, list[str]], dict[str, set[str]], dict[str, set[str]], dict]:
    """Context-gate aliases that are also protected ordinary content terms."""

    collision_compacts = protected_identity_collision_compacts(parent_policy)
    cfg = policy["clean_text_contract"][
        "seller_local_content_collision_handling"
    ]
    if (
        len(collision_compacts) != int(cfg["collision_compact_count"])
        or common.canonical_hash(sorted(collision_compacts))
        != cfg["collision_compact_registry_canonical_sha256"]
    ):
        raise ValueError("Step7-v4 protected local-collision registry drift")

    unconditional_literals = {}
    unconditional_phrases = {}
    contextual_collisions = {}
    affected_mapping = {}
    removed_literal_count = 0
    removed_phrase_count = 0
    for seller_uid in sorted(seller_uids):
        literals = list(seller_literals[seller_uid])
        phrases = set(seller_phrases[seller_uid])
        contextual = {
            source.compact_identifier(value)
            for value in literals
            if source.compact_identifier(value) in collision_compacts
        } | (phrases & collision_compacts)
        kept_literals = [
            value
            for value in literals
            if source.compact_identifier(value) not in collision_compacts
        ]
        kept_phrases = phrases - collision_compacts
        if any(
            source.compact_identifier(value) in collision_compacts
            for value in kept_literals
        ) or kept_phrases & collision_compacts:
            raise AssertionError("Step7-v4 local collision remained unconditional")
        unconditional_literals[seller_uid] = kept_literals
        unconditional_phrases[seller_uid] = kept_phrases
        contextual_collisions[seller_uid] = contextual
        removed_literal_count += len(literals) - len(kept_literals)
        removed_phrase_count += len(phrases) - len(kept_phrases)
        if contextual:
            affected_mapping[seller_uid] = sorted(contextual)

    if (
        len(affected_mapping) != int(cfg["expected_affected_seller_count"])
        or common.canonical_hash(affected_mapping)
        != cfg["expected_affected_seller_uid_to_tokens_canonical_sha256"]
        or cfg["unconditional_local_redaction_allowed"] is not False
        or cfg["identity_context_gated_redaction_enabled"] is not True
        or cfg["one_character_omission_matching_allowed"] is not False
    ):
        raise ValueError("Step7-v4 local collision seller mapping drift")
    audit = {
        "status": "pass_collision_prone_local_aliases_are_context_gated",
        "collision_compact_count": len(collision_compacts),
        "collision_compact_registry_canonical_sha256": common.canonical_hash(
            sorted(collision_compacts)
        ),
        "affected_seller_count": len(affected_mapping),
        "affected_seller_uid_to_tokens_canonical_sha256": common.canonical_hash(
            affected_mapping
        ),
        "affected_seller_uid_sha256_list": sorted(
            common.sha256_text(seller_uid) for seller_uid in affected_mapping
        ),
        "removed_from_unconditional_literal_registry_count": removed_literal_count,
        "removed_from_unconditional_phrase_registry_count": removed_phrase_count,
        "one_character_omission_matching_allowed": False,
        "uncued_ambiguous_occurrences_are_retained_and_audited": True,
        "labels_or_evidence_types_read": False,
    }
    return (
        unconditional_literals,
        unconditional_phrases,
        contextual_collisions,
        audit,
    )


def replay_parent_public(policy: dict) -> tuple[dict, dict, list[dict], list[dict]]:
    """Replay only the parent primitives v4 actually consumes.

    The v3.1 producer also regenerates its seven-megabyte compressed corpus and
    reruns two fixed-snapshot manual-review censuses.  Those frozen reports are
    not an input to v4.  Repeating them at every v4 entry point took minutes and
    added no boundary check, so v4 reconstructs the public pair projection,
    identity registries, clean seller primitives, and legacy18 rows directly,
    then compares the latter two artifacts byte-semantically with their pinned
    files.  No label or evidence file is opened.
    """

    public_input_records = common.verify_inputs(
        policy,
        (
            "parent_source_policy",
            "parent_pair_manifest",
            "parent_safe_features",
        ),
    )
    parent_policy = _parent_policy(policy)
    input_manifest = source.validate_input_hashes(
        parent_policy,
        ("seller_profiles", "item_identity_signals"),
    )
    profiles_path = common.resolve(parent_policy["inputs"]["seller_profiles"]["path"])
    signals_path = common.resolve(
        parent_policy["inputs"]["item_identity_signals"]["path"]
    )
    # The parent component-assignment file contains review_label. Public v4
    # preparation must never open it merely to discard that column. The
    # already-frozen parent pair manifest is its exact hash-pinned label-free
    # projection of pair/split/component/endpoints.
    frozen_pairs = common.load_csv(
        common.resolve(policy["inputs"]["parent_pair_manifest"]["path"])
    )
    source.validate_public_pair_rows(parent_policy, frozen_pairs)
    pairs = frozen_pairs
    isolation = parent_prepare.validate_split_isolation(frozen_pairs)
    boundary = parent_policy["supervision_boundary"]
    if (
        isolation["component_count_by_split"]
        != boundary["expected_component_count_by_split"]
        or isolation["seller_count_by_split"]
        != boundary["expected_seller_count_by_split"]
    ):
        raise ValueError("Step7-v4 parent split-isolation replay drift")

    profiles_list = source.load_jsonl(profiles_path)
    profiles = {str(row["seller_uid"]): row for row in profiles_list}
    if len(profiles) != len(profiles_list):
        raise ValueError("Step7-v4 parent seller profiles contain duplicate IDs")
    seller_split = {}
    for pair in pairs:
        for endpoint in ("seller_uid_left", "seller_uid_right"):
            seller_split[pair[endpoint]] = pair["split_name"]
    if not set(seller_split).issubset(profiles):
        raise ValueError("Step7-v4 parent seller profile universe is incomplete")

    literals, _signal_summary = source.signal_literals_by_seller(signals_path)
    global_tokens = source.global_identity_tokens(literals, profiles_list)
    audited_global_phrases = source.AUDITED_GLOBAL_IDENTITY_PHRASE_TOKENS
    contextual_aliases = source.contextual_global_alias_tokens(
        profiles_list, literals
    )
    contextual_alias_deletions = source.contextual_alias_deletion_tokens(
        contextual_aliases
    )
    seller_literals = {
        seller_uid: source.seller_identity_literals(profiles[seller_uid])
        for seller_uid in sorted(seller_split)
    }
    seller_phrases = {
        seller_uid: source.seller_identity_phrase_tokens(profiles[seller_uid])
        for seller_uid in sorted(seller_split)
    }
    seller_records = {}
    for seller_uid in sorted(seller_split):
        record, _diagnostics = source.build_clean_seller_record(
            profiles[seller_uid],
            parent_policy["clean_text_contract"],
            global_tokens,
            contextual_aliases,
            contextual_alias_deletions,
            audited_global_phrases,
        )
        seller_records[seller_uid] = record
    train_sellers = {
        pair[endpoint]
        for pair in pairs
        if pair["split_name"] == "train"
        for endpoint in ("seller_uid_left", "seller_uid_right")
    }
    reference = source.train_reference(seller_records, train_sellers)
    safe_rows = source.build_safe_pair_rows(pairs, seller_records, reference)
    source.validate_safe_pair_feature_rows(safe_rows)
    frozen_safe = common.load_csv(
        common.resolve(policy["inputs"]["parent_safe_features"]["path"])
    )
    _semantic_rows_match(safe_rows, frozen_safe, "safe features")
    (
        effective_pairs,
        effective_safe,
        quarantined_seller_uids,
        quarantine_audit,
    ) = project_label_free_parent_quarantine(policy, frozen_pairs, frozen_safe)

    effective_seller_split = {
        row[endpoint]: row["split_name"]
        for row in effective_pairs
        for endpoint in ("seller_uid_left", "seller_uid_right")
    }
    effective_isolation = parent_prepare.validate_split_isolation(effective_pairs)
    effective_train_sellers = {
        row[endpoint]
        for row in effective_pairs
        if row["split_name"] == "train"
        for endpoint in ("seller_uid_left", "seller_uid_right")
    }
    # The malformed continuation rows are not identities.  Keep the frozen
    # parent reconstruction long enough to verify the frozen artifacts, then
    # prevent those two bogus profiles/signals from contributing any v4
    # redaction rule.  The remaining records must still replay legacy18 exactly.
    raw_registry_profiles = [
        row
        for row in profiles_list
        if str(row["seller_uid"]) not in quarantined_seller_uids
    ]
    raw_registry_literals = {
        seller_uid: values
        for seller_uid, values in literals.items()
        if seller_uid not in quarantined_seller_uids
    }
    raw_global_tokens = source.global_identity_tokens(
        raw_registry_literals, raw_registry_profiles
    )
    raw_contextual_aliases = source.contextual_global_alias_tokens(
        raw_registry_profiles, raw_registry_literals
    )
    collision_compacts = protected_identity_collision_compacts(parent_policy)
    removed_global_contextual_collisions = sorted(
        raw_contextual_aliases & collision_compacts
    )
    collision_cfg = policy["clean_text_contract"][
        "seller_local_content_collision_handling"
    ]
    if (
        len(removed_global_contextual_collisions)
        != int(
            collision_cfg[
                "expected_removed_from_global_contextual_registry_count"
            ]
        )
        or common.canonical_hash(removed_global_contextual_collisions)
        != collision_cfg[
            "expected_removed_from_global_contextual_registry_canonical_sha256"
        ]
    ):
        raise ValueError("Step7-v4 global contextual content-collision drift")
    raw_contextual_aliases -= collision_compacts
    legacy_raw_contextual_alias_deletions = source.contextual_alias_deletion_tokens(
        raw_contextual_aliases
    )
    raw_contextual_alias_deletions = common.v4_contextual_alias_deletion_tokens(
        raw_contextual_aliases
    )
    removed_fuzzy_content_collisions = sorted(
        legacy_raw_contextual_alias_deletions - raw_contextual_alias_deletions
    )
    removed_fuzzy_source_anchors = sorted(
        token
        for token in raw_contextual_aliases
        if source.contextual_alias_deletion_tokens({token})
        & set(removed_fuzzy_content_collisions)
    )
    fuzzy_collision_cfg = policy["clean_text_contract"][
        "one_character_omission_content_collision_handling"
    ]
    if (
        removed_fuzzy_content_collisions
        != fuzzy_collision_cfg["denied_surface_tokens"]
        or len(removed_fuzzy_content_collisions)
        != int(fuzzy_collision_cfg["expected_removed_token_count"])
        or common.canonical_hash(removed_fuzzy_content_collisions)
        != fuzzy_collision_cfg["denied_surface_tokens_canonical_sha256"]
        or sorted(
            common.sha256_text(anchor)
            for anchor in removed_fuzzy_source_anchors
        )
        != fuzzy_collision_cfg["source_alias_anchor_sha256s"]
    ):
        raise ValueError(
            "Step7-v4 one-character-omission content-collision drift"
        )
    (
        raw_seller_literals,
        raw_seller_phrases,
        raw_seller_contextual_collisions,
        local_collision_audit,
    ) = split_seller_local_content_collisions(
        policy,
        parent_policy,
        set(effective_seller_split),
        seller_literals,
        seller_phrases,
    )
    local_collision_audit[
        "removed_from_global_contextual_registry_count"
    ] = len(removed_global_contextual_collisions)
    local_collision_audit[
        "removed_from_global_contextual_registry_canonical_sha256"
    ] = common.canonical_hash(removed_global_contextual_collisions)
    effective_seller_records = {}
    for seller_uid in sorted(effective_seller_split):
        record, _diagnostics = source.build_clean_seller_record(
            profiles[seller_uid],
            parent_policy["clean_text_contract"],
            raw_global_tokens,
            raw_contextual_aliases,
            legacy_raw_contextual_alias_deletions,
            audited_global_phrases,
        )
        effective_seller_records[seller_uid] = record
    effective_reference = source.train_reference(
        effective_seller_records, effective_train_sellers
    )
    filtered_registry_safe = source.build_safe_pair_rows(
        effective_pairs, effective_seller_records, effective_reference
    )
    source.validate_safe_pair_feature_rows(filtered_registry_safe)
    _semantic_rows_match(
        filtered_registry_safe,
        effective_safe,
        "quarantine-filtered legacy features",
    )
    quarantine_audit[
        "filtered_identity_registry_replays_frozen_legacy18"
    ] = True
    public = {
        "input_manifest": input_manifest,
        "pairs": effective_pairs,
        "isolation": effective_isolation,
        "safe_rows": effective_safe,
        "reference": effective_reference,
        "profiles": profiles,
        "seller_records": effective_seller_records,
        "global_identity_tokens": raw_global_tokens,
        "audited_global_identity_phrase_tokens": audited_global_phrases,
        "contextual_global_alias_tokens": raw_contextual_aliases,
        "contextual_alias_deletion_tokens": raw_contextual_alias_deletions,
        "seller_identity_literals": raw_seller_literals,
        "seller_identity_phrase_tokens": raw_seller_phrases,
        "seller_contextual_collision_tokens": raw_seller_contextual_collisions,
        "seller_local_content_collision_audit": local_collision_audit,
        "quarantined_seller_uids": quarantined_seller_uids,
        "quarantine_audit": quarantine_audit,
        "parent_full_report_regeneration_skipped": True,
    }
    return parent_policy, public, effective_pairs, effective_safe


def selected_item_manifest_rows(
    policy: dict,
    seller_uids: set[str],
    quarantined_seller_uids: set[str],
) -> tuple[dict[str, dict[int, dict]], dict[str, dict[int, dict]], dict]:
    """Stream the large Step2 manifest without loading rows outside the fixed universe."""

    if seller_uids & quarantined_seller_uids:
        raise ValueError("Step7-v4 effective and quarantined seller universes overlap")
    manifest_path = common.resolve(policy["inputs"]["item_manifest"]["path"])
    raw_cfg = policy["raw_item_boundary"]
    quarantine_cfg = policy["parent_fragment_quarantine"]
    allowed_sources = set(raw_cfg["allowed_source_datasets"])
    selected: dict[str, dict[int, dict]] = {
        source_name: {} for source_name in raw_cfg["allowed_source_datasets"]
    }
    quarantined: dict[str, dict[int, dict]] = {
        source_name: {} for source_name in raw_cfg["allowed_source_datasets"]
    }
    quarantine_by_seller_hash = {
        row["seller_uid_sha256"]: row for row in quarantine_cfg["raw_rows"]
    }
    if sorted(common.sha256_text(uid) for uid in quarantined_seller_uids) != sorted(
        quarantine_by_seller_hash
    ):
        raise ValueError("Step7-v4 quarantined seller/item-manifest universe drift")
    counts = Counter()
    ignored_pair_seller_rows = Counter()
    with manifest_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required_columns = {
            "item_uid",
            "seller_uid",
            "source_dataset",
            "source_row_number",
            "data_bucket",
            "eligibility_status",
        }
        if reader.fieldnames is None or not required_columns.issubset(reader.fieldnames):
            raise ValueError("Step7-v4 Step2 item manifest schema drift")
        for row in reader:
            seller_uid = str(row["seller_uid"])
            if seller_uid in quarantined_seller_uids:
                seller_hash = common.sha256_text(seller_uid)
                expected = quarantine_by_seller_hash[seller_hash]
                try:
                    row_number = int(row["source_row_number"])
                except (TypeError, ValueError) as error:
                    raise ValueError(
                        "Step7-v4 quarantined source row number is invalid"
                    ) from error
                source_name = row["source_dataset"]
                if (
                    source_name != quarantine_cfg["raw_source_dataset"]
                    or row_number != int(expected["source_row_number"])
                    or common.sha256_text(row["item_uid"])
                    != expected["item_uid_sha256"]
                    or row["data_bucket"] != raw_cfg["data_bucket"]
                    or row["eligibility_status"] != raw_cfg["eligibility_status"]
                    or row_number in quarantined[source_name]
                ):
                    raise ValueError(
                        "Step7-v4 quarantined Step2 fragment-row metadata drift"
                    )
                quarantined[source_name][row_number] = {
                    "item_uid": row["item_uid"],
                    "seller_uid": seller_uid,
                    "source_dataset": source_name,
                    "source_row_number": row_number,
                    "quarantine_contract": expected,
                }
                continue
            if seller_uid not in seller_uids:
                continue
            if (
                row["data_bucket"] != raw_cfg["data_bucket"]
                or row["eligibility_status"] != raw_cfg["eligibility_status"]
                or row["source_dataset"] not in allowed_sources
            ):
                ignored_pair_seller_rows[
                    f"{row['data_bucket']}|{row['eligibility_status']}|{row['source_dataset']}"
                ] += 1
                continue
            source_name = row["source_dataset"]
            try:
                row_number = int(row["source_row_number"])
            except (TypeError, ValueError) as error:
                raise ValueError("Step7-v4 source row number is invalid") from error
            if row_number < 2 or row_number in selected[source_name]:
                raise ValueError(
                    f"Step7-v4 duplicate/invalid source row: {source_name}:{row_number}"
                )
            selected[source_name][row_number] = {
                "item_uid": row["item_uid"],
                "seller_uid": seller_uid,
                "source_dataset": source_name,
                "source_row_number": row_number,
            }
            counts[source_name] += 1

    observed_total = sum(counts.values())
    expected_total = int(raw_cfg["expected_selected_item_count"])
    if observed_total != expected_total:
        raise ValueError(
            f"Step7-v4 selected raw item count drift: expected={expected_total} "
            f"observed={observed_total}"
        )
    expected_by_source = Counter(
        {
            key: int(value)
            for key, value in raw_cfg["expected_selected_item_count_by_source"].items()
        }
    )
    if counts != expected_by_source:
        raise ValueError(
            f"Step7-v4 selected source counts drift: expected={expected_by_source} "
            f"observed={counts}"
        )
    quarantined_rows = [
        row
        for source_rows in quarantined.values()
        for row in source_rows.values()
    ]
    if len(quarantined_rows) != len(quarantine_cfg["raw_rows"]):
        raise ValueError("Step7-v4 did not recover every quarantined Step2 row")
    return selected, quarantined, {
        "selected_item_count": observed_total,
        "selected_item_count_by_source": dict(sorted(counts.items())),
        "quarantined_fragment_item_count": len(quarantined_rows),
        "quarantined_fragment_source_row_contract_sha256": common.canonical_hash(
            quarantine_cfg["raw_rows"]
        ),
        "ignored_rows_for_pair_universe_sellers": int(
            sum(ignored_pair_seller_rows.values())
        ),
        "ignored_rows_by_boundary": dict(sorted(ignored_pair_seller_rows.items())),
    }


def _protected_occurrences(
    text: str, patterns: dict[str, re.Pattern[str]]
) -> Counter[str]:
    counts: Counter[str] = Counter()
    for term, pattern in patterns.items():
        count = sum(1 for _match in pattern.finditer(text))
        if count:
            counts[term] = count
    return counts


def _merge_spans(spans: list[tuple[int, int]]) -> list[tuple[int, int]]:
    merged: list[list[int]] = []
    for start, end in sorted(spans):
        if start < 0 or end <= start:
            raise ValueError("Step7-v4 intentional identity span is invalid")
        if merged and start <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], end)
        else:
            merged.append([start, end])
    return [(start, end) for start, end in merged]


def _intentional_identity_spans(
    text: str,
    *,
    seller_literals: list[str],
    seller_phrase_tokens: set[str],
    global_tokens: set[str],
    contextual_aliases: set[str],
    contextual_alias_deletions: set[str],
    seller_contextual_collision_tokens: set[str],
    audited_global_phrases: set[str],
) -> list[tuple[int, int]]:
    """Locate high-confidence raw spans excluded from content-fidelity counts."""

    spans: list[tuple[int, int]] = []
    logical_lines = []
    offset = 0
    for line in text.split("\n"):
        logical_lines.append((offset, line))
        offset += len(line) + 1
    for rule_name, pattern in (
        *source.GENERIC_IDENTIFIER_RULES,
        *source.OBFUSCATED_CONTACT_RULES,
        *common.V4_ADDITIONAL_IDENTITY_RULES,
    ):
        if rule_name in common.MULTILINE_IDENTITY_REDACTION_RULE_NAMES:
            spans.extend(
                (match.start(), match.end()) for match in pattern.finditer(text)
            )
        else:
            for line_offset, line in logical_lines:
                spans.extend(
                    (line_offset + match.start(), line_offset + match.end())
                    for match in pattern.finditer(line)
                )
    for literal in seller_literals:
        pattern = source.identity_literal_pattern(literal)
        for line_offset, line in logical_lines:
            spans.extend(
                (line_offset + match.start(), line_offset + match.end())
                for match in pattern.finditer(line)
            )
    for line_offset, line in logical_lines:
        spans.extend(
            (line_offset + start, line_offset + end)
            for start, end in source.unconditional_alias_spans(
                line, seller_phrase_tokens
            )
        )
        spans.extend(
            (line_offset + start, line_offset + end)
            for start, end in source.unconditional_alias_spans(
                line, audited_global_phrases
            )
        )
        spans.extend(
            (line_offset + start, line_offset + end)
            for start, end, _anchor in (
                common.seller_identity_bridged_audited_alias_spans(
                    line,
                    seller_literals,
                    seller_phrase_tokens,
                    audited_global_phrases,
                )
            )
        )
        for match in source.IDENTIFIER_TOKEN_RE.finditer(line):
            if source.matches_global_identity_token(match.group(0), global_tokens):
                spans.append(
                    (line_offset + match.start(), line_offset + match.end())
                )
        for match in source.contextual_alias_spans(
            line, contextual_aliases, contextual_alias_deletions
        ):
            spans.append(
                (
                    line_offset + int(match["redact_start"]),
                    line_offset + int(match["redact_end"]),
                )
            )
        for match in source.contextual_alias_spans(
            line, seller_contextual_collision_tokens, frozenset()
        ):
            spans.append(
                (
                    line_offset + int(match["redact_start"]),
                    line_offset + int(match["redact_end"]),
                )
            )
    return _merge_spans(spans)


def _protected_occurrences_overlapping_spans(
    text: str,
    patterns: dict[str, re.Pattern[str]],
    spans: list[tuple[int, int]],
) -> Counter[str]:
    counts: Counter[str] = Counter()
    if not spans:
        return counts
    for term, pattern in patterns.items():
        for match in pattern.finditer(text):
            if any(start < match.end() and end > match.start() for start, end in spans):
                counts[term] += 1
    return counts


class RawCorpusBuilder:
    def __init__(
        self,
        *,
        policy: dict,
        parent_policy: dict,
        public: dict,
        seller_split: dict[str, str],
    ) -> None:
        self.policy = policy
        self.parent_policy = parent_policy
        self.public = public
        self.seller_split = seller_split
        self.global_texts: dict[str, str] = {}
        self.mapping: dict[tuple[str, str, str], dict] = {}
        self.raw_lineage_rows: list[dict] = []
        self.redaction_counts: Counter[str] = Counter()
        self.global_identifier_counts: Counter[str] = Counter()
        self.audited_phrase_counts: Counter[str] = Counter()
        self.one_character_omission_surface_counts: Counter[str] = Counter()
        self.raw_character_count = 0
        self.clean_character_count = 0
        self.field_counts: Counter[str] = Counter()
        quality = parent_policy["clean_text_contract"]["quality_gates"]
        self.protected_terms = list(
            dict.fromkeys(
                [
                    *quality["protected_content_words"],
                    *quality["protected_identity_collision_terms"],
                ]
            )
        )
        self.protected_patterns = {
            term: re.compile(
                rf"(?<![a-z0-9]){re.escape(term)}(?![a-z0-9])",
                flags=re.IGNORECASE,
            )
            for term in self.protected_terms
        }
        self.protected_raw: Counter[str] = Counter()
        self.protected_clean: Counter[str] = Counter()
        self.protected_intentional_identity: Counter[str] = Counter()
        self.protected_eligible_content: Counter[str] = Counter()
        self.protected_unexplained_removed: Counter[str] = Counter()
        self.protected_created_surplus: Counter[str] = Counter()
        self._protected_raw_cache: dict[
            tuple[str, str], tuple[Counter[str], Counter[str]]
        ] = {}
        self._protected_clean_cache: dict[str, Counter[str]] = {}
        self.protected_candidate_pattern = re.compile(
            "|".join(
                f"(?:{re.escape(term)})"
                for term in sorted(self.protected_terms, key=len, reverse=True)
            ),
            flags=re.IGNORECASE,
        )

    def _add_clean_mapping(
        self,
        *,
        meta: dict,
        field_name: str,
        clean: str,
    ) -> str:
        text_uid = common.sha256_text(clean)
        existing_text = self.global_texts.setdefault(text_uid, clean)
        if existing_text != clean:
            raise ValueError("Step7-v4 SHA-256 collision in clean text corpus")
        key = (meta["seller_uid"], field_name, text_uid)
        record = self.mapping.setdefault(
            key,
            {
                "seller_uid": meta["seller_uid"],
                "split_name": self.seller_split[meta["seller_uid"]],
                "field_name": field_name,
                "text_uid": text_uid,
                "multiplicity": 0,
                "source_rows": defaultdict(list),
            },
        )
        record["multiplicity"] += 1
        record["source_rows"][meta["source_dataset"]].append(
            int(meta["source_row_number"])
        )
        return text_uid

    def _process_field(self, meta: dict, field_name: str, raw_value: object) -> str:
        seller_uid = meta["seller_uid"]
        try:
            clean, diagnostics = common.redact_raw_field(
                raw_value,
                seller_uid=seller_uid,
                seller_literals=self.public["seller_identity_literals"][seller_uid],
                seller_phrase_tokens=self.public["seller_identity_phrase_tokens"][seller_uid],
                global_tokens=self.public["global_identity_tokens"],
                contextual_aliases=self.public["contextual_global_alias_tokens"],
                contextual_alias_deletions=self.public["contextual_alias_deletion_tokens"],
                seller_contextual_collision_tokens=self.public[
                    "seller_contextual_collision_tokens"
                ][seller_uid],
                audited_global_phrases=self.public[
                    "audited_global_identity_phrase_tokens"
                ],
            )
        except ValueError as error:
            raise ValueError(
                "Step7-v4 raw field normalization/redaction failed: "
                f"source={meta['source_dataset']} row={meta['source_row_number']} "
                f"field={field_name} seller_hash={common.sha256_text(seller_uid)[:16]}"
            ) from error
        normalized_raw = common.normalize_raw_field(raw_value)
        self.raw_character_count += int(diagnostics["raw_character_count"])
        self.clean_character_count += int(diagnostics["clean_character_count"])
        raw_fidelity_key = (seller_uid, normalized_raw)
        raw_has_candidate = self.protected_candidate_pattern.search(normalized_raw)
        raw_fidelity = (
            self._protected_raw_cache.get(raw_fidelity_key)
            if raw_has_candidate
            else (Counter(), Counter())
        )
        if raw_fidelity is None:
            raw_protected = _protected_occurrences(
                normalized_raw, self.protected_patterns
            )
            intentional = Counter()
            if raw_protected:
                intentional_spans = _intentional_identity_spans(
                    normalized_raw,
                    seller_literals=self.public["seller_identity_literals"][
                        seller_uid
                    ],
                    seller_phrase_tokens=self.public[
                        "seller_identity_phrase_tokens"
                    ][seller_uid],
                    global_tokens=self.public["global_identity_tokens"],
                    contextual_aliases=self.public[
                        "contextual_global_alias_tokens"
                    ],
                    contextual_alias_deletions=self.public[
                        "contextual_alias_deletion_tokens"
                    ],
                    seller_contextual_collision_tokens=self.public[
                        "seller_contextual_collision_tokens"
                    ][seller_uid],
                    audited_global_phrases=self.public[
                        "audited_global_identity_phrase_tokens"
                    ],
                )
                intentional = _protected_occurrences_overlapping_spans(
                    normalized_raw, self.protected_patterns, intentional_spans
                )
            raw_fidelity = (raw_protected, intentional)
            self._protected_raw_cache[raw_fidelity_key] = raw_fidelity
        self.protected_raw.update(raw_fidelity[0])
        self.protected_intentional_identity.update(raw_fidelity[1])
        clean_has_candidate = self.protected_candidate_pattern.search(clean)
        clean_protected = (
            self._protected_clean_cache.get(clean)
            if clean_has_candidate
            else Counter()
        )
        if clean_protected is None:
            clean_protected = _protected_occurrences(clean, self.protected_patterns)
            self._protected_clean_cache[clean] = clean_protected
        self.protected_clean.update(clean_protected)
        for term in set(raw_fidelity[0]) | set(clean_protected):
            raw_count = int(raw_fidelity[0][term])
            intentional_count = int(raw_fidelity[1][term])
            clean_count = int(clean_protected[term])
            if not 0 <= intentional_count <= raw_count:
                raise ValueError(
                    f"Step7-v4 protected field identity-span drift: {term}"
                )
            eligible_count = raw_count - intentional_count
            self.protected_eligible_content[term] += eligible_count
            self.protected_unexplained_removed[term] += max(
                0, eligible_count - clean_count
            )
            self.protected_created_surplus[term] += max(
                0, clean_count - eligible_count
            )
        self.field_counts[f"{field_name}_source_occurrence_count"] += 1
        self.field_counts[f"{field_name}_raw_nonempty_count"] += int(
            bool(normalized_raw)
        )
        self.field_counts[f"{field_name}_clean_nonempty_count"] += int(bool(clean))
        self.field_counts[f"{field_name}_empty_after_redaction_count"] += int(
            bool(normalized_raw) and not clean
        )
        for name, value in diagnostics.items():
            if name == "global_identifier_token_counts":
                self.global_identifier_counts.update(value)
            elif name == "audited_global_identity_phrase_counts":
                self.audited_phrase_counts.update(value)
            elif name == "one_character_omission_surface_counts":
                self.one_character_omission_surface_counts.update(value)
            elif isinstance(value, bool):
                self.redaction_counts[name] += int(value)
            elif isinstance(value, int):
                if name == "redaction_pass_count":
                    self.redaction_counts["maximum_redaction_pass_count"] = max(
                        self.redaction_counts["maximum_redaction_pass_count"], value
                    )
                else:
                    self.redaction_counts[name] += value
        if not clean:
            return ""
        return self._add_clean_mapping(
            meta=meta, field_name=field_name, clean=clean
        )

    def add_item(self, meta: dict, *, title: object, description: object) -> None:
        title_uid = self._process_field(meta, "title", title)
        description_uid = self._process_field(meta, "description", description)
        self.raw_lineage_rows.append(
            {
                "source_dataset": meta["source_dataset"],
                "source_row_number": int(meta["source_row_number"]),
                "item_uid": meta["item_uid"],
                "seller_uid": meta["seller_uid"],
                "split_name": self.seller_split[meta["seller_uid"]],
                "title_text_uid": title_uid,
                "description_text_uid": description_uid,
            }
        )

    def finalize(self) -> tuple[list[dict], list[dict], list[dict], dict]:
        raw_lineage = sorted(
            self.raw_lineage_rows,
            key=lambda row: (row["source_dataset"], int(row["source_row_number"])),
        )
        unique_rows = [
            {"text_uid": text_uid, "text": text, "text_sha256": text_uid}
            for text_uid, text in sorted(self.global_texts.items())
        ]
        seller_rows = []
        for key in sorted(
            self.mapping,
            key=lambda item: (item[0], common.FIELD_NAMES.index(item[1]), item[2]),
        ):
            record = self.mapping[key]
            lineage = [
                {
                    "source_dataset": source_name,
                    "source_row_numbers": sorted(row_numbers),
                }
                for source_name, row_numbers in sorted(record["source_rows"].items())
            ]
            seller_rows.append(
                {
                    "seller_uid": record["seller_uid"],
                    "split_name": record["split_name"],
                    "field_name": record["field_name"],
                    "text_uid": record["text_uid"],
                    "multiplicity": int(record["multiplicity"]),
                    "source_lineage": lineage,
                }
            )

        represented_sellers = {row["seller_uid"] for row in seller_rows}
        missing_sellers = sorted(set(self.seller_split) - represented_sellers)
        if missing_sellers:
            raise ValueError(
                "Step7-v4 seller has no nonempty clean title or description: "
                + common.sha256_text(missing_sellers[0])[:16]
            )
        retention = (
            self.clean_character_count / self.raw_character_count
            if self.raw_character_count
            else 0.0
        )
        minimum_retention = float(
            self.policy["clean_text_contract"][
                "minimum_aggregate_character_retention"
            ]
        )
        if retention < minimum_retention:
            raise ValueError(
                f"Step7-v4 aggregate character retention is too low: {retention}"
            )
        quality = self.parent_policy["clean_text_contract"]["quality_gates"]
        protected_retention = {}
        minimum_protected = min(
            float(quality["minimum_protected_word_retention"]),
            float(quality["minimum_protected_identity_collision_term_retention"]),
        )
        for term in self.protected_terms:
            raw_count = int(self.protected_raw[term])
            clean_count = int(self.protected_clean[term])
            intentional_identity_count = int(
                self.protected_intentional_identity[term]
            )
            if not 0 <= intentional_identity_count <= raw_count:
                raise ValueError(
                    f"Step7-v4 protected identity-span count drift: {term}"
                )
            eligible_content_count = int(self.protected_eligible_content[term])
            if eligible_content_count != raw_count - intentional_identity_count:
                raise ValueError(
                    f"Step7-v4 protected field aggregation drift: {term}"
                )
            total_removed_count = raw_count - clean_count
            if total_removed_count < 0:
                raise ValueError(
                    f"Step7-v4 protected term count increased after redaction: {term}"
                )
            unexplained_removed_count = int(
                self.protected_unexplained_removed[term]
            )
            created_surplus_count = int(self.protected_created_surplus[term])
            if created_surplus_count:
                raise ValueError(
                    "Step7-v4 redaction created a protected surface occurrence: "
                    f"term={term} count={created_surplus_count}"
                )
            if not 0 <= unexplained_removed_count <= eligible_content_count:
                raise ValueError(
                    f"Step7-v4 protected unexplained-loss drift: {term}"
                )
            ratio = (
                1.0
                if eligible_content_count == 0
                else 1.0 - unexplained_removed_count / eligible_content_count
            )
            protected_retention[term] = {
                "raw_count": raw_count,
                "intentional_identity_span_count": intentional_identity_count,
                "eligible_content_count": eligible_content_count,
                "clean_count": clean_count,
                "total_removed_count": total_removed_count,
                "unexplained_removed_count": unexplained_removed_count,
                "created_surplus_count": created_surplus_count,
                "retention": ratio,
            }
            if eligible_content_count and ratio < minimum_protected:
                raise ValueError(
                    f"Step7-v4 protected content retention failed: term={term} "
                    f"retention={ratio}"
                )

        diagnostics = {
            "raw_character_count": self.raw_character_count,
            "clean_character_count_occurrence_weighted": self.clean_character_count,
            "aggregate_character_retention": retention,
            "field_counts": dict(sorted(self.field_counts.items())),
            "redaction_counts": dict(sorted(self.redaction_counts.items())),
            "removed_global_identifier_token_sha256_counts": {
                common.sha256_text(token): int(count)
                for token, count in sorted(self.global_identifier_counts.items())
            },
            "removed_audited_phrase_sha256_counts": {
                common.sha256_text(token): int(count)
                for token, count in sorted(self.audited_phrase_counts.items())
            },
            "removed_one_character_omission_surface_sha256_counts": {
                common.sha256_text(token): int(count)
                for token, count in sorted(
                    self.one_character_omission_surface_counts.items()
                )
            },
            "protected_content_occurrence_matching": self.policy[
                "clean_text_contract"
            ]["protected_content_occurrence_matching"],
            "protected_content_retention_aggregation": self.policy[
                "clean_text_contract"
            ]["protected_content_retention_aggregation"],
            "protected_content_retention": protected_retention,
        }
        return raw_lineage, unique_rows, seller_rows, diagnostics


def _source_path_for_dataset(policy: dict, source_dataset: str) -> Path:
    if source_dataset == "market_item.xlsx":
        return common.resolve(policy["inputs"]["market_item_snapshot"]["path"])
    if source_dataset == "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx":
        return common.resolve(policy["inputs"]["agora_snapshot"]["path"])
    raise ValueError(f"Step7-v4 unknown raw source: {source_dataset}")


def _step2_item_uid(prefix: str, *parts: object) -> str:
    payload = "\x1f".join("" if part is None else str(part) for part in parts)
    return f"{prefix}|{hashlib.sha1(payload.encode('utf-8')).hexdigest()}"


def _verify_step2_row_identity(
    meta: dict,
    *,
    vendor: object,
    market: object,
    title: object,
    description: object,
    price: object,
    category: object,
) -> None:
    source_dataset = meta["source_dataset"]
    row_number = int(meta["source_row_number"])
    if source_dataset == "market_item.xlsx":
        seller_raw = "" if vendor is None else str(vendor)
        market_raw = "" if market is None else str(market)
        expected_item_uid = _step2_item_uid(
            source_dataset,
            row_number,
            market_raw,
            seller_raw,
            title,
            description,
            price,
            category,
        )
    else:
        seller_raw = "" if vendor is None else str(vendor).strip()
        market_raw = "agora"
        expected_item_uid = _step2_item_uid(
            source_dataset,
            row_number,
            seller_raw,
            title,
            description,
            price,
            category,
        )
    seller_key = f"seller_raw:{seller_raw}" if seller_raw else "seller_id:"
    expected_seller_uid = f"{source_dataset}|{market_raw}|{seller_key}"
    if (
        meta["seller_uid"] != expected_seller_uid
        or meta["item_uid"] != expected_item_uid
    ):
        raise ValueError(
            "Step7-v4 raw workbook row does not replay its pinned Step2 "
            f"seller/item identity: source={source_dataset} row={row_number}"
        )


def _verify_quarantined_raw_workbook_values(
    meta: dict, values: tuple[object, ...]
) -> None:
    contract = meta["quarantine_contract"]
    canonical_values = ["" if value is None else str(value) for value in values]
    if (
        common.canonical_hash(canonical_values)
        != contract["canonical_cell_values_sha256"]
        or [len(value) for value in canonical_values]
        != contract["cell_string_lengths"]
    ):
        raise ValueError(
            "Step7-v4 quarantined raw fragment structure drift: "
            f"{meta['source_dataset']}:{meta['source_row_number']}"
        )


def extract_workbook_rows(
    policy: dict,
    selected: dict[str, dict[int, dict]],
    quarantined: dict[str, dict[int, dict]],
    builder: RawCorpusBuilder,
) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Step7-v4 raw workbook preparation requires openpyxl on Windows"
        ) from error
    raw_cfg = policy["raw_item_boundary"]
    found = Counter()
    found_quarantined = Counter()
    workbook_audit = {}
    for source_dataset in raw_cfg["allowed_source_datasets"]:
        path = _source_path_for_dataset(policy, source_dataset)
        expected_columns = (
            raw_cfg["market_item_column_order"]
            if source_dataset == "market_item.xlsx"
            else raw_cfg["agora_column_order"]
        )
        expected_header = (
            raw_cfg["market_item_exact_header"]
            if source_dataset == "market_item.xlsx"
            else raw_cfg["agora_exact_header"]
        )
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            if int(worksheet.max_column) != len(expected_columns):
                raise ValueError(
                    f"Step7-v4 raw workbook column-count drift: {source_dataset} "
                    f"expected={len(expected_columns)} observed={worksheet.max_column}"
                )
            observed_header = [
                "" if value is None else str(value)
                for value in next(
                    worksheet.iter_rows(
                        min_row=1, max_row=1, values_only=True
                    )
                )
            ]
            if observed_header != expected_header:
                raise ValueError(
                    f"Step7-v4 raw workbook header/order drift: {source_dataset}"
                )
            selected_rows = selected[source_dataset]
            for row_number, values in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                meta = selected_rows.get(row_number)
                quarantined_meta = quarantined[source_dataset].get(row_number)
                if meta is None and quarantined_meta is None:
                    continue
                if len(values) != len(expected_columns):
                    raise ValueError(
                        f"Step7-v4 workbook row width drift: {source_dataset}:{row_number}"
                    )
                if source_dataset == "market_item.xlsx":
                    vendor, _ship_from, title, description, price, category, market = values
                else:
                    (
                        vendor,
                        category,
                        title,
                        description,
                        price,
                        _origin,
                        _destination,
                        _rating,
                        _remarks,
                    ) = values
                    market = "agora"
                if quarantined_meta is not None:
                    if meta is not None:
                        raise ValueError(
                            "Step7-v4 raw row is both effective and quarantined"
                        )
                    _verify_step2_row_identity(
                        quarantined_meta,
                        vendor=vendor,
                        market=market,
                        title=title,
                        description=description,
                        price=price,
                        category=category,
                    )
                    _verify_quarantined_raw_workbook_values(
                        quarantined_meta, values
                    )
                    found_quarantined[source_dataset] += 1
                    continue
                _verify_step2_row_identity(
                    meta,
                    vendor=vendor,
                    market=market,
                    title=title,
                    description=description,
                    price=price,
                    category=category,
                )
                builder.add_item(meta, title=title, description=description)
                found[source_dataset] += 1
        finally:
            workbook.close()
        if found[source_dataset] != len(selected[source_dataset]):
            raise ValueError(
                f"Step7-v4 did not recover every selected workbook row: "
                f"{source_dataset} expected={len(selected[source_dataset])} "
                f"observed={found[source_dataset]}"
            )
        if found_quarantined[source_dataset] != len(quarantined[source_dataset]):
            raise ValueError(
                "Step7-v4 did not verify every quarantined workbook row: "
                f"{source_dataset} expected={len(quarantined[source_dataset])} "
                f"observed={found_quarantined[source_dataset]}"
            )
        workbook_audit[source_dataset] = {
            "worksheet_title": worksheet.title,
            "worksheet_max_row": int(worksheet.max_row),
            "worksheet_max_column": int(worksheet.max_column),
            "selected_row_count": int(found[source_dataset]),
            "quarantined_fragment_row_count": int(
                found_quarantined[source_dataset]
            ),
            "conceptual_column_order": list(expected_columns),
            "exact_header_verified": list(expected_header),
            "step2_seller_and_item_identity_replay_count": int(
                found[source_dataset] + found_quarantined[source_dataset]
            ),
        }
    return workbook_audit


def _identity_audit_rows(
    seller_rows: list[dict], unique_rows: list[dict]
) -> list[dict]:
    texts = {row["text_uid"]: row["text"] for row in unique_rows}
    return [
        {
            "seller_uid": row["seller_uid"],
            "model_text": texts[row["text_uid"]],
            "multiplicity": int(row["multiplicity"]),
        }
        for row in seller_rows
    ]


def prepare_public(policy: dict) -> dict:
    print("[Step7-v4] replaying label-free parent boundary", flush=True)
    public_input_records = common.verify_inputs(
        policy,
        (
            "parent_source_policy",
            "item_manifest",
            "market_item_snapshot",
            "agora_snapshot",
            "parent_pair_manifest",
            "parent_safe_features",
        ),
    )
    parent_policy, public, pair_rows, frozen_safe = replay_parent_public(policy)
    print("[Step7-v4] selecting pinned Step2 raw rows", flush=True)
    seller_split = {}
    for pair in pair_rows:
        for endpoint in ("seller_uid_left", "seller_uid_right"):
            seller_split[pair[endpoint]] = pair["split_name"]
    selected, quarantined, item_manifest_audit = selected_item_manifest_rows(
        policy,
        set(seller_split),
        public["quarantined_seller_uids"],
    )
    builder = RawCorpusBuilder(
        policy=policy,
        parent_policy=parent_policy,
        public=public,
        seller_split=seller_split,
    )
    print("[Step7-v4] extracting and identity-redacting raw item fields", flush=True)
    workbook_audit = extract_workbook_rows(
        policy, selected, quarantined, builder
    )
    raw_lineage, unique_rows, seller_rows, text_diagnostics = builder.finalize()
    if len(raw_lineage) != policy["raw_item_boundary"]["expected_selected_item_count"]:
        raise AssertionError("Step7-v4 raw lineage count drift after extraction")

    identity_rows = _identity_audit_rows(seller_rows, unique_rows)
    print("[Step7-v4] running exact high-precision identity residue scan", flush=True)
    final_identity_scan = common.exact_final_corpus_identity_residue_scan(
        identity_rows,
        public["seller_identity_literals"],
        public["global_identity_tokens"],
        public["contextual_global_alias_tokens"],
        public["contextual_alias_deletion_tokens"],
        public["seller_identity_phrase_tokens"],
        public["audited_global_identity_phrase_tokens"],
        seller_contextual_collision_tokens_by_uid=public[
            "seller_contextual_collision_tokens"
        ],
        fixed_content_collision_contract=policy["clean_text_contract"][
            "fixed_final_audit_content_collision_handling"
        ],
    )
    local_collision_census = common.seller_local_collision_residual_census(
        identity_rows, public["seller_contextual_collision_tokens"]
    )
    print("[Step7-v4] running full known-alias census", flush=True)
    known_alias_census = common.full_known_alias_residual_census(
        identity_rows, public["contextual_global_alias_tokens"]
    )
    print("[Step7-v4] running embedded-identity census", flush=True)
    embedded_identity_census = common.audited_identity_embedded_residual_census(
        identity_rows, public["audited_global_identity_phrase_tokens"]
    )

    print("[Step7-v4] building fixed stylometry and opaque GPU indices", flush=True)
    seller_stylometry = common.build_seller_stylometry_rows(
        seller_rows, unique_rows, seller_split
    )
    pair_stylometry = common.build_pair_stylometry_rows(
        pair_rows, seller_stylometry
    )
    legacy_rows = [
        {
            "pair_uid": row["pair_uid"],
            **{
                name: row[name]
                for name in (
                    *source.SHORTCUT_AUDIT_ONLY_FEATURE_NAMES,
                    *source.MODEL_ELIGIBLE_TRANSFER_FEATURE_NAMES,
                )
            },
        }
        for row in frozen_safe
    ]
    gpu_pair_rows, gpu_seller_rows = common.build_opaque_gpu_indices(
        pair_rows, seller_rows
    )

    outputs = policy["outputs"]
    common.write_csv_immutable(
        common.resolve(outputs["pair_manifest"]), pair_rows
    )
    print("[Step7-v4] writing and validating immutable public artifacts", flush=True)
    common.write_csv_immutable(
        common.resolve(outputs["gpu_pair_manifest"]), gpu_pair_rows
    )
    common.write_csv_immutable(
        common.resolve(outputs["raw_item_lineage"]), raw_lineage
    )
    common.write_jsonl_immutable(
        common.resolve(outputs["unique_text_corpus"]), unique_rows
    )
    common.write_jsonl_immutable(
        common.resolve(outputs["seller_text_index"]), seller_rows
    )
    common.write_jsonl_immutable(
        common.resolve(outputs["gpu_seller_text_index"]), gpu_seller_rows
    )
    common.write_csv_immutable(
        common.resolve(outputs["seller_stylometry"]), seller_stylometry
    )
    common.write_csv_immutable(
        common.resolve(outputs["pair_stylometry"]), pair_stylometry
    )
    common.write_csv_immutable(
        common.resolve(outputs["legacy_pair_features"]), legacy_rows
    )

    manifest = {
        "step": "step7_v4_prepare_source_data_public",
        "version": common.EXPECTED_VERSION,
        "labels_read": False,
        "evidence_types_read": False,
        "pair_label_or_evidence_bearing_input_files_opened": False,
        "labelled_component_assignment_file_opened": False,
        "frozen_label_free_parent_pair_projection_used": True,
        "historical_test_labels_read": False,
        "policy_sha256": common.sha256_file(common.DEFAULT_POLICY),
        "input_files": public_input_records,
        "parent_replayed_input_files": public["input_manifest"],
        "parent_full_report_regeneration_skipped": public[
            "parent_full_report_regeneration_skipped"
        ],
        "implementation": {
            "producer": common.file_record(PREPARATION_SCRIPT),
            "common": common.file_record(Path(common.__file__).resolve()),
            "redaction_module": common.file_record(Path(source.__file__).resolve()),
            "parent_preparation": common.file_record(
                Path(parent_prepare.__file__).resolve()
            ),
        },
        "counts": {
            "pair_count": len(pair_rows),
            "seller_count": len(seller_split),
            "raw_item_lineage_count": len(raw_lineage),
            "global_unique_clean_text_count": len(unique_rows),
            "seller_unique_text_mapping_count": len(seller_rows),
            "gpu_opaque_pair_count": len(gpu_pair_rows),
            "gpu_opaque_seller_text_mapping_count": len(gpu_seller_rows),
            "title_seller_unique_text_mapping_count": sum(
                row["field_name"] == "title" for row in seller_rows
            ),
            "description_seller_unique_text_mapping_count": sum(
                row["field_name"] == "description" for row in seller_rows
            ),
        },
        "item_manifest_audit": item_manifest_audit,
        "workbook_audit": workbook_audit,
        "parent_fragment_quarantine": {
            **public["quarantine_audit"],
            "step2_fragment_rows_verified": item_manifest_audit[
                "quarantined_fragment_item_count"
            ],
            "raw_workbook_fragment_rows_verified": sum(
                int(record["quarantined_fragment_row_count"])
                for record in workbook_audit.values()
            ),
            "raw_fragment_contract_sha256": item_manifest_audit[
                "quarantined_fragment_source_row_contract_sha256"
            ],
        },
        "text_diagnostics": text_diagnostics,
        "identity_audit": {
            "high_precision_final_serialized_scan": final_identity_scan,
            "retained_uncued_local_collision_census": local_collision_census,
            "full_known_alias_residual_census": known_alias_census,
            "audited_embedded_identity_residual_census": embedded_identity_census,
            "unknown_or_ambiguous_identifier_absence_proven": False,
        },
        "identity_registry_provenance": {
            "global_identity_token_count": len(public["global_identity_tokens"]),
            "global_identity_token_registry_sha256": common.canonical_hash(
                sorted(public["global_identity_tokens"])
            ),
            "contextual_alias_token_count": len(
                public["contextual_global_alias_tokens"]
            ),
            "contextual_alias_registry_sha256": common.canonical_hash(
                sorted(public["contextual_global_alias_tokens"])
            ),
            "contextual_alias_deletion_token_count": len(
                public["contextual_alias_deletion_tokens"]
            ),
            "contextual_alias_deletion_registry_sha256": common.canonical_hash(
                sorted(public["contextual_alias_deletion_tokens"])
            ),
            "audited_global_phrase_token_count": len(
                public["audited_global_identity_phrase_tokens"]
            ),
            "audited_global_phrase_registry_sha256": common.canonical_hash(
                sorted(public["audited_global_identity_phrase_tokens"])
            ),
            "seller_local_literal_registry_sha256": common.canonical_hash(
                public["seller_identity_literals"]
            ),
            "seller_local_phrase_registry_sha256": common.canonical_hash(
                {
                    seller_uid: sorted(tokens)
                    for seller_uid, tokens in sorted(
                        public["seller_identity_phrase_tokens"].items()
                    )
                }
            ),
            "seller_local_content_collision_registry": public[
                "seller_local_content_collision_audit"
            ],
            "quarantined_invalid_profile_count": len(
                public["quarantined_seller_uids"]
            ),
            "quarantined_invalid_profile_uid_sha256_list": sorted(
                common.sha256_text(value)
                for value in public["quarantined_seller_uids"]
            ),
            "labels_or_evidence_types_read": False,
        },
        "multiplicity_contract": {
            "source_rows_retained": True,
            "primary_encoding_global_exact_text_deduplicated": True,
            "primary_seller_aggregation_unique_text_equal_weighted": True,
            "multiplicity_used_by_selectable_features": False,
        },
        "outputs": {
            role: common.file_record(common.resolve(outputs[role]))
            for role in common.PUBLIC_OUTPUT_ROLES
        },
    }
    manifest["manifest_content_sha256"] = common.canonical_hash(manifest)
    common.write_json_immutable(
        common.resolve(outputs["preparation_manifest"]), manifest
    )
    validated_manifest, _bundle = common.validate_preparation_artifacts(policy)
    return validated_manifest


def prepare_private_labels(policy: dict) -> dict:
    """Project labels and diagnostic evidence into physically separate files."""

    print("[Step7-v4] projecting physically separated development supervision", flush=True)
    public_manifest, bundle = common.validate_preparation_artifacts(policy)
    del public_manifest
    input_records = common.verify_inputs(
        policy, ("parent_train_labels", "parent_valid_labels")
    )
    outputs = policy["outputs"]
    role_map = {"train": "parent_train_labels", "valid": "parent_valid_labels"}
    pair_uids_by_split = defaultdict(list)
    for row in bundle["pair_rows"]:
        pair_uids_by_split[row["split_name"]].append(row["pair_uid"])
    quarantine_cfg = policy["parent_fragment_quarantine"]
    projection_audit = {
        "decision_basis": quarantine_cfg["decision_basis"],
        "excluded_pair_uid_sha256": quarantine_cfg["pair_uid_sha256"],
        "excluded_pair_count_by_split": {},
        "labels_or_evidence_types_used_to_choose_exclusion": False,
    }
    for split, input_role in role_map.items():
        source_path = common.resolve(policy["inputs"][input_role]["path"])
        rows = common.load_csv(source_path)
        allowed_excluded_hashes = (
            {quarantine_cfg["pair_uid_sha256"]}
            if split == quarantine_cfg["split_name"]
            else set()
        )
        label_rows, evidence_rows = project_private_label_rows(
            rows,
            split,
            expected_pair_uids=pair_uids_by_split[split],
            allowed_excluded_pair_uid_sha256=allowed_excluded_hashes,
        )
        projection_audit["excluded_pair_count_by_split"][split] = (
            len(rows) - len(label_rows)
        )
        common.write_csv_immutable(
            common.resolve(outputs[f"{split}_labels"]),
            label_rows,
        )
        common.write_csv_immutable(
            common.resolve(outputs[f"{split}_evidence"]),
            evidence_rows,
        )
    manifest = {
        "step": "step7_v4_prepare_source_data_private_labels",
        "version": common.EXPECTED_VERSION,
        "public_preparation_manifest_sha256": common.sha256_file(
            common.resolve(outputs["preparation_manifest"])
        ),
        "label_inputs": input_records,
        "selection_label_columns": ["pair_uid", "review_label", "component_id"],
        "diagnostic_evidence_columns": ["pair_uid", "evidence_type"],
        "identity_rule_control_score_materialized": False,
        "evidence_is_physically_separate_from_selection_labels": True,
        "historical_test_labels_materialized": False,
        "parent_fragment_quarantine_projection": projection_audit,
        "policy_sha256": common.sha256_file(common.DEFAULT_POLICY),
        "producer_sha256": common.sha256_file(PREPARATION_SCRIPT),
        "outputs": {
            role: common.file_record(common.resolve(outputs[role]))
            for role in common.PRIVATE_OUTPUT_ROLES
        },
    }
    manifest["manifest_content_sha256"] = common.canonical_hash(manifest)
    common.write_json_immutable(
        common.resolve(outputs["development_labels_manifest"]), manifest
    )
    common.validate_private_label_artifacts(policy, bundle["pair_rows"])
    return manifest


def project_private_label_rows(
    rows: list[dict],
    split: str,
    *,
    expected_pair_uids: list[str] | None = None,
    allowed_excluded_pair_uid_sha256: set[str] | None = None,
) -> tuple[list[dict], list[dict]]:
    return common.project_private_label_rows(
        rows,
        split,
        expected_pair_uids=expected_pair_uids,
        allowed_excluded_pair_uid_sha256=allowed_excluded_pair_uid_sha256,
    )


def main() -> None:
    args = parse_args()
    policy = common.load_policy(args.policy, require_frozen=True)
    common.verify_implementation_files(policy)
    if args.validate_config_only:
        roles = (
            (
                "parent_source_policy",
                "item_manifest",
                "market_item_snapshot",
                "agora_snapshot",
                "parent_pair_manifest",
                "parent_safe_features",
            )
            if args.stage == "public"
            else (
                ("parent_train_labels", "parent_valid_labels")
                if args.stage == "private-labels"
                else tuple(policy["inputs"])
            )
        )
        common.verify_inputs(policy, roles)
        print(
            f"Step7-v4 frozen {args.stage} configuration is valid."
        )
        return
    if args.stage == "public":
        manifest = prepare_public(policy)
        print(
            "Prepared label-free Step7-v4 raw item corpus: "
            f"{manifest['counts']['raw_item_lineage_count']} items, "
            f"{manifest['counts']['global_unique_clean_text_count']} global unique texts."
        )
    elif args.stage == "private-labels":
        prepare_private_labels(policy)
        print("Prepared physically separate Step7-v4 train/valid development labels.")
    else:
        _manifest, bundle = common.validate_preparation_artifacts(policy)
        common.validate_private_label_artifacts(policy, bundle["pair_rows"])
        print("Step7-v4 existing preparation artifacts are valid.")


if __name__ == "__main__":
    main()
