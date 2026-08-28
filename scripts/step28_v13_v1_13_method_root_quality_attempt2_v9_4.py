#!/usr/bin/env python3
"""Run the complete V9.4 method-root quality audit attempt 2.

The hard text shortcut family is the frozen seven-view, two-model family from
the quality-audit C amendment.  Feature matrices are built before train or
development truth is opened.  Audit A/B private truth is never parsed.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import platform
import shutil
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

import numpy as np
import sklearn
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import average_precision_score, roc_auc_score
from sklearn.preprocessing import StandardScaler

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import step28_v13_profiles as profiles_module
import step28_v13_v1_13_balanced_world_schedule_v9_4 as schedule_v94
import step28_v13_v1_13_joint_noise_signatures_v9_4 as noise_v94
import step28_v13_v1_13_method_root_builder_v9_4 as builder
import step28_v13_v1_13_model_visible_prebuild_source_v9_4 as prebuild_v94
import step28_v13_v1_13_model_visible_public_replay_v9_4 as replay_v94
import step28_v13_v1_13_quality_probe_validator_v9 as probe_validator
import step28_v13_v1_13_quality_text_probe_views_v9 as text_views
import step28_v13_v1_13_scientific_common_v9 as scientific_common
import step28_v13_v1_13_style_derangement as style_derangement


VERSION = "2026-08-28-step28-v13-v1-13-method-root-quality-attempt2-v9-4-v1"
POLICY_PATH = (
    ROOT
    / "schema"
    / "step28_v13_v1_13_v9_4_method_root_quality_attempt2_policy.json"
)
BASE_POLICY_PATH = ROOT / "schema" / "step28_v13_synthetic_chinese_dataset_policy.json"
TEXT_TEMPLATE_PATH = ROOT / "schema" / "step28_v13_v1_13_candidate_text_templates_v9.json"
VISIBLE_ARTIFICIAL_CODE = builder.ARTIFICIAL_CODE_RE
ENDPOINT_FIELDS = (
    "canonical_pair_uid",
    "world_uid",
    "seller_uid_left",
    "seller_uid_right",
)


class MethodRootQualityAttempt2Error(ValueError):
    """Raised when the frozen audit cannot be completed exactly."""


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: Any) -> str:
    return hashlib.sha256(canonical_bytes(value)).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as stream:
        value = json.load(stream)
    if not isinstance(value, dict):
        raise MethodRootQualityAttempt2Error(f"JSON object required: {path}")
    return value


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    values: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as stream:
        for line in stream:
            if line.strip():
                value = json.loads(line)
                if not isinstance(value, dict):
                    raise MethodRootQualityAttempt2Error(
                        f"JSONL object required: {path}"
                    )
                values.append(value)
    return values


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def write_json_exclusive(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("x", encoding="utf-8", newline="\n") as stream:
        json.dump(value, stream, ensure_ascii=False, indent=2, sort_keys=True)
        stream.write("\n")
        stream.flush()
        os.fsync(stream.fileno())


def git_value(*arguments: str) -> str:
    result = subprocess.run(
        ["git", *arguments],
        cwd=ROOT,
        check=True,
        text=True,
        encoding="utf-8",
        capture_output=True,
    )
    return result.stdout.strip()


def verify_policy() -> dict[str, Any]:
    policy = read_json(POLICY_PATH)
    claimed = policy.get("canonical_self_hash")
    payload = dict(policy)
    payload.pop("canonical_self_hash", None)
    if claimed != canonical_sha256(payload):
        raise MethodRootQualityAttempt2Error("Quality attempt2 policy self-hash drift")
    expected_runtime = policy["runtime"]
    observed_runtime = {
        "python": platform.python_version(),
        "numpy": np.__version__,
        "scikit_learn": sklearn.__version__,
        "threads": 1,
        "float_dtype": "<f8",
    }
    if observed_runtime != expected_runtime:
        raise MethodRootQualityAttempt2Error(
            f"Frozen runtime drift: expected={expected_runtime} observed={observed_runtime}"
        )
    for pin in policy["pins"].values():
        path = ROOT / str(pin["path"])
        if (
            not path.is_file()
            or path.stat().st_size != int(pin["size_bytes"])
            or sha256_file(path) != str(pin["sha256"])
        ):
            raise MethodRootQualityAttempt2Error(f"Frozen input pin drift: {path}")
    family = policy["text_probe_family"]
    if (
        tuple(family["view_names"]) != text_views.VIEW_ORDER
        or tuple(family["feature_widths"]) != text_views.EXPECTED_WIDTHS
        or dict(family["feature_name_sha256"]) != text_views.EXPECTED_NAME_HASHES
        or int(family["single_feature_count"]) != sum(text_views.EXPECTED_WIDTHS)
        or int(family["total_model_count"]) != 14
    ):
        raise MethodRootQualityAttempt2Error("Frozen seven-view family drift")
    return policy


def verify_root_manifest(
    policy: Mapping[str, Any], public_root: Path, private_root: Path
) -> dict[str, Any]:
    manifest = read_json(public_root / "root_manifest.json")
    claimed = manifest.get("canonical_self_hash")
    payload = dict(manifest)
    payload.pop("canonical_self_hash", None)
    pin = policy["pins"]["root_manifest"]
    if (
        claimed != canonical_sha256(payload)
        or claimed != pin["canonical_self_hash"]
        or manifest.get("status") != "BUILT_NOT_TRAINING_QUALIFIED"
        or manifest.get("audit_truth_read_counts") != {"audit_a": 0, "audit_b": 0}
    ):
        raise MethodRootQualityAttempt2Error("Method-root manifest closure drift")
    for record in manifest["public_files"]:
        path = public_root / str(record["path"])
        if (
            not path.is_file()
            or path.stat().st_size != int(record["size_bytes"])
            or sha256_file(path) != str(record["sha256"])
        ):
            raise MethodRootQualityAttempt2Error(
                f"Public manifest payload drift: {record['path']}"
            )
    # This is raw-byte verification only.  In particular, audit A/B truth is
    # never parsed or represented as row objects by this process.
    for record in manifest["private_file_commitments"]:
        path = private_root / str(record["path"])
        if (
            not path.is_file()
            or path.stat().st_size != int(record["size_bytes"])
            or sha256_file(path) != str(record["sha256"])
        ):
            raise MethodRootQualityAttempt2Error(
                f"Private byte commitment drift: {record['path']}"
            )
    return manifest


def verify_formal_authorization(policy: Mapping[str, Any]) -> dict[str, Any]:
    path = ROOT / str(policy["authorization_path"])
    authorization = read_json(path)
    claimed = authorization.get("canonical_self_hash")
    payload = dict(authorization)
    payload.pop("canonical_self_hash", None)
    if claimed != canonical_sha256(payload):
        raise MethodRootQualityAttempt2Error("Quality authorization self-hash drift")
    expected = {
        "authorized": True,
        "policy_sha256": sha256_file(POLICY_PATH),
        "script_sha256": sha256_file(Path(__file__).resolve()),
        "root_manifest_sha256": policy["pins"]["root_manifest"]["sha256"],
        "root_manifest_canonical_self_hash": policy["pins"]["root_manifest"]["canonical_self_hash"],
        "output_root": policy["output_root"],
        "temporary_root": policy["temporary_root"],
        "git_commit": git_value("rev-parse", "HEAD"),
        "git_tree": git_value("rev-parse", "HEAD^{tree}"),
        "audit_a_truth_authorized": False,
        "audit_b_truth_authorized": False,
        "model_training_authorized": False,
        "formal_500x4_generation_authorized": False,
    }
    for name, expected_value in expected.items():
        if authorization.get(name) != expected_value:
            raise MethodRootQualityAttempt2Error(
                f"Quality authorization field drift: {name}"
            )
    return authorization


def consume_authorization(
    policy: Mapping[str, Any], authorization: Mapping[str, Any]
) -> dict[str, Any]:
    path = ROOT / str(policy["consumption_path"])
    if path.exists():
        raise MethodRootQualityAttempt2Error("Quality attempt2 authorization already consumed")
    receipt: dict[str, Any] = {
        "version": VERSION,
        "status": "CONSUMED_BEFORE_FIRST_DATA_ROW",
        "authorization_sha256": sha256_file(ROOT / str(policy["authorization_path"])),
        "authorization_canonical_self_hash": authorization["canonical_self_hash"],
        "policy_sha256": sha256_file(POLICY_PATH),
        "git_commit": authorization["git_commit"],
        "git_tree": authorization["git_tree"],
        "output_root": policy["output_root"],
        "audit_a_truth_authorized": False,
        "audit_b_truth_authorized": False,
        "model_training_authorized": False,
    }
    receipt["canonical_self_hash"] = canonical_sha256(receipt)
    write_json_exclusive(path, receipt)
    return receipt


def formal_key(name: str) -> bytes:
    build_policy = read_json(builder.POLICY_PATH)
    build_authorization = read_json(ROOT / str(build_policy["formal_authorization_path"]))
    spec = build_authorization["key_files"][name]
    path = ROOT / str(spec["path"])
    value = path.read_bytes() if path.is_file() else b""
    if len(value) != 32 or hashlib.sha256(value).hexdigest() != spec["commitment_sha256"]:
        raise MethodRootQualityAttempt2Error(f"Retained method-root authority drift: {name}")
    return value


def retained_time_key() -> bytes:
    value = builder.TIME_KEY_PATH.read_bytes() if builder.TIME_KEY_PATH.is_file() else b""
    expected = "b99fe117617313ec2cda0228d8d40d56ccea8f63891425fe5b2332dc5b338c82"
    if len(value) != 32 or hashlib.sha256(value).hexdigest() != expected:
        raise MethodRootQualityAttempt2Error("Retained V9.4 time-key commitment drift")
    return value


def grouped(rows: Sequence[Mapping[str, Any]], field: str) -> dict[str, list[dict[str, Any]]]:
    output: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        output[str(row[field])].append(dict(row))
    return dict(output)


def load_public_split(public_root: Path, split: str) -> dict[str, Any]:
    observed = public_root / split / "observed"
    worlds = read_jsonl(observed / "worlds.jsonl")
    sellers = read_jsonl(observed / "sellers.jsonl")
    items = read_jsonl(observed / "items.jsonl")
    replay_items = read_jsonl(observed / "model_visible_replay_items.jsonl")
    redacted = read_jsonl(observed / "redacted_items.jsonl")
    profiles = read_jsonl(observed / "model_seller_profiles.jsonl")
    endpoints = read_csv(observed / "complete_model_pair_endpoints.csv")
    identity33 = read_csv(observed / "identity33_all_pairs.csv")
    if any(tuple(row) != ENDPOINT_FIELDS for row in endpoints):
        raise MethodRootQualityAttempt2Error("Endpoint field order drift")
    seller_world = {
        str(row["seller_uid"]): str(row["world_uid"]) for row in sellers
    }
    profiles_by_world: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in profiles:
        seller_uid = str(row["seller_uid"])
        if seller_uid not in seller_world:
            raise MethodRootQualityAttempt2Error("Profile/seller join drift")
        profiles_by_world[seller_world[seller_uid]].append(dict(row))
    return {
        "worlds": worlds,
        "sellers": sellers,
        "items": items,
        "replay_items": replay_items,
        "redacted": redacted,
        "profiles": profiles,
        "endpoints": endpoints,
        "identity33": identity33,
        "by_world": {
            "sellers": grouped(sellers, "world_uid"),
            "items": grouped(items, "world_uid"),
            "replay_items": grouped(replay_items, "world_uid"),
            "redacted": grouped(redacted, "world_uid"),
            "profiles": dict(profiles_by_world),
            "endpoints": grouped(endpoints, "world_uid"),
            "identity33": grouped(identity33, "world_uid"),
        },
    }


def scheduled_worlds(split: str) -> list[builder.PublicWorld]:
    schedule = schedule_v94.build_split_schedule(split)
    return [
        builder.PublicWorld(
            split=split,
            ordinal=int(public["world_ordinal"]),
            world_uid=str(public["world_uid"]),
            seller_uids=tuple(public["seller_uids"]),
            noise_slots=tuple(public["noise_slot_by_seller_slot"]),
            controller_groups=tuple(tuple(group) for group in groups),
        )
        for public, groups in zip(
            schedule.public_worlds,
            schedule.controller_groups_by_world,
            strict=True,
        )
    ]


def normalized_identity_rows(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for row in rows:
        projected: dict[str, str] = {
            "canonical_pair_uid": str(row["canonical_pair_uid"]),
            "world_uid": str(row["world_uid"]),
        }
        for name in row:
            if name not in {"canonical_pair_uid", "world_uid"}:
                projected[name] = f"{float(row[name]):.12f}"
        output.append(projected)
    return output


def render_world_surface_once(
    *,
    world: builder.PublicWorld,
    split_data: Mapping[str, Any],
    generation_audit: Mapping[str, Any],
    base_policy: Mapping[str, Any],
    template: Mapping[str, Any],
    text_key: bytes,
    uid_key: bytes,
) -> dict[str, Any]:
    world_uid = world.world_uid
    source = split_data["by_world"]
    sellers = [dict(row) for row in source["sellers"][world_uid]]
    raw_items = [dict(row) for row in source["items"][world_uid]]
    replay_items = [dict(row) for row in source["replay_items"][world_uid]]
    persisted_redacted = [dict(row) for row in source["redacted"][world_uid]]
    persisted_profiles = [dict(row) for row in source["profiles"][world_uid]]
    endpoints = [dict(row) for row in source["endpoints"][world_uid]]
    persisted_identity33 = [dict(row) for row in source["identity33"][world_uid]]
    if (
        len(sellers) != 28
        or len(endpoints) != 378
        or tuple(row["seller_uid"] for row in sellers) != world.seller_uids
    ):
        raise MethodRootQualityAttempt2Error("World public key/order drift")
    styles = {
        str(row["seller_uid"]): {
            name: row[name]
            for name in (
                "seller_uid",
                "separator",
                "ending",
                "line_mode",
                "english_tag",
                "traditional_variant",
                "repeat_punctuation",
                "base_style_id",
                "perturbed_fields",
            )
        }
        for row in generation_audit["style_assignments"]
    }
    if set(styles) != set(world.seller_uids):
        raise MethodRootQualityAttempt2Error("Style-assignment seller universe drift")
    derangement = style_derangement.build_style_source_derangement(
        split=world.split,
        world_uid=world_uid,
        seller_uids=world.seller_uids,
    )
    source_style = derangement.as_mapping()
    if set(source_style) != set(world.seller_uids) or any(
        seller == source_style[seller] for seller in world.seller_uids
    ):
        raise MethodRootQualityAttempt2Error("Style derangement contract drift")
    noise_by_seller = {
        seller_uid: world.noise_slots[index]
        for index, seller_uid in enumerate(world.seller_uids)
    }
    clean_by_uid = {str(row["item_uid"]): row for row in persisted_redacted}
    raw_by_uid = {str(row["item_uid"]): row for row in raw_items}
    original_render: dict[str, tuple[str, str]] = {}
    counterfactual_render: dict[str, tuple[str, str]] = {}
    for row in replay_items:
        item_uid = str(row["item_uid"])
        seller_uid = str(row["seller_uid"])
        ordinal = int(row["logical_item_ordinal"])
        clean = clean_by_uid[item_uid]
        original_title, original_description, _category, _components = builder.render_base_item(
            world=world,
            seller_uid=seller_uid,
            noise_slot=noise_by_seller[seller_uid],
            ordinal=ordinal,
            title_nonempty=bool(clean["title"]),
            description_nonempty=bool(clean["description"]),
            style=styles[seller_uid],
            template=template,
            key=text_key,
        )
        cf_title, cf_description, _category, _components = builder.render_base_item(
            world=world,
            seller_uid=seller_uid,
            noise_slot=noise_by_seller[seller_uid],
            ordinal=ordinal,
            title_nonempty=bool(clean["title"]),
            description_nonempty=bool(clean["description"]),
            style=styles[source_style[seller_uid]],
            template=template,
            key=text_key,
        )
        original_render[item_uid] = (original_title, original_description)
        counterfactual_render[item_uid] = (cf_title, cf_description)
    controls = generation_audit["registered_negative_controls"]
    if len(controls) != 6:
        raise MethodRootQualityAttempt2Error("Registered text control count drift")
    for control in controls:
        source_uid = str(control["source_item_uid"])
        target_uid = str(control["target_item_uid"])
        control_type = str(control["control_type"])
        if control_type == "exact_title_clone_negative":
            suffix = ""
        elif control_type == "high_semantic_similarity_negative":
            suffix = " 配置另选"
        else:
            raise MethodRootQualityAttempt2Error("Unknown text-control type")
        original_render[target_uid] = (
            original_render[source_uid][0] + suffix,
            original_render[target_uid][1],
        )
        counterfactual_render[target_uid] = (
            counterfactual_render[source_uid][0] + suffix,
            counterfactual_render[target_uid][1],
        )
    for item_uid, clean in clean_by_uid.items():
        if original_render[item_uid] != (str(clean["title"]), str(clean["description"])):
            raise MethodRootQualityAttempt2Error("Original text renderer replay drift")
    counterfactual_redacted = [
        {
            "world_uid": str(row["world_uid"]),
            "seller_uid": str(row["seller_uid"]),
            "item_uid": str(row["item_uid"]),
            "title": counterfactual_render[str(row["item_uid"])][0],
            "description": counterfactual_render[str(row["item_uid"])][1],
        }
        for row in persisted_redacted
    ]
    original_safe_items: list[dict[str, Any]] = []
    counterfactual_safe_items: list[dict[str, Any]] = []
    counterfactual_raw_items: list[dict[str, Any]] = []
    for raw in raw_items:
        item_uid = str(raw["item_uid"])
        clean = clean_by_uid[item_uid]
        cf_title, cf_description = counterfactual_render[item_uid]
        raw_description = str(raw["description"])
        clean_description = str(clean["description"])
        if not raw_description.startswith(clean_description):
            raise MethodRootQualityAttempt2Error("Identity suffix prefix replay drift")
        identity_suffix = raw_description[len(clean_description) :]
        original_safe_items.append(
            {**raw, "title": str(clean["title"]), "description": clean_description}
        )
        counterfactual_safe_items.append(
            {**raw, "title": cf_title, "description": cf_description}
        )
        counterfactual_raw_items.append(
            {**raw, "title": cf_title, "description": cf_description + identity_suffix}
        )
    original_profiles, _original_profile_audit = profiles_module.build_world_profiles(
        base_policy,
        mode="formal",
        split=world.split,
        sellers=sellers,
        items=original_safe_items,
    )
    original_model_profiles = list(
        scientific_common.project_model_seller_profiles(original_profiles)
    )
    if canonical_bytes(original_model_profiles) != canonical_bytes(persisted_profiles):
        raise MethodRootQualityAttempt2Error("Original production profile replay drift")
    counterfactual_profiles, _cf_profile_audit = profiles_module.build_world_profiles(
        base_policy,
        mode="formal",
        split=world.split,
        sellers=sellers,
        items=counterfactual_safe_items,
    )
    counterfactual_model_profiles = list(
        scientific_common.project_model_seller_profiles(counterfactual_profiles)
    )
    parsed, history_rows, parser_audit = builder.production_parse_with_uid_aliases(
        world=world,
        sellers=sellers,
        items=counterfactual_raw_items,
        base_policy=base_policy,
        key=uid_key,
    )
    counterfactual_identity33 = builder.build_identity33(
        base_policy=base_policy,
        endpoints=endpoints,
        history_rows=history_rows,
    )
    if normalized_identity_rows(counterfactual_identity33) != normalized_identity_rows(
        persisted_identity33
    ):
        raise MethodRootQualityAttempt2Error("Counterfactual identity33 invariance drift")
    original_views, names = text_views.build_text_probe_views(
        items=persisted_redacted,
        profiles=original_model_profiles,
        endpoints=endpoints,
    )
    counterfactual_views, cf_names = text_views.build_text_probe_views(
        items=counterfactual_redacted,
        profiles=counterfactual_model_profiles,
        endpoints=endpoints,
    )
    if names != cf_names:
        raise MethodRootQualityAttempt2Error("Original/counterfactual view schema drift")
    return {
        "original_views": original_views,
        "counterfactual_views": counterfactual_views,
        "feature_names": names,
        "counterfactual_redacted": counterfactual_redacted,
        "counterfactual_profiles": counterfactual_model_profiles,
        "counterfactual_parsed": parsed,
        "counterfactual_history": history_rows,
        "counterfactual_identity33": counterfactual_identity33,
        "parser_audit": parser_audit,
        "derangement_mapping_sha256": derangement.mapping_sha256,
        "excluded_pair_uids": tuple(
            str(row["canonical_pair_uid"]) for row in controls
        ),
        "style_tuple_count": len(
            {
                tuple(styles[seller][name] for name in builder.STYLE_FIELDS)
                for seller in world.seller_uids
            }
        ),
        "base_style_count": len(
            {styles[seller]["base_style_id"] for seller in world.seller_uids}
        ),
        "seller_two_perturbation_count": sum(
            len(styles[seller]["perturbed_fields"]) == 2
            for seller in world.seller_uids
        ),
    }


def assert_counterfactual_replay(
    first: Mapping[str, Any], second: Mapping[str, Any]
) -> None:
    keys = (
        "counterfactual_redacted",
        "counterfactual_profiles",
        "counterfactual_parsed",
        "counterfactual_history",
        "counterfactual_identity33",
        "parser_audit",
        "derangement_mapping_sha256",
        "excluded_pair_uids",
    )
    for name in keys:
        if canonical_bytes(first[name]) != canonical_bytes(second[name]):
            raise MethodRootQualityAttempt2Error(
                f"Counterfactual independent replay drift: {name}"
            )


def allocate_matrix_files(
    temp_root: Path, split: str, surface: str, row_count: int
) -> dict[str, np.memmap]:
    output: dict[str, np.memmap] = {}
    for view, width in zip(text_views.VIEW_ORDER, text_views.EXPECTED_WIDTHS, strict=True):
        path = temp_root / f"{split}.{surface}.{view}.npy"
        output[view] = np.lib.format.open_memmap(
            path,
            mode="w+",
            dtype=np.dtype("<f8"),
            shape=(row_count, width),
            fortran_order=False,
        )
    return output


def build_split_matrices(
    *,
    split: str,
    split_data: Mapping[str, Any],
    private_root: Path,
    temp_root: Path,
    policy: Mapping[str, Any],
    base_policy: Mapping[str, Any],
    template: Mapping[str, Any],
    text_key: bytes,
    uid_key: bytes,
    world_limit: int | None = None,
) -> dict[str, Any]:
    worlds = scheduled_worlds(split)
    if world_limit is not None:
        worlds = worlds[:world_limit]
    public_world_uids = [str(row["world_uid"]) for row in split_data["worlds"]]
    if public_world_uids[: len(worlds)] != [world.world_uid for world in worlds]:
        raise MethodRootQualityAttempt2Error("Public/scheduled world order drift")
    audit_rows = read_jsonl(private_root / split / "generation_audit.jsonl")
    audit_by_world = {str(row["world_uid"]): row for row in audit_rows}
    if len(audit_by_world) != int(policy["splits"]["world_counts"][split]):
        raise MethodRootQualityAttempt2Error("Generation-audit world count drift")
    row_count = len(worlds) * int(policy["splits"]["eligible_pairs_per_world"])
    original = allocate_matrix_files(temp_root, split, "original", row_count)
    counterfactual = allocate_matrix_files(temp_root, split, "counterfactual", row_count)
    row_keys: list[tuple[str, str]] = []
    world_uids: list[str] = []
    feature_names: dict[str, tuple[str, ...]] | None = None
    derangement_hashes: list[tuple[str, str]] = []
    original_input_hashes: list[tuple[str, str]] = []
    counterfactual_input_hashes: list[tuple[str, str]] = []
    style_counts = Counter()
    cursor = 0
    for world_index, world in enumerate(worlds):
        audit = audit_by_world.get(world.world_uid)
        if audit is None:
            raise MethodRootQualityAttempt2Error("Generation-audit world join drift")
        first = render_world_surface_once(
            world=world,
            split_data=split_data,
            generation_audit=audit,
            base_policy=base_policy,
            template=template,
            text_key=text_key,
            uid_key=uid_key,
        )
        second = render_world_surface_once(
            world=world,
            split_data=split_data,
            generation_audit=audit,
            base_policy=base_policy,
            template=template,
            text_key=text_key,
            uid_key=uid_key,
        )
        assert_counterfactual_replay(first, second)
        endpoints = split_data["by_world"]["endpoints"][world.world_uid]
        excluded = set(first["excluded_pair_uids"])
        keep = np.asarray(
            [str(row["canonical_pair_uid"]) not in excluded for row in endpoints],
            dtype=bool,
        )
        if int(keep.sum()) != 372 or len(excluded) != 6:
            raise MethodRootQualityAttempt2Error("Mechanism-neutral text mask drift")
        stop = cursor + 372
        for view in text_views.VIEW_ORDER:
            original[view][cursor:stop] = first["original_views"][view][keep]
            counterfactual[view][cursor:stop] = first["counterfactual_views"][view][keep]
        kept_endpoints = [row for row, eligible in zip(endpoints, keep, strict=True) if eligible]
        row_keys.extend(
            (world.world_uid, str(row["canonical_pair_uid"]))
            for row in kept_endpoints
        )
        world_uids.extend([world.world_uid] * 372)
        feature_names = first["feature_names"]
        derangement_hashes.append(
            (world.world_uid, str(first["derangement_mapping_sha256"]))
        )
        original_input_hashes.append(
            (
                world.world_uid,
                canonical_sha256(
                    {
                        "items": split_data["by_world"]["redacted"][world.world_uid],
                        "profiles": split_data["by_world"]["profiles"][world.world_uid],
                    }
                ),
            )
        )
        counterfactual_input_hashes.append(
            (
                world.world_uid,
                canonical_sha256(
                    {
                        "items": first["counterfactual_redacted"],
                        "profiles": first["counterfactual_profiles"],
                    }
                ),
            )
        )
        style_counts["worlds"] += 1
        style_counts["minimum_style_tuple_count"] = (
            first["style_tuple_count"]
            if style_counts["minimum_style_tuple_count"] == 0
            else min(style_counts["minimum_style_tuple_count"], first["style_tuple_count"])
        )
        style_counts["minimum_base_style_count"] = (
            first["base_style_count"]
            if style_counts["minimum_base_style_count"] == 0
            else min(style_counts["minimum_base_style_count"], first["base_style_count"])
        )
        style_counts["seller_two_perturbation_count"] += int(
            first["seller_two_perturbation_count"]
        )
        cursor = stop
        if (world_index + 1) % 25 == 0 or world_index + 1 == len(worlds):
            print(
                f"[{split}] counterfactual replay and seven views: "
                f"{world_index + 1}/{len(worlds)} worlds",
                flush=True,
            )
    if cursor != row_count or feature_names is None:
        raise MethodRootQualityAttempt2Error("Frozen matrix row closure drift")
    for matrices in (original, counterfactual):
        for matrix in matrices.values():
            matrix.flush()
            if matrix.dtype != np.dtype("<f8") or not np.isfinite(matrix).all():
                raise MethodRootQualityAttempt2Error("Frozen matrix numerical drift")
    return {
        "original": original,
        "counterfactual": counterfactual,
        "feature_names": feature_names,
        "row_keys": tuple(row_keys),
        "world_uids": tuple(world_uids),
        "ordered_world_uids": tuple(world.world_uid for world in worlds),
        "matrix_commitments": {
            surface: {
                view: hashlib.sha256(
                    np.ascontiguousarray(matrix, dtype="<f8").tobytes(order="C")
                ).hexdigest()
                for view, matrix in matrices.items()
            }
            for surface, matrices in (
                ("original", original),
                ("counterfactual", counterfactual),
            )
        },
        "row_key_sha256": canonical_sha256(row_keys),
        "world_order_sha256": canonical_sha256(
            [world.world_uid for world in worlds]
        ),
        "derangement_mapping_commitment_sha256": canonical_sha256(
            derangement_hashes
        ),
        "original_input_commitment_sha256": canonical_sha256(original_input_hashes),
        "counterfactual_input_commitment_sha256": canonical_sha256(
            counterfactual_input_hashes
        ),
        "style_structure": dict(style_counts),
    }


def load_truth_once(
    *,
    private_root: Path,
    split: str,
    row_keys: Sequence[tuple[str, str]],
    expected_worlds: int,
) -> tuple[np.ndarray, dict[str, Any], dict[tuple[str, str], int]]:
    rows = read_csv(private_root / split / "pair_labels.csv")
    index: dict[tuple[str, str], int] = {}
    for row in rows:
        key = (str(row["world_uid"]), str(row["canonical_pair_uid"]))
        if key in index or str(row["label"]) not in {"0", "1"}:
            raise MethodRootQualityAttempt2Error("Private pair-label schema drift")
        index[key] = int(row["label"])
    labels = np.asarray([index[key] for key in row_keys], dtype=np.int8)
    if len(index) != expected_worlds * 378:
        raise MethodRootQualityAttempt2Error("Private pair-label count drift")
    positives_by_world = Counter()
    counts_by_world = Counter()
    for (world_uid, _pair_uid), label in zip(row_keys, labels, strict=True):
        positives_by_world[world_uid] += int(label)
        counts_by_world[world_uid] += 1
    if (
        set(counts_by_world.values()) != {372}
        or set(positives_by_world.values()) != {20}
        or len(counts_by_world) != expected_worlds
    ):
        raise MethodRootQualityAttempt2Error("Eligible truth per-world closure drift")
    receipt = {
        "split": split,
        "semantic_read_count": 1,
        "full_truth_row_count": len(rows),
        "eligible_truth_row_count": len(labels),
        "eligible_positive_count": int(labels.sum()),
        "eligible_negative_count": int(len(labels) - labels.sum()),
        "eligible_label_vector_sha256": hashlib.sha256(labels.tobytes()).hexdigest(),
        "row_key_sha256": canonical_sha256(row_keys),
    }
    return labels, receipt, index


def maximum_single_feature(
    matrices: Mapping[str, np.ndarray], labels: np.ndarray, names: Mapping[str, Sequence[str]]
) -> dict[str, Any]:
    best = -math.inf
    winners: list[str] = []
    evaluated = 0
    for view in text_views.VIEW_ORDER:
        matrix = matrices[view]
        for column_index, column_name in enumerate(names[view]):
            metric = probe_validator.symmetric_auc(labels, matrix[:, column_index])
            qualified = f"{view}::{column_name}"
            evaluated += 1
            if metric > best:
                best = metric
                winners = [qualified]
            elif metric == best:
                winners.append(qualified)
    return {
        "evaluated_feature_count": evaluated,
        "maximum_symmetric_roc_auc": float(best),
        "winner": min(winners, key=lambda value: value.encode("utf-8")),
        "tie_count": len(winners),
    }


def fit_model_family(
    *,
    train_matrices: Mapping[str, np.ndarray],
    development_matrices: Mapping[str, np.ndarray],
    train_labels: np.ndarray,
    development_labels: np.ndarray,
    policy: Mapping[str, Any],
    role: str,
) -> tuple[dict[str, Any], dict[str, np.ndarray]]:
    metrics: dict[str, Any] = {}
    scores: dict[str, np.ndarray] = {}
    baseline = float(policy["text_probe_family"]["average_precision_baseline"])
    for view_index, view in enumerate(text_views.VIEW_ORDER):
        fitted = probe_validator._fit_probe_models(
            train_x=train_matrices[view],
            train_y=train_labels,
            development_x=development_matrices[view],
            policy=policy,
        )
        for model_name, score in fitted.items():
            name = f"{view}::{model_name}"
            auc = float(roc_auc_score(development_labels, score))
            ap = float(average_precision_score(development_labels, score))
            scores[name] = score
            metrics[name] = {
                "view": view,
                "model": model_name,
                "qualification_role": role,
                "development_roc_auc": auc,
                "development_symmetric_roc_auc": max(auc, 1.0 - auc),
                "development_average_precision": ap,
                "development_average_precision_uplift": ap - baseline,
                "prediction_vector_sha256": hashlib.sha256(
                    np.asarray(score, dtype="<f8").tobytes(order="C")
                ).hexdigest(),
            }
        print(
            f"[{role}] fitted view {view_index + 1}/7: {view}",
            flush=True,
        )
    return metrics, scores


def identity_positive_control(
    *,
    public_root: Path,
    policy: Mapping[str, Any],
    truth_indexes: Mapping[str, Mapping[tuple[str, str], int]],
) -> dict[str, Any]:
    matrices: dict[str, np.ndarray] = {}
    labels: dict[str, np.ndarray] = {}
    for split in ("train", "development"):
        identity_rows = read_csv(
            public_root / split / "observed" / "identity33_all_pairs.csv"
        )
        truth = truth_indexes[split]
        feature_names = [
            name
            for name in identity_rows[0]
            if name not in {"canonical_pair_uid", "world_uid"}
        ]
        if len(feature_names) != 33:
            raise MethodRootQualityAttempt2Error("Identity33 width drift")
        keys = [(row["world_uid"], row["canonical_pair_uid"]) for row in identity_rows]
        matrices[split] = np.ascontiguousarray(
            [[float(row[name]) for name in feature_names] for row in identity_rows],
            dtype="<f8",
        )
        labels[split] = np.asarray([truth[key] for key in keys], dtype=np.int8)
    scaler = StandardScaler().fit(matrices["train"])
    model = LogisticRegression(
        C=1.0,
        penalty="l2",
        solver="lbfgs",
        max_iter=10000,
        tol=1e-10,
        random_state=281320828,
    ).fit(scaler.transform(matrices["train"]), labels["train"])
    score = model.predict_proba(scaler.transform(matrices["development"]))[:, 1]
    auc = float(roc_auc_score(labels["development"], score))
    ap = float(average_precision_score(labels["development"], score))
    gates = policy["gates"]
    return {
        "role": "POSITIVE_CONTROL_NOT_TEXT_SHORTCUT_GATE",
        "development_roc_auc": auc,
        "development_average_precision": ap,
        "minimum_roc_auc": gates["identity33_minimum_roc_auc"],
        "minimum_average_precision": gates["identity33_minimum_average_precision"],
        "passed": (
            auc >= float(gates["identity33_minimum_roc_auc"])
            and ap >= float(gates["identity33_minimum_average_precision"])
        ),
        "prediction_vector_sha256": hashlib.sha256(
            np.asarray(score, dtype="<f8").tobytes(order="C")
        ).hexdigest(),
    }


def exact_v94_public_replay(
    *, public_root: Path, time_key: bytes
) -> dict[str, Any]:
    signatures = [dict(row) for row in noise_v94.build_noise_signatures().rows]
    split_commitments: dict[str, str] = {}
    replayed = 0
    for split in ("train", "development"):
        schedule = schedule_v94.build_split_schedule(split)
        endpoints = grouped(
            read_csv(
                public_root
                / split
                / "observed"
                / "complete_model_pair_endpoints.csv"
            ),
            "world_uid",
        )
        items = grouped(
            read_jsonl(
                public_root / split / "observed" / "model_visible_replay_items.jsonl"
            ),
            "world_uid",
        )
        for public in schedule.public_worlds:
            world = {
                "split": split,
                "world_ordinal": int(public["world_ordinal"]),
                "world_uid": str(public["world_uid"]),
                "seller_uids": list(public["seller_uids"]),
                "noise_slot_by_seller_slot": list(public["noise_slot_by_seller_slot"]),
            }
            registered = prebuild_v94.build_truth_free_world_projection(
                world=world,
                noise_signatures=signatures,
                time_key_hex=time_key.hex(),
            )
            replay_v94.require_exact_replay(
                registered_rows=registered,
                public_endpoint_rows=endpoints[world["world_uid"]],
                public_item_rows=items[world["world_uid"]],
            )
            replayed += 1
        split_commitments[split] = str(
            schedule.commitment["split_schedule_commitment_sha256"]
        )
    return {
        "world_count": replayed,
        "split_schedule_commitments": split_commitments,
        "exact_public_14d_replay_passed": replayed == 1000,
    }


def static_visible_surface_audit(public_root: Path) -> dict[str, Any]:
    item_documents: set[str] = set()
    profile_documents: set[str] = set()
    seller_uids: set[str] = set()
    item_uids: set[str] = set()
    collisions = Counter()
    totals = Counter()
    for split in builder.SPLITS:
        observed = public_root / split / "observed"
        sellers = read_jsonl(observed / "sellers.jsonl")
        redacted = read_jsonl(observed / "redacted_items.jsonl")
        profiles = read_jsonl(observed / "model_seller_profiles.jsonl")
        for seller in sellers:
            seller_uid = str(seller["seller_uid"])
            collisions["seller_uid"] += int(seller_uid in seller_uids)
            seller_uids.add(seller_uid)
        for row in redacted:
            item_uid = str(row["item_uid"])
            collisions["item_uid"] += int(item_uid in item_uids)
            item_uids.add(item_uid)
            visible = str(row["title"]) + "\0" + str(row["description"])
            collisions["artificial_code"] += int(
                VISIBLE_ARTIFICIAL_CODE.search(visible) is not None
            )
            collisions["forbidden_marker"] += int(
                builder.FORBIDDEN_VISIBLE_RE.search(visible) is not None
            )
            digest = hashlib.sha256(visible.encode("utf-8")).hexdigest()
            collisions["redacted_document"] += int(digest in item_documents)
            item_documents.add(digest)
        for row in profiles:
            content = {name: value for name, value in row.items() if name != "seller_uid"}
            digest = canonical_sha256(content)
            collisions["model_profile"] += int(digest in profile_documents)
            profile_documents.add(digest)
        totals.update(sellers=len(sellers), items=len(redacted), profiles=len(profiles))
    return {
        "totals": dict(totals),
        "collision_and_forbidden_counts": dict(collisions),
        "passed": not any(collisions.values()),
    }


def readability_audit(public_root: Path) -> dict[str, Any]:
    counts = Counter()
    visible_sellers: set[str] = set()
    all_sellers: set[str] = set()
    title_lengths: list[int] = []
    description_lengths: list[int] = []
    for split in builder.SPLITS:
        for seller in read_jsonl(public_root / split / "observed" / "sellers.jsonl"):
            all_sellers.add(str(seller["seller_uid"]))
        for row in read_jsonl(
            public_root / split / "observed" / "redacted_items.jsonl"
        ):
            counts["item_row_count"] += 1
            seller_uid = str(row["seller_uid"])
            for field, lengths in (
                ("title", title_lengths),
                ("description", description_lengths),
            ):
                value = str(row[field])
                counts["visible_field_count"] += 1
                if value:
                    counts["nonempty_visible_field_count"] += 1
                    visible_sellers.add(seller_uid)
                    lengths.append(len(value))
                counts["unicode_replacement_character_count"] += value.count("\ufffd")
                counts["unresolved_template_brace_count"] += int(
                    "{" in value or "}" in value
                )
                counts["illegal_control_character_count"] += sum(
                    unicodedata.category(character) == "Cc"
                    and character not in {"\n", "\t"}
                    for character in value
                )
                counts["han_character_count"] += sum(
                    "CJK UNIFIED IDEOGRAPH" in unicodedata.name(character, "")
                    for character in value
                )
                counts["non_whitespace_character_count"] += sum(
                    not character.isspace() for character in value
                )
    missing_visible_sellers = all_sellers - visible_sellers
    denominator = max(1, counts["non_whitespace_character_count"])
    return {
        "counts": dict(counts),
        "seller_count": len(all_sellers),
        "seller_without_any_visible_text_count": len(missing_visible_sellers),
        "nonempty_title_length_minimum": min(title_lengths),
        "nonempty_title_length_maximum": max(title_lengths),
        "nonempty_description_length_minimum": min(description_lengths),
        "nonempty_description_length_maximum": max(description_lengths),
        "han_character_share": counts["han_character_count"] / denominator,
        "passed": (
            not missing_visible_sellers
            and counts["unicode_replacement_character_count"] == 0
            and counts["unresolved_template_brace_count"] == 0
            and counts["illegal_control_character_count"] == 0
        ),
    }


def sealed_literal_scan(policy: Mapping[str, Any]) -> dict[str, Any]:
    scanner_path = ROOT / str(policy["pins"]["sealed_literal_scanner"]["path"])
    environment = dict(os.environ)
    environment["PYTHONDONTWRITEBYTECODE"] = "1"
    result = subprocess.run(
        [sys.executable, "-B", str(scanner_path)],
        cwd=ROOT,
        env=environment,
        check=False,
        text=True,
        encoding="utf-8",
        capture_output=True,
    )
    if result.returncode != 0 or result.stderr or not result.stdout.strip():
        raise MethodRootQualityAttempt2Error("Sealed literal scanner execution drift")
    try:
        receipt = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise MethodRootQualityAttempt2Error(
            "Sealed literal scanner output drift"
        ) from exc
    if not isinstance(receipt, dict):
        raise MethodRootQualityAttempt2Error("Sealed literal scanner receipt type drift")
    claimed = receipt.get("canonical_self_hash")
    payload = dict(receipt)
    payload.pop("canonical_self_hash", None)
    if (
        claimed != canonical_sha256(payload)
        or receipt.get("status") != "SEALED_LITERAL_SCAN_COMPLETE"
        or receipt.get("private_values_returned") != 0
        or receipt.get("row_level_hits_returned") != 0
        or receipt.get("pair_labels_parsed") != 0
        or receipt.get("controller_relations_returned") != 0
        or receipt.get("qrels_returned") != 0
    ):
        raise MethodRootQualityAttempt2Error("Sealed literal scanner boundary drift")
    return receipt


def source_copy_audit(public_root: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    def clean(value: object) -> str:
        if value is None:
            return ""
        return " ".join(unicodedata.normalize("NFKC", str(value)).split())

    source_values: set[str] = set()
    workbook = load_workbook(ROOT / "market_item.xlsx", read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for value in (row[2], row[3]):
                normalized = clean(value)
                if normalized:
                    source_values.add(normalized)
    finally:
        workbook.close()
    copies = 0
    for split in builder.SPLITS:
        for row in read_jsonl(
            public_root / split / "observed" / "redacted_items.jsonl"
        ):
            for field in ("title", "description"):
                value = clean(row[field])
                copies += int(bool(value) and value in source_values)
    return {
        "real_source_field_count": len(source_values),
        "exact_real_source_text_copy_count": copies,
        "passed": copies == 0,
    }


def registered_private_summary(
    private_root: Path, world_counts: Mapping[str, Any]
) -> dict[str, Any]:
    mechanism_counts: Counter[str] = Counter()
    control_counts: Counter[str] = Counter()
    identity_types: set[str] = set()
    identity_hashes: set[str] = set()
    active_worlds = 0
    parsed_rows = 0
    for split in ("train", "development"):
        audits = read_jsonl(private_root / split / "generation_audit.jsonl")
        plans = read_jsonl(private_root / split / "identity_plan.jsonl")
        with (private_root / split / "parsed_identity_occurrences.jsonl").open(
            "r", encoding="utf-8"
        ) as stream:
            parsed_rows += sum(bool(line.strip()) for line in stream)
        if len(audits) != int(world_counts[split]):
            raise MethodRootQualityAttempt2Error("Private generation-audit count drift")
        for audit in audits:
            active_worlds += int(audit["identity33_active_pair_count"] > 0)
            mechanism_counts.update(
                str(row["mechanism"]) for row in audit["mechanism_assignments"]
            )
            control_counts.update(
                str(row["control_type"])
                for row in audit["registered_negative_controls"]
            )
        for plan in plans:
            identity_types.add(str(plan["identity_type"]))
            digest = str(plan["value_sha256"])
            if digest in identity_hashes:
                raise MethodRootQualityAttempt2Error("Identity value reuse drift")
            identity_hashes.add(digest)
    expected_worlds = int(world_counts["train"]) + int(world_counts["development"])
    return {
        "world_count": expected_worlds,
        "identity33_active_world_count": active_worlds,
        "identity_types": sorted(identity_types),
        "unique_identity_value_hash_count": len(identity_hashes),
        "mechanism_counts": dict(sorted(mechanism_counts.items())),
        "registered_negative_control_counts": dict(sorted(control_counts.items())),
        "production_parser_row_count": parsed_rows,
        "all_eight_identity_types_present": identity_types == set(builder.IDENTITY_TYPES),
        "all_worlds_identity33_non_degenerate": active_worlds == expected_worlds,
        "two_clone_and_four_semantic_controls_per_world": (
            control_counts["exact_title_clone_negative"] == 2 * expected_worlds
            and control_counts["high_semantic_similarity_negative"]
            == 4 * expected_worlds
        ),
    }


def run_audit(
    *, policy: Mapping[str, Any], public_root: Path, private_root: Path, temp_root: Path
) -> dict[str, Any]:
    manifest = verify_root_manifest(policy, public_root, private_root)
    static_surface = static_visible_surface_audit(public_root)
    readability = readability_audit(public_root)
    literal_scan = sealed_literal_scan(policy)
    source_copy = source_copy_audit(public_root)
    public_replay = exact_v94_public_replay(
        public_root=public_root, time_key=retained_time_key()
    )
    base_policy = read_json(BASE_POLICY_PATH)
    template = read_json(TEXT_TEMPLATE_PATH)
    text_key = formal_key("text")
    uid_key = formal_key("uid")
    split_data: dict[str, dict[str, Any]] = {}
    matrix_bundles: dict[str, dict[str, Any]] = {}
    for split in ("train", "development"):
        print(f"[{split}] loading public model surfaces", flush=True)
        split_data[split] = load_public_split(public_root, split)
        matrix_bundles[split] = build_split_matrices(
            split=split,
            split_data=split_data[split],
            private_root=private_root,
            temp_root=temp_root,
            policy=policy,
            base_policy=base_policy,
            template=template,
            text_key=text_key,
            uid_key=uid_key,
        )
        # Free the large row objects as soon as their matrices are frozen.
        split_data.pop(split)
    print("All original/counterfactual matrices frozen; opening train/development truth", flush=True)
    train_labels, train_truth_receipt, train_truth_index = load_truth_once(
        private_root=private_root,
        split="train",
        row_keys=matrix_bundles["train"]["row_keys"],
        expected_worlds=500,
    )
    development_labels, development_truth_receipt, development_truth_index = load_truth_once(
        private_root=private_root,
        split="development",
        row_keys=matrix_bundles["development"]["row_keys"],
        expected_worlds=500,
    )
    original_metrics, _original_scores = fit_model_family(
        train_matrices=matrix_bundles["train"]["original"],
        development_matrices=matrix_bundles["development"]["original"],
        train_labels=train_labels,
        development_labels=development_labels,
        policy=policy,
        role="DESCRIPTIVE_AUTHOR_STYLE_NOT_A_QUALITY_GATE",
    )
    counterfactual_metrics, counterfactual_scores = fit_model_family(
        train_matrices=matrix_bundles["train"]["counterfactual"],
        development_matrices=matrix_bundles["development"]["counterfactual"],
        train_labels=train_labels,
        development_labels=development_labels,
        policy=policy,
        role="HARD_TEXT_SHORTCUT_GATE",
    )
    single = maximum_single_feature(
        matrix_bundles["development"]["counterfactual"],
        development_labels,
        matrix_bundles["development"]["feature_names"],
    )
    bootstrap_policy = policy["bootstrap"]
    draws = probe_validator.generate_bootstrap_draws(
        replicates=int(bootstrap_policy["replicates"]),
        world_count=int(bootstrap_policy["world_count"]),
        seed=int(bootstrap_policy["seed"]),
    )
    draws_sha = hashlib.sha256(draws.tobytes(order="C")).hexdigest()
    if draws_sha != bootstrap_policy["expected_draws_sha256"]:
        raise MethodRootQualityAttempt2Error("Frozen bootstrap draw commitment drift")
    print("Running 9,999 paired world bootstrap replicates", flush=True)
    bootstrap = probe_validator._bootstrap_family_upper(
        labels=development_labels,
        row_world_uids=matrix_bundles["development"]["world_uids"],
        ordered_world_uids=matrix_bundles["development"]["ordered_world_uids"],
        score_family=counterfactual_scores,
        baseline=float(policy["text_probe_family"]["average_precision_baseline"]),
        draws=draws,
        batch_size=int(bootstrap_policy["batch_size"]),
    )
    identity = identity_positive_control(
        public_root=public_root,
        policy=policy,
        truth_indexes={
            "train": train_truth_index,
            "development": development_truth_index,
        },
    )
    private_summary = registered_private_summary(
        private_root, manifest["world_counts"]
    )
    gates = policy["gates"]
    model_max_auc = max(
        row["development_symmetric_roc_auc"]
        for row in counterfactual_metrics.values()
    )
    model_max_ap_uplift = max(
        row["development_average_precision_uplift"]
        for row in counterfactual_metrics.values()
    )
    style_pass = all(
        bundle["style_structure"]["minimum_style_tuple_count"] >= 2
        and bundle["style_structure"]["minimum_base_style_count"] >= 2
        and bundle["style_structure"]["seller_two_perturbation_count"]
        == 28 * 500
        for bundle in matrix_bundles.values()
    )
    hard_gates = {
        "root_manifest_and_payload_hashes": True,
        "static_visible_surface": static_surface["passed"],
        "all_visible_rows_readability": readability["passed"],
        "sealed_private_literal_leakage_zero": literal_scan["hard_gate_passed"],
        "real_source_exact_copy_zero": source_copy["passed"],
        "exact_v9_4_public_14d_replay": public_replay["exact_public_14d_replay_passed"],
        "style_structure_positive_control": style_pass,
        "counterfactual_independent_replay_and_identity33_invariance": True,
        "all_eight_identity_types": private_summary["all_eight_identity_types_present"],
        "identity33_all_worlds_non_degenerate": private_summary[
            "all_worlds_identity33_non_degenerate"
        ],
        "registered_text_controls": private_summary[
            "two_clone_and_four_semantic_controls_per_world"
        ],
        "identity33_positive_control": identity["passed"],
        "maximum_346_single_feature_symmetric_roc_auc": (
            single["maximum_symmetric_roc_auc"]
            <= float(gates["maximum_single_feature_symmetric_roc_auc"])
        ),
        "maximum_14_model_symmetric_roc_auc": (
            model_max_auc <= float(gates["maximum_model_family_symmetric_roc_auc"])
        ),
        "maximum_14_model_average_precision_uplift": (
            model_max_ap_uplift
            <= float(gates["maximum_model_family_average_precision_uplift"])
        ),
        "bootstrap_95_upper_symmetric_roc_auc": (
            bootstrap["symmetric_auc_95_upper"]
            <= float(gates["bootstrap_95_upper_symmetric_roc_auc"])
        ),
        "bootstrap_95_upper_average_precision_uplift": (
            bootstrap["average_precision_uplift_95_upper"]
            <= float(gates["bootstrap_95_upper_average_precision_uplift"])
        ),
        "audit_a_b_truth_unopened": manifest["audit_truth_read_counts"]
        == {"audit_a": 0, "audit_b": 0},
    }
    passed = all(hard_gates.values())
    result: dict[str, Any] = {
        "version": VERSION,
        "status": (
            "PASSED_METHOD_ROOT_QUALITY_ELIGIBLE_FOR_FORMAL_500X4_APPLICATION"
            if passed
            else "DATASET_INVALIDATED_PENDING_DOCUMENTATION_AND_CLEANUP"
        ),
        "claim_boundary": policy["claim_boundary"],
        "public_root": public_root.relative_to(ROOT).as_posix(),
        "private_root_commitment_only": private_root.relative_to(ROOT).as_posix(),
        "root_manifest_canonical_self_hash": manifest["canonical_self_hash"],
        "policy_canonical_self_hash": policy["canonical_self_hash"],
        "static_visible_surface_audit": static_surface,
        "visible_text_readability_audit": readability,
        "sealed_private_literal_scan": literal_scan,
        "real_source_copy_audit": source_copy,
        "exact_v9_4_public_replay": public_replay,
        "matrix_freeze": {
            split: {
                key: value
                for key, value in bundle.items()
                if key
                in {
                    "matrix_commitments",
                    "row_key_sha256",
                    "world_order_sha256",
                    "derangement_mapping_commitment_sha256",
                    "original_input_commitment_sha256",
                    "counterfactual_input_commitment_sha256",
                    "style_structure",
                }
            }
            for split, bundle in matrix_bundles.items()
        },
        "truth_access": {
            "train": train_truth_receipt,
            "development": development_truth_receipt,
            "audit_a_semantic_reads": 0,
            "audit_b_semantic_reads": 0,
            "opened_after_all_matrices_frozen": True,
        },
        "original_text_descriptive_14_models": original_metrics,
        "counterfactual_text_hard_gate": {
            "single_feature_maximum": single,
            "model_count": len(counterfactual_metrics),
            "maximum_model_symmetric_roc_auc": model_max_auc,
            "maximum_model_average_precision_uplift": model_max_ap_uplift,
            "models": counterfactual_metrics,
            "bootstrap": bootstrap,
        },
        "identity33_positive_control": identity,
        "registered_private_summary_train_development_only": private_summary,
        "hard_gates": hard_gates,
        "method_root_quality_passed": passed,
        "eligible_for_formal_500x4_generation_application": passed,
        "formal_500x4_generated": False,
        "training_qualified": False,
        "m0_m1_m2_m3_training_authorized": False,
        "row_level_labels_returned": 0,
        "row_level_predictions_returned": 0,
    }
    result["canonical_self_hash"] = canonical_sha256(result)
    return result


def preflight(policy: Mapping[str, Any]) -> None:
    public_root = ROOT / str(policy["dataset_root"])
    private_root = ROOT / str(policy["private_root"])
    verify_root_manifest(policy, public_root, private_root)
    draws = probe_validator.generate_bootstrap_draws(
        replicates=int(policy["bootstrap"]["replicates"]),
        world_count=int(policy["bootstrap"]["world_count"]),
        seed=int(policy["bootstrap"]["seed"]),
    )
    if hashlib.sha256(draws.tobytes(order="C")).hexdigest() != policy["bootstrap"][
        "expected_draws_sha256"
    ]:
        raise MethodRootQualityAttempt2Error("Preflight bootstrap binding drift")
    print("V9.4 quality attempt2 preflight passed; no truth row was read")


def one_world_preflight(policy: Mapping[str, Any]) -> None:
    public_root = ROOT / str(policy["dataset_root"])
    private_root = ROOT / str(policy["private_root"])
    verify_root_manifest(policy, public_root, private_root)
    temp_root = ROOT / "reports" / "step28_synthetic_chinese_dataset" / ".quality_attempt2_one_world_preflight"
    if temp_root.exists():
        raise MethodRootQualityAttempt2Error("One-world preflight temporary root exists")
    temp_root.mkdir(parents=True)
    try:
        base_policy = read_json(BASE_POLICY_PATH)
        template = read_json(TEXT_TEMPLATE_PATH)
        for split in ("train", "development"):
            split_data = load_public_split(public_root, split)
            build_split_matrices(
                split=split,
                split_data=split_data,
                private_root=private_root,
                temp_root=temp_root,
                policy=policy,
                base_policy=base_policy,
                template=template,
                text_key=formal_key("text"),
                uid_key=formal_key("uid"),
                world_limit=1,
            )
        print(
            "V9.4 quality attempt2 one-world train/development preflight passed; "
            "no pair-label or audit truth file was parsed"
        )
    finally:
        resolved = temp_root.resolve()
        expected_parent = (
            ROOT / "reports" / "step28_synthetic_chinese_dataset"
        ).resolve()
        if resolved.parent != expected_parent or not resolved.name.startswith(".quality_attempt2_"):
            raise MethodRootQualityAttempt2Error("Unsafe preflight cleanup target")
        shutil.rmtree(resolved, ignore_errors=False)


def formal_run(policy: Mapping[str, Any]) -> None:
    public_root = (ROOT / str(policy["dataset_root"])).resolve()
    private_root = (ROOT / str(policy["private_root"])).resolve()
    output_root = (ROOT / str(policy["output_root"])).resolve()
    temp_root = (ROOT / str(policy["temporary_root"])).resolve()
    if output_root.exists() or temp_root.exists():
        raise MethodRootQualityAttempt2Error("Formal output or temporary path already exists")
    authorization = verify_formal_authorization(policy)
    consumption = consume_authorization(policy, authorization)
    temp_root.mkdir(parents=True, exist_ok=False)
    stage = "quality_audit"
    try:
        result = run_audit(
            policy=policy,
            public_root=public_root,
            private_root=private_root,
            temp_root=temp_root,
        )
        stage = "result_publication"
        output_root.mkdir(parents=True, exist_ok=False)
        result["authorization_consumption"] = {
            "path": policy["consumption_path"],
            "canonical_self_hash": consumption["canonical_self_hash"],
            "sha256": sha256_file(ROOT / str(policy["consumption_path"])),
        }
        result["canonical_self_hash"] = canonical_sha256(
            {key: value for key, value in result.items() if key != "canonical_self_hash"}
        )
        write_json_exclusive(output_root / "quality_result.json", result)
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    except Exception as exc:
        if not output_root.exists():
            output_root.mkdir(parents=True, exist_ok=False)
        terminal: dict[str, Any] = {
            "version": VERSION,
            "status": "AUDITOR_EXECUTION_FAILED_NO_DATASET_CONCLUSION",
            "failure_stage": stage,
            "exception_type": type(exc).__name__,
            "exception_message_sha256": hashlib.sha256(str(exc).encode("utf-8")).hexdigest(),
            "input_dataset_retained": True,
            "temporary_matrices_deleted": True,
            "audit_a_semantic_reads": 0,
            "audit_b_semantic_reads": 0,
            "training_authorized": False,
        }
        terminal["canonical_self_hash"] = canonical_sha256(terminal)
        write_json_exclusive(output_root / "terminal.json", terminal)
        raise
    finally:
        if temp_root.exists():
            expected_parent = (
                ROOT / "reports" / "step28_synthetic_chinese_dataset"
            ).resolve()
            if temp_root.parent != expected_parent or not temp_root.name.startswith(".v9_4_"):
                raise MethodRootQualityAttempt2Error("Unsafe formal cleanup target")
            shutil.rmtree(temp_root, ignore_errors=False)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--preflight", action="store_true")
    mode.add_argument("--one-world-preflight", action="store_true")
    mode.add_argument("--formal", action="store_true")
    args = parser.parse_args()
    policy = verify_policy()
    if args.preflight:
        preflight(policy)
    elif args.one_world_preflight:
        one_world_preflight(policy)
    else:
        formal_run(policy)


if __name__ == "__main__":
    main()
