#!/usr/bin/env python3
"""Build the aggregate-only Chinese train style reference for Step 28-v13."""

from __future__ import annotations

import argparse
import csv
import math
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook

import step28_v13_common as common


WHITESPACE_RE = re.compile(r"\s+")
CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")
# Frozen, deliberately conservative set used only for an aggregate style ratio.
TRADITIONAL_ONLY_CHARS = frozenset(
    "萬與專業東絲兩嚴個豐臨為麗舉麼義烏樂喬習鄉書買亂乾爭於虧雲亞產畝親億"
    "僅從倉儀價眾優會傘偉傳傷倫偽體餘俠侶僥偵側僑儉債傾儲兒兌黨蘭關興養獸"
    "內岡冊寫軍農馮沖決況凍淨淺減湊凱劃劇劉則剛創刪別製剎劑劍辦務動勵勞勢"
    "勳區醫華協單賣盧衛卻廠歷厲壓厭廁廂廈廚廢廣莊慶廬庫應廟龐廳張強彈錄彥"
    "徑後徵德憂憑懷態慣慘慚慶憶總戀懸驚戶撲執擴擔擬擇擊擋據擰擱攔攝擺搖敗"
    "敵數齊斂斃曆書會機殺雜權條來楊極構標樣樓樹橋檔檢櫃欄歐歲歸殘殼毀氣漢"
    "湯溝滅滯滿濾濱瀏覽燈靈災爐點煉熱愛爺牆獨獲環現瑪產畫異當疊癡發盜盞監"
    "盤睏矚礙禮禱種稱穀穩窩竄竅競筆筍節範築簡糧糾紀約紅紋納紙級紛終組結絕"
    "統經綠網緊緒線練縣縮總績織繞繪繫繼續罰羅聽職聯聲聰肅脅脈腳腦臉臺與舊"
    "艙艱藝節莖萬葉著蔣蔥薑藥蘇蘋處虛號螞補裝複規覺覽觀觸訂計訊討訓記講設"
    "訪證評識詐詞詢試詩誠話該詳語誤說請諸諾謀謝譜貝負財貢貨販貪貫責貴貸貿"
    "費賀資賓賞賠賢賬買賣質賭贈趙趕趨車軟轉輪輸辦邊遼達遷過運還進遠違連遲"
    "選遺鄧鄭釋里鑒針鈔鐘鋼錄錢錯鍵鍋鎖鎮鏡長門閃閉問間閣閱隊陽陰陳陸險隨"
    "隱雙雖離難電霧靜頂項順須領頭頰頸頻題額顏風飛飯飲餘館馬馮驅驗驚鬥魚鳥"
    "鳴麥黃點龍龜"
)
ALLOWLIST_COLUMNS = ["seller_uid"]
MANIFEST_REQUIRED_COLUMNS = {
    "seller_uid",
    "source_dataset",
    "source_row_number",
    "data_bucket",
    "eligibility_status",
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", unicodedata.normalize("NFKC", str(value))).strip()


def item_style(title: object, description: object) -> dict[str, float]:
    title_text = clean_text(title)
    description_text = clean_text(description)
    raw_title = "" if title is None else str(title)
    raw_description = "" if description is None else str(description)
    combined = raw_title + " " + raw_description
    visible = [character for character in combined if not character.isspace()]
    cjk = CJK_RE.findall(combined)
    latin = [character for character in combined if character.isascii() and character.isalpha()]

    def ratio(numerator: int, denominator: int) -> float:
        return 0.0 if denominator == 0 else numerator / denominator

    return {
        "title_length": float(len(title_text)),
        "description_length": float(len(description_text)),
        "title_missing": float(not title_text),
        "description_missing": float(not description_text),
        "digit_ratio": ratio(sum(character.isdigit() for character in visible), len(visible)),
        "punctuation_ratio": ratio(
            sum(not character.isalnum() for character in visible),
            len(visible),
        ),
        "newline_count": float(raw_title.count("\n") + raw_description.count("\n")),
        "traditional_character_ratio": ratio(
            sum(character in TRADITIONAL_ONLY_CHARS for character in cjk),
            len(cjk),
        ),
        "latin_uppercase_ratio": ratio(
            sum(character.isupper() for character in latin),
            len(latin),
        ),
    }


def mean(values: list[float]) -> float:
    if not values:
        raise common.ContractError("Cannot average an empty style statistic")
    return math.fsum(values) / len(values)


def type7_quantile(values: list[float], probability: float) -> float:
    if not values:
        raise common.ContractError("Cannot take a quantile of an empty statistic")
    if not 0.0 <= probability <= 1.0:
        raise common.ContractError("Quantile probability is outside [0,1]")
    ordered = sorted(float(value) for value in values)
    index = (len(ordered) - 1) * probability
    lower = math.floor(index)
    upper = math.ceil(index)
    if lower == upper:
        return ordered[lower]
    fraction = index - lower
    return ordered[lower] + fraction * (ordered[upper] - ordered[lower])


def rounded(value: float) -> float:
    if not math.isfinite(value):
        raise common.ContractError("Non-finite style aggregate")
    return round(float(value), 12)


def read_allowlist(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if list(reader.fieldnames or []) != ALLOWLIST_COLUMNS:
            raise common.ContractError("Style allow-list schema drift")
        values = [row["seller_uid"].strip() for row in reader]
    if not values or any(not value for value in values):
        raise common.ContractError("Style allow-list is empty or contains an empty seller")
    if values != common.utf8_sort(set(values)):
        raise common.ContractError("Style allow-list must be unique and UTF-8-byte sorted")
    return values


def selected_manifest_rows(
    path: Path,
    allowed_sellers: set[str],
    *,
    expected_source_dataset: str,
    expected_bucket: str,
    expected_eligibility_status: str,
) -> tuple[dict[int, str], dict[str, int]]:
    source_rows: dict[int, str] = {}
    counts: Counter[str] = Counter()
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing = MANIFEST_REQUIRED_COLUMNS - set(reader.fieldnames or [])
        if missing:
            raise common.ContractError(f"Item manifest lacks columns: {sorted(missing)}")
        for row in reader:
            if (
                row["source_dataset"] != expected_source_dataset
                or row["data_bucket"] != expected_bucket
                or row["eligibility_status"] != expected_eligibility_status
                or row["seller_uid"] not in allowed_sellers
            ):
                continue
            source_row = int(row["source_row_number"])
            if source_row in source_rows:
                raise common.ContractError("Duplicate source row in selected item manifest")
            source_rows[source_row] = row["seller_uid"]
            counts[row["seller_uid"]] += 1
    missing_sellers = allowed_sellers - set(counts)
    if missing_sellers:
        raise common.ContractError(
            f"Train seller allow-list has {len(missing_sellers)} sellers without eligible items"
        )
    return source_rows, dict(counts)


def build_reference(
    *,
    workbook_path: Path,
    selected_rows: dict[int, str],
    allowed_sellers: list[str],
    policy: dict[str, Any],
) -> dict[str, Any]:
    seller_styles: dict[str, list[dict[str, float]]] = defaultdict(list)
    seller_categories: dict[str, Counter[str]] = defaultdict(Counter)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            seller_uid = selected_rows.get(row_number)
            if seller_uid is None:
                continue
            if len(row) != 7:
                raise common.ContractError("market_item workbook schema is not seven columns")
            _vendor, _ship_from, title, description, _price, category, _market = row
            seller_styles[seller_uid].append(item_style(title, description))
            normalized_category = clean_text(category).casefold() or "__missing__"
            seller_categories[seller_uid][normalized_category] += 1
    finally:
        workbook.close()

    expected = set(allowed_sellers)
    if set(seller_styles) != expected or set(seller_categories) != expected:
        raise common.ContractError("Workbook selection did not reconstruct every allowed seller")
    observed_item_rows = sum(len(values) for values in seller_styles.values())
    if observed_item_rows != len(selected_rows):
        raise common.ContractError("Workbook and manifest selected-item counts differ")

    boundary = policy["style_reference_boundary"]
    clip_min, clip_max = map(int, boundary["clip_item_count"])
    clipped_counts = Counter(
        min(clip_max, max(clip_min, len(seller_styles[seller_uid])))
        for seller_uid in allowed_sellers
    )
    seller_count = len(allowed_sellers)
    item_count_pmf = {
        str(item_count): rounded(clipped_counts[item_count] / seller_count)
        for item_count in range(clip_min, clip_max + 1)
    }

    per_seller: dict[str, list[float]] = defaultdict(list)
    for seller_uid in allowed_sellers:
        styles = seller_styles[seller_uid]
        per_seller["title_length"].append(median(row["title_length"] for row in styles))
        per_seller["description_length"].append(
            median(row["description_length"] for row in styles)
        )
        for field in (
            "title_missing",
            "description_missing",
            "digit_ratio",
            "punctuation_ratio",
            "newline_count",
            "traditional_character_ratio",
            "latin_uppercase_ratio",
        ):
            per_seller[field].append(mean([row[field] for row in styles]))

    quantiles: dict[str, dict[str, float]] = {}
    for field, values in sorted(per_seller.items()):
        quantiles[field] = {
            f"{float(probability):.2f}": rounded(
                type7_quantile(values, float(probability))
            )
            for probability in boundary["quantiles"]
        }

    support: Counter[str] = Counter()
    for category_counts in seller_categories.values():
        support.update(category_counts.keys())
    minimum_support = int(boundary["minimum_seller_support_per_category"])
    category_mass: Counter[str] = Counter()
    for seller_uid in allowed_sellers:
        counts = seller_categories[seller_uid]
        denominator = sum(counts.values())
        for category, count in counts.items():
            output_category = category if support[category] >= minimum_support else "__other__"
            category_mass[output_category] += count / denominator / seller_count
    ranked_mass = sorted(category_mass.values(), reverse=True)
    rank_count = int(boundary["anonymous_category_rank_count"])
    retained = ranked_mass[: rank_count - 1]
    retained.append(math.fsum(ranked_mass[rank_count - 1 :]))
    retained.extend([0.0] * (rank_count - len(retained)))
    total_mass = math.fsum(retained)
    if not math.isclose(total_mass, 1.0, rel_tol=0.0, abs_tol=1e-12):
        raise common.ContractError(f"Anonymous category mass is not one: {total_mass}")
    anonymous_category_rank_probability = [
        rounded(value / total_mass) for value in retained
    ]
    normalization_error = 1.0 - math.fsum(anonymous_category_rank_probability)
    anonymous_category_rank_probability[-1] = rounded(
        anonymous_category_rank_probability[-1] + normalization_error
    )

    return {
        "version": policy["version"],
        "statistical_unit": "seller",
        "seller_count": seller_count,
        "item_count_clip": [clip_min, clip_max],
        "item_count_pmf": item_count_pmf,
        "quantile_method": "Hyndman-Fan type 7 linear interpolation",
        "quantile_probabilities": [
            rounded(float(value)) for value in boundary["quantiles"]
        ],
        "seller_equal_weight_quantiles": quantiles,
        "anonymous_category_rank_probability": anonymous_category_rank_probability,
        "category_rank_count": rank_count,
        "minimum_category_seller_support": minimum_support,
        "contains_seller_uid": False,
        "contains_raw_text_or_fragment": False,
        "contains_real_category_name": False,
        "contains_identity_value": False,
    }


def run(policy: dict[str, Any], mode: str, policy_path: Path) -> dict[str, Any]:
    output_root = common.mode_output_root(policy, mode)
    boundary = policy["style_reference_boundary"]
    allowlist_path = output_root / boundary["membership_output"]
    if not allowlist_path.is_file():
        raise FileNotFoundError("Materialize the style-source seller allow-list first")
    allowed_sellers = read_allowlist(allowlist_path)

    item_spec = policy["frozen_inputs"]["item_manifest"]
    workbook_spec = policy["frozen_inputs"]["raw_chinese_items"]
    item_manifest = common.verify_file_pin(item_spec, label="item_manifest")
    workbook = common.verify_file_pin(workbook_spec, label="raw_chinese_items")
    selected_rows, _counts = selected_manifest_rows(
        item_manifest,
        set(allowed_sellers),
        expected_source_dataset=workbook_spec["source_dataset"],
        expected_bucket=policy["style_reference_boundary"]["eligible_data_bucket"],
        expected_eligibility_status=policy["style_reference_boundary"][
            "eligible_item_manifest_status"
        ],
    )
    profile = build_reference(
        workbook_path=workbook,
        selected_rows=selected_rows,
        allowed_sellers=allowed_sellers,
        policy=policy,
    )
    profile["lineage"] = {
        "allowlist_sha256": common.sha256_file(allowlist_path),
        "item_manifest_sha256": item_spec["sha256"],
        "raw_chinese_items_sha256": workbook_spec["sha256"],
        "producer_sha256": common.sha256_file(Path(__file__).resolve()),
        "policy_sha256": common.sha256_file(policy_path.resolve()),
    }

    profile_path = output_root / boundary["profile_output"]
    common.write_json(profile_path, profile)
    manifest = {
        "step": "step28_v13",
        "stage": "build_style_reference",
        "mode": mode,
        "run_id": policy["modes"][mode]["run_id"],
        "producer_sha256": common.sha256_file(Path(__file__).resolve()),
        "policy_sha256": common.sha256_file(policy_path.resolve()),
        "inputs": [
            common.artifact_record(
                allowlist_path,
                role="style_source_train_seller_allowlist",
                root=output_root,
            ),
            {
                "role": "item_manifest",
                "path": item_spec["path"],
                "sha256": item_spec["sha256"],
            },
            {
                "role": "raw_chinese_items",
                "path": workbook_spec["path"],
                "sha256": workbook_spec["sha256"],
            },
        ],
        "output": common.artifact_record(
            profile_path,
            role="chinese_train_style_profile",
            root=output_root,
        ),
        "forbidden_fields_absent": True,
    }
    manifest["canonical_self_hash"] = common.canonical_sha256(manifest)
    common.write_json(
        output_root / "manifests" / "style_reference_manifest.json",
        manifest,
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    common.add_policy_argument(parser)
    common.add_mode_argument(parser)
    parser.add_argument("--validate-config-only", action="store_true")
    args = parser.parse_args()
    policy = common.load_policy(args.policy, mode=args.mode)
    if args.validate_config_only:
        print("Step28-v13 style reference configuration is valid")
        return
    manifest = run(policy, args.mode, args.policy)
    print(
        "Step28-v13 style reference materialized:",
        manifest["output"]["path"],
        manifest["canonical_self_hash"],
    )


if __name__ == "__main__":
    main()
