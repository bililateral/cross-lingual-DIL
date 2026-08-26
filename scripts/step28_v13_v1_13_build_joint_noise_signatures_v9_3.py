#!/usr/bin/env python3
"""Build the label-free V9.3 seller-level joint noise-signature preflight.

The output contains aggregate presence masks only.  It never publishes seller
identifiers, source rows, raw text, identity relations, truth, or model output.
"""

from __future__ import annotations

import argparse
import copy
from collections import Counter, defaultdict
import csv
import hashlib
import json
from pathlib import Path
import unicodedata
import re
from typing import Any, Mapping

from openpyxl import load_workbook


VERSION = "2026-08-25-step28-v13-v1-13-joint-noise-signatures-v9-3"
ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "market_item.xlsx"
WORKBOOK_SHA256 = "3625a226974827a0441ec87c54688f85ed0ff93a1c4c687532ef550ec1640187"
MANIFEST_PATH = ROOT / "reports" / "step2_content_item_manifest.csv"
MANIFEST_SHA256 = "89307835c82683abe53561a56a39fb28419cbf46671d34463701d73179e3869f"
ALLOWLIST_PATH = (
    ROOT
    / "reports"
    / "step28_synthetic_chinese_dataset"
    / "v13_dev_smoke_v1_20260727"
    / "reference"
    / "style_source_train_sellers.csv"
)
ALLOWLIST_SHA256 = "b57cbfc3908ae3f36983c7c9e6d4dd974b59222f6617d214cec090524c3eb6de"
EXPECTED_SELLER_COUNT = 676
EXPECTED_SELECTED_ITEM_ROWS = 3439
SLOT_COUNT = 28
MINIMUM_ITEM_COUNT = 2
MAXIMUM_ITEM_COUNT = 8
SOURCE_DATASET = "market_item.xlsx"
DATA_BUCKET = "zh_target_strict"
ELIGIBILITY_STATUS = "target_eval_candidate"
WHITESPACE_RE = re.compile(r"\s+")
EXPECTED_STATUS = (
    "PASS_LABEL_FREE_AGGREGATE_PREFLIGHT_ONLY_NOT_METHOD_OR_TRAINING_QUALIFIED"
)
EXPECTED_STATISTICAL_UNIT = "seller_equal_weight"
FORBIDDEN_RECURSIVE_KEYS = {
    "label",
    "pair_label",
    "truth",
    "controller_uid",
    "controller_id",
    "seller_uid",
    "source_row_number",
    "raw_text",
    "raw_text_fragment",
    "model_output",
    "prediction",
}


class JointNoiseSignatureError(ValueError):
    """Raised when the label-free joint-signature boundary drifts."""


def canonical_json_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def canonical_sha256(value: object) -> str:
    return hashlib.sha256(canonical_json_bytes(value)).hexdigest()


def canonical_self_sha256(payload: Mapping[str, Any]) -> str:
    projection = copy.deepcopy(dict(payload))
    projection.pop("canonical_self_sha256", None)
    return canonical_sha256(projection)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _verify_pin(path: Path, expected_sha256: str, *, name: str) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"Pinned {name} is missing: {path}")
    observed = sha256_file(path)
    if observed != expected_sha256:
        raise JointNoiseSignatureError(
            f"Pinned {name} SHA-256 drift: expected={expected_sha256} observed={observed}"
        )


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(
        " ", unicodedata.normalize("NFKC", str(value))
    ).strip()


def _read_allowlist() -> list[str]:
    with ALLOWLIST_PATH.open("r", encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream)
        if list(reader.fieldnames or []) != ["seller_uid"]:
            raise JointNoiseSignatureError("Pinned seller allow-list schema drift")
        sellers = [row["seller_uid"].strip() for row in reader]
    if len(sellers) != EXPECTED_SELLER_COUNT or any(not value for value in sellers):
        raise JointNoiseSignatureError("Pinned seller allow-list cardinality drift")
    if sellers != sorted(set(sellers), key=lambda value: value.encode("utf-8")):
        raise JointNoiseSignatureError(
            "Pinned seller allow-list must be unique and UTF-8-byte sorted"
        )
    return sellers


def _read_selected_rows(allowed: set[str]) -> dict[int, str]:
    required = {
        "seller_uid",
        "source_dataset",
        "source_row_number",
        "data_bucket",
        "eligibility_status",
    }
    selected: dict[int, str] = {}
    counts: Counter[str] = Counter()
    with MANIFEST_PATH.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        if not required <= set(reader.fieldnames or []):
            raise JointNoiseSignatureError("Pinned item manifest schema drift")
        for row in reader:
            if (
                row["source_dataset"] != SOURCE_DATASET
                or row["data_bucket"] != DATA_BUCKET
                or row["eligibility_status"] != ELIGIBILITY_STATUS
                or row["seller_uid"] not in allowed
            ):
                continue
            source_row = int(row["source_row_number"])
            if source_row < 2 or source_row in selected:
                raise JointNoiseSignatureError(
                    "Pinned item manifest has an invalid or duplicate selected source row"
                )
            selected[source_row] = row["seller_uid"]
            counts[row["seller_uid"]] += 1
    if len(selected) != EXPECTED_SELECTED_ITEM_ROWS:
        raise JointNoiseSignatureError("Selected item-row cardinality drift")
    if set(counts) != allowed:
        raise JointNoiseSignatureError("Selected rows do not cover every allowed seller")
    return selected


def _read_presence_by_seller(
    selected_rows: Mapping[int, str], allowed_sellers: list[str]
) -> dict[str, list[tuple[int, bool, bool]]]:
    presence: dict[str, list[tuple[int, bool, bool]]] = defaultdict(list)
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            seller_uid = selected_rows.get(row_number)
            if seller_uid is None:
                continue
            if len(row) != 7:
                raise JointNoiseSignatureError(
                    "Pinned market-item workbook is not seven columns"
                )
            _vendor, _ship_from, title, description, _price, _category, _market = row
            presence[seller_uid].append(
                (row_number, bool(_clean_text(title)), bool(_clean_text(description)))
            )
    finally:
        workbook.close()
    if set(presence) != set(allowed_sellers):
        raise JointNoiseSignatureError("Workbook did not reconstruct every allowed seller")
    if sum(map(len, presence.values())) != EXPECTED_SELECTED_ITEM_ROWS:
        raise JointNoiseSignatureError("Workbook selected-row cardinality drift")
    for seller_uid in allowed_sellers:
        rows = presence[seller_uid]
        if rows != sorted(rows) or len({row[0] for row in rows}) != len(rows):
            raise JointNoiseSignatureError("Seller source-row order drift")
    return dict(presence)


def _signature(rows: list[tuple[int, bool, bool]]) -> dict[str, Any]:
    if not rows:
        raise JointNoiseSignatureError("Cannot construct a signature for an empty seller")
    states = [(title, description) for _row, title, description in rows]
    if len(states) == 1:
        states.append(states[0])
    states = states[:MAXIMUM_ITEM_COUNT]
    item_count = len(states)
    if not MINIMUM_ITEM_COUNT <= item_count <= MAXIMUM_ITEM_COUNT:
        raise JointNoiseSignatureError("Effective item count is outside 2..8")
    title_mask = "".join("1" if title else "0" for title, _description in states)
    description_mask = "".join(
        "1" if description else "0" for _title, description in states
    )
    joint_empty_mask = "".join(
        "1" if not title and not description else "0"
        for title, description in states
    )
    return {
        "item_count": item_count,
        "title_present_mask": title_mask,
        "description_present_mask": description_mask,
        "joint_empty_mask": joint_empty_mask,
    }


def _signature_key(signature: Mapping[str, Any]) -> bytes:
    return canonical_json_bytes(signature)


def _largest_remainder(
    counts: Counter[bytes], signatures: Mapping[bytes, dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if sum(counts.values()) != EXPECTED_SELLER_COUNT:
        raise JointNoiseSignatureError("Joint-signature seller total drift")
    rows: list[dict[str, Any]] = []
    allocated = 0
    for key in sorted(counts):
        seller_count = int(counts[key])
        numerator = seller_count * SLOT_COUNT
        floor_slots, remainder = divmod(numerator, EXPECTED_SELLER_COUNT)
        allocated += floor_slots
        rows.append(
            {
                "signature": signatures[key],
                "seller_count": seller_count,
                "probability_numerator": seller_count,
                "probability_denominator": EXPECTED_SELLER_COUNT,
                "floor_slot_count": floor_slots,
                "remainder_numerator": remainder,
                "allocated_slot_count": floor_slots,
            }
        )
    remaining = SLOT_COUNT - allocated
    if not 0 <= remaining <= len(rows):
        raise JointNoiseSignatureError("Largest-remainder residual cardinality drift")
    ranking = sorted(
        range(len(rows)),
        key=lambda index: (
            -int(rows[index]["remainder_numerator"]),
            _signature_key(rows[index]["signature"]),
        ),
    )
    for index in ranking[:remaining]:
        rows[index]["allocated_slot_count"] += 1
    slots: list[dict[str, Any]] = []
    for row in rows:
        for _copy in range(int(row["allocated_slot_count"])):
            slots.append(dict(row["signature"]))
    if len(slots) != SLOT_COUNT:
        raise JointNoiseSignatureError("Largest-remainder output is not exactly 28 slots")
    return rows, slots


def _validate_signature(signature: Mapping[str, Any]) -> None:
    if set(signature) != {
        "item_count",
        "title_present_mask",
        "description_present_mask",
        "joint_empty_mask",
    }:
        raise JointNoiseSignatureError("Joint-signature field drift")
    item_count = signature["item_count"]
    if (
        isinstance(item_count, bool)
        or not isinstance(item_count, int)
        or not MINIMUM_ITEM_COUNT <= item_count <= MAXIMUM_ITEM_COUNT
    ):
        raise JointNoiseSignatureError("Joint-signature item count drift")
    title = signature["title_present_mask"]
    description = signature["description_present_mask"]
    empty = signature["joint_empty_mask"]
    if not all(
        isinstance(value, str)
        and len(value) == item_count
        and set(value) <= {"0", "1"}
        for value in (title, description, empty)
    ):
        raise JointNoiseSignatureError("Joint-signature mask drift")
    expected_empty = "".join(
        "1" if title_bit == description_bit == "0" else "0"
        for title_bit, description_bit in zip(title, description)
    )
    if empty != expected_empty:
        raise JointNoiseSignatureError("Joint-empty mask is inconsistent")


def _reject_sensitive_nested_keys(value: object, *, path: str = "$") -> None:
    if isinstance(value, Mapping):
        overlap = FORBIDDEN_RECURSIVE_KEYS & set(value)
        if overlap:
            raise JointNoiseSignatureError(
                f"Sensitive nested key at {path}: {sorted(overlap)}"
            )
        for key, child in value.items():
            _reject_sensitive_nested_keys(child, path=f"{path}.{key}")
    elif isinstance(value, list):
        for index, child in enumerate(value):
            _reject_sensitive_nested_keys(child, path=f"{path}[{index}]")


def validate_payload(payload: Mapping[str, Any]) -> dict[str, Any]:
    required_keys = {
        "version",
        "status",
        "statistical_unit",
        "seller_count",
        "selected_item_row_count",
        "source_pins",
        "selection_boundary",
        "signature_definition",
        "raw_item_count_histogram",
        "observed_signature_count",
        "signature_frequency_and_integerization",
        "noise_slot_multiset",
        "slot_eligibility",
        "contains_seller_uid",
        "contains_source_row_number",
        "contains_raw_text_or_fragment",
        "contains_identity_or_pair_truth",
        "contains_model_output",
        "canonical_self_sha256",
    }
    if set(payload) != required_keys:
        raise JointNoiseSignatureError("Joint-signature payload key drift")
    if payload["version"] != VERSION:
        raise JointNoiseSignatureError("Joint-signature version drift")
    if payload["status"] != EXPECTED_STATUS:
        raise JointNoiseSignatureError("Joint-signature status drift")
    if payload["statistical_unit"] != EXPECTED_STATISTICAL_UNIT:
        raise JointNoiseSignatureError("Joint-signature statistical unit drift")
    if payload["seller_count"] != EXPECTED_SELLER_COUNT:
        raise JointNoiseSignatureError("Joint-signature seller count drift")
    if payload["selected_item_row_count"] != EXPECTED_SELECTED_ITEM_ROWS:
        raise JointNoiseSignatureError("Joint-signature item-row count drift")
    supplied_self = payload["canonical_self_sha256"]
    if (
        not isinstance(supplied_self, str)
        or len(supplied_self) != 64
        or supplied_self != canonical_self_sha256(payload)
    ):
        raise JointNoiseSignatureError("Joint-signature canonical self-hash drift")
    for flag in (
        "contains_seller_uid",
        "contains_source_row_number",
        "contains_raw_text_or_fragment",
        "contains_identity_or_pair_truth",
        "contains_model_output",
    ):
        if payload[flag] is not False:
            raise JointNoiseSignatureError(f"Forbidden-output flag drift: {flag}")
    expected_source_pins = {
        "raw_chinese_items": {
            "path": WORKBOOK_PATH.relative_to(ROOT).as_posix(),
            "sha256": WORKBOOK_SHA256,
        },
        "item_manifest": {
            "path": MANIFEST_PATH.relative_to(ROOT).as_posix(),
            "sha256": MANIFEST_SHA256,
        },
        "train_seller_allowlist": {
            "path": ALLOWLIST_PATH.relative_to(ROOT).as_posix(),
            "sha256": ALLOWLIST_SHA256,
        },
    }
    if payload["source_pins"] != expected_source_pins:
        raise JointNoiseSignatureError("Joint-signature source pins drift")
    expected_selection_boundary = {
        "source_dataset": SOURCE_DATASET,
        "data_bucket": DATA_BUCKET,
        "eligibility_status": ELIGIBILITY_STATUS,
        "within_seller_order": "ascending_source_row_number",
        "effective_item_count": "clip_raw_count_to_2_through_8",
        "singleton_padding": "repeat_the_only_items_presence_state_once",
        "over_eight_truncation": "retain_first_eight_selected_source_rows",
    }
    if payload["selection_boundary"] != expected_selection_boundary:
        raise JointNoiseSignatureError("Joint-signature selection boundary drift")
    expected_signature_definition = {
        "fields": [
            "item_count",
            "title_present_mask",
            "description_present_mask",
            "joint_empty_mask",
        ],
        "mask_order": "logical_item_ordinal_ascending",
        "present_bit": "1",
        "absent_bit": "0",
        "joint_empty_bit": (
            "1_exactly_when_title_and_description_are_both_absent"
        ),
    }
    if payload["signature_definition"] != expected_signature_definition:
        raise JointNoiseSignatureError("Joint-signature definition drift")
    _reject_sensitive_nested_keys(
        {
            key: value
            for key, value in payload.items()
            if key not in {"source_pins", "selection_boundary"}
        }
    )

    raw_histogram = payload["raw_item_count_histogram"]
    if not isinstance(raw_histogram, Mapping):
        raise JointNoiseSignatureError("Raw item-count histogram type drift")
    raw_total = 0
    raw_rows = 0
    for raw_key, raw_value in raw_histogram.items():
        if not str(raw_key).isdigit() or int(raw_key) < 1:
            raise JointNoiseSignatureError("Raw item-count histogram key drift")
        if isinstance(raw_value, bool) or not isinstance(raw_value, int) or raw_value < 1:
            raise JointNoiseSignatureError("Raw item-count histogram value drift")
        raw_total += raw_value
        raw_rows += int(raw_key) * raw_value
    if raw_total != EXPECTED_SELLER_COUNT or raw_rows != EXPECTED_SELECTED_ITEM_ROWS:
        raise JointNoiseSignatureError("Raw item-count histogram totals drift")

    frequency = payload["signature_frequency_and_integerization"]
    if not isinstance(frequency, list) or not frequency:
        raise JointNoiseSignatureError("Joint-signature frequency table is empty")
    counts: Counter[bytes] = Counter()
    signatures: dict[bytes, dict[str, Any]] = {}
    previous_key: bytes | None = None
    for row in frequency:
        if not isinstance(row, Mapping) or set(row) != {
            "signature",
            "seller_count",
            "probability_numerator",
            "probability_denominator",
            "floor_slot_count",
            "remainder_numerator",
            "allocated_slot_count",
        }:
            raise JointNoiseSignatureError("Joint-signature frequency row drift")
        signature = row["signature"]
        if not isinstance(signature, Mapping):
            raise JointNoiseSignatureError("Joint-signature row type drift")
        _validate_signature(signature)
        key = _signature_key(signature)
        if previous_key is not None and key <= previous_key:
            raise JointNoiseSignatureError("Joint-signature table is not canonical")
        previous_key = key
        seller_count = row["seller_count"]
        if isinstance(seller_count, bool) or not isinstance(seller_count, int) or seller_count < 1:
            raise JointNoiseSignatureError("Joint-signature seller frequency drift")
        if row["probability_numerator"] != seller_count:
            raise JointNoiseSignatureError("Joint-signature probability numerator drift")
        if row["probability_denominator"] != EXPECTED_SELLER_COUNT:
            raise JointNoiseSignatureError("Joint-signature probability denominator drift")
        counts[key] = seller_count
        signatures[key] = dict(signature)
    expected_frequency, expected_slots = _largest_remainder(counts, signatures)
    if frequency != expected_frequency:
        raise JointNoiseSignatureError("Largest-remainder integerization drift")
    if payload["observed_signature_count"] != len(frequency):
        raise JointNoiseSignatureError("Observed joint-signature count drift")

    slot_rows = payload["noise_slot_multiset"]
    if not isinstance(slot_rows, list) or len(slot_rows) != SLOT_COUNT:
        raise JointNoiseSignatureError("Noise-slot multiset cardinality drift")
    observed_slots: list[dict[str, Any]] = []
    for expected_slot, row in enumerate(slot_rows):
        if not isinstance(row, Mapping) or set(row) != {"noise_slot", "signature"}:
            raise JointNoiseSignatureError("Noise-slot row schema drift")
        if row["noise_slot"] != expected_slot or not isinstance(row["signature"], Mapping):
            raise JointNoiseSignatureError("Noise-slot ordinal drift")
        _validate_signature(row["signature"])
        observed_slots.append(dict(row["signature"]))
    if observed_slots != expected_slots:
        raise JointNoiseSignatureError("Noise-slot signature multiset drift")

    expected_eligibility = []
    for noise_slot, signature in enumerate(expected_slots):
        title = signature["title_present_mask"]
        description = signature["description_present_mask"]
        expected_eligibility.append(
            {
                "noise_slot": noise_slot,
                "title_present_logical_item_ordinals": [
                    index for index, bit in enumerate(title) if bit == "1"
                ],
                "title_and_description_present_logical_item_ordinals": [
                    index
                    for index, (title_bit, description_bit) in enumerate(
                        zip(title, description)
                    )
                    if title_bit == description_bit == "1"
                ],
            }
        )
    if payload["slot_eligibility"] != expected_eligibility:
        raise JointNoiseSignatureError("Noise-slot eligibility drift")
    return {
        "version": VERSION,
        "canonical_self_sha256": supplied_self,
        "seller_count": EXPECTED_SELLER_COUNT,
        "selected_item_row_count": EXPECTED_SELECTED_ITEM_ROWS,
        "observed_signature_count": len(frequency),
        "noise_slot_count": SLOT_COUNT,
        "title_eligible_noise_slot_count": sum(
            bool(row["title_present_logical_item_ordinals"])
            for row in expected_eligibility
        ),
        "title_and_description_eligible_noise_slot_count": sum(
            bool(row["title_and_description_present_logical_item_ordinals"])
            for row in expected_eligibility
        ),
        "status": "PASS_LABEL_FREE_AGGREGATE_VALIDATION_ONLY_NOT_METHOD_OR_TRAINING_QUALIFIED",
    }


def build_payload() -> dict[str, Any]:
    for path, expected, name in (
        (WORKBOOK_PATH, WORKBOOK_SHA256, "raw Chinese item workbook"),
        (MANIFEST_PATH, MANIFEST_SHA256, "item manifest"),
        (ALLOWLIST_PATH, ALLOWLIST_SHA256, "train-side seller allow-list"),
    ):
        _verify_pin(path, expected, name=name)
    sellers = _read_allowlist()
    selected = _read_selected_rows(set(sellers))
    presence = _read_presence_by_seller(selected, sellers)
    counts: Counter[bytes] = Counter()
    signatures: dict[bytes, dict[str, Any]] = {}
    raw_item_count_histogram: Counter[int] = Counter()
    for seller_uid in sellers:
        raw_item_count_histogram[len(presence[seller_uid])] += 1
        signature = _signature(presence[seller_uid])
        key = _signature_key(signature)
        signatures.setdefault(key, signature)
        counts[key] += 1
    frequency, slots = _largest_remainder(counts, signatures)

    slot_eligibility = []
    for noise_slot, signature in enumerate(slots):
        title = signature["title_present_mask"]
        description = signature["description_present_mask"]
        title_ordinals = [index for index, bit in enumerate(title) if bit == "1"]
        joint_ordinals = [
            index
            for index, (title_bit, description_bit) in enumerate(zip(title, description))
            if title_bit == description_bit == "1"
        ]
        slot_eligibility.append(
            {
                "noise_slot": noise_slot,
                "title_present_logical_item_ordinals": title_ordinals,
                "title_and_description_present_logical_item_ordinals": joint_ordinals,
            }
        )

    payload: dict[str, Any] = {
        "version": VERSION,
        "status": EXPECTED_STATUS,
        "statistical_unit": EXPECTED_STATISTICAL_UNIT,
        "seller_count": EXPECTED_SELLER_COUNT,
        "selected_item_row_count": EXPECTED_SELECTED_ITEM_ROWS,
        "source_pins": {
            "raw_chinese_items": {
                "path": WORKBOOK_PATH.relative_to(ROOT).as_posix(),
                "sha256": WORKBOOK_SHA256,
            },
            "item_manifest": {
                "path": MANIFEST_PATH.relative_to(ROOT).as_posix(),
                "sha256": MANIFEST_SHA256,
            },
            "train_seller_allowlist": {
                "path": ALLOWLIST_PATH.relative_to(ROOT).as_posix(),
                "sha256": ALLOWLIST_SHA256,
            },
        },
        "selection_boundary": {
            "source_dataset": SOURCE_DATASET,
            "data_bucket": DATA_BUCKET,
            "eligibility_status": ELIGIBILITY_STATUS,
            "within_seller_order": "ascending_source_row_number",
            "effective_item_count": "clip_raw_count_to_2_through_8",
            "singleton_padding": "repeat_the_only_items_presence_state_once",
            "over_eight_truncation": "retain_first_eight_selected_source_rows",
        },
        "signature_definition": {
            "fields": [
                "item_count",
                "title_present_mask",
                "description_present_mask",
                "joint_empty_mask",
            ],
            "mask_order": "logical_item_ordinal_ascending",
            "present_bit": "1",
            "absent_bit": "0",
            "joint_empty_bit": "1_exactly_when_title_and_description_are_both_absent",
        },
        "raw_item_count_histogram": {
            str(key): value for key, value in sorted(raw_item_count_histogram.items())
        },
        "observed_signature_count": len(frequency),
        "signature_frequency_and_integerization": frequency,
        "noise_slot_multiset": [
            {"noise_slot": index, "signature": signature}
            for index, signature in enumerate(slots)
        ],
        "slot_eligibility": slot_eligibility,
        "contains_seller_uid": False,
        "contains_source_row_number": False,
        "contains_raw_text_or_fragment": False,
        "contains_identity_or_pair_truth": False,
        "contains_model_output": False,
    }
    payload["canonical_self_sha256"] = canonical_self_sha256(payload)
    validate_payload(payload)
    return payload


def write_new(path: Path, payload: Mapping[str, Any]) -> None:
    if path.exists():
        raise JointNoiseSignatureError("Refusing to overwrite a joint-signature artifact")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("xb") as stream:
        stream.write(canonical_json_bytes(payload))
        stream.write(b"\n")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--output", type=Path)
    group.add_argument("--validate-existing", type=Path)
    args = parser.parse_args()
    if args.validate_existing is not None:
        value = json.loads(args.validate_existing.read_text(encoding="utf-8"))
        if not isinstance(value, dict):
            raise JointNoiseSignatureError("Joint-signature root must be an object")
        print(canonical_json_bytes(validate_payload(value)).decode("utf-8"))
        return
    payload = build_payload()
    if args.output is None:
        raise AssertionError("argparse did not provide the required output path")
    write_new(args.output, payload)
    print(canonical_json_bytes({
        "canonical_self_sha256": payload["canonical_self_sha256"],
        "observed_signature_count": payload["observed_signature_count"],
        "seller_count": payload["seller_count"],
        "selected_item_row_count": payload["selected_item_row_count"],
        "status": payload["status"],
    }).decode("utf-8"))


if __name__ == "__main__":
    main()
