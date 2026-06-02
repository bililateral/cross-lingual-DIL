from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median, quantiles
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SCHEMA_PATH = ROOT / "schema" / "source_schema_map.json"
OUTPUT_PATH = ROOT / "reports" / "step1_schema_audit.json"
TARGET_MARKETS = {"中文暗网交易市场", "茶马古道"}
CJK_RE = re.compile(r"[\u4e00-\u9fff]")
VENDOR_SQL_RE = re.compile(
    r"^\((\d+), '((?:[^'\\]|\\.)*)', (\d+), (\d+), (\d+), '(.*?)', "
    r"'((?:[^'\\]|\\.)*)', (\d+), (\d+), (\d+), (\d+)\),?$"
)
PGP_SQL_TAIL_RE = re.compile(r", '([^']*)', '([^']*)', (\d+), ([0-9.]+)\),?;?$")


def load_schema() -> dict:
    with SCHEMA_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def csv_headers(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or [])


def workbook_headers(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    workbook.close()
    return [str(value) for value in row]


def sql_insert_columns(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            if line.startswith("INSERT INTO"):
                return re.findall(r"`([^`]+)`", line)
    raise RuntimeError(f"Could not find INSERT column list in {path}")


def validate_mappings(schema: dict) -> dict:
    validators = {
        ".csv": csv_headers,
        ".xlsx": workbook_headers,
        ".sql": sql_insert_columns,
    }
    report = {}
    for source_name, source_spec in schema["source_mappings"].items():
        path = ROOT / source_name
        suffix = path.suffix.lower()
        actual_fields = validators[suffix](path)
        mapped_fields = [field for field in source_spec["field_map"].values() if field]
        missing = sorted(set(mapped_fields) - set(actual_fields))
        report[source_name] = {
            "exists": path.exists(),
            "actual_field_count": len(actual_fields),
            "mapped_field_count": len(mapped_fields),
            "missing_fields": missing,
            "status": "ok" if not missing else "missing_fields",
        }
    return report


def activity_summary(counter: Counter) -> dict:
    values = list(counter.values())
    if not values:
        return {
            "sellers": 0,
            "items": 0,
            "median_items_per_seller": 0,
            "p90_items_per_seller": 0,
            "vendors_ge_2_items": 0,
            "vendors_ge_5_items": 0,
        }
    p90 = quantiles(values, n=10)[8] if len(values) >= 10 else max(values)
    return {
        "sellers": len(values),
        "items": sum(values),
        "median_items_per_seller": median(values),
        "p90_items_per_seller": p90,
        "vendors_ge_2_items": sum(1 for value in values if value >= 2),
        "vendors_ge_5_items": sum(1 for value in values if value >= 5),
    }


def audit_products_data() -> dict:
    path = ROOT / "products_data.csv"
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    seller_counts = Counter(row["卖家ID"] for row in rows if row.get("卖家ID"))
    categories = Counter(row["类别"] for row in rows if row.get("类别"))
    return {
        "row_count": len(rows),
        "seller_activity": activity_summary(seller_counts),
        "nonempty_description_rows": sum(1 for row in rows if row.get("商品描述")),
        "nonempty_price_rows": sum(1 for row in rows if row.get("单价(USD)")),
        "top_categories": categories.most_common(12),
    }


def audit_market_item() -> tuple[dict, set[str], set[str]]:
    path = ROOT / "market_item.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    market_counts = Counter()
    target_vendor_counts = defaultdict(Counter)
    non_target_vendor_counts = Counter()
    target_categories = defaultdict(Counter)
    non_target_cjk_rows = 0
    non_target_vendors = set()
    target_vendors = set()
    total_rows = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        vendor, ship_from, title, description, price, category, market = row
        total_rows += 1
        market_str = str(market)
        vendor_str = str(vendor) if vendor is not None else ""
        market_counts[market_str] += 1
        text = f"{title or ''} {description or ''} {category or ''}"
        if market_str in TARGET_MARKETS:
            target_vendor_counts[market_str][vendor_str] += 1
            target_vendors.add(vendor_str.lower())
            if category:
                target_categories[market_str][str(category)] += 1
        else:
            non_target_vendor_counts[vendor_str] += 1
            non_target_vendors.add(vendor_str.lower())
            if CJK_RE.search(text):
                non_target_cjk_rows += 1

    workbook.close()

    target_market_stats = {}
    for market_name in sorted(TARGET_MARKETS):
        target_market_stats[market_name] = {
            "row_count": market_counts[market_name],
            "seller_activity": activity_summary(target_vendor_counts[market_name]),
            "top_categories": target_categories[market_name].most_common(12),
        }

    non_target_rows = total_rows - sum(market_counts[m] for m in TARGET_MARKETS)
    return (
        {
            "row_count": total_rows,
            "market_count": len(market_counts),
            "rows_by_market": market_counts.most_common(),
            "target_markets": target_market_stats,
            "non_target_seller_activity": activity_summary(non_target_vendor_counts),
            "non_target_cjk_rows": non_target_cjk_rows,
            "non_target_cjk_ratio": round(non_target_cjk_rows / max(1, non_target_rows), 6),
        },
        non_target_vendors,
        target_vendors,
    )


def audit_agora() -> tuple[dict, set[str]]:
    path = ROOT / "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    vendor_counts = Counter()
    total_rows = 0
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        vendor = row[0]
        total_rows += 1
        if vendor is not None:
            vendor_counts[str(vendor).strip()] += 1
    workbook.close()
    vendor_set = {vendor.lower() for vendor in vendor_counts}
    return {
        "row_count": total_rows,
        "seller_activity": activity_summary(vendor_counts),
    }, vendor_set


def audit_vendor_sql() -> tuple[dict, set[str]]:
    path = ROOT / "tijkc3xx.sql"
    market_counts = Counter()
    market_domains = defaultdict(Counter)
    alias_set = set()
    total_rows = 0
    imposter_rows = 0

    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            line = line.strip()
            if not line.startswith("("):
                continue
            match = VENDOR_SQL_RE.match(line)
            if not match:
                continue
            total_rows += 1
            user_name = match.group(2)
            market_id = match.group(3)
            vendor_link = match.group(7)
            imposter_flag = match.group(11)
            market_counts[market_id] += 1
            alias_set.add(user_name.lower())
            host = urlparse(vendor_link).netloc.lower()
            market_domains[market_id][host] += 1
            if imposter_flag == "1":
                imposter_rows += 1

    return {
        "row_count": total_rows,
        "unique_alias_count": len(alias_set),
        "imposter_rows": imposter_rows,
        "rows_by_market_id": market_counts.most_common(),
        "market_id_to_top_hosts": {
            market_id: counter.most_common(3)
            for market_id, counter in sorted(market_domains.items(), key=lambda item: int(item[0]))
        },
    }, alias_set


def audit_pgp_sql() -> dict:
    path = ROOT / "3z669jwe.sql"
    row_count = 0
    review_nonzero = 0
    vendor_span = Counter()

    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            line = line.strip()
            if not line.startswith("("):
                continue
            row_count += 1
            match = PGP_SQL_TAIL_RE.search(line)
            if not match:
                continue
            vendor_ids = match.group(1)
            review_count = int(match.group(3))
            span = 0 if not vendor_ids else vendor_ids.count(",") + 1
            vendor_span[span] += 1
            if review_count > 0:
                review_nonzero += 1

    return {
        "row_count": row_count,
        "rows_with_nonzero_reviews": review_nonzero,
        "group_size_distribution_top": vendor_span.most_common(12),
        "groups_ge_2_vendors": sum(count for span, count in vendor_span.items() if span >= 2),
        "groups_ge_3_vendors": sum(count for span, count in vendor_span.items() if span >= 3),
    }


def audit_suspected_csv(path_name: str, group_field: str) -> dict:
    path = ROOT / path_name
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    market_counts = Counter(row["market_id"] for row in rows if row.get("market_id"))
    key_alias_counts = Counter(row["key_alias"] for row in rows if row.get("key_alias"))
    return {
        "row_count": len(rows),
        "unique_user_names": len({row["user_name"] for row in rows if row.get("user_name")}),
        "top_market_ids": market_counts.most_common(10),
        "group_count": len(key_alias_counts),
        "groups_ge_3": sum(1 for _, count in key_alias_counts.items() if count >= 3),
        "group_size_field": group_field,
    }


def audit_imposter_csv() -> dict:
    path = ROOT / "suspected_imposter_rows.csv"
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    return {
        "row_count": len(rows),
        "unique_user_names": len({row["user_name"] for row in rows if row.get("user_name")}),
    }


def overlap_report(
    en_gold_aliases: set[str],
    market_item_en_vendors: set[str],
    market_item_zh_vendors: set[str],
    agora_vendors: set[str],
) -> dict:
    return {
        "en_gold_aliases": len(en_gold_aliases),
        "market_item_en_vendors": len(market_item_en_vendors),
        "market_item_zh_vendors": len(market_item_zh_vendors),
        "agora_vendors": len(agora_vendors),
        "overlap_gold_market_item_en": len(en_gold_aliases & market_item_en_vendors),
        "overlap_gold_agora": len(en_gold_aliases & agora_vendors),
        "overlap_market_item_en_agora": len(market_item_en_vendors & agora_vendors),
    }


def build_warnings(overlaps: dict, market_item_audit: dict) -> list[str]:
    warnings = []
    if overlaps["overlap_gold_agora"] > 0:
        warnings.append(
            f"Exact lowercased alias overlap exists between EN-Gold and Agora content: {overlaps['overlap_gold_agora']} aliases."
        )
    if overlaps["overlap_gold_market_item_en"] > 0:
        warnings.append(
            f"Exact lowercased alias overlap exists between EN-Gold and market_item English content: {overlaps['overlap_gold_market_item_en']} aliases."
        )
    if market_item_audit["non_target_cjk_rows"] > 0:
        warnings.append(
            f"market_item.xlsx contains {market_item_audit['non_target_cjk_rows']} non-target rows with Chinese characters; language filtering is mandatory."
        )
    warnings.append(
        "products_data.csv does not expose an explicit market field; keep its market provenance unresolved until independently verified."
    )
    warnings.append(
        "html-rips and 2017-12-09-grams-christopher appear to be mirrored HTML snapshots and should be deduplicated before downstream parsing."
    )
    return warnings


def main() -> None:
    schema = load_schema()
    mapping_validation = validate_mappings(schema)
    products_audit = audit_products_data()
    market_item_audit, market_item_en_vendors, market_item_zh_vendors = audit_market_item()
    agora_audit, agora_vendors = audit_agora()
    vendor_sql_audit, en_gold_aliases = audit_vendor_sql()
    pgp_audit = audit_pgp_sql()
    strong_audit = audit_suspected_csv("suspected_sockpuppet_strong.csv", "strong_group_distinct_user_count")
    weak_audit = audit_suspected_csv("suspected_sockpuppet_weak.csv", "weak_group_distinct_user_count")
    imposter_audit = audit_imposter_csv()
    overlaps = overlap_report(en_gold_aliases, market_item_en_vendors, market_item_zh_vendors, agora_vendors)
    warnings = build_warnings(overlaps, market_item_audit)

    report = {
        "schema_path": str(SCHEMA_PATH.relative_to(ROOT)),
        "mapping_validation": mapping_validation,
        "source_audit": {
            "products_data.csv": products_audit,
            "market_item.xlsx": market_item_audit,
            "2017-12-05-philipjames11-darknetmarketplacedataagora20142015.xlsx": agora_audit,
            "tijkc3xx.sql": vendor_sql_audit,
            "3z669jwe.sql": pgp_audit,
            "suspected_sockpuppet_strong.csv": strong_audit,
            "suspected_sockpuppet_weak.csv": weak_audit,
            "suspected_imposter_rows.csv": imposter_audit
        },
        "overlap_audit": overlaps,
        "warnings": warnings
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open("w", encoding="utf-8") as handle:
        json.dump(report, handle, ensure_ascii=False, indent=2)

    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
